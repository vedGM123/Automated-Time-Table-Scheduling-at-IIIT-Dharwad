import pytest
import os
import pandas as pd
from openpyxl import load_workbook
from io import StringIO
import builtins

# Import your functions
from generate_timetable1 import has_conflict, get_session_counts, schedule_course, time_to_minutes, LECTURE_SLOTS, TUTORIAL_SLOTS, LAB_SLOTS, days, room_list

# ------------------------- Fixtures -------------------------
@pytest.fixture
def sample_course():
    return {
        'Course Code': 'CS101',
        'Course-Name': 'Intro to CS',
        'Instructor': 'Dr. A',
        'Semester': 1,
        'Branch': 'CSE',
        'L-T-P-S-C': '2-1-0-0-3',
        'ScheduledDays': set()
    }

@pytest.fixture
def sample_schedule():
    return [
        {
            'Course Code': 'CS100',
            'Course-Name': 'Test Course',
            'Instructor': 'Dr. B',
            'Room': 'R1',
            'Day': 'Mon',
            'Start-Time': '09:00',
            'End-Time': '10:30',
            'Semester': 1
        }
    ]

# ------------------------- Unit Tests -------------------------
def test_time_to_minutes_converts_correctly():
    """Test time_to_minutes conversion"""
    assert time_to_minutes("00:00") == 0
    assert time_to_minutes("01:30") == 90
    assert time_to_minutes("23:59") == 1439

def test_has_conflict_returns_false_for_no_conflict(sample_schedule):
    """Test has_conflict with no overlapping courses"""
    new_class = {
        'Course Code': 'CS101',
        'Course-Name': 'Intro to CS',
        'Instructor': 'Dr. A',
        'Room': 'R2',
        'Day': 'Mon',
        'Start-Time': '10:40',
        'End-Time': '12:10',
        'Semester': 1
    }
    assert has_conflict(sample_schedule, new_class) == False

def test_has_conflict_detects_room_conflict(sample_schedule):
    """Test has_conflict detects room overlap"""
    new_class = {
        'Course Code': 'CS101',
        'Course-Name': 'Intro to CS',
        'Instructor': 'Dr. A',
        'Room': 'R1',
        'Day': 'Mon',
        'Start-Time': '10:00',
        'End-Time': '11:00',
        'Semester': 1
    }
    assert has_conflict(sample_schedule, new_class) == True

def test_has_conflict_detects_instructor_conflict(sample_schedule):
    """Test has_conflict detects instructor overlap"""
    new_class = {
        'Course Code': 'CS101',
        'Course-Name': 'Intro to CS',
        'Instructor': 'Dr. B',
        'Room': 'R2',
        'Day': 'Mon',
        'Start-Time': '09:30',
        'End-Time': '10:30',
        'Semester': 1
    }
    assert has_conflict(sample_schedule, new_class) == True

def test_has_conflict_blocks_lunch_break():
    """Test has_conflict blocks lunch slot"""
    new_class = {
        'Course Code': 'CS101',
        'Course-Name': 'Intro to CS',
        'Instructor': 'Dr. A',
        'Room': 'R1',
        'Day': 'Tue',
        'Start-Time': '13:15',
        'End-Time': '14:15',
        'Semester': 1
    }
    assert has_conflict([], new_class) == True

def test_get_session_counts_calculates_correctly(sample_course):
    """Test get_session_counts returns correct lecture/tutorial/lab counts"""
    lectures, tutorials, labs = get_session_counts(sample_course)
    assert lectures == 2
    assert tutorials == 1
    assert labs == 0

def test_schedule_course_adds_to_timetable(sample_course):
    """Test schedule_course adds a new session to the timetable"""
    timetable = []
    scheduled_day = schedule_course(sample_course, LECTURE_SLOTS, " (Lecture)", timetable)
    assert scheduled_day in days
    assert len(timetable) == 1
    assert timetable[0]['Course Code'] == "CS101"

# ------------------------- Integration Test -------------------------
def test_csv_and_excel_generation(tmp_path, monkeypatch):
    """Test timetable CSV and Excel generation end-to-end"""
    csv_content = """Course Code,Course-Name,Instructor,Semester,Branch,L-T-P-S-C,Type
CS101,Intro to CS,Dr. A,1,CSE,2-1-0-0-3,Core
CS102,Data Structures,Dr. B,1,CSE,1-1-0-0-3,Core
"""
    csv_file = tmp_path / "course_data.csv"
    csv_file.write_text(csv_content)

    import main
    monkeypatch.setattr(main, "COURSE_FILE", str(csv_file))
    monkeypatch.setattr(main, "OUTPUT_FLAT", str(tmp_path / "out.csv"))
    monkeypatch.setattr(main, "OUTPUT_EXCEL", str(tmp_path / "out.xlsx"))

    # Minimal timetable scheduling
    main.__dict__['courses'] = pd.read_csv(str(csv_file))
    main.timetable = []
    for _, course in main.courses.iterrows():
        main.schedule_course(course, main.LECTURE_SLOTS, " (Lecture)", main.timetable)

    # Export CSV
    df = pd.DataFrame(main.timetable)
    df.to_csv(main.OUTPUT_FLAT, index=False)
    assert os.path.exists(main.OUTPUT_FLAT)

    # Export Excel
    df['Slot'] = df['Start-Time'] + " - " + df['End-Time']
    df['Display'] = df['Course Code'] + "\n" + df['Course-Name']
    structured_df = df.pivot_table(index='Day', columns='Slot', values='Display', aggfunc=lambda x: "\n---\n".join(x))
    structured_df.to_excel(main.OUTPUT_EXCEL)
    assert os.path.exists(main.OUTPUT_EXCEL)

    # Verify Excel content
    wb = load_workbook(main.OUTPUT_EXCEL)
    ws = wb.active
    found_course = any("CS101" in str(cell.value) for row in ws.iter_rows(values_only=True) for cell in row if cell)
    assert found_course
