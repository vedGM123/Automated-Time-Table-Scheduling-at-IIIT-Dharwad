"""
ExamTimetableGenerator.py - Complete Exam Timetable Generation System
Combines scheduling logic, seating arrangements, and execution in one file

Folder Structure:
    Project Root/
    ├── data/                          (All input files)
    ├── output/Exam_timetable/         (Generated output)
    └── src/                           (This script)
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill, Font, Border, Side
from openpyxl.utils import get_column_letter
import math
from datetime import datetime, timedelta
import copy
from collections import defaultdict
import os


def generate_seating_matrix(room_name, capacity, student_groups):
    """
    Generate STRICT anti-cheating seating arrangement for a room.
    NEVER places two students from the same group on the same bench.
    Uses empty seats as buffer when groups are exhausted.
    
    Args:
        room_name (str): Name of the room
        capacity (int): Total capacity of the room
        student_groups (list): List of dicts with keys 'course_code', 'student_ids'
                               e.g., [{'course_code': 'CSE301', 'student_ids': ['CS001', 'CS002', ...]}, ...]
    
    Returns:
        pd.DataFrame: Seating matrix with rows and columns representing physical layout
    """
    
    # Calculate room dimensions
    # 2 columns (Left and Right), 2 students per bench, so 4 students per row
    num_rows = math.ceil(capacity / 4)
    total_benches = num_rows * 2  # 2 benches per row
    
    # Prepare student queues with metadata
    student_queues = []
    total_students = 0
    
    for group in student_groups:
        student_list = group.get('student_ids', [])
        course_code = group.get('course_code', 'UNKNOWN')
        total_students += len(student_list)
        
        student_queues.append({
            'course': course_code,
            'students': list(student_list),  # Keep as list for easy indexing
            'index': 0,  # Current position in the list
            'remaining': len(student_list)
        })
    
    num_groups = len(student_queues)
    
    # Capacity Warning Check
    if num_groups < 2:
        print(f"⚠️  WARNING: Room {room_name} has only {num_groups} group(s).")
        print(f"    Anti-cheating seating requires at least 2 groups for bench-mate mixing.")
    
    # Theoretical minimum benches needed (1 student per bench worst case)
    if total_students > capacity:
        print(f"⚠️  CRITICAL: Room {room_name} - {total_students} students exceed capacity {capacity}!")
        print(f"    Some students cannot be seated.")
    
    # ========================================
    # STRICT SPACING ALGORITHM
    # ========================================
    benches = []  # Each bench = [Seat1, Seat2]
    
    for bench_num in range(total_benches):
        # Try to find two students from DIFFERENT groups
        seat1 = None
        seat2 = None
        group1_idx = None
        group2_idx = None
        
        # Strategy: Find first available group, then find a different group
        for i in range(num_groups):
            if student_queues[i]['remaining'] > 0:
                group1_idx = i
                student_id = student_queues[i]['students'][student_queues[i]['index']]
                seat1 = f"{student_id}\n({student_queues[i]['course']})"
                break
        
        # Now find a DIFFERENT group for seat 2
        if group1_idx is not None:
            for i in range(num_groups):
                if i != group1_idx and student_queues[i]['remaining'] > 0:
                    group2_idx = i
                    student_id = student_queues[i]['students'][student_queues[i]['index']]
                    seat2 = f"{student_id}\n({student_queues[i]['course']})"
                    break
            
            # If we found a different group, consume both students
            if seat2 is not None:
                student_queues[group1_idx]['index'] += 1
                student_queues[group1_idx]['remaining'] -= 1
                student_queues[group2_idx]['index'] += 1
                student_queues[group2_idx]['remaining'] -= 1
            else:
                # Only one group left - enforce empty seat rule
                student_queues[group1_idx]['index'] += 1
                student_queues[group1_idx]['remaining'] -= 1
                seat2 = "EMPTY\n(Anti-Cheat)"
        else:
            # No students left at all
            seat1 = "EMPTY\n(Anti-Cheat)"
            seat2 = "EMPTY\n(Anti-Cheat)"
        
        benches.append([seat1, seat2])
    
    # ========================================
    # ARRANGE INTO PHYSICAL LAYOUT
    # ========================================
    # Layout: Row -> Left Bench [Seat1, Seat2] | Right Bench [Seat1, Seat2]
    
    seating_matrix = []
    
    for row_idx in range(num_rows):
        left_bench_idx = row_idx * 2
        right_bench_idx = row_idx * 2 + 1
        
        # Get benches (or create empty if beyond range)
        left_bench = benches[left_bench_idx] if left_bench_idx < len(benches) else ["EMPTY\n(Anti-Cheat)", "EMPTY\n(Anti-Cheat)"]
        right_bench = benches[right_bench_idx] if right_bench_idx < len(benches) else ["EMPTY\n(Anti-Cheat)", "EMPTY\n(Anti-Cheat)"]
        
        row_data = {
            'Row': f"Row {row_idx + 1}",
            'Left Bench - Seat A': left_bench[0],
            'Left Bench - Seat B': left_bench[1],
            'Right Bench - Seat A': right_bench[0],
            'Right Bench - Seat B': right_bench[1]
        }
        seating_matrix.append(row_data)
    
    df_seating = pd.DataFrame(seating_matrix)
    
    # ========================================
    # POST-GENERATION WARNINGS
    # ========================================
    # Count students actually seated vs empty seats
    total_cells = num_rows * 4
    empty_count = 0
    seated_count = 0
    
    for row in seating_matrix:
        for col in ['Left Bench - Seat A', 'Left Bench - Seat B', 'Right Bench - Seat A', 'Right Bench - Seat B']:
            if "EMPTY" in row[col]:
                empty_count += 1
            else:
                seated_count += 1
    
    efficiency = (seated_count / total_cells) * 100 if total_cells > 0 else 0
    
    print(f"\n📊 Seating Report for {room_name}:")
    print(f"   Total Students: {total_students}")
    print(f"   Seated: {seated_count}")
    print(f"   Empty Seats (Anti-Cheat): {empty_count}")
    print(f"   Room Efficiency: {efficiency:.1f}%")
    
    if efficiency < 60:
        print(f"   ⚠️  LOW EFFICIENCY WARNING: Only {efficiency:.1f}% of seats used.")
        print(f"   Consider redistributing students to other rooms.")
    
    # Check if anyone was left unseated
    unseated = sum(q['remaining'] for q in student_queues)
    if unseated > 0:
        print(f"   🚨 CRITICAL: {unseated} student(s) could NOT be seated!")
        for q in student_queues:
            if q['remaining'] > 0:
                print(f"      - {q['course']}: {q['remaining']} students unseated")
    
    return df_seating


def generate_timetable(start_date, end_date, branch_slot_allocation, max_credits_per_day=5):
    """
    Generate exam timetable with given parameters including detailed seating arrangements
    
    Args:
        start_date (str): Start date in 'YYYY-MM-DD' format
        end_date (str): End date in 'YYYY-MM-DD' format
        branch_slot_allocation (dict): Branch allocation per year and slot
        max_credits_per_day (int): Maximum credits allowed per day per branch (default: 5)
    
    Returns:
        str: Path to generated Excel file
    """
    
    # Get the directory of the current script
    script_dir = os.path.dirname(os.path.abspath(__file__))
    project_root = os.path.dirname(script_dir)  # Go up one level from src/
    
    # Define paths
    data_dir = os.path.join(project_root, "data")
    output_dir = os.path.join(project_root, "output", "Exam_timetable")
    
    # Create output directory if it doesn't exist
    os.makedirs(output_dir, exist_ok=True)
    
    # Convert string dates to datetime
    start_date = datetime.strptime(start_date, "%Y-%m-%d")
    end_date = datetime.strptime(end_date, "%Y-%m-%d")
    
    # Public holidays in India (example, update as needed)
    public_holidays = [
        "2025-01-26", "2025-03-14", "2025-03-31", "2025-04-06", "2025-04-18",
        "2025-05-12", "2025-07-06", "2025-08-15", "2025-10-02", "2025-10-14", "2025-10-20"
    ]
    public_holidays = [datetime.strptime(d, "%Y-%m-%d") for d in public_holidays]
    
    # Generate list of valid exam dates (exclude Sundays & holidays)
    exam_dates = []
    current_date = start_date
    while current_date <= end_date:
        if current_date.weekday() != 6 and current_date not in public_holidays:
            exam_dates.append(current_date)
        current_date += timedelta(days=1)
    
    # -----------------------------
    # Step 1: Read Inputs
    # -----------------------------
    print(f"📂 Loading data from: {data_dir}")
    
    df_strength = pd.read_excel(os.path.join(data_dir, "BranchStrength.xlsx"))
    df_courses = pd.read_excel(os.path.join(data_dir, "CoursesPerYear.xlsx"))
    df_common = pd.read_excel(os.path.join(data_dir, "CommonCourse.xlsx"))
    df_settings = pd.read_excel(os.path.join(data_dir, "Settings.xlsx"))
    df_faculty = pd.read_csv(os.path.join(data_dir, "FACULTY.csv"))
    
    # ========================================
    # SMART CSV LOADING FOR ROOMS
    # ========================================
    print("🔧 Loading rooms.csv with smart column detection...")
    try:
        df_room = pd.read_csv(os.path.join(data_dir, "rooms.csv"))
    except FileNotFoundError:
        print("⚠️  rooms.csv not found, trying rooms.xlsx as fallback...")
        df_room = pd.read_excel(os.path.join(data_dir, "rooms.xlsx"))
    
    # Normalize column names to lowercase for consistent access
    df_room.columns = df_room.columns.str.strip().str.lower()
    
    # Find the room and capacity columns intelligently
    room_col = None
    capacity_col = None
    
    for col in df_room.columns:
        if 'room' in col:
            room_col = col
        if 'capacity' in col or 'cap' in col:
            capacity_col = col
    
    if not room_col or not capacity_col:
        raise ValueError(f"❌ Could not find 'room' or 'capacity' columns in rooms.csv. Found columns: {df_room.columns.tolist()}")
    
    print(f"✅ Detected columns: Room='{room_col}', Capacity='{capacity_col}'")
    
    # Extract room data
    room_names = df_room[room_col].dropna().astype(str).tolist()
    room_capacities = dict(zip(df_room[room_col].astype(str), df_room[capacity_col].astype(int)))
    
    # ========================================
    # DYNAMIC TOTAL_ROOMS CALCULATION
    # ========================================
    total_rooms = len(room_names)  # Calculate dynamically from CSV
    print(f"📊 Dynamically detected {total_rooms} rooms from rooms.csv")
    
    course_book = pd.read_excel(os.path.join(data_dir, "courselist.xlsx"), sheet_name=None)
    
    # -----------------------------
    # Step 2: Clean Inputs
    # -----------------------------
    df_strength["Year"] = df_strength["Year"].astype(str).str.strip().str.title()
    df_strength["Branch"] = df_strength["Branch"].astype(str).str.strip().str.upper()
    df_courses["Year"] = df_courses["Year"].astype(str).str.strip().str.title()
    faculty_list = df_faculty["Name"].astype(str).tolist()
    
    # -----------------------------
    # Step 2: Map Sheets to Years Automatically
    # -----------------------------
    sheet_year_map = {}
    year_names = ["1St Year", "2Nd Year", "3Rd Year", "4Th Year", "5Th Year"]
    
    for i, sheet_name in enumerate(course_book.keys()):
        if i < len(year_names):
            sheet_year_map[sheet_name] = year_names[i]
        else:
            sheet_year_map[sheet_name] = f"{i+1}Th Year"
    
    # -----------------------------
    # Step 3: Prepare Branch Courses
    # -----------------------------
    branch_courses = {}
    for sheet_name, df in course_book.items():
        year = sheet_year_map[sheet_name]
        branch_courses[year] = {}
        df.columns = [str(c).strip().upper() for c in df.columns]
        for branch in ["CSE", "DSAI", "ECE"]:
            branch_courses[year][branch] = []
            for val in df.get(branch, pd.Series()).dropna().astype(str):
                try:
                    course_code, credits = val.split(",")
                    branch_courses[year][branch].append({
                        "course_code": course_code.strip(),
                        "credits": int(credits.strip())
                    })
                except:
                    print(f"Skipping invalid course entry: {val}")
    
    # -----------------------------
    # Step 3: Extract Settings
    # -----------------------------
    settings = dict(zip(df_settings["SettingName"], df_settings["Value"]))
    credits_per_course = int(settings["CreditsPerCourse"])
    max_students_per_slot = int(settings["MaxStudentsPerSlot"])
    max_courses_per_slot = int(settings["MaxCoursesPerSlot"])
    room_capacity_per_course = int(settings["RoomCapacityPerCourse"])
    
    # -----------------------------
    # Step 4: Python Structures
    # -----------------------------
    branches = df_strength["Branch"].unique().tolist()
    years = df_strength["Year"].unique().tolist()
    
    branch_strength = {
        year: dict(zip(
            df_strength[df_strength["Year"]==year]["Branch"],
            df_strength[df_strength["Year"]==year]["Strength"]
        ))
        for year in years
    }
    
    courses_per_year = dict(zip(df_courses["Year"], df_courses["CoursesPerYear"]))
    
    common_course = {
        "course_code": df_common.loc[0, "CourseCode"],
        "credits": int(df_common.loc[0, "Credits"])
    }
    
    # -----------------------------
    # Step 5: Initialize Schedule
    # -----------------------------
    days = [d.strftime("%Y-%m-%d") for d in exam_dates]
    slots = ["Morning", "Evening"]
    schedule = {day: {slot: {year: [] for year in years} for slot in slots} for day in days}
    day_year_credits = {day: {year: {branch: 0 for branch in branches} for year in years} for day in days}
    day_slot_total_students = {day: {slot:0 for slot in slots} for day in days}
    day_slot_total_courses = {day: {slot:0 for slot in slots} for day in days}
    remaining_courses = {year: {branch: courses_per_year[year] for branch in branches} for year in years}
    
    def normalize_year(year_text):
        text = str(year_text).lower().replace(" ", "")
        if "1st" in text or "first" in text or text == "1":
            return "1St Year"
        elif "2nd" in text or "second" in text or text == "2":
            return "2Nd Year"
        elif "3rd" in text or "third" in text or text == "3":
            return "3Rd Year"
        else:
            return year_text
    
    slot_max_students = total_rooms * 48
    common_course_map = {}
    for _, row in df_common.iterrows():
        year_norm = normalize_year(str(row['Year']))
        branches_cell = row['Branches']
        if pd.isna(branches_cell):
            branches_for_course = []
        else:
            branches_for_course = [b.strip() for b in str(branches_cell).split(",")]
        
        common_course_map[row['CourseCode']] = {
            "credits": row['Credits'],
            "Year": year_norm,
            "Branches": branches_for_course
        }
    
    common_assigned = {code: False for code in common_course_map}
    
    # -----------------------------
    # Step 6b: Assign Common Courses
    # -----------------------------
    for day in days:
        for course_code, info in common_course_map.items():
            year_norm = info['Year']
            branches_to_block = info['Branches']
            if common_assigned[course_code]:
                continue
            
            for slot in slots:
                if year_norm not in schedule[day][slot]:
                    schedule[day][slot][year_norm] = []
                if year_norm not in day_year_credits[day]:
                    day_year_credits[day][year_norm] = {b: 0 for b in branches}
                
                conflict = False
                for branch in branches_to_block:
                    if day_year_credits[day][year_norm].get(branch, 0) + info['credits'] > max_credits_per_day:
                        conflict = True
                        break
                    if any(c.get('branch') == branch for c in schedule[day][slot][year_norm]):
                        conflict = True
                        break
                if conflict:
                    continue
                
                total_students = sum(branch_strength[year_norm][b] for b in branches_to_block)
                if day_slot_total_students[day][slot] + total_students > slot_max_students:
                    continue
                
                for branch in branches_to_block:
                    schedule[day][slot][year_norm].append({
                        "course_code": course_code,
                        "credits": info['credits'],
                        "branch": branch,
                        "year": year_norm,
                        "students": branch_strength[year_norm][branch],
                        "type": "Common"
                    })
                    day_year_credits[day][year_norm][branch] += info['credits']
                    day_slot_total_students[day][slot] += branch_strength[year_norm][branch]
                
                day_slot_total_courses[day][slot] += 1
                common_assigned[course_code] = True
                break
    
    # -----------------------------
    # Step 7: Assign Main Courses
    # -----------------------------
    default_branch_slot_allocation = branch_slot_allocation
    
    # Validate against slot capacity
    for year in default_branch_slot_allocation:
        normalized_year = normalize_year(year)
        for slot in default_branch_slot_allocation[year]:
            branches_in_slot = default_branch_slot_allocation[year][slot]
            if normalized_year not in branch_strength:
                continue
            total_strength = sum(branch_strength[normalized_year][b] for b in branches_in_slot if b in branch_strength[normalized_year])
            if total_strength > slot_max_students:
                print(f"⚠️ Warning: {normalized_year} {slot} total ({total_strength}) exceeds slot capacity ({slot_max_students}).")
    
    branch_slot_allocation_day = {day: copy.deepcopy(default_branch_slot_allocation) for day in days}
    
    for day in days:
        for slot in slots:
            for year in years:
                normalized_year = normalize_year(year)
                allowed_branches = branch_slot_allocation_day[day][normalized_year][slot]
                
                running_total = day_slot_total_students[day][slot]
                
                final_branches = []
                for b in allowed_branches:
                    b_strength = branch_strength[normalized_year].get(b, 0)
                    if running_total + b_strength <= slot_max_students:
                        final_branches.append(b)
                        running_total += b_strength
                    else:
                        print(f"Skipping branch {b} for {normalized_year} {slot} on {day} – would exceed slot capacity")
                
                branch_slot_allocation_day[day][normalized_year][slot] = final_branches
    
    def total_remaining():
        return sum(remaining_courses[y][b] for y in years for b in branches)
    
    # Find ENV day
    env_day = None
    env_code = common_course["course_code"]
    for d in days:
        if any(
            c.get("course_code") == env_code
            for s in slots
            for y in years
            for c in schedule[d][s][y]
        ):
            env_day = d
            break
    
    after_env_days = days[days.index(env_day)+1:] if env_day in days else days[:]
    
    for day in days:
        day_allocation = branch_slot_allocation_day.get(day, default_branch_slot_allocation)
        
        for slot in slots:
            blocked_branches = set()
            for year in years:
                for c in schedule[day][slot][year]:
                    if c.get("type") == "Common":
                        if isinstance(c["branch"], str):
                            for b in c["branch"].split(","):
                                blocked_branches.add(b.strip())
            
            for year in years:
                normalized_year = normalize_year(year)
                
                if any(c.get("type") == "Common" for c in schedule[day][slot][year]):
                    continue
                
                allowed_branches = day_allocation.get(normalized_year, {}).get(slot, [])
                
                candidates = [
                    b for b in allowed_branches
                    if remaining_courses[year].get(b, 0) > 0
                    and b not in blocked_branches
                ]
                
                candidates.sort(key=lambda b: branch_strength[year][b], reverse=True)
                
                for b in candidates:
                    if day_year_credits[day][year][b] + credits_per_course > max_credits_per_day:
                        continue
                    if day_slot_total_students[day][slot] + branch_strength[year][b] > slot_max_students:
                        continue
                    
                    course_index = courses_per_year[year] - remaining_courses[year][b]
                    year_key = normalize_year(year)
                    if course_index < len(branch_courses.get(year_key, {}).get(b, [])):
                        course_info = branch_courses[year_key][b][course_index]
                    else:
                        course_info = {"course_code": f"{b}{year[0]}X", "credits": credits_per_course}
                    
                    schedule[day][slot][year].append({
                        "course_code": course_info["course_code"],
                        "credits": course_info["credits"],
                        "branch": b,
                        "year": year,
                        "students": branch_strength[year][b],
                        "type": "Main"
                    })
                    
                    remaining_courses[year][b] -= 1
                    day_year_credits[day][year][b] += course_info["credits"]
                    day_slot_total_students[day][slot] += branch_strength[year][b]
                    day_slot_total_courses[day][slot] += 1
            
            # Fill empty slots after ENV day
            if day in after_env_days and total_remaining() > 0:
                slot_has_main = any(
                    len([c for c in schedule[day][slot][y] if c.get("type") != "Common"]) > 0
                    for y in years
                )
                if not slot_has_main:
                    placed = False
                    year_order = sorted(
                        years, key=lambda Y: sum(remaining_courses[Y][b] for b in branches), reverse=True
                    )
                    for year in year_order:
                        normalized_year = normalize_year(year)
                        if any(c.get("type") == "Common" for c in schedule[day][slot][year]):
                            continue
                        allowed_branches = day_allocation.get(normalized_year, {}).get(slot, [])
                        for b in allowed_branches:
                            if remaining_courses[year].get(b, 0) <= 0:
                                continue
                            if day_year_credits[day][year][b] + credits_per_course > max_credits_per_day:
                                continue
                            if day_slot_total_students[day][slot] + branch_strength[year][b] > slot_max_students:
                                continue
                            
                            course_index = courses_per_year[year] - remaining_courses[year][b]
                            if course_index < len(branch_courses.get(normalized_year, {}).get(b, [])):
                                course_info = branch_courses[normalized_year][b][course_index]
                            else:
                                course_info = {"course_code": f"{b}{year[0]}X", "credits": credits_per_course}
                            
                            schedule[day][slot][year].append({
                                "course_code": course_info["course_code"],
                                "credits": course_info["credits"],
                                "branch": b,
                                "year": year,
                                "students": branch_strength[year][b],
                                "type": "Main"
                            })
                            
                            remaining_courses[year][b] -= 1
                            day_year_credits[day][year][b] += course_info["credits"]
                            day_slot_total_students[day][slot] += branch_strength[year][b]
                            day_slot_total_courses[day][slot] += 1
                            placed = True
                            break
                        if placed:
                            break
    
    # -----------------------------
    # Step 8: Prepare Exam Schedule DataFrame
    # -----------------------------
    rows = []
    for day in days:
        row = {"Day": day}
        for slot in slots:
            for year in years:
                course_list = [f"{c['course_code']} ({c['students']} students)" for c in schedule[day][slot][year]]
                row[f"{slot} - {year}"] = ", ".join(course_list) if course_list else "Empty"
        for year in years:
            for branch in branches:
                row[f"Credits {year}-{branch}"] = day_year_credits[day][year][branch]
        for slot in slots:
            row[f"Total Students - {slot}"] = day_slot_total_students[day][slot]
        rows.append(row)
    
    df_schedule = pd.DataFrame(rows)
    
    # -----------------------------
    # Step 9 & 10: Room Allocation, Faculty Assignment, and SEATING
    # -----------------------------
    student_book = pd.read_excel(os.path.join(data_dir, "students.xlsx"), sheet_name=None)
    student_ids = {}
    
    for sheet_name, df in student_book.items():
        year_key = normalize_year(sheet_name.replace(" Year", ""))
        for branch in ["CSE", "DSAI", "ECE"]:
            if branch in df.columns:
                ids = df[branch].dropna().astype(str).tolist()
                student_ids[(year_key, branch)] = ids
    
    faculty_count = len(faculty_list)
    faculty_index = 0
    df_rooms_rows = []
    
    # Dictionary to store seating arrangements for each room/day/slot
    seating_arrangements = {}
    
    for day in days:
        used_faculty_today = set()
        for slot in slots:
            assigned_faculty = []
            for _ in range(total_rooms):
                # Assign unique faculty or reuse if exhausted
                attempts = 0
                while faculty_list[faculty_index % faculty_count] in used_faculty_today and attempts < faculty_count:
                    faculty_index += 1
                    attempts += 1
                
                assigned_faculty.append(faculty_list[faculty_index % faculty_count])
                used_faculty_today.add(faculty_list[faculty_index % faculty_count])
                faculty_index += 1
            
            room_faculty_mapping = {room_names[i]: assigned_faculty[i] for i in range(total_rooms)}
            
            room_courses = {room: [] for room in room_names}
            room_students = {room: [] for room in room_names}
            room_student_groups = {room: [] for room in room_names}  # Track groups for seating
            
            courses_in_slot = []
            for year in years:
                courses_in_slot.extend(schedule[day][slot][year])
            
            for course in courses_in_slot:
                branch = course["branch"]
                year_label = normalize_year(course["year"])
                ids = student_ids.get((year_label, branch), [])
                batch_size = room_capacity_per_course
                
                if not ids:
                    student_ranges = [f"{course['students']} students"]
                    student_id_ranges = [[f"Student_{i}" for i in range(course['students'])]]
                else:
                    student_ranges = []
                    student_id_ranges = []
                    for i in range(0, len(ids), batch_size):
                        batch = ids[i:min(i + batch_size, len(ids))]
                        student_ranges.append(f"{batch[0]}–{batch[-1]}")
                        student_id_ranges.append(batch)
                
                for idx, sr in enumerate(student_ranges):
                    target_room = None
                    for r in room_names:
                        if len(room_courses[r]) < 2 and course["course_code"] not in room_courses[r]:
                            target_room = r
                            break
                    if not target_room:
                        continue
                    
                    room_courses[target_room].append(course["course_code"])
                    room_students[target_room].append(sr)
                    
                    # Store student group info for seating generation
                    room_student_groups[target_room].append({
                        'course_code': course["course_code"],
                        'student_ids': student_id_ranges[idx]
                    })
            
            for r in room_names:
                for i, c in enumerate(room_courses[r]):
                    df_rooms_rows.append({
                        "Day": day,
                        "Slot": slot,
                        "Course": c,
                        "Branch": "".join([x for x in c if not x.isdigit()])[:4],
                        "Students": room_students[r][i],
                        "Rooms Assigned": r,
                        "Faculty": room_faculty_mapping[r]
                    })
                
                # ========================================
                # GENERATE STRICT ANTI-CHEAT SEATING
                # ========================================
                if room_student_groups[r]:
                    room_capacity = room_capacities.get(r, 48)
                    seating_df = generate_seating_matrix(r, room_capacity, room_student_groups[r])
                    seating_key = f"{day}_{slot}_{r}"
                    seating_arrangements[seating_key] = {
                        'df': seating_df,
                        'day': day,
                        'slot': slot,
                        'room': r,
                        'faculty': room_faculty_mapping[r]
                    }
    
    df_rooms = pd.DataFrame(df_rooms_rows)
    
    # Merge courses in same room
    merged_rows = []
    room_dict = defaultdict(list)
    
    for row in df_rooms_rows:
        key = (row["Day"], row["Slot"], row["Rooms Assigned"], row["Faculty"])
        course_info = f"{row['Course']} ({row['Students']})"
        room_dict[key].append(course_info)
    
    for key, course_list in room_dict.items():
        day, slot, room, faculty = key
        merged_rows.append({
            "Day": day,
            "Slot": slot,
            "Rooms Assigned": room,
            "Faculty": faculty,
            "Courses + Students": ", ".join(course_list)
        })
    
    df_rooms_merged = pd.DataFrame(merged_rows)
    
    # -----------------------------
    # Step 11: Save to Excel
    # -----------------------------
    output_file = os.path.join(output_dir, "exam_schedule_with_rooms_faculty.xlsx")
    print(f"💾 Saving output to: {output_file}")
    
    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        df_schedule.to_excel(writer, sheet_name="Exam Schedule", index=False)
        for day in days:
            df_day = df_rooms[df_rooms["Day"]==day]
            df_day.to_excel(writer, sheet_name=f"Rooms-{day}", index=False)
        
        # Add seating arrangement sheets
        for seating_key, seating_info in seating_arrangements.items():
            sheet_name = f"Seating_{seating_info['room']}_{seating_info['day']}_{seating_info['slot']}"[:31]
            seating_info['df'].to_excel(writer, sheet_name=sheet_name, index=False)
    
    # -----------------------------
    # Step 12: Excel Formatting
    # -----------------------------
    wb = load_workbook(output_file)
    
    # Format main schedule
    ws_main = wb["Exam Schedule"]
    
    for i, col in enumerate(ws_main.columns, start=1):
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws_main.column_dimensions[get_column_letter(i)].width = max_length + 5
    
    for row in ws_main.iter_rows(min_row=2, max_row=ws_main.max_row):
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws_main.row_dimensions[cell.row].height = 30
    
    branch_colors = {
        "CSE": "FFC7CE",
        "DSAI": "C6EFCE",
        "ECE": "FFEB9C",
        "All": "BDD7EE"
    }
    
    for row in ws_main.iter_rows(min_row=2, max_row=ws_main.max_row):
        for cell in row:
            for branch, color in branch_colors.items():
                if branch in str(cell.value):
                    cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                    cell.font = Font(bold=True)
    
    # Format seating arrangement sheets
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for sheet_name in wb.sheetnames:
        if sheet_name.startswith("Seating_"):
            ws = wb[sheet_name]
            
            # Set column widths
            ws.column_dimensions['A'].width = 12
            for col in ['B', 'C', 'D', 'E']:
                ws.column_dimensions[col].width = 25
            
            # Format cells
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
                for cell in row:
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    cell.border = thin_border
                    
                    # Color code by course
                    cell_value = str(cell.value)
                    if "CSE" in cell_value:
                        cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                        cell.font = Font(size=9, bold=True)
                    elif "DSAI" in cell_value:
                        cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                        cell.font = Font(size=9, bold=True)
                    elif "ECE" in cell_value:
                        cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                        cell.font = Font(size=9, bold=True)
                    elif "EMPTY" in cell_value:
                        cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
                        cell.font = Font(size=9, italic=True, color="808080")
                    
                    # Make header row bold
                    if cell.row == 1:
                        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                        cell.font = Font(bold=True, size=11, color="FFFFFF")
                
                # Set row heights
                ws.row_dimensions[row[0].row].height = 40
    
    wb.save(output_file)
    
    # -----------------------------
    # Step 13: Create Free Slot Sheet
    # -----------------------------
    free_rows = []
    
    for day in days:
        for slot in slots:
            rooms_filled = df_rooms[(df_rooms["Day"]==day) & (df_rooms["Slot"]==slot)]["Rooms Assigned"].unique().tolist()
            available_rooms = [r for r in room_names if r not in rooms_filled]
            
            total_students_in_slot = day_slot_total_students[day][slot]
            remaining_capacity = max_students_per_slot - total_students_in_slot
            
            for year in years:
                for branch in branches:
                    courses_assigned = schedule[day][slot][year]
                    assigned_branches = [c["branch"] for c in courses_assigned]
                    has_common_course = any(c["branch"] == "All" for c in courses_assigned)
                    is_free = branch not in assigned_branches and not has_common_course
                    
                    free_rows.append({
                        "Day": day,
                        "Slot": slot,
                        "Year": year,
                        "Branch": branch,
                        "Status": "Free" if is_free else "Engaged",
                        "Available Rooms": ", ".join(available_rooms) if year==years[0] and branch==branches[0] else "",
                        "Remaining Capacity": remaining_capacity if year==years[0] and branch==branches[0] else ""
                    })
    
    df_free_slots = pd.DataFrame(free_rows)
    
    with pd.ExcelWriter(output_file, engine="openpyxl", mode="a") as writer:
        df_free_slots.to_excel(writer, sheet_name="Free Slots", index=False)
    
    # Final formatting for all sheets
    wb = load_workbook(output_file)
    branch_colors = {
        "CSE": "FFC7CE",
        "DSAI": "C6EFCE",
        "ECE": "FFEB9C",
        "ALL": "BDD7EE"
    }
    
    sheet_names = wb.sheetnames
    
    for sheet_name in sheet_names:
        if sheet_name.startswith("Seating_"):
            continue  # Already formatted
            
        ws = wb[sheet_name]
        
        for i, col in enumerate(ws.columns, start=1):
            max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
            ws.column_dimensions[get_column_letter(i)].width = max_length + 5
        
        for row_cells in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row_cells:
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                ws.row_dimensions[cell.row].height = 30
        
        for row_cells in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row_cells:
                cell_value_upper = str(cell.value).upper()
                for branch, color in branch_colors.items():
                    if branch in cell_value_upper:
                        cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                        cell.font = Font(bold=True)
    
    wb.save(output_file)
    print(f"\n✅ Exam timetable with STRICT anti-cheating seating generated successfully!")
    print(f"📄 Output file: {output_file}")
    print(f"📊 Generated {len(seating_arrangements)} seating arrangement sheets")
    print(f"🏫 Total rooms used: {total_rooms}")
    
    return output_file


# =============================================================================
# MAIN EXECUTION SECTION
# =============================================================================

if __name__ == "__main__":
    print("=" * 80)
    print("EXAM TIMETABLE GENERATOR - Complete System")
    print("=" * 80)
    print("\n🔧 FEATURES:")
    print("  ✅ Smart CSV/Excel loading with automatic column detection")
    print("  ✅ Dynamic room calculation from input files")
    print("  ✅ STRICT anti-cheating seating (never seats same-group students together)")
    print("  ✅ Comprehensive exam scheduling with conflict resolution")
    print("  ✅ Faculty assignment and room allocation")
    print("  ✅ Beautiful Excel output with color coding")
    
    # Display folder structure
    script_dir = os.path.dirname(os.path.abspath(__file__))
    project_root = os.path.dirname(script_dir)
    data_dir = os.path.join(project_root, "data")
    output_dir = os.path.join(project_root, "output", "Exam_timetable")
    
    print(f"\n📁 FOLDER STRUCTURE:")
    print(f"  📂 Project Root: {project_root}")
    print(f"  📂 Data Files: {data_dir}")
    print(f"  📂 Output: {output_dir}")
    print("\n" + "=" * 80)
    
    # --- DEFINE YOUR INPUTS HERE ---
    
    # 1. Set your start and end dates
    start_date = '2025-11-20'
    end_date = '2025-12-10'
    
    # 2. Define how you want branches allocated to slots
    branch_slot_allocation = {
        "1St Year": {"Morning": ["CSE", "DSAI", "ECE"], "Evening": []},
        "2Nd Year": {"Morning": ["DSAI"], "Evening": ["ECE", "CSE"]},
        "3Rd Year": {"Morning": ["CSE", "DSAI"], "Evening": ["ECE"]},
        "4Th Year": {"Morning": ["CSE"], "Evening": ["DSAI", "ECE"]}
    }
    
    # --- RUN THE GENERATOR ---
    print("\n🚀 Starting timetable generation...")
    print(f"📅 Period: {start_date} to {end_date}")
    print(f"🎓 Branch allocation configured for 4 years\n")
    
    try:
        output_file = generate_timetable(
            start_date=start_date,
            end_date=end_date,
            branch_slot_allocation=branch_slot_allocation,
            max_credits_per_day=5 
        )
        print(f"\n{'=' * 80}")
        print(f"✅ SUCCESS! Output file created: {output_file}")
        print(f"{'=' * 80}")
        
    except ImportError as e:
        print(f"❌ ERROR: Missing required Python library: {e}")
        print("   Install with: pip install pandas openpyxl")
        
    except FileNotFoundError as e:
        print(f"❌ ERROR: Input file not found")
        print(f"   Missing file: {e.filename}")
        print("\n   Make sure all required files are in the 'data' folder:")
        print(f"   {data_dir}")
        print("\n   Required files:")
        print("   - BranchStrength.xlsx")
        print("   - CoursesPerYear.xlsx")
        print("   - CommonCourse.xlsx")
        print("   - Settings.xlsx")
        print("   - FACULTY.csv")
        print("   - rooms.csv (or rooms.xlsx)")
        print("   - courselist.xlsx")
        print("   - students.xlsx")
        
    except Exception as e:
        print(f"❌ An unexpected error occurred: {e}")
        import traceback
        print("\nFull error details:")
        traceback.print_exc()