# TT_gen.py -- Complete Fixed Timetable Generator
# Fixes applied:
# 1. Removed 10:30-10:45 break
# 2. Changed lunch break to 13:00-14:00
# 3. Better time slot generation
# 4. Fixed elective basket scheduling
# 5. Supports increased lab capacity (80 students)
# 6. *** NEW: Added room capacity check against student strength ***

import pandas as pd
import random
from datetime import datetime, time, timedelta
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter
from dataclasses import dataclass
from pathlib import Path
import traceback
import os
import json

# ---------------------------
# Configuration
# ---------------------------
BASE_DIR = Path(__file__).resolve().parent
PROJECT_DIR = BASE_DIR.parent
DATA_DIR = BASE_DIR.parent / "data"
CONFIG_PATH = DATA_DIR / "config.json"

try:
    with open(CONFIG_PATH, "r", encoding="utf-8") as f:
        config = json.load(f)
except Exception:
    config = {}

DAYS = config.get("days", ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday'])
LECTURE_MIN = config.get("LECTURE_MIN", 90)
LAB_MIN = config.get("LAB_MIN", 120)
TUTORIAL_MIN = config.get("TUTORIAL_MIN", 60)
SELF_STUDY_MIN = config.get("SELF_STUDY_MIN", 60)
MIN_GAP_BETWEEN_LECTURES = 10

# Updated break windows - REMOVED morning break, extended lunch
LUNCH_BREAK_START = time(13, 0)
LUNCH_BREAK_END = time(14, 0)

@dataclass
class UnscheduledComponent:
    department: str
    semester: int
    code: str
    name: str
    faculty: str
    component_type: str
    sessions: int
    section: int
    reason: str

INPUT_DIR = PROJECT_DIR / "data"
OUTPUT_DIR = PROJECT_DIR / "output"

os.makedirs(OUTPUT_DIR, exist_ok=True)

# Meta entries collected while writing timetables so teacher generator can
# pick up faculty/room info even when cells hide them for basket display.
META_ENTRIES = []

# ---------------------------
# Load CSVs
# ---------------------------
try:
    df = pd.read_csv(os.path.join(INPUT_DIR, 'combined.csv'))
    print(f"✅ Loaded {len(df)} courses from combined.csv")
except FileNotFoundError:
    raise SystemExit("Error: 'combined.csv' not found in data directory.")

# --- FIX: Ensure 'total_students' column exists ---
if "total_students" not in df.columns:
    print("⚠️ Warning: 'total_students' column not in combined.csv. Defaulting to 50.")
    df["total_students"] = 50
df["total_students"] = pd.to_numeric(df["total_students"], errors='coerce').fillna(50).astype(int)
print(f"✅ Processed student strengths (defaulting to 50)")
# --- END FIX ---

try:
    rooms_df = pd.read_csv(os.path.join(INPUT_DIR, 'rooms.csv'))
    print(f"✅ Loaded {len(rooms_df)} rooms from rooms.csv")
except FileNotFoundError:
    print("⚠️ Warning: rooms.csv not found. Using empty rooms list.")
    rooms_df = pd.DataFrame(columns=['roomNumber', 'type', 'capacity'])

# --- FIX: Process and store room capacities ---
if 'capacity' not in rooms_df.columns:
    print("⚠️ Warning: 'capacity' column not in rooms.csv. Defaulting all rooms to 50.")
    rooms_df['capacity'] = 50
if 'type' not in rooms_df.columns:
    rooms_df['type'] = 'LECTURE_ROOM' # Default type

rooms_df['capacity'] = pd.to_numeric(rooms_df['capacity'], errors='coerce').fillna(50).astype(int)
rooms_df['type'] = rooms_df['type'].str.upper()
rooms_df['roomNumber'] = rooms_df['roomNumber'].astype(str)

# Create a global lookup dictionary
ROOM_DATA = {}
for _, row in rooms_df.iterrows():
    ROOM_DATA[row['roomNumber']] = {
        'type': row['type'],
        'capacity': row['capacity']
    }

print(f"\n📊 Room inventory: {len(ROOM_DATA)} rooms loaded with capacity data.")
room_types_summary = rooms_df.groupby('type').size().to_dict()
for rtype, count in room_types_summary.items():
    print(f"   - {rtype}: {count}")
# --- END FIX ---


# ---------------------------
# Time Slot Generation - UPDATED
# ---------------------------
def generate_time_slots():
    """
    Generate continuous 60 and 30 minute time slots for flexibility.
    - NO morning break (removed 10:30-10:45)
    - Lunch break: 13:00-14:00
    """
    slots = []
    
    slots = []
    
    slots.append((time(7, 30), time(9, 00)))    # 60 min
    # slots.append((time(9, 30), time(9, 30)))    # 60 min
    slots.append((time(9, 30), time(10, 30)))  # 60 min
    slots.append((time(10, 30), time(11, 30))) # 60 min
    slots.append((time(11, 30), time(12, 30))) # 60 min
    slots.append((time(12, 30), time(13, 0)))    # 30 min
    
    # Lunch break: 13:00 - 14:00
    slots.append((time(13, 0), time(14, 0)))    # BREAK
    
    # Afternoon session: 14:00 - 18:30 (continuous)
    slots.append((time(14, 0), time(15, 0)))    # 60 min
    slots.append((time(15, 0), time(16, 0)))    # 60 min
    slots.append((time(16, 0), time(17, 0)))    # 60 min
    slots.append((time(17, 0), time(18, 0)))    # 60 min
    slots.append((time(18, 0), time(18, 30)))  # 30 min
    slots.append((time(18, 30), time(20, 00)))  # 30 min

    
    return slots

TIME_SLOTS = generate_time_slots()
print(f"⏰ Generated {len(TIME_SLOTS)} time slots (no morning break, lunch 13:00-14:00)")

# ---------------------------
# Helper functions
# ---------------------------
def slot_minutes(slot):
    s, e = slot
    s_m = s.hour * 60 + s.minute
    e_m = e.hour * 60 + e.minute
    if e_m < s_m:
        e_m += 24 * 60
    return e_m - s_m

def overlaps(a_start, a_end, b_start, b_end):
    a_s_min = a_start.hour * 60 + a_start.minute
    a_e_min = a_end.hour * 60 + a_end.minute
    b_s_min = b_start.hour * 60 + b_start.minute
    b_e_min = b_end.hour * 60 + b_end.minute
    return (a_s_min < b_e_min) and (b_s_min < a_e_min)

def is_break_time_slot(slot, semester=None):
    """Updated to only check lunch break (13:00-14:00)"""
    start, end = slot
    # Only lunch break now
    if start == LUNCH_BREAK_START and end == LUNCH_BREAK_END:
        return True
    if overlaps(start, end, LUNCH_BREAK_START, LUNCH_BREAK_END):
        return True
    return False

def is_minor_slot(slot):
    """Check if slot is early morning or late evening"""
    start, end = slot
    if start.hour < 8:
        return True
    
    # Updated check for 18:30 (6:30 PM) or later
    if start.hour > 18 or (start.hour == 18 and start.minute >= 30):
        return True
    
    return False

def select_faculty(faculty_field):
    if pd.isna(faculty_field) or str(faculty_field).strip().lower() in ['nan', 'none', '']:
        return "TBD"
    s = str(faculty_field).strip()
    for sep in ['/', ',', '&', ';']:
        if sep in s:
            return s.split(sep)[0].strip()
    return s

def get_course_priority(row):
    try:
        l = int(row.get('L', 0)) if pd.notna(row.get('L', 0)) else 0
        t = int(row.get('T', 0)) if pd.notna(row.get('T', 0)) else 0
        p = int(row.get('P', 0)) if pd.notna(row.get('P', 0)) else 0
        return -(l + t + p)
    except Exception:
        return 0

def calculate_required_minutes(course_row):
    l = int(course_row['L']) if ('L' in course_row and pd.notna(course_row['L'])) else 0
    t = int(course_row['T']) if ('T' in course_row and pd.notna(course_row['T'])) else 0
    p = int(course_row['P']) if ('P' in course_row and pd.notna(course_row['P'])) else 0
    
    # *** MODIFIED FIX ***
    # Removed the "is_half_semester" logic based on 'C' column.
    # We will trust the L/T/P values in the CSV as the correct
    # weekly hours, regardless of the 'C' column.
    # The B1/B2 logic (before/after midsem) is handled
    # by the global scheduler.
    
    lec_count = l
    tut_count = t
    lab_count = p
    
    return (lec_count, tut_count, lab_count, 0)

def get_required_room_type(component_type):
    if component_type == 'LAB':
        return 'COMPUTER_LAB'
    return 'LECTURE_ROOM'

# ---------------------------
# Elective basket helpers
# ---------------------------
def extract_elective_basket(course_code):
    if pd.isna(course_code):
        return None
    code = str(course_code).strip().upper()
    import re
    match = re.match(r'^(B\d+)-', code)
    return match.group(1) if match else None

def get_base_course_code(course_code):
    if pd.isna(course_code):
        return str(course_code)
    code = str(course_code).strip()
    # This now correctly handles codes like "B2-MA161-C004"
    # It will return "MA161-C004"
    if '-' in code and code.split('-')[0].upper().startswith('B'):
        return code.split('-', 1)[1]
    return code

def is_elective(course_row):
    code = str(course_row.get('Course Code', '')).strip()
    return extract_elective_basket(code) is not None

def has_component_on_day(timetable, day, course_code, component_type):
    # Check all slots for the given day
    for slot_idx in range(len(TIME_SLOTS)):
        slot_data = timetable[day][slot_idx]
        
        # Get the base code from the slot (e.g., "MA161-C004" from "B2\nMA161-C004")
        slot_code = slot_data['code']
        if '\n' in slot_code:
            slot_code = slot_code.split('\n')[-1]
            
        if slot_code == course_code and slot_data['type'] == component_type:
            return True
    return False

# ---------------------------
# Room allocation
# ---------------------------
def find_suitable_room_for_slot(course_code, room_type, day, slot_indices, room_schedule,
                                  course_room_mapping, component_type, student_strength):
    """
    Find suitable room with proper type checking AND capacity checking.
    Tries to find the smallest room that fits the student strength.
    """
    mapping_key = f"{course_code}_{component_type}"
    
    # *** NEW: Check for forced room in course code (e.g., "-C004") ***
    if "-C004" in course_code.upper():
        forced_room = "C004"
        
        # Check if C004 is available
        if forced_room not in room_schedule:
            room_schedule[forced_room] = {d: set() for d in range(len(DAYS))}
        
        # If this is the first time booking this course, check if C004 is free
        if mapping_key not in course_room_mapping:
            if all(si not in room_schedule[forced_room][day] for si in slot_indices):
                course_room_mapping[mapping_key] = forced_room # Book it
                return forced_room
            else:
                # C004 is busy at this time, this schedule attempt fails
                return None
        else:
            # This course is already mapped to C004, just return it
            return course_room_mapping[mapping_key]
    # *** END NEW ***
    
    # This is the logic for all *other* courses (including shared electives)
    if mapping_key in course_room_mapping:
        fixed_room = course_room_mapping[mapping_key]
        # We found a pre-assigned room for this course.
        # Return it immediately. This allows multiple departments
        # to be mapped to the same room for the same course.
        return fixed_room
    
    # If no room is mapped, find a new one...
    
    # 1. Get all room names and shuffle them
    all_room_names = list(ROOM_DATA.keys())
    random.shuffle(all_room_names)
    
    # 2. Find the smallest suitable room
    best_room = None
    min_suitable_capacity = float('inf')

    for room_name in all_room_names:
        if room_name not in ROOM_DATA:
            continue # Should not happen, but safe check
            
        room = ROOM_DATA[room_name]
        
        # 3. Check suitability
        is_type_ok = False
        if room_type == 'COMPUTER_LAB':
            # Allow COMPUTER_LAB or HARDWARE_LAB
            is_type_ok = room['type'] in ['COMPUTER_LAB', 'HARDWARE_LAB']
        else:
            # Allow LECTURE_ROOM or SEATER_120 (or other large rooms)
            is_type_ok = room['type'] in ['LECTURE_ROOM', 'SEATER_120']
        
        # --- THIS IS THE KEY FIX ---
        is_capacity_ok = (room['capacity'] >= student_strength)
        
        if not (is_type_ok and is_capacity_ok):
            continue # This room is not suitable (wrong type or too small)

        # 4. Check availability
        if room_name not in room_schedule:
            room_schedule[room_name] = {d: set() for d in range(len(DAYS))}
            
        is_available = all(si not in room_schedule[room_name][day] for si in slot_indices)
        
        if is_available:
            # This room works. Is it the *best* (smallest) one so far?
            if room['capacity'] < min_suitable_capacity:
                best_room = room_name
                min_suitable_capacity = room['capacity']
    
    # 5. If we found a suitable room, map it and return it
    if best_room:
        course_room_mapping[mapping_key] = best_room
        return best_room

    # Special preference: if the course has more than 120 students and room C004
    # exists (240 seater), prefer assigning C004 for non-lab components and do
    # NOT fall back to combining smaller rooms. This keeps large courses in
    # the single large hall instead of splitting across multiple rooms.
    try:
        if student_strength > 120 and 'C004' in ROOM_DATA and room_type != 'COMPUTER_LAB':
            c004 = ROOM_DATA['C004']
            if c004['capacity'] >= student_strength:
                # Ensure room schedule entry exists
                if 'C004' not in room_schedule:
                    room_schedule['C004'] = {d: set() for d in range(len(DAYS))}
                if all(si not in room_schedule['C004'][day] for si in slot_indices):
                    course_room_mapping[mapping_key] = 'C004'
                    for si in slot_indices:
                        room_schedule['C004'][day].add(si)
                    print(f"    ✅ Assigned C004 for large course {course_code} (needs {student_strength})")
                    return 'C004'
    except Exception:
        # If anything goes wrong with the C004 attempt, fall through to normal logic
        pass

    # If this is a LAB component, try to find two lab rooms whose combined
    # capacity meets the student strength (preferred over leaving unscheduled).
    if room_type == 'COMPUTER_LAB':
        lab_room_names = [rn for rn, info in ROOM_DATA.items() if info['type'] in ('COMPUTER_LAB', 'HARDWARE_LAB')]
        best_pair = None
        best_pair_cap = float('inf')

        # Ensure schedule entries exist
        for rn in lab_room_names:
            if rn not in room_schedule:
                room_schedule[rn] = {d: set() for d in range(len(DAYS))}

        for i in range(len(lab_room_names)):
            for j in range(i+1, len(lab_room_names)):
                r1 = lab_room_names[i]
                r2 = lab_room_names[j]
                # both must be available for all slot indices
                avail1 = all(si not in room_schedule[r1][day] for si in slot_indices)
                avail2 = all(si not in room_schedule[r2][day] for si in slot_indices)
                if not (avail1 and avail2):
                    continue
                cap_sum = ROOM_DATA[r1]['capacity'] + ROOM_DATA[r2]['capacity']
                if cap_sum >= student_strength and cap_sum < best_pair_cap:
                    best_pair = (r1, r2)
                    best_pair_cap = cap_sum

        if best_pair:
            r1, r2 = best_pair
            combined_name = f"{r1}+{r2}"
            course_room_mapping[mapping_key] = combined_name
            for si in slot_indices:
                room_schedule[r1][day].add(si)
                room_schedule[r2][day].add(si)
            print(f"    ✅ Assigned combined labs {combined_name} for {course_code} (combined capacity {best_pair_cap}, needs {student_strength})")
            return combined_name

    # No single free and suitable room found
    return None

def find_consecutive_slots_for_minutes(timetable, day, start_idx, required_minutes,
                                       semester, professor_schedule, faculty,
                                       room_schedule, room_type, course_code, 
                                       course_room_mapping, component_type,
                                       course_day_components, student_strength):
    """Find consecutive free slots"""
    n = len(TIME_SLOTS)
    slot_indices = []
    i = start_idx
    accumulated = 0
    
    # *** MODIFIED: Re-enabled this constraint per user request ***
    # Check lecture-tutorial same-day constraint
    # We check against the base_code (e.g., "MA161-CSE")
    if component_type == 'LEC' and has_component_on_day(timetable, day, course_code, 'TUT'):
        return None, None
    if component_type == 'TUT' and has_component_on_day(timetable, day, course_code, 'LEC'):
        return None, None
    
    # *** NEW: Added constraint for two lectures on the same day ***
    if component_type == 'LEC' and has_component_on_day(timetable, day, course_code, 'LEC'):
        return None, None
    
    while i < n and accumulated < required_minutes:
        if is_minor_slot(TIME_SLOTS[i]) or is_break_time_slot(TIME_SLOTS[i], semester):
            return None, None
        
        if timetable[day][i]['type'] is not None:
            return None, None
        
        if faculty in professor_schedule and i in professor_schedule[faculty][day]:
            return None, None
        
        slot_indices.append(i)
        accumulated += slot_minutes(TIME_SLOTS[i])
        i += 1
    
    if accumulated >= required_minutes:
        # *** MODIFIED: Pass the *full code* for C004 check ***
        # For non-elective core courses, course_code is the base_code (e.g., "MA161-CSE")
        # For electives, this will be the full code (e.g., "B1-PHD151")
        room = find_suitable_room_for_slot(course_code, room_type, day, slot_indices, 
                                           room_schedule, course_room_mapping, component_type,
                                           student_strength) # <-- ADDED
        if room is not None:
            return slot_indices, room
    
    return None, None

def get_all_possible_start_indices():
    idxs = list(range(len(TIME_SLOTS)))
    random.shuffle(idxs)
    return idxs

def check_professor_availability(professor_schedule, faculty, day, start_idx, duration_slots):
    if faculty not in professor_schedule:
        return True
    if not professor_schedule[faculty][day]:
        return True
    
    new_slots = set(range(start_idx, start_idx + duration_slots))
    existing_slots = professor_schedule[faculty][day]
    
    return not (new_slots & existing_slots)

def load_batch_data():
    batch_info = {}
    for _, r in df.iterrows():
        dept = str(r['Department'])
        sem = int(r['Semester'])
        batch_info[(dept, sem)] = {'num_sections': 1}
    return batch_info

# ---------------------------
# Global elective basket scheduling
# ---------------------------
def schedule_global_elective_baskets(df_input, professor_schedule, room_schedule, course_room_mapping):
    """Pre-schedule elective baskets globally"""
    print("\n" + "="*80)
    print("🎓 GLOBAL ELECTIVE BASKET SCHEDULING")
    print("="*80)
    
    basket_groups = {}
    
    # *** MODIFIED FIX: B1/B2 OVERLAP ***
    # This dictionary will hold separate time locks for each basket type
    # e.g., {'B1': {0: set(), 1: set()}, 'B2': {0: set(), 1: set()}}
    # This allows B1 and B2 to be scheduled at the same time.
    global_basket_slots_by_type = {}

    for _, course in df_input.iterrows():
        code = str(course.get('Course Code', '')).strip()
        basket = extract_elective_basket(code)
        
        if basket and pd.notna(basket):
            if 'Schedule' in course and str(course.get('Schedule', 'Yes')).strip().upper() != 'YES':
                continue
            
            # *** NEW: Initialize the lock for this basket type if it's new ***
            if basket not in global_basket_slots_by_type:
                global_basket_slots_by_type[basket] = {d: set() for d in range(len(DAYS))}

            semester = int(course.get('Semester', 0))
            key = (semester, basket)
            
            if key not in basket_groups:
                basket_groups[key] = []
            basket_groups[key].append(course)
    
    global_schedule = {}
    
    for (semester, basket_name), basket_courses in sorted(basket_groups.items()):
        print(f"\n📚 Semester {semester}, Basket {basket_name}: {len(basket_courses)} courses")
        
        # Get the correct global slot lock for this basket type (e.g., 'B1's lock)
        current_basket_global_slots = global_basket_slots_by_type[basket_name]

        first_course = basket_courses[0]
        lec_count, tut_count, lab_count, _ = calculate_required_minutes(first_course)
        
        print(f"    Structure: L={lec_count}h, T={tut_count}h, P={lab_count}h")
        
        lec_sessions = int(lec_count * 60 / LECTURE_MIN) if lec_count > 0 else 0
        tut_sessions = int(tut_count * 60 / TUTORIAL_MIN) if tut_count > 0 else 0
        lab_sessions = int(lab_count * 60 / LAB_MIN) if lab_count > 0 else 0
        
        basket_schedule = []
        
        # Schedule lectures
        for session_num in range(lec_sessions):
            scheduled = False
            for attempt in range(5000):
                day = random.randint(0, len(DAYS)-1)
                start_idx = random.randint(0, max(0, len(TIME_SLOTS)-3))
                
                conflict = False
                for prev_day, prev_slots, prev_type in basket_schedule:
                    if prev_day == day and prev_type in ['LEC', 'TUT']:
                        if any(s in prev_slots for s in range(start_idx, start_idx + 2)):
                            conflict = True
                            break
                
                if conflict:
                    continue
                
                slot_indices = []
                accumulated = 0
                valid = True
                
                for i in range(start_idx, len(TIME_SLOTS)):
                    if is_minor_slot(TIME_SLOTS[i]) or is_break_time_slot(TIME_SLOTS[i], semester):
                        valid = False
                        break
                    
                    slot_indices.append(i)
                    accumulated += slot_minutes(TIME_SLOTS[i])
                    
                    if accumulated >= LECTURE_MIN:
                        break
                
                # *** MODIFIED: Check only against the current basket type's global slots ***
                if valid and accumulated >= LECTURE_MIN and len(slot_indices) > 0:
                    if any(s in current_basket_global_slots[day] for s in slot_indices):
                        continue # Slot already taken by another 'B1' course, try again
                    
                    # This slot is free. Book it.
                    basket_schedule.append((day, slot_indices, 'LEC'))
                    current_basket_global_slots[day].update(slot_indices) # Add to 'B1' lock
                    scheduled = True
                    slot_time = TIME_SLOTS[slot_indices[0]][0].strftime('%H:%M')
                    print(f"    ✅ Lecture {session_num+1}/{lec_sessions}: {DAYS[day]} at {slot_time}")
                    break
            
            if not scheduled:
                print(f"    ⚠️ Could not schedule Lecture {session_num+1}/{lec_sessions}")
        
        # Schedule tutorials
        for session_num in range(tut_sessions):
            scheduled = False
            for attempt in range(5000):
                day = random.randint(0, len(DAYS)-1)
                start_idx = random.randint(0, max(0, len(TIME_SLOTS)-2))
                
                lec_on_day = any(d == day and ct == 'LEC' for d, _, ct in basket_schedule)
                if lec_on_day:
                    continue
                
                conflict = False
                for prev_day, prev_slots, prev_type in basket_schedule:
                    if prev_day == day and prev_type == 'TUT':
                        if any(s in prev_slots for s in range(start_idx, start_idx + 2)):
                            conflict = True
                            break
                
                if conflict:
                    continue
                
                slot_indices = []
                accumulated = 0
                valid = True
                
                for i in range(start_idx, len(TIME_SLOTS)):
                    if is_minor_slot(TIME_SLOTS[i]) or is_break_time_slot(TIME_SLOTS[i], semester):
                        valid = False
                        break
                    
                    slot_indices.append(i)
                    accumulated += slot_minutes(TIME_SLOTS[i])
                    
                    if accumulated >= TUTORIAL_MIN:
                        break
                
                # *** MODIFIED: Check only against the current basket type's global slots ***
                if valid and accumulated >= TUTORIAL_MIN and len(slot_indices) > 0:
                    if any(s in current_basket_global_slots[day] for s in slot_indices):
                        continue # Slot already taken by another 'B1' course, try again

                    # This slot is free. Book it.
                    basket_schedule.append((day, slot_indices, 'TUT'))
                    current_basket_global_slots[day].update(slot_indices) # Add to 'B1' lock
                    scheduled = True
                    slot_time = TIME_SLOTS[slot_indices[0]][0].strftime('%H:%M')
                    print(f"    ✅ Tutorial {session_num+1}/{tut_sessions}: {DAYS[day]} at {slot_time}")
                    break
            
            if not scheduled:
                print(f"    ⚠️ Could not schedule Tutorial {session_num+1}/{tut_sessions}")
        
        # Schedule labs
        for session_num in range(lab_sessions):
            scheduled = False
            for attempt in range(5000):
                day = random.randint(0, len(DAYS)-1)
                start_idx = random.randint(0, max(0, len(TIME_SLOTS)-3))
                
                conflict = False
                for prev_day, prev_slots, prev_type in basket_schedule:
                    if prev_day == day:
                        if any(s in prev_slots for s in range(start_idx, start_idx + 3)):
                            conflict = True
                            break
                
                if conflict:
                    continue
                
                slot_indices = []
                accumulated = 0
                valid = True
                
                for i in range(start_idx, len(TIME_SLOTS)):
                    if is_minor_slot(TIME_SLOTS[i]) or is_break_time_slot(TIME_SLOTS[i], semester):
                        valid = False
                        break
                    
                    slot_indices.append(i)
                    accumulated += slot_minutes(TIME_SLOTS[i])
                    
                    if accumulated >= LAB_MIN:
                        break
                
                # *** MODIFIED: Check only against the current basket type's global slots ***
                if valid and accumulated >= LAB_MIN and len(slot_indices) > 0:
                    if any(s in current_basket_global_slots[day] for s in slot_indices):
                        continue # Slot already taken by another 'B1' course, try again
                        
                    # This slot is free. Book it.
                    basket_schedule.append((day, slot_indices, 'LAB'))
                    current_basket_global_slots[day].update(slot_indices) # Add to 'B1' lock
                    scheduled = True
                    slot_time = TIME_SLOTS[slot_indices[0]][0].strftime('%H:%M')
                    print(f"    ✅ Lab {session_num+1}/{lab_sessions}: {DAYS[day]} at {slot_time}")
                    break
            
            if not scheduled:
                print(f"    ⚠️ Could not schedule Lab {session_num+1}/{lab_sessions}")
        
        global_schedule[(semester, basket_name)] = basket_schedule
        print(f"    📋 Total sessions scheduled: {len(basket_schedule)}")
    
    print("\n" + "="*80)
    print(f"✅ Global basket scheduling complete: {len(global_schedule)} baskets")
    print("="*80 + "\n")
    
    return global_schedule

def add_unscheduled_course(unscheduled_components, department, semester, code, name,
                           faculty, comp_type, section, reason):
    existing = next((u for u in unscheduled_components if u.code == code), None)
    if existing:
        if comp_type not in existing.component_type:
            existing.component_type += f", {comp_type}"
        if reason not in existing.reason:
            existing.reason += f"; {reason}"
    else:
        unscheduled_components.append(
            UnscheduledComponent(department, semester, code, name, faculty,
                                 comp_type, 1, section, reason)
        )

def is_7th_semester(department, semester):
    """Check if this is 7th semester from CSE/DSAI/ECE"""
    dept_upper = str(department).strip().upper()
    return int(semester) == 7 and dept_upper in ['CSE', 'DSAI', 'ECE']

# ---------------------------
# Main generation function
# ---------------------------
def generate_all_timetables():
    global TIME_SLOTS
    TIME_SLOTS = generate_time_slots()
    
    batch_info = load_batch_data()
    room_schedule = {}
    professor_schedule = {}
    course_room_mapping = {}
    
    # Pre-schedule elective baskets
    global_basket_schedule = schedule_global_elective_baskets(df, professor_schedule, room_schedule, course_room_mapping)
    
    if global_basket_schedule is None:
        global_basket_schedule = {}

    # *** NEW: List to store B1 course-room mappings ***
    b1_schedule_list = []

    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    overview = wb.create_sheet("Overview")
    overview.append(["Combined Timetable for All Departments and Semesters"])
    overview.append(["Generated on:", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    overview.append([])
    overview.append(["Department", "Semester", "Sheet Name"])
    row_index = 5

    unscheduled_components = []

    SUBJECT_COLORS = [
        "FF6B6B", "4ECDC4", "FF9F1C", "5D5FEF", "45B7D1",
        "F72585", "7209B7", "3A0CA3", "4361EE", "4CC9F0",
        "06D6A0", "FFD166", "EF476F", "118AB2", "073B4C"
    ]

    seventh_sem_processed = False
    seventh_sem_course_data = []

    for department in df['Department'].unique():
        sems = sorted(df[df['Department'] == department]['Semester'].unique())
        for semester in sems:
            if is_7th_semester(department, semester):
                if not seventh_sem_processed:
                    for dept in ['CSE', 'DSAI', 'ECE']:
                        dept_courses = df[(df['Department'] == dept) & (df['Semester'] == 7)]
                        if not dept_courses.empty:
                            seventh_sem_course_data.append(dept_courses)
                    seventh_sem_processed = True
                continue

            dept_upper = str(department).strip().upper()
            num_sections = 2 if (dept_upper == "CSE" and int(semester) in [2, 4, 6]) else 1

            courses = df[(df['Department'] == department) & (df['Semester'] == semester)]
            if 'Schedule' in courses.columns:
                courses = courses[(courses['Schedule'].fillna('Yes').str.upper() == 'YES') | (courses['Schedule'].isna())]
            if courses.empty:
                continue

            if 'P' in courses.columns:
                lab_courses = courses[courses['P'] > 0].copy()
                non_lab_courses = courses[courses['P'] == 0].copy()
            else:
                lab_courses = courses.head(0)
                non_lab_courses = courses.copy()

            if not lab_courses.empty:
                lab_courses['priority'] = lab_courses.apply(get_course_priority, axis=1)
                lab_courses = lab_courses.sort_values('priority', ascending=False)
            non_lab_courses['priority'] = non_lab_courses.apply(get_course_priority, axis=1)
            non_lab_courses = non_lab_courses.sort_values('priority', ascending=False)

            combined = pd.concat([lab_courses, non_lab_courses])
            combined['is_elective'] = combined.apply(is_elective, axis=1)
            combined['elective_basket'] = combined['Course Code'].apply(extract_elective_basket)

            courses_combined = combined.sort_values(
                by=['is_elective', 'priority'], 
                ascending=[False, False]
            ).drop_duplicates()

            for section in range(num_sections):
                section_title = f"{department}_{semester}" if num_sections == 1 else f"{department}_{semester}_{chr(65 + section)}"
                ws = wb.create_sheet(title=section_title)

                overview.cell(row=row_index, column=1, value=department)
                overview.cell(row=row_index, column=2, value=str(semester))
                overview.cell(row=row_index, column=3, value=section_title)
                row_index += 1

                timetable = {d: {s: {'type': None, 'code': '', 'name': '', 'faculty': '', 'classroom': ''} for s in range(len(TIME_SLOTS))} for d in range(len(DAYS))}
                course_day_components = {}

                section_subject_color = {}
                color_iter = iter(SUBJECT_COLORS)
                course_faculty_map = {}
                basket_scheduled_courses = set()

                for _, c in courses_combined.iterrows():
                    code = str(c.get('Course Code', '')).strip()
                    if code and code not in section_subject_color:
                        try:
                            section_subject_color[code] = next(color_iter)
                        except StopIteration:
                            section_subject_color[code] = random.choice(SUBJECT_COLORS)
                        course_faculty_map[code] = select_faculty(c.get('Faculty', 'TBD'))

                # Apply global basket schedules
                print(f"\n📋 Applying basket schedules for {section_title}...")
                
                # This now includes B1/B2 core courses AND electives
                elective_courses_in_section = courses_combined[courses_combined['is_elective'] == True]
                
                for _, course in elective_courses_in_section.iterrows():
                    code = str(course.get('Course Code', '')).strip()
                    basket = course.get('elective_basket')
                    
                    if not basket or pd.isna(basket):
                        continue
                    
                    basket_key = (semester, basket)
                    if basket_key not in global_basket_schedule:
                        print(f"    ⚠️ No global schedule found for {basket} (Course: {code})")
                        continue
                    
                    basket_schedule = global_basket_schedule[basket_key]
                    base_code = get_base_course_code(code)
                    name = str(course.get('Course Name', '')).strip()
                    faculty = select_faculty(course.get('Faculty', 'TBD'))
                    
                    if faculty not in professor_schedule:
                        professor_schedule[faculty] = {d: set() for d in range(len(DAYS))}
                    
                    if base_code not in course_day_components:
                        course_day_components[base_code] = {}
                    
                    for day, slot_indices, comp_type in basket_schedule:
                        room_type = get_required_room_type(comp_type)
                        
                        # *** MODIFIED FIX ***
                        # Use the full 'code' (e.g., "B1-CS366" or "B2-MA161-C004") to find the room
                        student_strength = int(course.get('total_students', 50)) # <-- ADDED
                        candidate_room = find_suitable_room_for_slot(
                            code, room_type, day, slot_indices,  # <-- Use full 'code'
                            room_schedule, course_room_mapping, comp_type,
                            student_strength # <-- ADDED
                        )
                        
                        if candidate_room is None:
                            print(f"    ⚠️ No room available for {code} {comp_type}")
                            add_unscheduled_course(unscheduled_components, department, semester, code, name, faculty, comp_type, section, f"No suitable room found (Needs {student_strength} capacity)")
                            continue
                        
                        # *** NEW: Add to B1 list ***
                        if basket == "B1" and (code, candidate_room) not in b1_schedule_list:
                            # Add to list only once per course
                            b1_schedule_list.append((code, candidate_room))
                        # *** END NEW ***
                        
                        for si_idx, si in enumerate(slot_indices):
                            # Only fill the timetable if the slot is empty
                            if timetable[day][si]['type'] is None:
                                timetable[day][si]['type'] = comp_type
                                # Display the base code (e.g., MA161-C004)
                                timetable[day][si]['code'] = f"{basket}\n{base_code}" if si_idx == 0 else ''
                                timetable[day][si]['name'] = name if si_idx == 0 else ''
                                timetable[day][si]['faculty'] = faculty if si_idx == 0 else ''
                                timetable[day][si]['classroom'] = candidate_room if si_idx == 0 else ''
                                
                                # Also book the professor and room
                                professor_schedule[faculty][day].add(si)
                                if candidate_room not in room_schedule:
                                    room_schedule[candidate_room] = {d: set() for d in range(len(DAYS))}
                                room_schedule[candidate_room][day].add(si)
                        
                        if day not in course_day_components[base_code]:
                            course_day_components[base_code][day] = []
                        course_day_components[base_code][day].append(comp_type)
                    
                    basket_scheduled_courses.add(code)
                    print(f"    ✅ Applied {basket} schedule to {code}")

                # Schedule non-elective courses
                print(f"\n📖 Scheduling non-elective courses for {section_title}...")
                for _, course in courses_combined.iterrows():
                    code = str(course.get('Course Code', '')).strip()
                    
                    if code in basket_scheduled_courses:
                        continue
                    
                    # This check is now correct: it skips B1/B2 electives AND B1/B2 core courses
                    if course.get('is_elective'):
                        continue
                    
                    base_code = get_base_course_code(code)
                    name = str(course.get('Course Name', '')).strip()
                    faculty = select_faculty(course.get('Faculty', 'TBD'))
                    student_strength = int(course.get('total_students', 50)) # <-- ADDED

                    if faculty not in professor_schedule:
                        professor_schedule[faculty] = {d: set() for d in range(len(DAYS))}
                    
                    if base_code not in course_day_components:
                        course_day_components[base_code] = {}

                    lec_count, tut_count, lab_count, ss_count = calculate_required_minutes(course)
                    lec_sessions_needed = int(lec_count * 60 / LECTURE_MIN) if lec_count > 0 else 0
                    tut_sessions_needed = int(tut_count * 60 / TUTORIAL_MIN) if tut_count > 0 else 0
                    lab_sessions_needed = int(lab_count * 60 / LAB_MIN) if lab_count > 0 else 0

                    def schedule_component(required_minutes, comp_type, student_strength, attempts_limit=5000):
                        room_type = get_required_room_type(comp_type)
                        
                        for attempt in range(attempts_limit):
                            day = random.randint(0, len(DAYS)-1)
                            starts = get_all_possible_start_indices()
                            
                            for start_idx in starts:
                                slot_indices, candidate_room = find_consecutive_slots_for_minutes(
                                    timetable, day, start_idx, required_minutes, semester,
                                    professor_schedule, faculty, room_schedule, room_type,
                                    base_code, course_room_mapping, comp_type, course_day_components,
                                    student_strength # <-- ADDED
                                )

                                if slot_indices is None:
                                    continue
                                if not check_professor_availability(professor_schedule, faculty, day, slot_indices[0], len(slot_indices)):
                                    continue
                                if candidate_room is None:
                                    continue
                                
                                for si_idx, si in enumerate(slot_indices):
                                    timetable[day][si]['type'] = comp_type
                                    timetable[day][si]['code'] = code if si_idx == 0 else ''
                                    timetable[day][si]['name'] = name if si_idx == 0 else ''
                                    timetable[day][si]['faculty'] = faculty if si_idx == 0 else ''
                                    timetable[day][si]['classroom'] = candidate_room if si_idx == 0 else ''
                                    professor_schedule[faculty][day].add(si)
                                    if candidate_room not in room_schedule:
                                        room_schedule[candidate_room] = {d: set() for d in range(len(DAYS))}
                                    room_schedule[candidate_room][day].add(si)
                                
                                if day not in course_day_components[base_code]:
                                    course_day_components[base_code][day] = []
                                course_day_components[base_code][day].append(comp_type)
                                
                                return True
                        return False

                    # Schedule lectures
                    for _ in range(lec_sessions_needed):
                        ok = schedule_component(LECTURE_MIN, 'LEC', student_strength)
                        if not ok:
                            add_unscheduled_course(unscheduled_components, department, semester, code, name, faculty, 'LEC', section, f"Could not find suitable slot (Needs {student_strength} capacity)")

                    # Schedule tutorials
                    for _ in range(tut_sessions_needed):
                        ok = schedule_component(TUTORIAL_MIN, 'TUT', student_strength)
                        if not ok:
                            add_unscheduled_course(unscheduled_components, department, semester, code, name, faculty, 'TUT', section, f"Could not find suitable slot (Needs {student_strength} capacity)")

                    # Schedule labs
                    for _ in range(lab_sessions_needed):
                        ok = schedule_component(LAB_MIN, 'LAB', student_strength)
                        if not ok:
                            add_unscheduled_course(unscheduled_components, department, semester, code, name, faculty, 'LAB', section, f"No computer lab available (Needs {student_strength} capacity)")

                # Write timetable to sheet
                write_timetable_to_sheet(ws, timetable, section_subject_color, course_faculty_map, 
                                         courses_combined, course_room_mapping, semester)

    # *** NEW: Print B1 course list ***
    if b1_schedule_list:
        print("\n" + "="*80)
        print("B1 Basket Room Allocations")
        print("="*80)
        
        # Use a dictionary to store unique code-room pairs
        unique_b1_courses = {}
        for code, room in b1_schedule_list:
            if code not in unique_b1_courses:
                unique_b1_courses[code] = room
                
        for code, room in sorted(unique_b1_courses.items()):
            print(f"    - Course: {code.ljust(25)} Room: {room}")
        print("="*80 + "\n")
    # *** END NEW ***

    # Generate common 7th semester timetable
    if seventh_sem_course_data:
        generate_7th_sem_common_timetable(wb, seventh_sem_course_data, overview, row_index, 
                                          unscheduled_components, professor_schedule, 
                                          room_schedule, course_room_mapping, SUBJECT_COLORS)

    # Format overview sheet
    format_overview_sheet(overview, row_index)

    # Write meta entries into a hidden sheet so teacher generator can read
    if META_ENTRIES:
        try:
            meta_ws = wb.create_sheet('_META')
            headers = ['sheet', 'row', 'start_col', 'end_col', 'faculty', 'classroom', 'typ', 'code']
            meta_ws.append(headers)
            for me in META_ENTRIES:
                meta_ws.append([
                    me.get('sheet'), me.get('row'), me.get('start_col'), me.get('end_col'),
                    me.get('faculty'), me.get('classroom'), me.get('typ'), me.get('code')
                ])
            meta_ws.sheet_state = 'hidden'
        except Exception:
            pass

    # Save workbook
    out_filename = os.path.join(OUTPUT_DIR, "timetable_all_departments.xlsx")
    try:
        wb.save(out_filename)
        print(f"\n✅ Combined timetable saved as {out_filename}")
    except Exception as e:
        print(f"❌ Failed to save timetable: {e}")
        traceback.print_exc()

    # Generate teacher and unscheduled workbooks
    try:
        create_teacher_and_unscheduled_from_combined(out_filename, unscheduled_components)
    except Exception as e:
        print("❌ Failed to generate teacher/unscheduled workbooks:", e)
        traceback.print_exc()

    return out_filename

# ---------------------------
# 7th semester common timetable
# ---------------------------
def generate_7th_sem_common_timetable(wb, course_data_list, overview, row_index, 
                                      unscheduled_components, professor_schedule, 
                                      room_schedule, course_room_mapping, SUBJECT_COLORS):
    """Generate common timetable for 7th semester"""
    
    all_7th_courses = pd.concat(course_data_list, ignore_index=True)
    if 'Schedule' in all_7th_courses.columns:
        all_7th_courses = all_7th_courses[(all_7th_courses['Schedule'].fillna('Yes').str.upper() == 'YES') | 
                                          (all_7th_courses['Schedule'].isna())]
    
    if all_7th_courses.empty:
        return
    
    section_title = "Common_7th_Semester"
    ws = wb.create_sheet(title=section_title)
    
    overview.cell(row=row_index, column=1, value="CSE/DSAI/ECE")
    overview.cell(row=row_index, column=2, value="7")
    overview.cell(row=row_index, column=3, value=section_title)
    
    if 'P' in all_7th_courses.columns:
        lab_courses = all_7th_courses[all_7th_courses['P'] > 0].copy()
        non_lab_courses = all_7th_courses[all_7th_courses['P'] == 0].copy()
    else:
        lab_courses = all_7th_courses.head(0)
        non_lab_courses = all_7th_courses.copy()
    
    if not lab_courses.empty:
        lab_courses['priority'] = lab_courses.apply(get_course_priority, axis=1)
        lab_courses = lab_courses.sort_values('priority', ascending=False)
    non_lab_courses['priority'] = non_lab_courses.apply(get_course_priority, axis=1)
    non_lab_courses = non_lab_courses.sort_values('priority', ascending=False)
    
    combined = pd.concat([lab_courses, non_lab_courses])
    combined['is_elective'] = combined.apply(is_elective, axis=1)
    combined['elective_basket'] = combined['Course Code'].apply(extract_elective_basket)
    
    courses_combined = combined.sort_values(
        by=['is_elective', 'priority'], 
        ascending=[False, False]
    ).drop_duplicates()
    
    timetable = {d: {s: {'type': None, 'code': '', 'name': '', 'faculty': '', 'classroom': ''} 
                       for s in range(len(TIME_SLOTS))} for d in range(len(DAYS))}
    course_day_components = {}
    
    section_subject_color = {}
    color_iter = iter(SUBJECT_COLORS)
    course_faculty_map = {}
    
    for _, c in courses_combined.iterrows():
        code = str(c.get('Course Code', '')).strip()
        if code and code not in section_subject_color:
            try:
                section_subject_color[code] = next(color_iter)
            except StopIteration:
                section_subject_color[code] = random.choice(SUBJECT_COLORS)
            course_faculty_map[code] = select_faculty(c.get('Faculty', 'TBD'))
    
    # Schedule non-electives as before, and schedule elective baskets together
    non_electives = courses_combined[courses_combined['is_elective'] == False]
    electives = courses_combined[courses_combined['is_elective'] == True]

    # Helper to schedule a single course (reuse existing pattern)
    def schedule_single_course(course_row):
        code = str(course_row.get('Course Code', '')).strip()
        base_code = get_base_course_code(code)
        name = str(course_row.get('Course Name', '')).strip()
        faculty = select_faculty(course_row.get('Faculty', 'TBD'))
        student_strength = int(course_row.get('total_students', 50))

        if faculty not in professor_schedule:
            professor_schedule[faculty] = {d: set() for d in range(len(DAYS))}
        if base_code not in course_day_components:
            course_day_components[base_code] = {}

        lec_count, tut_count, lab_count, _ = calculate_required_minutes(course_row)
        lec_sessions_needed = int(lec_count * 60 / LECTURE_MIN) if lec_count > 0 else 0
        tut_sessions_needed = int(tut_count * 60 / TUTORIAL_MIN) if tut_count > 0 else 0
        lab_sessions_needed = int(lab_count * 60 / LAB_MIN) if lab_count > 0 else 0

        def schedule_component(required_minutes, comp_type, student_strength_local, faculty_local, code_local, name_local):
            room_type = get_required_room_type(comp_type)
            for attempt in range(5000):
                day = random.randint(0, len(DAYS)-1)
                starts = get_all_possible_start_indices()
                for start_idx in starts:
                    slot_indices, candidate_room = find_consecutive_slots_for_minutes(
                        timetable, day, start_idx, required_minutes, 7,
                        professor_schedule, faculty_local, room_schedule, room_type,
                        code_local, course_room_mapping, comp_type, course_day_components,
                        student_strength_local
                    )
                    if slot_indices is None or candidate_room is None:
                        continue
                    if not check_professor_availability(professor_schedule, faculty_local, day, slot_indices[0], len(slot_indices)):
                        continue
                    for si_idx, si in enumerate(slot_indices):
                        timetable[day][si]['type'] = comp_type
                        timetable[day][si]['code'] = code_local if si_idx == 0 else ''
                        timetable[day][si]['name'] = name_local if si_idx == 0 else ''
                        timetable[day][si]['faculty'] = faculty_local if si_idx == 0 else ''
                        timetable[day][si]['classroom'] = candidate_room if si_idx == 0 else ''
                        professor_schedule[faculty_local][day].add(si)
                        if candidate_room not in room_schedule:
                            room_schedule[candidate_room] = {d: set() for d in range(len(DAYS))}
                        room_schedule[candidate_room][day].add(si)
                    if day not in course_day_components[get_base_course_code(code_local)]:
                        course_day_components[get_base_course_code(code_local)][day] = []
                    course_day_components[get_base_course_code(code_local)][day].append(comp_type)
                    return True
            return False

        for _ in range(lec_sessions_needed):
            ok = schedule_component(LECTURE_MIN, 'LEC', student_strength, faculty, code, name)
            if not ok:
                add_unscheduled_course(unscheduled_components, "Common_7th", 7, code, name, faculty, 'LEC', 0, f"Could not find suitable slot (Needs {student_strength} capacity)")
        for _ in range(tut_sessions_needed):
            ok = schedule_component(TUTORIAL_MIN, 'TUT', student_strength, faculty, code, name)
            if not ok:
                add_unscheduled_course(unscheduled_components, "Common_7th", 7, code, name, faculty, 'TUT', 0, f"Could not find suitable slot (Needs {student_strength} capacity)")
        for _ in range(lab_sessions_needed):
            ok = schedule_component(LAB_MIN, 'LAB', student_strength, faculty, code, name)
            if not ok:
                add_unscheduled_course(unscheduled_components, "Common_7th", 7, code, name, faculty, 'LAB', 0, f"No computer lab available (Needs {student_strength} capacity)")

    # Schedule non-elective courses normally
    for _, course in non_electives.iterrows():
        schedule_single_course(course)

    # Group electives by basket and schedule each basket as a unit
    if not electives.empty:
        baskets = {}
        for _, row in electives.iterrows():
            basket = extract_elective_basket(str(row.get('Course Code', '')).strip())
            if not basket:
                continue
            baskets.setdefault(basket, []).append(row)

        for basket_label, rows in baskets.items():
            # Use first row as representative for LTPS structure
            first = rows[0]
            lec_count, tut_count, lab_count, _ = calculate_required_minutes(first)
            lec_sessions_needed = int(lec_count * 60 / LECTURE_MIN) if lec_count > 0 else 0
            tut_sessions_needed = int(tut_count * 60 / TUTORIAL_MIN) if tut_count > 0 else 0
            lab_sessions_needed = int(lab_count * 60 / LAB_MIN) if lab_count > 0 else 0

            # Aggregate faculties and student strengths
            facs = [select_faculty(r.get('Faculty', 'TBD')) for r in rows]
            facs = [f for f in facs if f]
            unique_facs = list(dict.fromkeys(facs))
            agg_faculty = '/'.join(unique_facs) if unique_facs else 'TBD'
            total_strength = sum(int(r.get('total_students', 50)) for r in rows)

            # Representative base code (used for meta); prefer first course's base
            rep_code = str(first.get('Course Code', '')).strip()
            rep_base = get_base_course_code(rep_code)

            # Ensure professor_schedule entries exist for all faculties
            for f in unique_facs:
                if f not in professor_schedule:
                    professor_schedule[f] = {d: set() for d in range(len(DAYS))}

            def schedule_basket_component(required_minutes, comp_type, student_strength_local, faculties_local):
                room_type = get_required_room_type(comp_type)
                for attempt in range(5000):
                    day = random.randint(0, len(DAYS)-1)
                    starts = get_all_possible_start_indices()
                    for start_idx in starts:
                        # Build slot_indices manually (same slot for all courses)
                        slot_indices_local = []
                        acc = 0
                        i = start_idx
                        while i < len(TIME_SLOTS) and acc < required_minutes:
                            if is_minor_slot(TIME_SLOTS[i]) or is_break_time_slot(TIME_SLOTS[i], 7):
                                break
                            if timetable[day][i]['type'] is not None:
                                break
                            slot_indices_local.append(i)
                            acc += slot_minutes(TIME_SLOTS[i])
                            i += 1
                        if acc < required_minutes:
                            continue

                        # Try to find a separate room for each course in this basket for the same slot_indices
                        course_room_map = {}
                        failed = False
                        rooms_assigned = []
                        faculties_assigned = []

                        for row in rows:
                            course_code = str(row.get('Course Code', '')).strip()
                            base = get_base_course_code(course_code)
                            faculty_c = select_faculty(row.get('Faculty', 'TBD'))
                            strength_c = int(row.get('total_students', 50))
                            room_type_c = get_required_room_type(comp_type)

                            # Check professor availability
                            if not check_professor_availability(professor_schedule, faculty_c, day, slot_indices_local[0], len(slot_indices_local)):
                                failed = True
                                break

                            # Find a suitable room for this course for these exact slot indices
                            candidate_room_c = find_suitable_room_for_slot(course_code, room_type_c, day, slot_indices_local, room_schedule, course_room_mapping, comp_type, strength_c)
                            if candidate_room_c is None:
                                failed = True
                                break

                            # Tentatively record assignment
                            course_room_map[course_code] = (candidate_room_c, faculty_c, base)
                            rooms_assigned.append(candidate_room_c)
                            faculties_assigned.append(faculty_c)

                        if failed:
                            # revert any tentative course_room_mapping done by find_suitable_room_for_slot? (mapping is persistent)
                            continue

                        # All courses in basket can be placed in these slot_indices; commit assignments
                        for si_idx, si in enumerate(slot_indices_local):
                            timetable[day][si]['type'] = comp_type
                            # Visible code shows basket label and a representative base (first course)
                            timetable[day][si]['code'] = f"{basket_label}\n{rep_base}" if si_idx == 0 else ''
                            timetable[day][si]['name'] = basket_label if si_idx == 0 else ''
                            # Aggregate faculties and rooms for meta/display
                            timetable[day][si]['faculty'] = '/'.join(unique_facs) if si_idx == 0 else ''
                            timetable[day][si]['classroom'] = '/'.join(rooms_assigned) if si_idx == 0 else ''

                            # Mark professors and rooms busy per course
                            for course_code, (room_c, fac_c, base_c) in course_room_map.items():
                                professor_schedule[fac_c][day].add(si)
                                if room_c not in room_schedule:
                                    room_schedule[room_c] = {d: set() for d in range(len(DAYS))}
                                room_schedule[room_c][day].add(si)

                        if rep_base not in course_day_components:
                            course_day_components[rep_base] = {}
                        if day not in course_day_components[rep_base]:
                            course_day_components[rep_base][day] = []
                        course_day_components[rep_base][day].append(comp_type)
                        return True
                return False

            for _ in range(lec_sessions_needed):
                ok = schedule_basket_component(LECTURE_MIN, 'LEC', total_strength, unique_facs)
                if not ok:
                    add_unscheduled_course(unscheduled_components, "Common_7th", 7, basket_label, basket_label, agg_faculty, 'LEC', 0, f"Could not schedule basket {basket_label} (Needs {total_strength} capacity)")
            for _ in range(tut_sessions_needed):
                ok = schedule_basket_component(TUTORIAL_MIN, 'TUT', total_strength, unique_facs)
                if not ok:
                    add_unscheduled_course(unscheduled_components, "Common_7th", 7, basket_label, basket_label, agg_faculty, 'TUT', 0, f"Could not schedule basket {basket_label} (Needs {total_strength} capacity)")
            for _ in range(lab_sessions_needed):
                ok = schedule_basket_component(LAB_MIN, 'LAB', total_strength, unique_facs)
                if not ok:
                    add_unscheduled_course(unscheduled_components, "Common_7th", 7, basket_label, basket_label, agg_faculty, 'LAB', 0, f"Could not schedule basket {basket_label} labs (Needs {total_strength} capacity)")
    
    write_timetable_to_sheet(ws, timetable, section_subject_color, course_faculty_map, 
                             courses_combined, course_room_mapping, 7)

# ---------------------------
# Write timetable to Excel
# ---------------------------
def write_timetable_to_sheet(ws, timetable, section_subject_color, course_faculty_map, 
                             courses_combined, course_room_mapping, semester):
    """Write formatted timetable to worksheet"""
    
    # Header row
    header = ['Day'] + [f"{slot[0].strftime('%H:%M')}-{slot[1].strftime('%H:%M')}" for slot in TIME_SLOTS]
    ws.append(header)
    
    # Styling
    header_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
    header_font = Font(bold=True)
    header_alignment = Alignment(horizontal='center', vertical='center')
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                    top=Side(style='thin'), bottom=Side(style='thin'))
    lec_fill_default = PatternFill(start_color="FA8072", end_color="FA8072", fill_type="solid")
    lab_fill_default = PatternFill(start_color="7CFC00", end_color="7CFC00", fill_type="solid")
    tut_fill_default = PatternFill(start_color="87CEEB", end_color="87CEEB", fill_type="solid")
    break_fill = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")
    minor_fill = PatternFill(start_color="9ACD32", end_color="9ACD32", fill_type="solid")
    
    # Write timetable rows
    for day_idx, day_name in enumerate(DAYS):
        ws.append([day_name] + [''] * len(TIME_SLOTS))
        row_num = ws.max_row
        merges = []
        
        for slot_idx in range(len(TIME_SLOTS)):
            cell_obj = ws.cell(row=row_num, column=slot_idx + 2)
            
            if is_minor_slot(TIME_SLOTS[slot_idx]):
                cell_obj.value = "Minor Slot"
                cell_obj.fill = minor_fill
                cell_obj.font = Font(bold=True)
                cell_obj.alignment = Alignment(horizontal='center', vertical='center')
                cell_obj.border = border
                continue
            
            if is_break_time_slot(TIME_SLOTS[slot_idx], semester):
                cell_obj.value = "LUNCH BREAK"  # Updated label
                cell_obj.fill = break_fill
                cell_obj.font = Font(bold=True)
                cell_obj.alignment = Alignment(horizontal='center', vertical='center')
                cell_obj.border = border
                continue
            
            if timetable[day_idx][slot_idx]['type'] is None:
                cell_obj.border = border
                continue
            
            typ = timetable[day_idx][slot_idx]['type']
            code = timetable[day_idx][slot_idx]['code']
            cls = timetable[day_idx][slot_idx]['classroom']
            fac = timetable[day_idx][slot_idx]['faculty']
            
            if code:
                span = [slot_idx]
                j = slot_idx + 1
                while j < len(TIME_SLOTS) and timetable[day_idx][j]['type'] is not None and timetable[day_idx][j]['code'] == '':
                    span.append(j)
                    j += 1
                import re
                # Detect basket entries even when stored as 'B1\nBASECODE'
                raw_code = str(code) if code is not None else ''
                parts = [p.strip() for p in raw_code.splitlines() if p.strip()]
                basket_label = None
                base_code = ''
                if parts and re.match(r'^B\d+$', parts[0].upper()):
                    basket_label = parts[0].upper()
                    if len(parts) > 1:
                        base_code = parts[1]
                elif re.match(r'^B\d+$', raw_code.strip().upper()):
                    basket_label = raw_code.strip().upper()

                if basket_label:
                    # Record metadata so teacher generator can still pick up
                    # faculty and room info, but display only basket+type.
                    start_col = slot_idx + 2
                    end_col = slot_idx + 2 + len(span) - 1
                    META_ENTRIES.append({
                        'sheet': ws.title,
                        'row': row_num,
                        'start_col': start_col,
                        'end_col': end_col,
                        'faculty': fac,
                        'classroom': cls,
                        'typ': typ,
                        'code': base_code or basket_label
                    })
                    display = f"{basket_label}\n{typ}"
                else:
                    display = f"{code}\n{typ}\nRoom: {cls}\n{fac}"
                
                # Use the full code (including B1/B2) for coloring
                full_code = code.replace('\n', '-') if '\n' in code else code
                
                if full_code in section_subject_color:
                    subj_color = section_subject_color.get(full_code)
                    fill = PatternFill(start_color=subj_color, end_color=subj_color, fill_type="solid")
                elif code.split('\n')[-1] in section_subject_color:
                     # Fallback for base code
                    actual_code = code.split('\n')[-1]
                    subj_color = section_subject_color.get(actual_code)
                    fill = PatternFill(start_color=subj_color, end_color=subj_color, fill_type="solid")
                else:
                    fill = {'LEC': lec_fill_default, 'LAB': lab_fill_default, 
                            'TUT': tut_fill_default}.get(typ, lec_fill_default)
                
                cell_obj.value = display
                cell_obj.fill = fill
                merges.append((slot_idx + 2, slot_idx + 2 + len(span) - 1, display, fill))
            
            cell_obj.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
            cell_obj.border = border
        
        # Apply merges
        for start_col, end_col, val, fill in merges:
            if end_col > start_col:
                rng = f"{get_column_letter(start_col)}{row_num}:{get_column_letter(end_col)}{row_num}"
                try:
                    ws.merge_cells(rng)
                    mc = ws[f"{get_column_letter(start_col)}{row_num}"]
                    mc.value = val
                    mc.fill = fill
                    mc.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
                    mc.border = border
                except Exception:
                    pass
    
    # Set column widths
    for col_idx in range(1, len(TIME_SLOTS)+2):
        try:
            ws.column_dimensions[get_column_letter(col_idx)].width = 15
        except Exception:
            pass
    
    # Set row heights
    for row in ws.iter_rows(min_row=2, max_row=len(DAYS)+1):
        ws.row_dimensions[row[0].row].height = 40
    
    # Add self-study courses section
    current_row = len(DAYS) + 4
    ss_courses = []
    for _, course in courses_combined.iterrows():
        l = int(course['L']) if pd.notna(course['L']) else 0
        t = int(course['T']) if pd.notna(course['T']) else 0
        p = int(course['P']) if pd.notna(course['P']) else 0
        s = int(course['S']) if pd.notna(course['S']) else 0
        if s > 0 and l == 0 and t == 0 and p == 0:
            ss_courses.append({
                'code': str(course['Course Code']),
                'name': str(course['Course Name']),
                'faculty': str(course['Faculty'])
            })
    
    if ss_courses:
        ws.cell(row=current_row, column=1, value="Self-Study Only Courses")
        ws.cell(row=current_row, column=1).font = Font(bold=True)
        current_row += 1
        
        headers = ['Course Code', 'Course Name', 'Faculty']
        for col, header in enumerate(headers, 1):
            ws.cell(row=current_row, column=col, value=header)
            ws.cell(row=current_row, column=col).font = Font(bold=True)
        current_row += 1
        
        for course in ss_courses:
            ws.cell(row=current_row, column=1, value=course['code'])
            ws.cell(row=current_row, column=2, value=course['name'])
            ws.cell(row=current_row, column=3, value=course['faculty'])
            current_row += 1
        
        current_row += 2
    
    # Add legend
    legend_title = ws.cell(row=current_row, column=1, value="Legend")
    legend_title.font = Font(bold=True, size=12)
    current_row += 2
    
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    
    legend_headers = ['Subject Code', 'Color', 'Subject Name', 'Faculty', 'LTPS', 'Room']
    for col, header in enumerate(legend_headers, 1):
        cell = ws.cell(row=current_row, column=col, value=header)
        cell.font = Font(bold=True)
        cell.border = border
        cell.fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    current_row += 1
    
    # Add legend entries
    for code, color in section_subject_color.items():
        # Use the code itself (which is unique, e.g., "B1-MA161")
        assigned_room = course_room_mapping.get(f"{code}_LEC") or \
                        course_room_mapping.get(f"{code}_LAB") or \
                        course_room_mapping.get(f"{code}_TUT") or "—"
        
        if not assigned_room or assigned_room == "—":
            # Try to find a room for the base code if full code fails
            base_code = get_base_course_code(code)
            assigned_room = course_room_mapping.get(f"{base_code}_LEC") or \
                            course_room_mapping.get(f"{base_code}_LAB") or \
                            course_room_mapping.get(f"{base_code}_TUT") or "—"
            if not assigned_room or assigned_room == "—":
                continue # Skip if no room is mapped at all
        
        ws.row_dimensions[current_row].height = 30
        
        ltps_value = ""
        course_name = ''
        fac_name = ''
        for _, course_row in courses_combined.iterrows():
            if str(course_row['Course Code']) == code:
                l = str(int(course_row['L'])) if pd.notna(course_row['L']) else "0"
                t = str(int(course_row['T'])) if pd.notna(course_row['T']) else "0"
                p = str(int(course_row['P'])) if pd.notna(course_row['P']) else "0"
                s = str(int(course_row['S'])) if pd.notna(course_row['S']) and 'S' in course_row else "0"
                ltps_value = f"{l}-{t}-{p}-{s}"
                course_name = str(course_row['Course Name'])
                fac_name = course_faculty_map.get(code, '')
                break
        
        cells = [
            (code, None),
            ('', PatternFill(start_color=color, end_color=color, fill_type="solid")),
            (course_name, None),
            (fac_name, None),
            (ltps_value, None),
            (assigned_room, None)
        ]
        
        for col, (value, fill) in enumerate(cells, 1):
            cell = ws.cell(row=current_row, column=col, value=value)
            cell.border = border
            if fill:
                cell.fill = fill
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True, indent=2)
        
        current_row += 1

def format_overview_sheet(overview, row_index):
    """Format overview sheet"""
    for col in range(1, 4):
        overview.column_dimensions[get_column_letter(col)].width = 20
    
    for row_ in overview.iter_rows(min_row=1, max_row=4):
        for cell in row_:
            cell.font = Font(bold=True)
    
    for cell in overview[4]:
        cell.fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
        cell.font = Font(bold=True)
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                            top=Side(style='thin'), bottom=Side(style='thin'))

# ---------------------------
# Teacher and unscheduled workbooks
# ---------------------------
def split_faculty_names(fac_str):
    """Split faculty string into separate names"""
    if fac_str is None:
        return []
    s = str(fac_str).strip()
    if s == '' or s.lower() in ['nan', 'none']:
        return []
    parts = [s]
    for sep in ['/', ',', '&', ';']:
        if sep in s:
            parts = [p.strip() for p in s.split(sep) if p.strip()]
            break
    return parts if parts else [s]

def parse_cell_for_course(cell_value):
    """Parse cell value to extract course info"""
    if cell_value is None:
        return ('', '', '', '')
    text = str(cell_value).strip()
    if text == '':
        return ('', '', '', '')
    
    lines = [ln.strip() for ln in text.splitlines() if ln.strip()]
    faculty = ''
    room = ''
    code = ''
    typ = ''
    
    for ln in lines:
        if 'room' in ln.lower():
            parts = ln.split(':')
            if len(parts) >= 2:
                room = parts[-1].strip()
    
    if len(lines) >= 1:
        last = lines[-1]
        if 'room' not in last.lower() and ':' not in last:
            faculty = last
    
    first = lines[0] if lines else ''
    if first:
        # Handle basket format (B1\nCODE)
        if '\n' in text and first.startswith('B'):
            code = lines[1] if len(lines) > 1 else first # This gets the base code
        else:
            code = first
        
        for t in ['LEC', 'LAB', 'TUT']:
            if t in text.upper():
                typ = t
                break
    
    return (code, typ, room, faculty)

def create_teacher_and_unscheduled_from_combined(timetable_filename, unscheduled_components):
    """Build teacher_timetables.xlsx and unscheduled_courses.xlsx"""
    try:
        wb = load_workbook(timetable_filename, data_only=True)
    except Exception as e:
        print(f"❌ Failed to open {timetable_filename}: {e}")
        return
    
    teacher_slots = {}
    slot_headers = []

    # Load meta information if present so teacher workbook can be built
    meta_map = {}
    if '_META' in wb.sheetnames:
        meta_ws = wb['_META']
        try:
            for r in range(2, meta_ws.max_row + 1):
                m_sheet = meta_ws.cell(r, 1).value
                m_row = meta_ws.cell(r, 2).value
                m_start = meta_ws.cell(r, 3).value
                m_end = meta_ws.cell(r, 4).value
                m_fac = meta_ws.cell(r, 5).value
                m_room = meta_ws.cell(r, 6).value
                m_typ = meta_ws.cell(r, 7).value
                m_code = meta_ws.cell(r, 8).value
                if not m_sheet or not m_row or not m_start:
                    continue
                for col in range(int(m_start), int(m_end) + 1):
                    meta_map[(m_sheet, int(m_row), int(col))] = {
                        'faculty': m_fac,
                        'room': m_room,
                        'typ': m_typ,
                        'code': m_code
                    }
        except Exception:
            meta_map = {}
    
    for sheetname in wb.sheetnames:
        if sheetname.lower() == 'overview':
            continue
        ws = wb[sheetname]
        header = [str(ws.cell(1, c).value).strip() if ws.cell(1, c).value else '' for c in range(2, ws.max_column + 1)]
        if len(header) > len(slot_headers):
            slot_headers = header

        for r in range(2, ws.max_row + 1):
            day = ws.cell(r, 1).value
            if not day or str(day) not in DAYS:
                break
            day_idx = DAYS.index(day)

            for c in range(2, ws.max_column + 1):
                # If there's meta info for this exact cell use it
                if (sheetname, r, c) in meta_map:
                    m = meta_map[(sheetname, r, c)]
                    faculty = m.get('faculty')
                    room = m.get('room')
                    typ = m.get('typ')
                    code = m.get('code')
                    fac_list = split_faculty_names(faculty)
                    for f in fac_list:
                        if not f or str(f).strip().upper() in ["BREAK", "MINOR SLOT", "NAN", "NONE", "", "LUNCH BREAK"]:
                            continue
                        teacher_slots.setdefault(f, {d: {i: '' for i in range(len(slot_headers))} for d in range(len(DAYS))})
                        teacher_slots[f][day_idx][c - 2] = f"{code} {typ}\n({sheetname})\nRoom: {room}" if code else ''
                    continue

                code, typ, room, faculty = parse_cell_for_course(ws.cell(r, c).value)
                for f in split_faculty_names(faculty):
                    if not f or str(f).strip().upper() in ["BREAK", "MINOR SLOT", "NAN", "NONE", "", "LUNCH BREAK"]:
                        continue
                    teacher_slots.setdefault(f, {d: {i: '' for i in range(len(slot_headers))} for d in range(len(DAYS))})
                    teacher_slots[f][day_idx][c - 2] = f"{code} {typ}\n({sheetname})\nRoom: {room}" if code else ''
    
    # Create teacher workbook
    twb = Workbook()
    if "Sheet" in twb.sheetnames:
        twb.remove(twb["Sheet"])
    
    header_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
    alt_fill = PatternFill(start_color="FFF8DC", end_color="FFF8DC", fill_type="solid")
    border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
    header_font = Font(bold=True, size=12)
    title_font = Font(bold=True, size=14)
    cell_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    for teacher in sorted(teacher_slots.keys()):
        safe_name = teacher[:31] or "Unknown"
        ws = twb.create_sheet(title=safe_name)
        
        ws.merge_cells("A1:{}1".format(get_column_letter(len(slot_headers) + 1)))
        title_cell = ws.cell(row=1, column=1, value=f"{teacher} — Weekly Timetable")
        title_cell.font = title_font
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        
        ws.append(["Day"] + slot_headers)
        for cell in ws[2]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
        
        for d, day in enumerate(DAYS):
            row = [day] + [teacher_slots[teacher][d][i] for i in range(len(slot_headers))]
            ws.append(row)
            row_idx = ws.max_row
            if d % 2 == 0:
                for cell in ws[row_idx]:
                    cell.fill = alt_fill
            for cell in ws[row_idx]:
                cell.alignment = cell_align
                cell.border = border
            ws.row_dimensions[row_idx].height = 35
        
        ws.column_dimensions["A"].width = 15
        for col in range(2, len(slot_headers) + 2):
            ws.column_dimensions[get_column_letter(col)].width = 20
    
    twb.save(os.path.join(OUTPUT_DIR, "teacher_timetables.xlsx"))
    print("✅ Saved teacher_timetables.xlsx")
    
    # Create unscheduled workbook
    uwb = Workbook()
    ws = uwb.active
    ws.title = "Unscheduled Courses"
    
    headers = ["Course Code", "Department", "Semester", "Component Type", "Reason"]
    ws.append(headers)
    
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    unscheduled_unique = {}
    
    for u in unscheduled_components:
        if u.code not in unscheduled_unique:
            if hasattr(u, "reason") and u.reason and len(str(u.reason).strip()) > 0:
                reason_text = str(u.reason).strip()
            else:
                reason_text = "Unspecified scheduling issue"
            
            unscheduled_unique[u.code] = {
                "Course Code": u.code,
                "Department": u.department,
                "Semester": u.semester,
                "Component Type": u.component_type,
                "Reason": reason_text
            }
    
    for entry in unscheduled_unique.values():
        ws.append([entry["Course Code"], entry["Department"], entry["Semester"], 
                   entry["Component Type"], entry["Reason"]])
    
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 50
    
    uwb.save(os.path.join(OUTPUT_DIR, "unscheduled_courses.xlsx"))
    print(f"✅ Saved unscheduled_courses.xlsx with {len(unscheduled_unique)} courses")

# ---------------------------
# Exam generator
# ---------------------------
def exam_generator():
    """Generate exam timetable with room allocation and invigilation"""
    exam_file = "Exam_timetable.xlsx"
    
    try:
        df_courses = pd.read_csv(os.path.join(INPUT_DIR, "combined.csv"))
        df_rooms = pd.read_csv(os.path.join(INPUT_DIR, "rooms.csv"))
    except FileNotFoundError as e:
        print(f"❌ Missing file: {e}")
        return None
    
    df_courses = df_courses.dropna(subset=["Course Code", "Course Name", "Faculty", "Department", "Semester"])
    if "total_students" not in df_courses.columns:
        df_courses["total_students"] = 50
    df_courses["total_students"] = df_courses["total_students"].fillna(50).astype(int)
    
    df_rooms.columns = [c.strip().lower() for c in df_rooms.columns]
    
    def find_col(keywords):
        for c in df_rooms.columns:
            if any(k in c for k in keywords):
                return c
        return None
    
    room_col = find_col(["room", "num", "id"])
    cap_col = find_col(["cap", "seat"])
    type_col = find_col(["type"])
    
    if not room_col or not cap_col:
        print("❌ rooms.csv must have columns for room number and capacity.")
        return None
    
    df_rooms = df_rooms.rename(columns={room_col: "room", cap_col: "capacity"})
    if type_col:
        df_rooms = df_rooms.rename(columns={type_col: "type"})
    else:
        df_rooms["type"] = "LECTURE_ROOM"
    
    df_rooms["room"] = df_rooms["room"].astype(str).str.strip()
    df_rooms["capacity"] = pd.to_numeric(df_rooms["capacity"], errors="coerce").fillna(0).astype(int)
    df_rooms = df_rooms[df_rooms["capacity"] > 0].sort_values(by="capacity").reset_index(drop=True)
    
    print(f"✅ Loaded {len(df_rooms)} rooms for exam scheduling")
    
    faculty_list = list(set(sum([str(f).replace(" and ", "/").replace(",", "/").split("/") for f in df_courses["Faculty"]], [])))
    faculty_list = [f.strip() for f in faculty_list if f.strip()]
    
    session_title = "Jan-April 03:00 PM to 04:30 PM"
    start_date = datetime(2025, 11, 20)
    num_days = min(10, len(df_courses))
    dates = [start_date + timedelta(days=i) for i in range(num_days)]
    days = [d.strftime("%A") for d in dates]
    
    shuffled = df_courses.sample(frac=1, random_state=42).reset_index(drop=True)
    course_date_map = {row["Course Code"]: dates[i % len(dates)] for i, row in shuffled.iterrows()}
    
    date_room_usage = {d.strftime("%d-%b-%Y"): set() for d in dates}
    invigilation_entries = []
    
    for date in dates:
        date_str = date.strftime("%d-%b-%Y")
        today_courses = shuffled[shuffled["Course Code"].isin(
            [c for c, dt in course_date_map.items() if dt == date]
        )]
        
        for _, course in today_courses.iterrows():
            code = course["Course Code"]
            name = course["Course Name"]
            dept = course["Department"]
            sem = course["Semester"]
            teacher = str(course["Faculty"]).strip()
            students = int(course["total_students"])
            time_slot = "03:00 PM—04:30 PM"
            
            assigned_rooms = []
            remaining = students
            
            available = df_rooms[~df_rooms["room"].isin(date_room_usage[date_str])].copy()
            available = available.sort_values(by="capacity", ascending=True)
            
            suitable = available[available["capacity"] >= remaining]
            if not suitable.empty:
                best = suitable.iloc[0]
                assigned_rooms = [best["room"]]
                date_room_usage[date_str].add(best["room"])
            else:
                total_cap = 0
                for _, room_row in available.iterrows():
                    assigned_rooms.append(room_row["room"])
                    total_cap += room_row["capacity"]
                    date_room_usage[date_str].add(room_row["room"])
                    if total_cap >= remaining:
                        break
                if total_cap < remaining:
                    print(f"⚠️ Not enough capacity for {code}")
            
            available_teachers = [f for f in faculty_list if f.lower() not in teacher.lower()]
            for room in assigned_rooms:
                invigilator = random.choice(available_teachers) if available_teachers else "TBD"
                invigilation_entries.append({
                    "Faculty": invigilator,
                    "Date": date_str,
                    "Time": time_slot,
                    "Course Code": code,
                    "Course Name": name,
                    "Department": dept,
                    "Semester": sem,
                    "Room": room,
                    "Strength": students
                })
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Exam Timetable"
    
    bold_center = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    header_fill = PatternFill(start_color="BFBFBF", end_color="BFBFBF", fill_type="solid")
    border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
    
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(dates)+1)
    title = ws.cell(row=1, column=1, value=session_title)
    title.font = Font(bold=True, size=14)
    title.alignment = center
    title.fill = header_fill
    title.border = border
    
    ws.cell(row=2, column=1, value="Date").font = bold_center
    for i, d in enumerate(dates):
        c = ws.cell(row=2, column=i+2, value=d.strftime("%d-%b-%Y"))
        c.font = bold_center
        c.alignment = center
        c.fill = header_fill
        c.border = border
    
    ws.cell(row=3, column=1, value="Days").font = bold_center
    for i, day in enumerate(days):
        c = ws.cell(row=3, column=i+2, value=day)
        c.font = bold_center
        c.alignment = center
        c.fill = header_fill
        c.border = border
    
    grouped_by_date = {}
    for e in invigilation_entries:
        grouped_by_date.setdefault(e["Date"], set()).add(e["Course Code"])
    max_rows = max(len(v) for v in grouped_by_date.values()) if grouped_by_date else 0
    
    for r in range(max_rows):
        for i, d in enumerate(dates):
            code_list = list(grouped_by_date.get(d.strftime("%d-%b-%Y"), []))
            val = code_list[r] if r < len(code_list) else ""
            cell = ws.cell(row=r+4, column=i+2, value=val)
            cell.alignment = center
            cell.border = border
    
    ws.column_dimensions["A"].width = 15
    for col in range(2, len(dates)+2):
        ws.column_dimensions[get_column_letter(col)].width = 16
    
    ws2 = wb.create_sheet("Exam Invigilation Schedule")
    headers = ["Faculty", "Date", "Time", "Course Code", "Course Name", "Department", "Semester", "Room", "Strength"]
    ws2.append(headers)
    
    for i, h in enumerate(headers, 1):
        cell = ws2.cell(row=1, column=i, value=h)
        cell.font = bold_center
        cell.alignment = center
        cell.fill = header_fill
        cell.border = border
    
    for entry in invigilation_entries:
        ws2.append([entry[h] for h in headers])
    
    for col in range(1, len(headers)+1):
        ws2.column_dimensions[get_column_letter(col)].width = 22
    for r in range(2, ws2.max_row+1):
        for c in ws2[r]:
            c.alignment = center
            c.border = border
    
    exam_file = os.path.join(OUTPUT_DIR, "Exam_timetable.xlsx")
    wb.save(exam_file)
    print(f"✅ Exam timetable saved → {exam_file}")
    return exam_file

# ---------------------------
# Main execution
# ---------------------------
if __name__ == "__main__":
    try:
        print("\n" + "="*80)
        print("🎓 IIIT DHARWAD TIMETABLE GENERATOR")
        print("="*80)
        print("\n🔧 Configuration:")
        print(f"    - No morning break (removed 10:30-10:45)")
        print(f"    - Lunch break: 13:00-14:00 (extended)")
        print(f"    - Lecture duration: {LECTURE_MIN} minutes")
        print(f"    - Tutorial duration: {TUTORIAL_MIN} minutes")
        print(f"    - Lab duration: {LAB_MIN} minutes")
        print(f"    - Min gap between lectures: {MIN_GAP_BETWEEN_LECTURES} minutes")
        print("="*80 + "\n")
        
        generate_all_timetables()
        exam_generator()
        
        print("\n" + "="*80)
        print("✅ TIMETABLE GENERATION COMPLETE!")
        print("="*80)
        print(f"\n📁 Output files saved in: {OUTPUT_DIR}")
        print("    1. timetable_all_departments.xlsx - Main timetable")
        print("    2. teacher_timetables.xlsx - Faculty schedules")
        print("    3. unscheduled_courses.xlsx - Courses that couldn't be scheduled")
        print("    4. Exam_timetable.xlsx - Exam schedule")
        print("\n💡 Tips:")
        print("    - Check unscheduled_courses.xlsx to see which courses failed")
        print("    - If many courses are unscheduled, consider:")
        print("      * Adding more computer labs or larger lecture rooms")
        print("      * Relaxing some constraints")
        print("      * Extending time slots into evening")
        print("="*80 + "\n")
        
    except Exception as e:
        print("\n" + "="*80)
        print("❌ ERROR DURING TIMETABLE GENERATION")
        print("="*80)
        print(f"Error: {e}")
        traceback.print_exc()
        print("="*80 + "\n")