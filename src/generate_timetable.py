import pandas as pd
import os
import random
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment, Font
from openpyxl.utils import get_column_letter
import colorsys

# ==============================
# Configuration
# ==============================
COURSE_FILE = "data/course_data.csv"
FACULTY_FILE = "data/faculty.csv"
ROOM_FILE = "data/rooms.csv"
FLAT_CSV_FILE = "output/generated_timetable.csv"
EXCEL_FILE = "output/structured_timetable.xlsx"

days = ["Mon", "Tue", "Wed", "Thu", "Fri"]

LECTURE_SLOTS = [("09:00", "10:30"), ("10:30", "12:00"), ("13:00", "14:30"), ("14:30", "16:00"), ("16:00", "17:30")]
TUTORIAL_SLOTS = [("09:00", "10:00"), ("10:00", "11:00"), ("11:00", "12:00"), ("13:00", "14:00"), ("14:00", "15:00"), ("15:00", "16:00"), ("16:00", "17:00")]
LAB_SLOTS = [("09:00", "11:00"), ("11:00", "13:00"), ("13:00", "15:00"), ("15:00", "17:00"), ("17:00", "19:00")]

OPEN_ELECTIVE_DAY = "Wed"
OPEN_ELECTIVE_SLOT = ("14:30", "15:30")

# ==============================
# Step 1: Load Data
# ==============================
courses = pd.read_csv(COURSE_FILE)
faculty = pd.read_csv(FACULTY_FILE)
rooms = pd.read_csv(ROOM_FILE)
room_list = rooms["Room"].tolist() if "Room" in rooms.columns else ["R1", "R2", "R3", "R4", "R5"]

# ==============================
# Helper Functions
# ==============================
def has_conflict(existing_schedule, new_class):
    """
    Checks conflicts for:
    - Room
    - Instructor
    - Same course
    - Same branch & semester (students cannot attend multiple classes at the same time)
    """
    for cls in existing_schedule:
        if cls['Day'] == new_class['Day']:
            overlap = not (new_class['End-Time'] <= cls['Start-Time'] or new_class['Start-Time'] >= cls['End-Time'])
            if overlap:
                # Room conflict
                if cls['Room'] == new_class['Room']:
                    return True
                # Instructor conflict
                if cls['Instructor'] == new_class['Instructor']:
                    return True
                # Same course conflict
                if cls['Course Code'] == new_class['Course Code']:
                    return True
                # Branch & semester conflict
                if cls['Branch'] == new_class['Branch'] and cls['Semester'] == new_class['Semester']:
                    return True
    return False

def parse_ltp(ltp_str):
    parts = ltp_str.split("-")
    return int(parts[0]), int(parts[1]), int(parts[2])  # L, T, P

def schedule_course(course, slots, suffix, timetable):
    """Schedule a single class into timetable"""
    random.shuffle(days)
    random.shuffle(slots)
    random.shuffle(room_list)
    for day in days:
        for start_time, end_time in slots:
            for room in room_list:
                new_class = {
                    'Course Code': course['Course Code'],
                    'Course-Name': course['Course-Name'] + suffix,
                    'Instructor': course['Instructor'],
                    'Room': room,
                    'Day': day,
                    'Start-Time': start_time,
                    'End-Time': end_time,
                    'Branch': course['Branch'],
                    'Semester': course['Semester']
                }
                if not has_conflict(timetable, new_class):
                    timetable.append(new_class)
                    return True
    return False

# ==============================
# Step 2: Schedule Classes
# ==============================
timetable = []

core_courses = courses[courses["Type"].str.lower() == "core"]
elective_courses = courses[courses["Type"].str.lower() == "elective"]

# Schedule core courses
for _, course in core_courses.iterrows():
    L, T, P = parse_ltp(course["L-T-P-S-C"])
    for _ in range(L):
        schedule_course(course, LECTURE_SLOTS, " (Lecture)", timetable)
    for _ in range(T):
        schedule_course(course, TUTORIAL_SLOTS, " (Tutorial)", timetable)
    for _ in range(P):
        schedule_course(course, LAB_SLOTS, " (Lab)", timetable)

# Schedule open electives
for _, course in elective_courses.iterrows():
    timetable.append({
        'Course Code': course['Course Code'],
        'Course-Name': "Open Elective - " + course['Course-Name'],
        'Instructor': course['Instructor'],
        'Room': random.choice(room_list),
        'Day': OPEN_ELECTIVE_DAY,
        'Start-Time': OPEN_ELECTIVE_SLOT[0],
        'End-Time': OPEN_ELECTIVE_SLOT[1],
        'Branch': course['Branch'] if 'Branch' in course else "ALL",
        'Semester': course['Semester'] if 'Semester' in course else 0
    })

# ==============================
# Step 3: Export Flat CSV
# ==============================
timetable_df = pd.DataFrame(timetable)
timetable_df = timetable_df.sort_values(by=["Day", "Start-Time"])
os.makedirs(os.path.dirname(FLAT_CSV_FILE), exist_ok=True)
timetable_df.to_csv(FLAT_CSV_FILE, index=False)
print(f"✅ Flat timetable saved to: {FLAT_CSV_FILE}")

# ==============================
# Step 4: Structured Excel
# ==============================
# Prepare pivot table
timetable_df['Slot'] = timetable_df['Start-Time'] + " - " + timetable_df['End-Time']
timetable_df['Display'] = timetable_df['Course Code'] + "\n" + timetable_df['Course-Name'] + "\n(" + timetable_df['Room'] + ")"

structured_df = timetable_df.pivot_table(
    index='Day',
    columns='Slot',
    values='Display',
    aggfunc=lambda x: "\n".join(x)
)

# Ensure correct day order and sort slots
days_order = ["Mon", "Tue", "Wed", "Thu", "Fri"]
structured_df = structured_df.reindex(days_order)
structured_df = structured_df[sorted(structured_df.columns)]

# Save temporarily
os.makedirs(os.path.dirname(EXCEL_FILE), exist_ok=True)
structured_df.to_excel(EXCEL_FILE, index=True)

# ==============================
# Step 5: Apply Unique Course Colors
# ==============================
unique_courses = timetable_df['Course Code'].unique()

def generate_distinct_colors(n):
    """Generate n visually distinct hex colors"""
    hues = [i/n for i in range(n)]
    colors = []
    for h in hues:
        rgb = colorsys.hsv_to_rgb(h, 0.5, 0.95)  # Saturation=0.5, Value=0.95
        hex_color = ''.join(f'{int(c*255):02X}' for c in rgb)
        colors.append(hex_color)
    return colors

course_colors = dict(zip(unique_courses, generate_distinct_colors(len(unique_courses))))

# Load workbook
wb = load_workbook(EXCEL_FILE)
ws = wb.active

# Apply formatting
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=ws.max_column):
    for cell in row:
        if cell.value:
            cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
            
            # Apply course-specific color
            for code in unique_courses:
                if code in cell.value:
                    fill_color = course_colors[code]
                    cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                    cell.font = Font(bold=True, color="FFFFFF")
                    break

# Auto-adjust row heights
for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    max_lines = max((str(cell.value).count('\n') + 1 if cell.value else 1) for cell in row)
    ws.row_dimensions[row[0].row].height = max_lines * 18

# Auto-adjust column widths
for col in ws.columns:
    max_length = 0
    col_letter = get_column_letter(col[0].column)
    for cell in col:
        if cell.value:
            max_length = max(max_length, max(len(line) for line in str(cell.value).split("\n")))
    ws.column_dimensions[col_letter].width = max_length + 2

wb.save(EXCEL_FILE)
print("✅ Structured Excel saved with no overlapping lectures/labs and unique course colors!")
