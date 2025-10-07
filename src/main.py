"""
Main module for timetable generation
Contains the main generation logic and Excel export functionality
"""

import pandas as pd
import random
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter

import config
from utils import (initialize_time_slots, calculate_lunch_breaks, load_rooms, 
                   load_batch_data, load_reserved_slots, load_faculty_preferences,
                   is_basket_course, get_basket_group, get_basket_group_slots,
                   is_break_time, is_slot_reserved, select_faculty, 
                   calculate_required_slots, get_required_room_type, 
                   get_course_priority, UnscheduledComponent)
from generate_timetable import (find_suitable_room, is_lecture_scheduled, 
                       check_faculty_daily_components, check_faculty_course_gap,
                       get_best_slots)


def load_course_data():
    """Load and validate course data from CSV"""
    try:
        encodings_to_try = ['utf-8-sig', 'utf-8', 'cp1252']
        df = None
        last_error = None
        
        for encoding in encodings_to_try:
            try:
                df = pd.read_csv('combined.csv', encoding=encoding)
                df = df.replace(r'^\s*$', pd.NA, regex=True)
                df = df.replace('nan', pd.NA)
                break
            except UnicodeDecodeError:
                continue
            except Exception as e:
                last_error = e
                continue
                
        if df is None:
            print(f"Error: Unable to read combined.csv. Details: {str(last_error)}")
            exit()
            
    except Exception as e:
        print(f"Error: Failed to load combined.csv. Details: {str(e)}")
        exit()

    if df.empty:
        print("Error: No data found in combined.csv")
        exit()
        
    return df


def schedule_lectures(courses, timetable, professor_schedule, rooms, batch_info, 
                      reserved_slots, department, semester, section, unscheduled_components):
    """Schedule all lecture sessions"""
    for _, course in courses.iterrows():
        code = str(course['Course Code'])
        name = str(course['Course Name'])
        faculty = select_faculty(str(course['Faculty']))
        
        lecture_sessions, _, _, _ = calculate_required_slots(course)
        
        if faculty not in professor_schedule:
            professor_schedule[faculty] = {day: set() for day in range(len(config.DAYS))}

        for _ in range(lecture_sessions):
            scheduled = False
            attempts = 0
            while not scheduled and attempts < 1000:
                day = random.randint(0, len(config.DAYS)-1)
                start_slot = random.randint(0, len(config.TIME_SLOTS)-config.LECTURE_DURATION)
                
                if not check_faculty_course_gap(professor_schedule, timetable, faculty, code, day, start_slot):
                    attempts += 1
                    continue
                
                slots_reserved = any(is_slot_reserved(config.TIME_SLOTS[start_slot + i], 
                                                    config.DAYS[day], semester, department, reserved_slots) 
                                   for i in range(config.LECTURE_DURATION))
                
                if slots_reserved:
                    attempts += 1
                    continue
                
                if not check_faculty_daily_components(professor_schedule, faculty, day, 
                                                   department, semester, section, timetable, code, 'LEC'):
                    attempts += 1
                    continue
                    
                slots_free = True
                for i in range(config.LECTURE_DURATION):
                    current_slot = start_slot + i
                    if (current_slot in professor_schedule[faculty][day] or 
                        timetable[day][current_slot]['type'] is not None or
                        is_break_time(config.TIME_SLOTS[current_slot], semester)):
                        slots_free = False
                        break
                    
                    if current_slot > 0:
                        if is_lecture_scheduled(timetable, day, 
                                             max(0, current_slot - config.BREAK_DURATION), 
                                             current_slot):
                            slots_free = False
                            break
                    
                    if current_slot < len(config.TIME_SLOTS) - 1:
                        if is_lecture_scheduled(timetable, day, current_slot + 1,
                                             min(len(config.TIME_SLOTS), 
                                                 current_slot + config.BREAK_DURATION + 1)):
                            slots_free = False
                            break
                
                if slots_free:
                    room_id = find_suitable_room('LECTURE_ROOM', department, semester, 
                                              day, start_slot, config.LECTURE_DURATION, 
                                              rooms, batch_info, timetable, code)
                    
                    if room_id:
                        classroom = room_id
                        
                        for i in range(config.LECTURE_DURATION):
                            professor_schedule[faculty][day].add(start_slot+i)
                            timetable[day][start_slot+i]['type'] = 'LEC'
                            timetable[day][start_slot+i]['code'] = code if i == 0 else ''
                            timetable[day][start_slot+i]['name'] = name if i == 0 else ''
                            timetable[day][start_slot+i]['faculty'] = faculty if i == 0 else ''
                            timetable[day][start_slot+i]['classroom'] = classroom if i == 0 else ''
                        scheduled = True
                attempts += 1
            if not scheduled:
                unscheduled_components.add(
                    UnscheduledComponent(department, semester, code, name, 
                                       faculty, 'LEC', 1, section)
                )


def schedule_tutorials(courses, timetable, professor_schedule, rooms, batch_info,
                       reserved_slots, department, semester, section, unscheduled_components):
    """Schedule all tutorial sessions"""
    for _, course in courses.iterrows():
        code = str(course['Course Code'])
        name = str(course['Course Name'])
        faculty = select_faculty(str(course['Faculty']))
        
        _, tutorial_sessions, _, _ = calculate_required_slots(course)
        
        if faculty not in professor_schedule:
            professor_schedule[faculty] = {day: set() for day in range(len(config.DAYS))}

        for _ in range(tutorial_sessions):
            scheduled = False
            attempts = 0
            while not scheduled and attempts < 1000:
                day = random.randint(0, len(config.DAYS)-1)
                start_slot = random.randint(0, len(config.TIME_SLOTS)-config.TUTORIAL_DURATION)
                
                if not check_faculty_course_gap(professor_schedule, timetable, faculty, code, day, start_slot):
                    attempts += 1
                    continue
                
                if not check_faculty_daily_components(professor_schedule, faculty, day,
                                                   department, semester, section, timetable, code, 'TUT'):
                    attempts += 1
                    continue
                
                slots_reserved = any(is_slot_reserved(config.TIME_SLOTS[start_slot + i], 
                                                    config.DAYS[day], semester, department, reserved_slots) 
                                   for i in range(config.TUTORIAL_DURATION))
                
                if slots_reserved:
                    attempts += 1
                    continue
                
                slots_free = True
                for i in range(config.TUTORIAL_DURATION):
                    if (start_slot+i in professor_schedule[faculty][day] or 
                        timetable[day][start_slot+i]['type'] is not None or
                        is_break_time(config.TIME_SLOTS[start_slot+i], semester)):
                        slots_free = False
                        break
                
                if slots_free:
                    room_id = find_suitable_room('LECTURE_ROOM', department, semester, 
                                              day, start_slot, config.TUTORIAL_DURATION, 
                                              rooms, batch_info, timetable, code)
                    
                    if room_id:
                        classroom = room_id
                        
                        for i in range(config.TUTORIAL_DURATION):
                            professor_schedule[faculty][day].add(start_slot+i)
                            timetable[day][start_slot+i]['type'] = 'TUT'
                            timetable[day][start_slot+i]['code'] = code if i == 0 else ''
                            timetable[day][start_slot+i]['name'] = name if i == 0 else ''
                            timetable[day][start_slot+i]['faculty'] = faculty if i == 0 else ''
                            timetable[day][start_slot+i]['classroom'] = classroom if i == 0 else ''
                        scheduled = True
                attempts += 1
            if not scheduled:
                unscheduled_components.add(
                    UnscheduledComponent(department, semester, code, name,
                                       faculty, 'TUT', 1, section)
                )


def schedule_labs(courses, timetable, professor_schedule, rooms, batch_info,
                  reserved_slots, faculty_preferences, department, semester, 
                  section, unscheduled_components):
    """Schedule all lab sessions"""
    for _, course in courses.iterrows():
        code = str(course['Course Code'])
        name = str(course['Course Name'])
        faculty = select_faculty(str(course['Faculty']))
        
        _, _, lab_sessions, _ = calculate_required_slots(course)
        
        if lab_sessions > 0:
            if faculty not in professor_schedule:
                professor_schedule[faculty] = {day: set() for day in range(len(config.DAYS))}
            
            room_type = get_required_room_type(course)
            
            for _ in range(lab_sessions):
                scheduled = False
                days = list(range(len(config.DAYS)))
                random.shuffle(days)
                
                for day in days:
                    possible_slots = get_best_slots(timetable, professor_schedule, 
                                                  faculty, day, config.LAB_DURATION, 
                                                  reserved_slots, semester, department, 
                                                  faculty_preferences)
                    
                    for start_slot in possible_slots:
                        room_id = find_suitable_room(room_type, department, semester,
                                                   day, start_slot, config.LAB_DURATION,
                                                   rooms, batch_info, timetable, code)
                        
                        if room_id:
                            classroom = room_id if ',' not in str(room_id) else \
                                       f"{room_id.split(',')[0]}+{room_id.split(',')[1]}"
                            
                            for i in range(config.LAB_DURATION):
                                professor_schedule[faculty][day].add(start_slot+i)
                                timetable[day][start_slot+i]['type'] = 'LAB'
                                timetable[day][start_slot+i]['code'] = code if i == 0 else ''
                                timetable[day][start_slot+i]['name'] = name if i == 0 else ''
                                timetable[day][start_slot+i]['faculty'] = faculty if i == 0 else ''
                                timetable[day][start_slot+i]['classroom'] = classroom if i == 0 else ''
                            scheduled = True
                            break
                    
                    if scheduled:
                        break
                    
                if not scheduled:
                    unscheduled_components.add(
                        UnscheduledComponent(department, semester, code, name,
                                           faculty, 'LAB', 1, section,
                                           "Could not find suitable room and time slot")
                    )


def schedule_self_study(courses, timetable, professor_schedule, rooms, batch_info,
                        reserved_slots, department, semester):
    """Schedule all self-study sessions"""
    for _, course in courses.iterrows():
        code = str(course['Course Code'])
        name = str(course['Course Name'])
        faculty = select_faculty(str(course['Faculty']))
        
        _, _, _, self_study_sessions = calculate_required_slots(course)
        
        if self_study_sessions > 0:
            if faculty not in professor_schedule:
                professor_schedule[faculty] = {day: set() for day in range(len(config.DAYS))}
            
            for _ in range(self_study_sessions):
                scheduled = False
                attempts = 0
                while not scheduled and attempts < 1000:
                    day = random.randint(0, len(config.DAYS)-1)
                    start_slot = random.randint(0, len(config.TIME_SLOTS)-config.SELF_STUDY_DURATION)
                    
                    slots_reserved = any(is_slot_reserved(config.TIME_SLOTS[start_slot + i], 
                                                        config.DAYS[day], semester, department, reserved_slots) 
                                       for i in range(config.SELF_STUDY_DURATION))
                    
                    if slots_reserved:
                        attempts += 1
                        continue
                    
                    slots_free = True
                    for i in range(config.SELF_STUDY_DURATION):
                        if (start_slot+i in professor_schedule[faculty][day] or 
                            timetable[day][start_slot+i]['type'] is not None or
                            is_break_time(config.TIME_SLOTS[start_slot+i], semester)):
                            slots_free = False
                            break
                    
                    if slots_free:
                        room_id = find_suitable_room('LECTURE_ROOM', department, semester, 
                                                  day, start_slot, config.SELF_STUDY_DURATION, 
                                                  rooms, batch_info, timetable, code)
                        
                        if room_id:
                            classroom = room_id
                            
                            for i in range(config.SELF_STUDY_DURATION):
                                professor_schedule[faculty][day].add(start_slot+i)
                                timetable[day][start_slot+i]['type'] = 'SS'
                                timetable[day][start_slot+i]['code'] = code if i == 0 else ''
                                timetable[day][start_slot+i]['name'] = name if i == 0 else ''
                                timetable[day][start_slot+i]['faculty'] = faculty if i == 0 else ''
                                timetable[day][start_slot+i]['classroom'] = classroom if i == 0 else ''
                            scheduled = True
                    attempts += 1


def write_timetable_to_excel(ws, timetable, semester, subject_color_map, 
                             course_faculty_map, self_study_courses, 
                             unscheduled_components, department):
    """Write timetable data to Excel worksheet with formatting"""
    # Write header
    header = ['Day'] + [f"{slot[0].strftime('%H:%M')}-{slot[1].strftime('%H:%M')}" 
                       for slot in config.TIME_SLOTS]
    ws.append(header)
    
    # Style definitions
    header_fill = PatternFill(start_color=config.HEADER_FILL_COLOR, 
                             end_color=config.HEADER_FILL_COLOR, fill_type="solid")
    header_font = Font(bold=True)
    header_alignment = Alignment(horizontal='center', vertical='center')
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    lec_fill = PatternFill(start_color=config.LEC_FILL_COLOR, 
                          end_color=config.LEC_FILL_COLOR, fill_type="solid")
    lab_fill = PatternFill(start_color=config.LAB_FILL_COLOR, 
                          end_color=config.LAB_FILL_COLOR, fill_type="solid")
    tut_fill = PatternFill(start_color=config.TUT_FILL_COLOR, 
                          end_color=config.TUT_FILL_COLOR, fill_type="solid")
    ss_fill = PatternFill(start_color=config.SS_FILL_COLOR, 
                         end_color=config.SS_FILL_COLOR, fill_type="solid")
    break_fill = PatternFill(start_color=config.BREAK_FILL_COLOR, 
                            end_color=config.BREAK_FILL_COLOR, fill_type="solid")
    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                  top=Side(style='thin'), bottom=Side(style='thin'))
    
    # Write timetable rows
    for day_idx, day in enumerate(config.DAYS):
        row_num = day_idx + 2
        ws.append([day])
        
        merge_ranges = []
        
        for slot_idx in range(len(config.TIME_SLOTS)):
            cell_value = ''
            cell_fill = None
            
            if is_break_time(config.TIME_SLOTS[slot_idx], semester):
                cell_value = "BREAK"
                cell_fill = break_fill
            elif timetable[day_idx][slot_idx]['type']:
                activity_type = timetable[day_idx][slot_idx]['type']
                code = timetable[day_idx][slot_idx]['code']
                classroom = timetable[day_idx][slot_idx]['classroom']
                faculty = timetable[day_idx][slot_idx]['faculty']
                
                if code:
                    duration = {
                        'LEC': config.LECTURE_DURATION,
                        'LAB': config.LAB_DURATION,
                        'TUT': config.TUTORIAL_DURATION,
                        'SS': config.SELF_STUDY_DURATION
                    }.get(activity_type, 1)
                    
                    if code in subject_color_map:
                        cell_fill = PatternFill(start_color=subject_color_map[code],
                                              end_color=subject_color_map[code],
                                              fill_type="solid")
                    else:
                        cell_fill = {
                            'LAB': lab_fill, 'TUT': tut_fill,
                            'SS': ss_fill, 'LEC': lec_fill
                        }.get(activity_type, lec_fill)
                    
                    if code and is_basket_course(code):
                        basket_group = get_basket_group(code)
                        basket_codes = set()
                        basket_details = {}
                        
                        for slot_id, slot_data in timetable[day_idx].items():
                            slot_code = slot_data.get('code', '')
                            if (slot_data.get('type') == activity_type and 
                                get_basket_group(slot_code) == basket_group):
                                basket_codes.add(slot_code)
                                if slot_code not in basket_details:
                                    basket_details[slot_code] = {
                                        'faculty': slot_data['faculty'],
                                        'room': slot_data['classroom']
                                    }
                        
                        if basket_codes:
                            basket_header = f"{basket_group} Courses\n"
                            codes_str = ', '.join(sorted(basket_codes))
                            course_details = [
                                f"{code}: {details['faculty']} ({details['room']})"
                                for code, details in sorted(basket_details.items())
                                if code and details['faculty'] and details['room']
                            ]
                            
                            cell_value = f"{basket_header}{codes_str}\n" + "\n".join(course_details)
                    else:
                        cell_value = f"{code} {activity_type}\n{classroom}\n{faculty}"
                    
                    if duration > 1:
                        start_col = get_column_letter(slot_idx + 2)
                        end_col = get_column_letter(slot_idx + duration + 1)
                        merge_range = f"{start_col}{row_num}:{end_col}{row_num}"
                        merge_ranges.append((merge_range, cell_fill))
            
            cell = ws.cell(row=row_num, column=slot_idx+2, value=cell_value)
            if cell_fill:
                cell.fill = cell_fill
            cell.border = border
            cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
        
        for merge_range, fill in merge_ranges:
            ws.merge_cells(merge_range)
            merged_cell = ws[merge_range.split(':')[0]]
            merged_cell.fill = fill
            merged_cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')

    # Set column widths
    for col_idx in range(1, len(config.TIME_SLOTS)+2):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = 15
    
    for row in ws.iter_rows(min_row=2, max_row=len(config.DAYS)+1):
        ws.row_dimensions[row[0].row].height = 40

    # Add sections for self-study courses, unscheduled components, and legend
    current_row = len(config.DAYS) + 4
    
    # Self-study only courses
    if self_study_courses:
        ws.cell(row=current_row, column=1, value="Self-Study Only Courses")
        ws.cell(row=current_row, column=1).font = Font(bold=True)
        current_row += 1
        
        headers = ['Course Code', 'Course Name', 'Faculty']
        for col, header in enumerate(headers, 1):
            ws.cell(row=current_row, column=col, value=header)
            ws.cell(row=current_row, column=col).font = Font(bold=True)
        current_row += 1
        
        for course in self_study_courses:
            if course['department'] == department and course['semester'] == semester:
                ws.cell(row=current_row, column=1, value=course['code'])
                ws.cell(row=current_row, column=2, value=course['name'])
                ws.cell(row=current_row, column=3, value=course['faculty'])
                current_row += 1
        
        current_row += 2

    # Unscheduled components
    if unscheduled_components:
        current_row += 2
        unsch_title = ws.cell(row=current_row, column=1, value="Unscheduled Components")
        unsch_title.font = Font(bold=True, size=12, color="FF0000")
        current_row += 2

        headers = ['Course Code', 'Course Name', 'Faculty', 'Component', 'Sessions', 'Reason']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col, value=header)
            cell.font = Font(bold=True)
            cell.border = border
            cell.fill = PatternFill(start_color=config.UNSCHEDULED_FILL_COLOR, 
                                   end_color=config.UNSCHEDULED_FILL_COLOR, fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')
            ws.column_dimensions[get_column_letter(col)].width = 20
        current_row += 1

        for comp in unscheduled_components:
            cells = [
                (comp.code, None), (comp.name, None), (comp.faculty, None),
                (comp.component_type, None), (comp.sessions, None),
                (comp.reason or "Could not find suitable slot", None)
            ]
            
            for col, (value, fill) in enumerate(cells, 1):
                cell = ws.cell(row=current_row, column=col, value=value)
                cell.border = border
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            current_row += 1
        
        current_row += 2

    # Legend
    legend_title = ws.cell(row=current_row, column=1, value="Legend")
    legend_title.font = Font(bold=True, size=12)
    current_row += 2

    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 15

    legend_headers = ['Subject Code', 'Subject Name', 'Faculty', 'Color']
    for col, header in enumerate(legend_headers, 1):
        cell = ws.cell(row=current_row, column=col, value=header)
        cell.font = Font(bold=True)
        cell.border = border
        cell.fill = PatternFill(start_color=config.LEGEND_FILL_COLOR, 
                               end_color=config.LEGEND_FILL_COLOR, fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center')
    current_row += 1

    for code, color in subject_color_map.items():
        if code in course_faculty_map:
            ws.row_dimensions[current_row].height = 25
            
            cells = [
                (code, None),
                (course_faculty_map[code]['name'], None),
                (course_faculty_map[code]['faculty'], None),
                ('', PatternFill(start_color=color, end_color=color, fill_type="solid"))
            ]
            
            for col, (value, fill) in enumerate(cells, 1):
                cell = ws.cell(row=current_row, column=col, value=value)
                cell.border = border
                if fill:
                    cell.fill = fill
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            
            current_row += 1


def generate_all_timetables():
    """Main function to generate all timetables"""
    initialize_time_slots()
    
    # Load all necessary data
    df = load_course_data()
    reserved_slots = load_reserved_slots()
    faculty_preferences = load_faculty_preferences()
    rooms = load_rooms()
    batch_info = load_batch_data()
    
    workbooks = {}
    professor_schedule = {}
    unscheduled_components = set()
    self_study_courses = []
    
    # Calculate lunch breaks
    all_semesters = sorted(set(int(str(sem)[0]) for sem in df['Semester'].unique()))
    config.lunch_breaks = calculate_lunch_breaks(all_semesters)
    
    # Process each department
    for department in df['Department'].unique():
        wb = Workbook()
        wb.remove(wb.active)
        workbooks[department] = wb
        
        course_faculty_assignments = {}
        
        for semester in df[df['Department'] == department]['Semester'].unique():
            courses = df[(df['Department'] == department) & 
                        (df['Semester'] == semester) &
                        ((df['Schedule'].fillna('Yes').str.upper() == 'YES') | 
                         (df['Schedule'].isna()))].copy()
            
            if courses.empty:
                continue

            # Identify self-study only courses
            for _, course in courses.iterrows():
                l = int(course['L']) if pd.notna(course['L']) else 0
                t = int(course['T']) if pd.notna(course['T']) else 0
                p = int(course['P']) if pd.notna(course['P']) else 0
                s = int(course['S']) if pd.notna(course['S']) else 0
                
                if s > 0 and l == 0 and t == 0 and p == 0:
                    self_study_courses.append({
                        'code': str(course['Course Code']),
                        'name': str(course['Course Name']),
                        'faculty': str(course['Faculty']),
                        'department': department,
                        'semester': semester
                    })

            # Separate lab and non-lab courses
            lab_courses = courses[courses['P'] > 0].copy()
            lab_courses['priority'] = lab_courses.apply(get_course_priority, axis=1)
            lab_courses = lab_courses.sort_values('priority', ascending=False)

            non_lab_courses = courses[courses['P'] == 0].copy()
            non_lab_courses['priority'] = non_lab_courses.apply(get_course_priority, axis=1)
            non_lab_courses = non_lab_courses.sort_values('priority', ascending=False)

            courses = pd.concat([lab_courses, non_lab_courses])

            # Get section info
            dept_info = batch_info.get((department, semester))
            num_sections = dept_info['num_sections'] if dept_info else 1

            for section in range(num_sections):
                section_title = f"{department}_{semester}" if num_sections == 1 else \
                               f"{department}_{semester}_{chr(65+section)}"
                ws = wb.create_sheet(title=section_title)
                
                # Initialize timetable structure
                timetable = {day: {slot: {'type': None, 'code': '', 'name': '', 
                                         'faculty': '', 'classroom': ''} 
                             for slot in range(len(config.TIME_SLOTS))} 
                            for day in range(len(config.DAYS))}
                
                # Create subject color mapping
                subject_color_map = {}
                course_faculty_map = {}
                color_idx = 0
                
                for _, course in courses.iterrows():
                    code = str(course['Course Code'])
                    if code not in subject_color_map and code and code != 'nan':
                        if is_basket_course(code):
                            basket_group = get_basket_group(code)
                            subject_color_map[code] = config.BASKET_GROUP_COLORS.get(
                                basket_group, 
                                config.SUBJECT_COLORS[color_idx % len(config.SUBJECT_COLORS)]
                            )
                        else:
                            subject_color_map[code] = config.SUBJECT_COLORS[color_idx % len(config.SUBJECT_COLORS)]
                        course_faculty_map[code] = {
                            'name': str(course['Course Name']),
                            'faculty': str(course['Faculty'])
                        }
                        color_idx += 1

                # Sort courses by priority
                courses['priority'] = courses.apply(get_course_priority, axis=1)
                courses = courses.sort_values('priority', ascending=False)

                # Handle faculty assignments for non-basket courses
                for _, course in courses.iterrows():
                    code = str(course['Course Code'])
                    faculty = str(course['Faculty'])
                    
                    if not any(code.startswith(f'B{i}') for i in range(1, 10)):
                        if code in course_faculty_assignments:
                            if '/' in faculty:
                                faculty_options = [f.strip() for f in faculty.split('/')] 
                                available_faculty = [f for f in faculty_options 
                                                     if f not in course_faculty_assignments[code]]
                                if available_faculty:
                                    faculty = available_faculty[0]
                                else:
                                    faculty = select_faculty(faculty)
                        else:
                            faculty = select_faculty(faculty)
                            course_faculty_assignments[code] = [faculty]

                # Schedule all components
                schedule_lectures(courses, timetable, professor_schedule, rooms, 
                                batch_info, reserved_slots, department, semester, 
                                section, unscheduled_components)
                
                schedule_tutorials(courses, timetable, professor_schedule, rooms, 
                                 batch_info, reserved_slots, department, semester, 
                                 section, unscheduled_components)
                
                schedule_labs(courses, timetable, professor_schedule, rooms, 
                            batch_info, reserved_slots, faculty_preferences, 
                            department, semester, section, unscheduled_components)
                
                schedule_self_study(courses, timetable, professor_schedule, rooms, 
                                  batch_info, reserved_slots, department, semester)

                # Filter unscheduled components for this section
                dept_unscheduled = [c for c in unscheduled_components 
                                    if c.department == department and 
                                    c.semester == semester and
                                    (c.section == section if num_sections > 1 else True)]

                # Write to Excel
                write_timetable_to_excel(ws, timetable, semester, subject_color_map,
                                        course_faculty_map, self_study_courses,
                                        dept_unscheduled, department)

    # Save workbooks
    for department, wb in workbooks.items():
        filename = f"timetable_{department}.xlsx"
        wb.save(filename)
        print(f"Timetable for {department} saved as {filename}")

    return [f"timetable_{dept}.xlsx" for dept in workbooks.keys()]


if __name__ == "__main__":
    generate_all_timetables()