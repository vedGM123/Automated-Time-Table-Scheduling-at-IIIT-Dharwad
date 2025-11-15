# TT_gen.py -- Timetable generator with room allocation and global room conflict avoidance
# Run: python TT_gen.py
# Requires: pandas, openpyxl


import pandas as pd
import random
from datetime import datetime, time, timedelta
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter
from dataclasses import dataclass
from pathlib import Path
import traceback
import os
import json


# ---------------------------
# Constants and durations (minutes)
# ---------------------------

from pathlib import Path
BASE_DIR = Path(__file__).resolve().parent
PROJECT_DIR = BASE_DIR.parent
DATA_DIR = BASE_DIR.parent / "data"
CONFIG_PATH = DATA_DIR / "config.json"
try:
    with open(CONFIG_PATH, "r", encoding="utf-8") as f:
        config = json.load(f)
except Exception:
    config = {}

DAYS = config.get("days", ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday'])
LECTURE_MIN = config.get("LECTURE_MIN", 90)   # 1.5 hours
LAB_MIN = config.get("LAB_MIN", 120)       # 2 hours (as used in your code)
TUTORIAL_MIN = config.get("TUTORIAL_MIN", 60)  # 1 hour
SELF_STUDY_MIN = config.get("SELF_STUDY_MIN", 60)

# Break windows
MORNING_BREAK_START = time(10, 30)
MORNING_BREAK_END = time(10, 45)
LUNCH_BREAK_START = time(13, 0)
LUNCH_BREAK_END = time(13, 45)

# ---------------------------
# Dataclasses
# ---------------------------
@dataclass
class UnscheduledComponent:
    department: str
    semester: int
    code: str
    name: str
    faculty: str
    component_type: str
    sessions: int
    section: int
    reason: str


INPUT_DIR = PROJECT_DIR / "data"
OUTPUT_DIR = PROJECT_DIR / "output"
# ---------------------------
# Load CSVs
# ---------------------------
try:
    df = pd.read_csv(os.path.join(INPUT_DIR, 'combined.csv'))
except FileNotFoundError:
    raise SystemExit("Error: 'combined.csv' not found in working directory.")

try:
    rooms_df = pd.read_csv(os.path.join(INPUT_DIR, 'rooms.csv'))
except FileNotFoundError:
    rooms_df = pd.DataFrame(columns=['roomNumber', 'type'])

# Normalize rooms lists (case-insensitive)
lecture_rooms = rooms_df[rooms_df.get('type', '').str.upper() == 'LECTURE_ROOM']['roomNumber'].tolist()
computer_lab_rooms = rooms_df[rooms_df.get('type', '').str.upper() == 'COMPUTER_LAB']['roomNumber'].tolist()
large_rooms = rooms_df[rooms_df.get('type', '').str.upper() == 'SEATER_120']['roomNumber'].tolist()

# ---------------------------
# Custom time slot definitions (irregular)
# ---------------------------
def generate_time_slots():
    slots = [
        (time(7, 30), time(9, 0)),     # Minor Slot (morning)
        (time(9, 0), time(10, 0)),
        (time(10, 0), time(10, 30)),
        (time(10, 30), time(10, 45)),  # Short break
        (time(10, 45), time(11, 0)),
        (time(11, 0), time(11, 30)),
        (time(11, 30), time(12, 0)),
        (time(12, 0), time(12, 15)),
        (time(12, 15), time(12, 30)),
        (time(12, 30), time(13, 15)),
        (time(13, 15), time(13, 30)),
        (time(13, 30), time(14, 0)),
        (time(14, 0), time(14, 30)),
        (time(14, 30), time(15, 30)),
        (time(15, 30), time(15, 40)),
        (time(16, 0), time(16, 30)),
        (time(16, 30), time(17, 10)),
        (time(17, 10), time(17, 30)),
        (time(17, 30), time(18, 30)),
        (time(18, 30), time(23, 59)),  # Minor Slot (evening)
    ]
    return slots

TIME_SLOTS = generate_time_slots()

# ---------------------------
# Helpers
# ---------------------------
def slot_minutes(slot):
    s, e = slot
    s_m = s.hour*60 + s.minute
    e_m = e.hour*60 + e.minute
    if e_m < s_m:
        e_m += 24*60
    return e_m - s_m

def overlaps(a_start, a_end, b_start, b_end):
    # input times or time objects - convert to minutes-of-day where needed
    a_s_min = a_start.hour*60 + a_start.minute
    a_e_min = a_end.hour*60 + a_end.minute
    b_s_min = b_start.hour*60 + b_start.minute
    b_e_min = b_end.hour*60 + b_end.minute
    return (a_s_min < b_e_min) and (b_s_min < a_e_min)

def is_break_time_slot(slot, semester=None):
    start, end = slot
    if overlaps(start, end, MORNING_BREAK_START, MORNING_BREAK_END):
        return True
    if overlaps(start, end, LUNCH_BREAK_START, LUNCH_BREAK_END):
        return True
    return False

def is_minor_slot(slot):
    start, end = slot
    if start == time(7, 30) and end == time(9, 0):
        return True
    if start == time(18, 30):
        return True
    return False

def select_faculty(faculty_field):
    if pd.isna(faculty_field) or str(faculty_field).strip().lower() in ['nan', 'none', '']:
        return "TBD"
    s = str(faculty_field).strip()
    for sep in ['/', ',', '&', ';']:
        if sep in s:
            return s.split(sep)[0].strip()
    return s

def get_course_priority(row):
    try:
        l = int(row.get('L', 0)) if pd.notna(row.get('L', 0)) else 0
        t = int(row.get('T', 0)) if pd.notna(row.get('T', 0)) else 0
        p = int(row.get('P', 0)) if pd.notna(row.get('P', 0)) else 0
        return -(l + t + p)
    except Exception:
        return 0

def calculate_required_minutes(course_row):
    l = int(course_row['L']) if ('L' in course_row and pd.notna(course_row['L'])) else 0
    t = int(course_row['T']) if ('T' in course_row and pd.notna(course_row['T'])) else 0
    p = int(course_row['P']) if ('P' in course_row and pd.notna(course_row['P'])) else 0
    s = int(course_row['S']) if ('S' in course_row and pd.notna(course_row['S'])) else 0
    c = int(course_row['C']) if ('C' in course_row and pd.notna(course_row['C'])) else 0
    is_half_semester = (c < 3 and c > 0)
    
    if is_half_semester:
        # For half semester, courses run for ~8 weeks, so divide by 2
        lec_count = l // 2 if l > 0 else 0
        tut_count = t // 2 if t > 0 else 0
        lab_count = p // 2 if p > 0 else 0
    else:
        # For full semester courses
        lec_count = l
        tut_count = t
        lab_count = p
    
    # S is for self-study, not scheduled
    return (lec_count, tut_count, lab_count, 0)  # Return 0 for S

def get_required_room_type(course_row):
    try:
        p = int(course_row['P']) if ('P' in course_row and pd.notna(course_row['P'])) else 0
        return 'COMPUTER_LAB' if p > 0 else 'LECTURE_ROOM'
    except Exception:
        return 'LECTURE_ROOM'

# ---------------------------
# Room allocation helpers (global room_schedule)
# ---------------------------
def find_suitable_room_for_slot(course_code, room_type, day, slot_indices, room_schedule, course_room_mapping):
    """
    Finds a suitable room for the given course and slots.
    Ensures:
    - Each course always gets the same room across all its sessions.
    - No two courses overlap in the same room at the same time.
    """
    # if this course already has a room, reuse it (if it's free for these slots)
    if course_code in course_room_mapping:
        fixed_room = course_room_mapping[course_code]
        for si in slot_indices:
            if si in room_schedule[fixed_room][day]:
                # room occupied -> cannot use, fail
                return None
        return fixed_room

    # else assign a new room (first time this course is scheduled)
    pool = computer_lab_rooms if room_type == 'COMPUTER_LAB' else lecture_rooms
    if not pool:
        return None
    random.shuffle(pool)
    for room in pool:
        if room not in room_schedule:
            room_schedule[room] = {d: set() for d in range(len(DAYS))}
        # check if room is free for all these slots
        if all(si not in room_schedule[room][day] for si in slot_indices):
            # assign this room permanently to this course
            course_room_mapping[course_code] = room
            return room
    return None

def find_consecutive_slots_for_minutes(timetable, day, start_idx, required_minutes,
                                       semester, professor_schedule, faculty,
                                       room_schedule, room_type, course_code, course_room_mapping):
    """
    Find consecutive TIME_SLOTS starting at start_idx whose total minutes >= required_minutes,
    respecting minor slots, breaks, existing timetable occupancy, professor schedule, and
    room availability. Returns (slot_indices, room) or (None, None).
    """
    n = len(TIME_SLOTS)
    slot_indices = []
    i = start_idx
    accumulated = 0

    # accumulate consecutive slots
    while i < n and accumulated < required_minutes:
        # can't schedule in minor slot or break
        if is_minor_slot(TIME_SLOTS[i]) or is_break_time_slot(TIME_SLOTS[i], semester):
            return None, None
        # slot already occupied in this timetable
        if timetable[day][i]['type'] is not None:
            return None, None
        # professor busy
        if faculty in professor_schedule and i in professor_schedule[faculty][day]:
            return None, None
        # if no rooms of required type exist at all, fail early
        if room_type == 'COMPUTER_LAB' and not computer_lab_rooms:
            return None, None
        if room_type == 'LECTURE_ROOM' and not lecture_rooms:
            return None, None

        slot_indices.append(i)
        accumulated += slot_minutes(TIME_SLOTS[i])
        i += 1

    # if we gathered enough minutes, find a room free across these slots
    if accumulated >= required_minutes:
        room = find_suitable_room_for_slot(course_code, room_type, day, slot_indices, room_schedule, course_room_mapping)
        if room is not None:
            return slot_indices, room

    return None, None

def get_all_possible_start_indices_for_duration():
    idxs = list(range(len(TIME_SLOTS)))
    random.shuffle(idxs)
    return idxs

def check_professor_availability(professor_schedule, faculty, day, start_idx, duration_slots):
    if faculty not in professor_schedule:
        return True
    if not professor_schedule[faculty][day]:
        return True
    new_start = TIME_SLOTS[start_idx][0]
    new_start_m = new_start.hour*60 + new_start.minute
    MIN_GAP = 90
    for s in professor_schedule[faculty][day]:
        exist_start = TIME_SLOTS[s][0]
        exist_m = exist_start.hour*60 + exist_start.minute
        if abs(exist_m - new_start_m) < MIN_GAP:
            return False
    return True

def load_rooms():
    return {'lecture_rooms': lecture_rooms, 'computer_lab_rooms': computer_lab_rooms, 'large_rooms': large_rooms}

def load_batch_data():
    batch_info = {}
    for _, r in df.iterrows():
        dept = str(r['Department'])
        sem = int(r['Semester'])
        batch_info[(dept, sem)] = {'num_sections': 1}
    return batch_info

# ---------------------------
# Main generation function
# ---------------------------
def record_unscheduled(unscheduled_dict, code, dept, sem, reason):
    """
    Adds an unscheduled course to a dictionary only once.
    Prevents duplicate entries and stores the reason.
    """
    if code not in unscheduled_dict:
        unscheduled_dict[code] = {
            "Course Code": code,
            "Department": dept,
            "Semester": sem,
            "Reason": reason
        }
def add_unscheduled_course(unscheduled_components, department, semester, code, name, faculty, comp_type, section, reason):
    """
    Prevent duplicate unscheduled entries for the same course.
    If the course already exists, append component info to its reason.
    """
    existing = next((u for u in unscheduled_components if u.code == code), None)
    if existing:
        if comp_type not in existing.component_type:
            existing.component_type += f", {comp_type}"
        if reason not in existing.reason:
            existing.reason += f"; {reason}"
    else:
        unscheduled_components.append(UnscheduledComponent(department, semester, code, name, faculty, comp_type, 1, section, reason))

def generate_all_timetables():
    global TIME_SLOTS
    unscheduled_dict = {}
    TIME_SLOTS = generate_time_slots()
    rooms = load_rooms()
    batch_info = load_batch_data()

    room_schedule = {}
    professor_schedule = {}
    course_room_mapping = {}

    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    overview = wb.create_sheet("Overview")
    overview.append(["Combined Timetable for All Departments and Semesters"])
    overview.append(["Generated on:", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    overview.append([])
    overview.append(["Department", "Semester", "Sheet Name"])
    row_index = 5

    unscheduled_components = []

    SUBJECT_COLORS = [
        "FF6B6B", "4ECDC4", "FF9F1C", "5D5FEF", "45B7D1",
        "F72585", "7209B7", "3A0CA3", "4361EE", "4CC9F0",
        "06D6A0", "FFD166", "EF476F", "118AB2", "073B4C"
    ]

    for department in df['Department'].unique():
        sems = sorted(df[df['Department'] == department]['Semester'].unique())
        for semester in sems:
# ---------------------------
# Section and Priority Rules
# ---------------------------

# Give 2 sections for CSE semesters 2, 4, 6

            dept_upper = str(department).strip().upper()
            num_sections = 2 if (dept_upper == "CSE" and int(semester) in [2, 4, 6]) else 1


            courses = df[(df['Department'] == department) & (df['Semester'] == semester)]
            if 'Schedule' in courses.columns:
                courses = courses[(courses['Schedule'].fillna('Yes').str.upper() == 'YES') | (courses['Schedule'].isna())]
            if courses.empty:
                continue

            # Split into lab and non-lab courses
            if 'P' in courses.columns:
                lab_courses = courses[courses['P'] > 0].copy()
                non_lab_courses = courses[courses['P'] == 0].copy()
            else:
                lab_courses = courses.head(0)
                non_lab_courses = courses.copy()

            # Priority by total workload (L + T + P)
            if not lab_courses.empty:
                lab_courses['priority'] = lab_courses.apply(get_course_priority, axis=1)
                lab_courses = lab_courses.sort_values('priority', ascending=False)
            non_lab_courses['priority'] = non_lab_courses.apply(get_course_priority, axis=1)
            non_lab_courses = non_lab_courses.sort_values('priority', ascending=False)

            # --- ELECTIVE PRIORITY LOGIC ---
            def is_elective(course_row):
                name = str(course_row.get('Course Name', '')).lower()
                code = str(course_row.get('Course Code', '')).lower()
                keywords = ["elective", "oe", "open elective", "pe", "program elective"]
                return any(k in name for k in keywords) or any(k in code for k in keywords)

            combined = pd.concat([lab_courses, non_lab_courses])
            combined['is_elective'] = combined.apply(is_elective, axis=1)

            # Electives first, then core — higher total workload within each group
            courses_combined = combined.sort_values(by=['is_elective', 'priority'], ascending=[False, False]).drop_duplicates()

            for section in range(num_sections):
                section_title = f"{department}_{semester}" if num_sections == 1 else f"{department}_{semester}_{chr(65 + section)}"
                ws = wb.create_sheet(title=section_title)

                overview.cell(row=row_index, column=1, value=department)
                overview.cell(row=row_index, column=2, value=str(semester))
                overview.cell(row=row_index, column=3, value=section_title)
                row_index += 1

                timetable = {d: {s: {'type': None, 'code': '', 'name': '', 'faculty': '', 'classroom': ''} for s in range(len(TIME_SLOTS))} for d in range(len(DAYS))}

                section_subject_color = {}
                color_iter = iter(SUBJECT_COLORS)
                course_faculty_map = {}

                for _, c in courses_combined.iterrows():
                    code = str(c.get('Course Code', '')).strip()
                    if code and code not in section_subject_color:
                        try:
                            section_subject_color[code] = next(color_iter)
                        except StopIteration:
                            section_subject_color[code] = random.choice(SUBJECT_COLORS)
                        course_faculty_map[code] = select_faculty(c.get('Faculty', 'TBD'))
                # --- PRIORITIZE ELECTIVES FIRST ---
                def is_elective(course_name):
                    if pd.isna(course_name):
                        return False
                    name = str(course_name).lower()
                    keywords = ["elective", "oe", "open elective", "pe", "program elective"]
                    return any(k in name for k in keywords)

                # Separate electives and non-electives
                elective_courses = courses_combined[courses_combined['Course Name'].apply(is_elective)]
                core_courses = courses_combined[~courses_combined['Course Name'].apply(is_elective)]

                # Recombine — electives first, then core
                courses_combined = pd.concat([elective_courses, core_courses])

                for _, course in courses_combined.iterrows():
                    code = str(course.get('Course Code', '')).strip()
                    name = str(course.get('Course Name', '')).strip()
                    faculty = select_faculty(course.get('Faculty', 'TBD'))

                    if faculty not in professor_schedule:
                        professor_schedule[faculty] = {d: set() for d in range(len(DAYS))}

                    lec_count, tut_count, lab_count, ss_count = calculate_required_minutes(course)
                    lec_sessions_needed = int(lec_count * 60 / LECTURE_MIN) if lec_count > 0 else 0
                    tut_sessions_needed = int(tut_count * 60 / TUTORIAL_MIN) if tut_count > 0 else 0
                    lab_sessions_needed = int(lab_count * 60 / LAB_MIN) if lab_count > 0 else 0

                    room_type = get_required_room_type(course)

                    def schedule_component(required_minutes, comp_type, attempts_limit=800):
                        for attempt in range(attempts_limit):
                            day = random.randint(0, len(DAYS)-1)
                            starts = get_all_possible_start_indices_for_duration()
                            for start_idx in starts:
                                slot_indices, candidate_room = find_consecutive_slots_for_minutes(
    timetable, day, start_idx, required_minutes, semester,
    professor_schedule, faculty, room_schedule, room_type,
    code, course_room_mapping
)

                                if slot_indices is None:
                                    continue
                                if not check_professor_availability(professor_schedule, faculty, day, slot_indices[0], len(slot_indices)):
                                    continue
                                if candidate_room is None:
                                    continue
                                for si_idx, si in enumerate(slot_indices):
                                    timetable[day][si]['type'] = 'LEC' if comp_type == 'LEC' else ('LAB' if comp_type == 'LAB' else ('TUT' if comp_type == 'TUT' else 'SS'))
                                    timetable[day][si]['code'] = code if si_idx == 0 else ''
                                    timetable[day][si]['name'] = name if si_idx == 0 else ''
                                    timetable[day][si]['faculty'] = faculty if si_idx == 0 else ''
                                    timetable[day][si]['classroom'] = candidate_room if si_idx == 0 else ''
                                    professor_schedule[faculty][day].add(si)
                                    if candidate_room not in room_schedule:
                                        room_schedule[candidate_room] = {d: set() for d in range(len(DAYS))}
                                    room_schedule[candidate_room][day].add(si)
                                return True
                        return False

                    for _ in range(lec_sessions_needed):
                        ok = schedule_component(LECTURE_MIN, 'LEC', attempts_limit=5000)
                        if not ok:
                            add_unscheduled_course(unscheduled_components, department, semester, code, name, faculty, 'LEC', section, "Number of collisions exceeded limit")

                    for _ in range(tut_sessions_needed):
                        ok = schedule_component(TUTORIAL_MIN, 'TUT', attempts_limit=5000)
                        if not ok:
                            add_unscheduled_course(unscheduled_components, department, semester, code, name, faculty, 'TUT', section, "No slot available")

                    for _ in range(lab_sessions_needed):
                        ok = schedule_component(LAB_MIN, 'LAB', attempts_limit=5000)
                        if not ok:
                            add_unscheduled_course(unscheduled_components, department, semester, code, name, faculty, 'LAB', section, "Lab not scheduled (slot unavailable)")


                # Write sheet
                header = ['Day'] + [f"{slot[0].strftime('%H:%M')}-{slot[1].strftime('%H:%M')}" for slot in TIME_SLOTS]
                ws.append(header)
                header_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
                header_font = Font(bold=True)
                header_alignment = Alignment(horizontal='center', vertical='center')
                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = header_alignment

                border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                lec_fill_default = PatternFill(start_color="FA8072", end_color="FA8072", fill_type="solid")
                lab_fill_default = PatternFill(start_color="7CFC00", end_color="7CFC00", fill_type="solid")
                tut_fill_default = PatternFill(start_color="87CEEB", end_color="87CEEB", fill_type="solid")
                ss_fill_default = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
                break_fill = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")
                minor_fill = PatternFill(start_color="9ACD32", end_color="9ACD32", fill_type="solid")

                for day_idx, day_name in enumerate(DAYS):
                    ws.append([day_name] + [''] * len(TIME_SLOTS))
                    row_num = ws.max_row
                    merges = []
                    for slot_idx in range(len(TIME_SLOTS)):
                        cell_obj = ws.cell(row=row_num, column=slot_idx + 2)
                        if is_minor_slot(TIME_SLOTS[slot_idx]):
                            cell_obj.value = "Minor Slot"
                            cell_obj.fill = minor_fill
                            cell_obj.font = Font(bold=True)
                            cell_obj.alignment = Alignment(horizontal='center', vertical='center')
                            cell_obj.border = border
                            continue

                        if is_break_time_slot(TIME_SLOTS[slot_idx], semester):
                            cell_obj.value = "BREAK"
                            cell_obj.fill = break_fill
                            cell_obj.font = Font(bold=True)
                            cell_obj.alignment = Alignment(horizontal='center', vertical='center')
                            cell_obj.border = border
                            continue

                        if timetable[day_idx][slot_idx]['type'] is None:
                            cell_obj.border = border
                            continue

                        typ = timetable[day_idx][slot_idx]['type']
                        code = timetable[day_idx][slot_idx]['code']
                        cls = timetable[day_idx][slot_idx]['classroom']
                        fac = timetable[day_idx][slot_idx]['faculty']

                        if code:
                            span = [slot_idx]
                            j = slot_idx + 1
                            while j < len(TIME_SLOTS) and timetable[day_idx][j]['type'] is not None and timetable[day_idx][j]['code'] == '':
                                span.append(j)
                                j += 1
                            display = f"{typ}\nroom no. :{cls}\n{fac}"

                            if code in section_subject_color:
                                subj_color = section_subject_color[code]
                                fill = PatternFill(start_color=subj_color, end_color=subj_color, fill_type="solid")
                            else:
                                fill = {'LEC': lec_fill_default, 'LAB': lab_fill_default, 'TUT': tut_fill_default, 'SS': ss_fill_default}.get(typ, lec_fill_default)

                            cell_obj.value = display
                            cell_obj.fill = fill
                            merges.append((slot_idx + 2, slot_idx + 2 + len(span) - 1, display, fill))
                        cell_obj.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
                        cell_obj.border = border

                    for start_col, end_col, val, fill in merges:
                        if end_col > start_col:
                            rng = f"{get_column_letter(start_col)}{row_num}:{get_column_letter(end_col)}{row_num}"
                            try:
                                ws.merge_cells(rng)
                                mc = ws[f"{get_column_letter(start_col)}{row_num}"]
                                mc.value = val
                                mc.fill = fill
                                mc.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
                                mc.border = border
                            except Exception:
                                pass

                for col_idx in range(1, len(TIME_SLOTS)+2):
                    try:
                        ws.column_dimensions[get_column_letter(col_idx)].width = 15
                    except Exception:
                        pass

                for row in ws.iter_rows(min_row=2, max_row=len(DAYS)+1):
                    ws.row_dimensions[row[0].row].height = 40

                # Add Self-Study Only Courses section
                current_row = len(DAYS) + 4  # Initialize current_row here, before any sections

                # Build a list of self-study-only courses for this dept/sem
                ss_courses_for_this_section = []
                for _, course in courses_combined.iterrows():
                    l = int(course['L']) if pd.notna(course['L']) else 0
                    t = int(course['T']) if pd.notna(course['T']) else 0
                    p = int(course['P']) if pd.notna(course['P']) else 0
                    s = int(course['S']) if pd.notna(course['S']) else 0
                    if s > 0 and l == 0 and t == 0 and p == 0:
                        ss_courses_for_this_section.append({
                            'code': str(course['Course Code']),
                            'name': str(course['Course Name']),
                            'faculty': str(course['Faculty'])
                        })

                if ss_courses_for_this_section:
                    ws.cell(row=current_row, column=1, value="Self-Study Only Courses")
                    ws.cell(row=current_row, column=1).font = Font(bold=True)
                    current_row += 1

                    headers = ['Course Code', 'Course Name', 'Faculty']
                    for col, header in enumerate(headers, 1):
                        ws.cell(row=current_row, column=col, value=header)
                        ws.cell(row=current_row, column=col).font = Font(bold=True)
                    current_row += 1

                    for course in ss_courses_for_this_section:
                        ws.cell(row=current_row, column=1, value=course['code'])
                        ws.cell(row=current_row, column=2, value=course['name'])
                        ws.cell(row=current_row, column=3, value=course['faculty'])
                        current_row += 1

                    current_row += 2  # Add extra spacing after self-study courses

                

                # Improved legend formatting
                legend_title = ws.cell(row=current_row, column=1, value="Legend")
                legend_title.font = Font(bold=True, size=12)
                current_row += 2

                # Wider columns for legend
                ws.column_dimensions['A'].width = 20  # Subject Code
                ws.column_dimensions['B'].width = 10  # Color (moved next to code)
                ws.column_dimensions['C'].width = 40  # Subject Name
                ws.column_dimensions['D'].width = 30  # Faculty
                ws.column_dimensions['E'].width = 15  # LTPS
                ws.column_dimensions['F'].width = 15  # Room

                # Add legend headers with better formatting
                # Add legend headers with better formatting (added Room column)
                legend_headers = ['Subject Code', 'Color', 'Subject Name', 'Faculty', 'LTPS', 'Room']
                for col, header in enumerate(legend_headers, 1):
                    cell = ws.cell(row=current_row, column=col, value=header)
                    cell.font = Font(bold=True)
                    cell.border = border
                    cell.fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                current_row += 1

                # Add subject entries with improved spacing and color next to code
                for code, color in section_subject_color.items():
    # Skip courses that have no assigned room
                    assigned_room = course_room_mapping.get(code, "—")
                    if not assigned_room or assigned_room == "—":
                        continue  # ❌ Skip this entry

                    ws.row_dimensions[current_row].height = 30

                    # Get LTPS values for this course
                    ltps_value = ""
                    for _, course_row in courses_combined.iterrows():
                        if str(course_row['Course Code']) == code:
                            l = str(int(course_row['L'])) if pd.notna(course_row['L']) else "0"
                            t = str(int(course_row['T'])) if pd.notna(course_row['T']) else "0"
                            p = str(int(course_row['P'])) if pd.notna(course_row['P']) else "0"
                            s = str(int(course_row['S'])) if pd.notna(course_row['S']) and 'S' in course_row else "0"
                            ltps_value = f"{l}-{t}-{p}-{s}"
                            break

                    course_name = ''
                    fac_name = ''
                    if code in course_faculty_map:
                        fac_name = course_faculty_map[code]
                        # find name from courses_combined
                        for _, cr in courses_combined.iterrows():
                            if str(cr['Course Code']) == code:
                                course_name = str(cr['Course Name'])
                                break

                    cells = [
                        (code, None),
                        ('', PatternFill(start_color=color, end_color=color, fill_type="solid")),
                        (course_name, None),
                        (fac_name, None),
                        (ltps_value, None),
                        (assigned_room, None)
                    ]

                    for col, (value, fill) in enumerate(cells, 1):
                        cell = ws.cell(row=current_row, column=col, value=value)
                        cell.border = border
                        if fill:
                            cell.fill = fill
                        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True, indent=2)

                    current_row += 1


    # Format the overview sheet
    for col in range(1, 4):
        overview.column_dimensions[get_column_letter(col)].width = 20

    for row_ in overview.iter_rows(min_row=1, max_row=4):
        for cell in row_:
            cell.font = Font(bold=True)

    # Apply formatting to the overview table headers
    for cell in overview[4]:
        cell.fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
        cell.font = Font(bold=True)
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                           top=Side(style='thin'), bottom=Side(style='thin'))

    # Apply borders to the overview data
    for row_ in overview.iter_rows(min_row=5, max_row=row_index-1):
        for cell in row_:
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                               top=Side(style='thin'), bottom=Side(style='thin'))

    # Save the workbook
    out_filename = os.path.join(OUTPUT_DIR, "timetable_all_departments.xlsx")
    try:
        wb.save(out_filename)
        print(f"Combined timetable for all departments and semesters saved as {out_filename}")
    except Exception as e:
        print(f"Failed to save combined timetable: {e}")
        traceback.print_exc()

    # After saving combined workbook, create teacher and unscheduled Excels
    try:
        create_teacher_and_unscheduled_from_combined(out_filename, unscheduled_components)
    except Exception as e:
        print("Failed to generate teacher/unscheduled workbooks:", e)
        traceback.print_exc()

    return out_filename

# ---------------------------
# Teacher + Unscheduled helper
# ---------------------------
def split_faculty_names(fac_str):
    """Split faculty string into separate names using common separators."""
    if fac_str is None:
        return []
    s = str(fac_str).strip()
    if s == '' or s.lower() in ['nan', 'none']:
        return []
    # treat common separators
    parts = [s]
    for sep in ['/', ',', '&', ';']:
        if sep in s:
            parts = [p.strip() for p in s.split(sep) if p.strip()]
            break
    return parts if parts else [s]

def parse_cell_for_course(cell_value):
    """
    Expected formats in timetable cells (examples):
      "CS101 LEC\nroom no. :A101\nDr. Name"
      "B1 Courses\nCS101, CS102\nCS101: Dr. X (A101)\nCS102: Dr. Y (A102)"
    This function tries to extract:
      code (first token), type (LEC/TUT/LAB/SS), room (if 'room no.' present), faculty (last line)
    Returns tuple (code, typ, room, faculty) — many may be '' if not found.
    """
    if cell_value is None:
        return ('', '', '', '')
    text = str(cell_value).strip()
    if text == '':
        return ('', '', '', '')

    lines = [ln.strip() for ln in text.splitlines() if ln.strip()]
    faculty = ''
    room = ''
    code = ''
    typ = ''

    # detect 'room no.' pattern anywhere
    for ln in lines:
        if 'room no' in ln.lower():
            # attempt to extract room after colon or :
            parts = ln.split(':')
            if len(parts) >= 2:
                room = parts[-1].strip()
            else:
                room = ln.strip()

    # last line probably faculty if it doesn't include 'room' or 'Courses' keywords
    if len(lines) >= 1:
        last = lines[-1]
        if 'room no' not in last.lower() and 'courses' not in last.lower() and ':' not in last:
            faculty = last

    # Try extracting code and type from the first line if it looks like "CODE TYPE"
    first = lines[0] if lines else ''
    if first:
        tokens = first.split()
        if len(tokens) >= 2 and tokens[1].upper() in ['LEC', 'LAB', 'TUT', 'SS']:
            code = tokens[0].strip()
            typ = tokens[1].strip().upper()
        else:
            # fallback: take first token as code
            code = tokens[0].strip() if tokens else ''
            # try to detect type anywhere
            for t in ['LEC', 'LAB', 'TUT', 'SS']:
                if t in text.upper():
                    typ = t
                    break

    # If faculty still empty try to find a token pattern in other lines
    if not faculty and len(lines) >= 2:
        for cand in lines[1:]:
            if any(ch.isalpha() for ch in cand) and 'room no' not in cand.lower() and 'courses' not in cand.lower() and ':' not in cand:
                faculty = cand
                break

    return (code, typ, room, faculty)

def create_teacher_and_unscheduled_from_combined(timetable_filename, unscheduled_components):
    """
    Builds teacher_timetables.xlsx with clean, formatted sheets per teacher.
    Also writes unscheduled_courses.xlsx.
    """
    try:
        wb = load_workbook(timetable_filename, data_only=True)
    except Exception as e:
        print(f"Failed to open {timetable_filename}: {e}")
        return

    teacher_slots = {}
    slot_headers = []

    for sheetname in wb.sheetnames:
        if sheetname.lower() == 'overview':
            continue
        ws = wb[sheetname]
        header = [str(ws.cell(1, c).value).strip() if ws.cell(1, c).value else '' for c in range(2, ws.max_column + 1)]
        if len(header) > len(slot_headers):
            slot_headers = header
        for r in range(2, ws.max_row + 1):
            day = ws.cell(r, 1).value
            if not day or str(day) not in DAYS:
                break
            day_idx = DAYS.index(day)
            for c in range(2, ws.max_column + 1):
                code, typ, room, faculty = parse_cell_for_course(ws.cell(r, c).value)
                for f in split_faculty_names(faculty):
                    if not f:
                        continue
                    # Skip any fake or unwanted teacher names
                    if str(f).strip().upper() in ["BREAK", "MINOR SLOT", "NAN", "NONE", ""]:
                        continue

                    teacher_slots.setdefault(f, {d: {i: '' for i in range(len(slot_headers))} for d in range(len(DAYS))})
                    teacher_slots[f][day_idx][c - 2] = f"{code} {typ}\n({sheetname})\nRoom: {room}" if code else ''

    # --- Create formatted teacher workbook ---
    twb = Workbook()
    if "Sheet" in twb.sheetnames:
        twb.remove(twb["Sheet"])

    # Define styles
    header_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
    alt_fill = PatternFill(start_color="FFF8DC", end_color="FFF8DC", fill_type="solid")
    border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
    header_font = Font(bold=True, size=12)
    title_font = Font(bold=True, size=14)
    cell_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for teacher in sorted(teacher_slots.keys()):
        safe_name = teacher[:31] or "Unknown"
        ws = twb.create_sheet(title=safe_name)

        # Add title row
        ws.merge_cells("A1:{}1".format(get_column_letter(len(slot_headers) + 1)))
        title_cell = ws.cell(row=1, column=1, value=f"{teacher} — Weekly Timetable")
        title_cell.font = title_font
        title_cell.alignment = Alignment(horizontal="center", vertical="center")

        # Header row
        ws.append(["Day"] + slot_headers)
        for cell in ws[2]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

        # Fill timetable rows
        for d, day in enumerate(DAYS):
            row = [day] + [teacher_slots[teacher][d][i] for i in range(len(slot_headers))]
            ws.append(row)
            row_idx = ws.max_row
            # alternating day color
            if d % 2 == 0:
                for cell in ws[row_idx]:
                    cell.fill = alt_fill
            for cell in ws[row_idx]:
                cell.alignment = cell_align
                cell.border = border
            ws.row_dimensions[row_idx].height = 35

        # Adjust column widths
        ws.column_dimensions["A"].width = 15
        for col in range(2, len(slot_headers) + 2):
            ws.column_dimensions[get_column_letter(col)].width = 20

    twb.save(os.path.join(OUTPUT_DIR, "teacher_timetables.xlsx"))
    print("✅ Saved formatted teacher_timetables.xlsx")

    # --- Unscheduled Courses workbook (unchanged) ---
    # --- Improved Unscheduled Courses Workbook (Unique + Reason) ---
    # --- Improved Unscheduled Courses Workbook (Unique + Dynamic Reason) ---
    uwb = Workbook()
    ws = uwb.active
    ws.title = "Unscheduled Courses"

    headers = ["Course Code", "Department", "Semester", "Reason"]
    ws.append(headers)

    # Dictionary to avoid duplicates
    unscheduled_unique = {}

    for u in unscheduled_components:
        if u.code not in unscheduled_unique:
            # --- Detect Reason ---
            if hasattr(u, "reason") and u.reason and len(str(u.reason).strip()) > 0:
                reason_text = str(u.reason).strip()
            else:
                # Intelligent guess based on context
                if "collision" in str(u.component_type).lower():
                    reason_text = "Number of collisions exceeded limit"
                elif "slot" in str(u.component_type).lower() or "no slot" in str(u.reason).lower():
                    reason_text = "No slot available"
                elif "faculty" in str(u.reason).lower():
                    reason_text = "Faculty unavailable"
                else:
                    reason_text = "Unspecified scheduling issue"

            # Store one entry per course
            unscheduled_unique[u.code] = {
                "Course Code": u.code,
                "Department": u.department,
                "Semester": u.semester,
                "Reason": reason_text
            }

    # Write data
    for entry in unscheduled_unique.values():
        ws.append([entry[h] for h in headers])

    uwb.save(os.path.join(OUTPUT_DIR, "unscheduled_courses.xlsx"))
    print(f"✅ Saved unscheduled_courses.xlsx with {len(unscheduled_unique)} unique courses and detailed reasons")


def allocate_exam_rooms(course_code, students, date_str, df_rooms, date_room_usage):
    """
    Allocate rooms for an exam based on number of students and room capacities.
    Ensures:
    - No room clashes on same day
    - Uses smallest number of rooms to fit all students
    - Chooses rooms with minimal extra capacity
    """
    assigned_rooms = []
    total_cap = 0

    # Filter available rooms (not used on this date)
    available = df_rooms[~df_rooms["roomNumt"].isin(date_room_usage[date_str])].copy()
    available = available.sort_values(by="capacity", ascending=True)

    # Step 1: try to find smallest single room that fits all students
    suitable = available[available["capacity"] >= students]
    if not suitable.empty:
        best = suitable.iloc[0]  # smallest room that fits
        assigned_rooms = [best["roomNumt"]]
        total_cap = best["capacity"]
        date_room_usage[date_str].add(best["roomNumt"])
    else:
        # Step 2: combine multiple smaller rooms until enough capacity
        remaining = students
        for _, r in available.iterrows():
            assigned_rooms.append(r["roomNumt"])
            total_cap += r["capacity"]
            date_room_usage[date_str].add(r["roomNumt"])
            remaining -= r["capacity"]
            if remaining <= 0:
                break
        if remaining > 0:
            print(f"⚠️ Not enough total room capacity for {course_code} ({students} students). Assigned all available rooms.")

    return assigned_rooms

# ---------------------------
# Exam generator and invigilation sheet
# ---------------------------
def exam_generator():
    global INPUT_DIR
    """
    Generates an exam timetable with:
    - Room assignment strictly by capacity (closest fit)
    - No room clashes for the same day
    - Automatic split for large courses
    - Smart room reuse across days
    - Invigilation schedule with course + room + faculty
    """
    exam_file = "Exam_timetable.xlsx"

    try:
        df_courses = pd.read_csv(os.path.join(INPUT_DIR, "combined.csv"))
        df_rooms = pd.read_csv(os.path.join(INPUT_DIR, "rooms.csv"))
    except FileNotFoundError as e:
        print(f"❌ Missing file: {e}")
        return None

    # --- Clean and validate ---
    df_courses = df_courses.dropna(subset=["Course Code", "Course Name", "Faculty", "Department", "Semester"])
    if "total_students" not in df_courses.columns:
        df_courses["total_students"] = 50
    df_courses["total_students"] = df_courses["total_students"].fillna(50).astype(int)

    # Normalize rooms.csv headers
    df_rooms.columns = [c.strip().lower() for c in df_rooms.columns]

    def find_col(keywords):
        for c in df_rooms.columns:
            if any(k in c for k in keywords):
                return c
        return None

    room_col = find_col(["room", "num", "id"])
    cap_col = find_col(["cap", "seat"])
    type_col = find_col(["type"])

    if not room_col or not cap_col:
        print("❌ rooms.csv must have columns for room number and capacity.")
        print("Detected columns:", df_rooms.columns)
        return None

    df_rooms = df_rooms.rename(columns={room_col: "room", cap_col: "capacity"})
    if type_col:
        df_rooms = df_rooms.rename(columns={type_col: "type"})
    else:
        df_rooms["type"] = "LECTURE_ROOM"

    df_rooms["room"] = df_rooms["room"].astype(str).str.strip()
    df_rooms["capacity"] = pd.to_numeric(df_rooms["capacity"], errors="coerce").fillna(0).astype(int)
    df_rooms = df_rooms[df_rooms["capacity"] > 0].sort_values(by="capacity").reset_index(drop=True)

    print(f"✅ Loaded {len(df_rooms)} rooms for exam scheduling")

    # --- Faculty list ---
    faculty_list = list(set(sum([str(f).replace(" and ", "/").replace(",", "/").split("/") for f in df_courses["Faculty"]], [])))
    faculty_list = [f.strip() for f in faculty_list if f.strip()]

    # --- Dates setup ---
    session_title = "Jan-April 03:00 PM to 04:30 PM"
    start_date = datetime(2025, 11, 20)
    num_days = min(10, len(df_courses))
    dates = [start_date + timedelta(days=i) for i in range(num_days)]
    days = [d.strftime("%A") for d in dates]

    shuffled = df_courses.sample(frac=1, random_state=42).reset_index(drop=True)
    course_date_map = {row["Course Code"]: dates[i % len(dates)] for i, row in shuffled.iterrows()}

    # --- Track rooms used per date ---
    date_room_usage = {d.strftime("%d-%b-%Y"): set() for d in dates}
    invigilation_entries = []

    for date in dates:
        date_str = date.strftime("%d-%b-%Y")
        today_courses = shuffled[shuffled["Course Code"].isin(
            [c for c, dt in course_date_map.items() if dt == date]
        )]

        for _, course in today_courses.iterrows():
            code = course["Course Code"]
            name = course["Course Name"]
            dept = course["Department"]
            sem = course["Semester"]
            teacher = str(course["Faculty"]).strip()
            students = int(course["total_students"])
            time_slot = "03:00 PM–04:30 PM"

            assigned_rooms = []
            remaining = students

            # --- Select rooms based on capacity ---
            available = df_rooms[~df_rooms["room"].isin(date_room_usage[date_str])].copy()
            available = available.sort_values(by="capacity", ascending=True)

            # Try exact fit
            suitable = available[available["capacity"] >= remaining]
            if not suitable.empty:
                best = suitable.iloc[0]
                assigned_rooms = [best["room"]]
                date_room_usage[date_str].add(best["room"])
            else:
                # Use multiple smaller rooms
                total_cap = 0
                for _, room_row in available.iterrows():
                    assigned_rooms.append(room_row["room"])
                    total_cap += room_row["capacity"]
                    date_room_usage[date_str].add(room_row["room"])
                    if total_cap >= remaining:
                        break
                if total_cap < remaining:
                    print(f"⚠️ Not enough total capacity for {code} ({students} students). Assigned all available rooms.")

            # --- Assign invigilators ---
            available_teachers = [f for f in faculty_list if f.lower() not in teacher.lower()]
            for room in assigned_rooms:
                invigilator = random.choice(available_teachers) if available_teachers else "TBD"
                invigilation_entries.append({
                    "Faculty": invigilator,
                    "Date": date_str,
                    "Time": time_slot,
                    "Course Code": code,
                    "Course Name": name,
                    "Department": dept,
                    "Semester": sem,
                    "Room": room,
                    "Strength": students
                })

    # --- Excel output ---
    wb = Workbook()
    ws = wb.active
    ws.title = "Exam Timetable"

    bold_center = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    header_fill = PatternFill(start_color="BFBFBF", end_color="BFBFBF", fill_type="solid")
    border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(dates)+1)
    title = ws.cell(row=1, column=1, value=session_title)
    title.font = Font(bold=True, size=14)
    title.alignment = center
    title.fill = header_fill
    title.border = border

    ws.cell(row=2, column=1, value="Date").font = bold_center
    for i, d in enumerate(dates):
        c = ws.cell(row=2, column=i+2, value=d.strftime("%d-%b-%Y"))
        c.font = bold_center
        c.alignment = center
        c.fill = header_fill
        c.border = border

    ws.cell(row=3, column=1, value="Days").font = bold_center
    for i, day in enumerate(days):
        c = ws.cell(row=3, column=i+2, value=day)
        c.font = bold_center
        c.alignment = center
        c.fill = header_fill
        c.border = border

    grouped_by_date = {}
    for e in invigilation_entries:
        grouped_by_date.setdefault(e["Date"], set()).add(e["Course Code"])
    max_rows = max(len(v) for v in grouped_by_date.values())

    for r in range(max_rows):
        for i, d in enumerate(dates):
            code_list = list(grouped_by_date.get(d.strftime("%d-%b-%Y"), []))
            val = code_list[r] if r < len(code_list) else ""
            cell = ws.cell(row=r+4, column=i+2, value=val)
            cell.alignment = center
            cell.border = border

    ws.column_dimensions["A"].width = 15
    for col in range(2, len(dates)+2):
        ws.column_dimensions[get_column_letter(col)].width = 16

    # Invigilation schedule
    ws2 = wb.create_sheet("Exam Invigilation Schedule")
    headers = ["Faculty", "Date", "Time", "Course Code", "Course Name", "Department", "Semester", "Room", "Strength"]
    ws2.append(headers)

    for i, h in enumerate(headers, 1):
        cell = ws2.cell(row=1, column=i, value=h)
        cell.font = bold_center
        cell.alignment = center
        cell.fill = header_fill
        cell.border = border

    for entry in invigilation_entries:
        ws2.append([entry[h] for h in headers])

    for col in range(1, len(headers)+1):
        ws2.column_dimensions[get_column_letter(col)].width = 22
    for r in range(2, ws2.max_row+1):
        for c in ws2[r]:
            c.alignment = center
            c.border = border

    exam_file = os.path.join(OUTPUT_DIR, "Exam_timetable.xlsx")
    wb.save(exam_file)
    print(f"✅ Exam timetable and invigilation schedule saved → {exam_file}")
    return exam_file


# ---------------------------
# Run main if executed
# ---------------------------
if __name__ == "__main__":
    try:
        # Generate all class and teacher timetables
        generate_all_timetables()

        # Generate exam timetable (includes invigilation schedule automatically)
        exam_generator()

    except Exception as e:
        print("Error running TT_gen:", e)
        traceback.print_exc()