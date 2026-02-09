import pandas as pd
import random
from datetime import datetime, time, timedelta
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter
from dataclasses import dataclass
from pathlib import Path
import traceback
import os
import json

# ---------------------------
# Configuration
# ---------------------------
BASE_DIR = Path(__file__).resolve().parent
PROJECT_DIR = BASE_DIR.parent
DATA_DIR = BASE_DIR.parent / "data"
CONFIG_PATH = DATA_DIR / "config.json"

try:
    with open(CONFIG_PATH, "r", encoding="utf-8") as f:
        config = json.load(f)
except Exception:
    config = {}

DAYS = config.get("days", ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday'])
LECTURE_MIN = config.get("LECTURE_MIN", 90)
LAB_MIN = config.get("LAB_MIN", 120)
TUTORIAL_MIN = config.get("TUTORIAL_MIN", 60)
SELF_STUDY_MIN = config.get("SELF_STUDY_MIN", 60)
MIN_GAP_BETWEEN_LECTURES = 10

# Updated break windows - REMOVED morning break, extended lunch
LUNCH_BREAK_START = time(13, 15)
LUNCH_BREAK_END = time(14, 0)

@dataclass
class UnscheduledComponent:
    department: str
    semester: int
    code: str
    name: str
    faculty: str
    component_type: str
    sessions: int
    section: int
    reason: str

INPUT_DIR = PROJECT_DIR / "data"
OUTPUT_DIR = PROJECT_DIR / "output"

os.makedirs(OUTPUT_DIR, exist_ok=True)

# Meta entries collected while writing timetables so teacher generator can
# pick up faculty/room info even when cells hide them for basket display.
META_ENTRIES = []

# ---------------------------
# Load CSVs
# ---------------------------
try:
    df = pd.read_csv(os.path.join(INPUT_DIR, 'combined.csv'))
    print(f"✅ Loaded {len(df)} courses from combined.csv")
except FileNotFoundError:
    raise SystemExit("Error: 'combined.csv' not found in data directory.")

# --- FIX: Ensure 'total_students' column exists ---
if "total_students" not in df.columns:
    print("⚠️ Warning: 'total_students' column not in combined.csv. Defaulting to 50.")
    df["total_students"] = 50
df["total_students"] = pd.to_numeric(df["total_students"], errors='coerce').fillna(50).astype(int)
print(f"✅ Processed student strengths (defaulting to 50)")
# --- END FIX ---

try:
    rooms_df = pd.read_csv(os.path.join(INPUT_DIR, 'rooms.csv'))
    print(f"✅ Loaded {len(rooms_df)} rooms from rooms.csv")
except FileNotFoundError:
    print("⚠️ Warning: rooms.csv not found. Using empty rooms list.")
    rooms_df = pd.DataFrame(columns=['roomNumber', 'type', 'capacity'])

# --- FIX: Process and store room capacities ---
if 'capacity' not in rooms_df.columns:
    print("⚠️ Warning: 'capacity' column not in rooms.csv. Defaulting all rooms to 50.")
    rooms_df['capacity'] = 50
if 'type' not in rooms_df.columns:
    rooms_df['type'] = 'LECTURE_ROOM' # Default type

rooms_df['capacity'] = pd.to_numeric(rooms_df['capacity'], errors='coerce').fillna(50).astype(int)
rooms_df['type'] = rooms_df['type'].str.upper()
rooms_df['roomNumber'] = rooms_df['roomNumber'].astype(str)

# Create a global lookup dictionary
ROOM_DATA = {}
for _, row in rooms_df.iterrows():
    ROOM_DATA[row['roomNumber']] = {
        'type': row['type'],
        'capacity': row['capacity']
    }

print(f"\n📊 Room inventory: {len(ROOM_DATA)} rooms loaded with capacity data.")
room_types_summary = rooms_df.groupby('type').size().to_dict()
for rtype, count in room_types_summary.items():
    print(f"   - {rtype}: {count}")
# --- END FIX ---


# ---------------------------
# Time Slot Generation - UPDATED
# ---------------------------
def generate_time_slots():
    """
    Generate continuous 60 and 30 minute time slots for flexibility.
    - NO morning break (removed 10:30-10:45)
    - Lunch break: 13:15-14:00
    """
    slots = []
    
    slots = []
    
    slots.append((time(7, 30), time(9, 00)))    # 60 min
    # slots.append((time(9, 30), time(9, 30)))    # 60 min
    slots.append((time(9, 00), time(10, 00)))  # 60 min
    slots.append((time(10, 00), time(10, 30))) # 60 min
    slots.append((time(10, 30), time(10, 45))) # 60 min
    slots.append((time(10, 45), time(11, 00)))    # 30 min
    slots.append((time(11, 00), time(11, 30)))
    slots.append((time(11, 30), time(12, 00)))
    slots.append((time(12, 00), time(12, 15)))
    slots.append((time(12, 15), time(12, 30)))
    slots.append((time(12, 30), time(13, 15)))
    # Lunch break: 13:15 - 14:00
    slots.append((time(13, 15), time(14, 0)))    # BREAK
    
    # Afternoon session: 14:00 - 18:30 (continuous)
    slots.append((time(14, 0), time(14, 30)))    # 60 min
    slots.append((time(14, 30), time(15, 30)))    # 60 min
    slots.append((time(15, 30), time(15, 40)))    # 60 min
    slots.append((time(15, 40), time(16, 00)))    # 60 min
    slots.append((time(16, 00), time(16, 30)))  # 30 min
    slots.append((time(16, 30), time(17, 10)))
    slots.append((time(17, 10), time(17, 30)))
    slots.append((time(17, 30), time(18, 30)))  # 60/60 min
    # Evening minor slot added: 18:30 - 20:00
    slots.append((time(18, 30), time(20, 0)))

    
    return slots

TIME_SLOTS = generate_time_slots()
print(f"⏰ Generated {len(TIME_SLOTS)} time slots (no morning break, lunch 13:15-14:00)")

# ---------------------------
# Helper functions
# ---------------------------
def slot_minutes(slot):
    s, e = slot
    s_m = s.hour * 60 + s.minute
    e_m = e.hour * 60 + e.minute
    if e_m < s_m:
        e_m += 24 * 60
    return e_m - s_m

def overlaps(a_start, a_end, b_start, b_end):
    a_s_min = a_start.hour * 60 + a_start.minute
    a_e_min = a_end.hour * 60 + a_end.minute
    b_s_min = b_start.hour * 60 + b_start.minute
    b_e_min = b_end.hour * 60 + b_end.minute
    return (a_s_min < b_e_min) and (b_s_min < a_e_min)

def is_break_time_slot(slot, semester=None):
    """Updated to only check lunch break (13:00-14:00)"""
    start, end = slot
    # Only lunch break now
    if start == LUNCH_BREAK_START and end == LUNCH_BREAK_END:
        return True
    if overlaps(start, end, LUNCH_BREAK_START, LUNCH_BREAK_END):
        return True
    return False

def is_minor_slot(slot):
    """Check if slot is early morning or late evening"""
    start, end = slot
    if start.hour < 8:
        return True
    
    # Updated check for 18:30 (6:30 PM) or later
    if start.hour > 18 or (start.hour == 18 and start.minute >= 30):
        return True
    
    return False

def select_faculty(faculty_field):
    if pd.isna(faculty_field) or str(faculty_field).strip().lower() in ['nan', 'none', '']:
        return "TBD"
    s = str(faculty_field).strip()
    for sep in ['/', ',', '&', ';']:
        if sep in s:
            return s.split(sep)[0].strip()
    return s

def get_course_priority(row):
    try:
        l = int(row.get('L', 0)) if pd.notna(row.get('L', 0)) else 0
        t = int(row.get('T', 0)) if pd.notna(row.get('T', 0)) else 0
        p = int(row.get('P', 0)) if pd.notna(row.get('P', 0)) else 0
        return -(l + t + p)
    except Exception:
        return 0

def calculate_required_minutes(course_row):
    l = int(course_row['L']) if ('L' in course_row and pd.notna(course_row['L'])) else 0
    t = int(course_row['T']) if ('T' in course_row and pd.notna(course_row['T'])) else 0
    p = int(course_row['P']) if ('P' in course_row and pd.notna(course_row['P'])) else 0
    
    # *** MODIFIED FIX ***
    # Removed the "is_half_semester" logic based on 'C' column.
    # We will trust the L/T/P values in the CSV as the correct
    # weekly hours, regardless of the 'C' column.
    # The B1/B2 logic (before/after midsem) is handled
    # by the global scheduler.
    
    lec_count = l
    tut_count = t
    lab_count = p
    
    return (lec_count, tut_count, lab_count, 0)

def get_required_room_type(component_type):
    if component_type == 'LAB':
        return 'COMPUTER_LAB'
    return 'LECTURE_ROOM'

# ---------------------------
# Elective basket helpers
# ---------------------------
def extract_elective_basket(course_code):
    if pd.isna(course_code):
        return None
    code = str(course_code).strip().upper()
    import re
    match = re.match(r'^(B\d+)-', code)
    return match.group(1) if match else None

def get_base_course_code(course_code):
    if pd.isna(course_code):
        return str(course_code)
    code = str(course_code).strip()
    # This now correctly handles codes like "B2-MA161-C004"
    # It will return "MA161-C004"
    if '-' in code and code.split('-')[0].upper().startswith('B'):
        return code.split('-', 1)[1]
    return code

def is_elective(course_row):
    code = str(course_row.get('Course Code', '')).strip()
    return extract_elective_basket(code) is not None

def has_component_on_day(timetable, day, course_code, component_type):
    # Check all slots for the given day
    for slot_idx in range(len(TIME_SLOTS)):
        slot_data = timetable[day][slot_idx]
        
        # Get the base code from the slot (e.g., "MA161-C004" from "B2\nMA161-C004")
        slot_code = slot_data['code']
        if '\n' in slot_code:
            slot_code = slot_code.split('\n')[-1]
            
        if slot_code == course_code and slot_data['type'] == component_type:
            return True
    return False

# ---------------------------
# Room allocation
# ---------------------------
def find_suitable_room_for_slot(course_code, room_type, day, slot_indices, room_schedule,
                                  course_room_mapping, component_type, student_strength):
    """
    Find suitable room with proper type checking AND capacity checking.
    Tries to find the smallest room that fits the student strength.
    """
    mapping_key = f"{course_code}_{component_type}"
    
    # *** NEW: Check for forced room in course code (e.g., "-C004") ***
    if "-C004" in course_code.upper():
        forced_room = "C004"
        
        # Check if C004 is available
        if forced_room not in room_schedule:
            room_schedule[forced_room] = {d: set() for d in range(len(DAYS))}
        
        # If this is the first time booking this course, check if C004 is free
        if mapping_key not in course_room_mapping:
            if all(si not in room_schedule[forced_room][day] for si in slot_indices):
                course_room_mapping[mapping_key] = forced_room # Book it
                return forced_room
            else:
                # C004 is busy at this time, this schedule attempt fails
                return None
        else:
            # This course is already mapped to C004, just return it
            return course_room_mapping[mapping_key]
    # *** END NEW ***
    
    # This is the logic for all *other* courses (including shared electives)
    if mapping_key in course_room_mapping:
        fixed_room = course_room_mapping[mapping_key]
        # We found a pre-assigned room for this course.
        # Use it only if it's actually free for these slot indices.
        if fixed_room not in room_schedule:
            room_schedule[fixed_room] = {d: set() for d in range(len(DAYS))}
        if all(si not in room_schedule[fixed_room][day] for si in slot_indices):
            return fixed_room
        # Room is occupied; fall through to find another suitable room.
    
    # If no room is mapped, find a new one...
    
    # 1. Get all room names and shuffle them
    all_room_names = list(ROOM_DATA.keys())
    random.shuffle(all_room_names)
    
    # 2. Find the smallest suitable room
    best_room = None
    min_suitable_capacity = float('inf')

    for room_name in all_room_names:
        if room_name not in ROOM_DATA:
            continue # Should not happen, but safe check
            
        room = ROOM_DATA[room_name]
        
        # 3. Check suitability
        is_type_ok = False
        if room_type == 'COMPUTER_LAB':
            # Allow COMPUTER_LAB or HARDWARE_LAB
            is_type_ok = room['type'] in ['COMPUTER_LAB', 'HARDWARE_LAB']
        else:
            # Allow LECTURE_ROOM or SEATER_120 (or other large rooms)
            is_type_ok = room['type'] in ['LECTURE_ROOM', 'SEATER_120']
        
        # --- THIS IS THE KEY FIX ---
        is_capacity_ok = (room['capacity'] >= student_strength)
        
        if not (is_type_ok and is_capacity_ok):
            continue # This room is not suitable (wrong type or too small)

        # 4. Check availability
        if room_name not in room_schedule:
            room_schedule[room_name] = {d: set() for d in range(len(DAYS))}
            
        is_available = all(si not in room_schedule[room_name][day] for si in slot_indices)
        
        if is_available:
            # This room works. Is it the *best* (smallest) one so far?
            if room['capacity'] < min_suitable_capacity:
                best_room = room_name
                min_suitable_capacity = room['capacity']
    
    # 5. If we found a suitable room, map it and return it
    if best_room:
        course_room_mapping[mapping_key] = best_room
        return best_room

    # Special preference: if the course has more than 120 students and room C004
    # exists (240 seater), prefer assigning C004 for non-lab components and do
    # NOT fall back to combining smaller rooms. This keeps large courses in
    # the single large hall instead of splitting across multiple rooms.
    try:
        if student_strength > 120 and 'C004' in ROOM_DATA and room_type != 'COMPUTER_LAB':
            c004 = ROOM_DATA['C004']
            if c004['capacity'] >= student_strength:
                # Ensure room schedule entry exists
                if 'C004' not in room_schedule:
                    room_schedule['C004'] = {d: set() for d in range(len(DAYS))}
                if all(si not in room_schedule['C004'][day] for si in slot_indices):
                    course_room_mapping[mapping_key] = 'C004'
                    for si in slot_indices:
                        room_schedule['C004'][day].add(si)
                    print(f"    ✅ Assigned C004 for large course {course_code} (needs {student_strength})")
                    return 'C004'
    except Exception:
        # If anything goes wrong with the C004 attempt, fall through to normal logic
        pass

    # If this is a LAB component, try to find two lab rooms whose combined
    # capacity meets the student strength (preferred over leaving unscheduled).
    if room_type == 'COMPUTER_LAB':
        lab_room_names = [rn for rn, info in ROOM_DATA.items() if info['type'] in ('COMPUTER_LAB', 'HARDWARE_LAB')]
        best_pair = None
        best_pair_cap = float('inf')

        # Ensure schedule entries exist
        for rn in lab_room_names:
            if rn not in room_schedule:
                room_schedule[rn] = {d: set() for d in range(len(DAYS))}

        for i in range(len(lab_room_names)):
            for j in range(i+1, len(lab_room_names)):
                r1 = lab_room_names[i]
                r2 = lab_room_names[j]
                # both must be available for all slot indices
                avail1 = all(si not in room_schedule[r1][day] for si in slot_indices)
                avail2 = all(si not in room_schedule[r2][day] for si in slot_indices)
                if not (avail1 and avail2):
                    continue
                cap_sum = ROOM_DATA[r1]['capacity'] + ROOM_DATA[r2]['capacity']
                if cap_sum >= student_strength and cap_sum < best_pair_cap:
                    best_pair = (r1, r2)
                    best_pair_cap = cap_sum

        if best_pair:
            r1, r2 = best_pair
            combined_name = f"{r1}+{r2}"
            course_room_mapping[mapping_key] = combined_name
            for si in slot_indices:
                room_schedule[r1][day].add(si)
                room_schedule[r2][day].add(si)
            print(f"    ✅ Assigned combined labs {combined_name} for {course_code} (combined capacity {best_pair_cap}, needs {student_strength})")
            return combined_name

    # No single free and suitable room found
    return None

def find_consecutive_slots_for_minutes(timetable, day, start_idx, required_minutes,
                                       semester, professor_schedule, faculty,
                                       room_schedule, room_type, course_code, 
                                       course_room_mapping, component_type,
                                       course_day_components, student_strength):
    """Find consecutive free slots"""
    n = len(TIME_SLOTS)
    slot_indices = []
    i = start_idx
    accumulated = 0
    
    # *** MODIFIED: Re-enabled this constraint per user request ***
    # Check lecture-tutorial same-day constraint
    # We check against the base_code (e.g., "MA161-CSE")
    if component_type == 'LEC' and has_component_on_day(timetable, day, course_code, 'TUT'):
        return None, None
    if component_type == 'TUT' and has_component_on_day(timetable, day, course_code, 'LEC'):
        return None, None
    
    # *** NEW: Added constraint for two lectures on the same day ***
    if component_type == 'LEC' and has_component_on_day(timetable, day, course_code, 'LEC'):
        return None, None
    
    while i < n and accumulated < required_minutes:
        if is_minor_slot(TIME_SLOTS[i]) or is_break_time_slot(TIME_SLOTS[i], semester):
            return None, None
        
        if timetable[day][i]['type'] is not None:
            return None, None
        
        if faculty in professor_schedule and i in professor_schedule[faculty][day]:
            return None, None
        
        slot_indices.append(i)
        accumulated += slot_minutes(TIME_SLOTS[i])
        if accumulated > required_minutes:
            return None, None
        i += 1
    
    if accumulated == required_minutes:
        # *** MODIFIED: Pass the *full code* for C004 check ***
        # For non-elective core courses, course_code is the base_code (e.g., "MA161-CSE")
        # For electives, this will be the full code (e.g., "B1-PHD151")
        # Enforce a minimum break between two lectures by rejecting placements
        # that are immediately adjacent to another lecture in the same
        # timetable (student/section) or adjacent to a professor's lecture.
        # This approximates a 10-minute gap by preventing contiguous LEC slots.
        try:
            # previous slot
            prev_idx = slot_indices[0] - 1
            if prev_idx >= 0:
                prev_type = timetable[day][prev_idx]['type']
                if prev_type == 'LEC':
                    return None, None
                # Prevent two LABs back-to-back for the same timetable/section
                if component_type == 'LAB' and prev_type == 'LAB':
                    return None, None
                # professor adjacent constraint
                if faculty in professor_schedule and prev_idx in professor_schedule[faculty][day]:
                    return None, None
            # next slot
            next_idx = slot_indices[-1] + 1
            if next_idx < len(TIME_SLOTS):
                next_type = timetable[day][next_idx]['type']
                if next_type == 'LEC':
                    return None, None
                # Prevent two LABs back-to-back for the same timetable/section
                if component_type == 'LAB' and next_type == 'LAB':
                    return None, None
                if faculty in professor_schedule and next_idx in professor_schedule[faculty][day]:
                    return None, None
        except Exception:
            # If any issue evaluating adjacency, fall back to usual behavior
            pass

        room = find_suitable_room_for_slot(course_code, room_type, day, slot_indices, 
                                           room_schedule, course_room_mapping, component_type,
                                           student_strength) # <-- ADDED
        if room is not None:
            return slot_indices, room
    
    return None, None

def get_all_possible_start_indices():
    idxs = list(range(len(TIME_SLOTS)))
    random.shuffle(idxs)
    return idxs

def check_professor_availability(professor_schedule, faculty, day, start_idx, duration_slots):
    if faculty not in professor_schedule:
        return True
    if not professor_schedule[faculty][day]:
        return True
    
    new_slots = set(range(start_idx, start_idx + duration_slots))
    existing_slots = professor_schedule[faculty][day]
    
    return not (new_slots & existing_slots)

def load_batch_data():
    batch_info = {}
    for _, r in df.iterrows():
        dept = str(r['Department'])
        sem = int(r['Semester'])
        batch_info[(dept, sem)] = {'num_sections': 1}
    return batch_info

def normalize_section_info(courses_df):
    df_local = courses_df.copy()
    if 'Section' not in df_local.columns:
        df_local['Section'] = ''
    if 'SectionMode' not in df_local.columns:
        df_local['SectionMode'] = ''
    df_local['Section'] = df_local['Section'].fillna('').astype(str).str.strip()
    df_local['SectionMode'] = df_local['SectionMode'].fillna('').astype(str).str.strip()
    return df_local

def normalize_crossdept_info(courses_df):
    df_local = courses_df.copy()
    if 'CrossDeptGroup' not in df_local.columns:
        df_local['CrossDeptGroup'] = ''
    if 'CrossDeptMode' not in df_local.columns:
        df_local['CrossDeptMode'] = ''
    df_local['CrossDeptGroup'] = df_local['CrossDeptGroup'].fillna('').astype(str).str.strip()
    df_local['CrossDeptMode'] = df_local['CrossDeptMode'].fillna('').astype(str).str.strip()
    return df_local

def filter_courses_for_section(courses_df, section_label):
    if not section_label:
        return courses_df
    sec = courses_df['Section'].fillna('').astype(str).str.strip().str.upper()
    mode = courses_df['SectionMode'].fillna('').astype(str).str.strip().str.upper()
    mode = mode.where(mode != '', 'COMBINED')
    sec = sec.where(sec != '', 'ALL')
    combined_mask = (mode == 'COMBINED') & (sec.isin(['ALL', section_label]))
    split_mask = (mode == 'SPLIT') & (sec == section_label)
    other_mask = (~mode.isin(['COMBINED', 'SPLIT'])) & (sec.isin(['ALL', section_label]))
    return courses_df[combined_mask | split_mask | other_mask]

def get_combined_courses_all_sections(courses_df):
    sec = courses_df['Section'].fillna('').astype(str).str.strip().str.upper()
    mode = courses_df['SectionMode'].fillna('').astype(str).str.strip().str.upper()
    mode = mode.where(mode != '', 'COMBINED')
    sec = sec.where(sec != '', 'ALL')
    return courses_df[(mode == 'COMBINED') & (sec.isin(['ALL']))]

def get_crossdept_groups(courses_df):
    group = courses_df['CrossDeptGroup'].fillna('').astype(str).str.strip()
    mode = courses_df['CrossDeptMode'].fillna('').astype(str).str.strip().str.upper()
    valid = (group != '') & (mode == 'COMBINED')
    groups = {}
    for g in group[valid].unique():
        groups[g] = courses_df[group == g]
    return groups

def place_course_on_slots(course_row, timetable, day, slot_indices, comp_type,
                          professor_schedule, room_schedule, course_room_mapping,
                          course_day_components, room_override=None, skip_prof_check=False):
    code = str(course_row.get('Course Code', '')).strip()
    if not code:
        return False
    base_code = get_base_course_code(code)
    name = str(course_row.get('Course Name', '')).strip()
    faculty = select_faculty(course_row.get('Faculty', 'TBD'))
    student_strength = int(course_row.get('total_students', 50))

    if not skip_prof_check:
        if not check_professor_availability(professor_schedule, faculty, day, slot_indices[0], len(slot_indices)):
            return False
    if any(timetable[day][si]['type'] is not None for si in slot_indices):
        return False

    room_type = get_required_room_type(comp_type)
    candidate_room = room_override or find_suitable_room_for_slot(
        code, room_type, day, slot_indices, room_schedule, course_room_mapping, comp_type, student_strength
    )
    if candidate_room is None:
        return False

    if faculty not in professor_schedule:
        professor_schedule[faculty] = {d: set() for d in range(len(DAYS))}
    if base_code not in course_day_components:
        course_day_components[base_code] = {}

    for si_idx, si in enumerate(slot_indices):
        timetable[day][si]['type'] = comp_type
        timetable[day][si]['code'] = code if si_idx == 0 else ''
        timetable[day][si]['name'] = name if si_idx == 0 else ''
        timetable[day][si]['faculty'] = faculty if si_idx == 0 else ''
        timetable[day][si]['classroom'] = candidate_room if si_idx == 0 else ''
        professor_schedule[faculty][day].add(si)
        if candidate_room not in room_schedule:
            room_schedule[candidate_room] = {d: set() for d in range(len(DAYS))}
        room_schedule[candidate_room][day].add(si)

    if day not in course_day_components[base_code]:
        course_day_components[base_code][day] = []
    course_day_components[base_code][day].append(comp_type)

    return True

def schedule_crossdept_group(timetable, group_courses, semester, professor_schedule,
                             room_schedule, course_room_mapping, course_day_components,
                             unscheduled_components, department, basket_slots_by_semester=None):
    if group_courses is None or len(group_courses) == 0:
        return []
    rep_course = group_courses.iloc[0]
    code = str(rep_course.get('Course Code', '')).strip()
    base_code = get_base_course_code(code)
    faculty = select_faculty(rep_course.get('Faculty', 'TBD'))
    student_strength = int(rep_course.get('total_students', 50))

    if faculty not in professor_schedule:
        professor_schedule[faculty] = {d: set() for d in range(len(DAYS))}
    if base_code not in course_day_components:
        course_day_components[base_code] = {}

    lec_count, tut_count, lab_count, _ = calculate_required_minutes(rep_course)
    lec_sessions_needed = int(lec_count * 60 / LECTURE_MIN) if lec_count > 0 else 0
    tut_sessions_needed = int(tut_count * 60 / TUTORIAL_MIN) if tut_count > 0 else 0
    lab_sessions_needed = int(lab_count * 60 / LAB_MIN) if lab_count > 0 else 0

    schedule_entries = []

    def schedule_component(required_minutes, comp_type):
        room_type = get_required_room_type(comp_type)
        for attempt in range(5000):
            day = random.randint(0, len(DAYS)-1)
            starts = get_all_possible_start_indices()
            for start_idx in starts:
                slot_indices, candidate_room = find_consecutive_slots_for_minutes(
                    timetable, day, start_idx, required_minutes, semester,
                    professor_schedule, faculty, room_schedule, room_type,
                    base_code, course_room_mapping, comp_type, course_day_components,
                    student_strength
                )
                if slot_indices is None or candidate_room is None:
                    continue
                if not check_professor_availability(professor_schedule, faculty, day, slot_indices[0], len(slot_indices)):
                    continue
                if basket_slots_by_semester:
                    locked = basket_slots_by_semester.get(int(semester), {}).get(day, set())
                    if any(si in locked for si in slot_indices):
                        continue
                # place representative course with the chosen room
                if not place_course_on_slots(
                    rep_course, timetable, day, slot_indices, comp_type,
                    professor_schedule, room_schedule, course_room_mapping,
                    course_day_components, room_override=candidate_room
                ):
                    continue
                # place any additional rows for this dept in the same group
                if len(group_courses) > 1:
                    for _, row in group_courses.iterrows():
                        if row is rep_course:
                            continue
                        place_course_on_slots(
                            row, timetable, day, slot_indices, comp_type,
                            professor_schedule, room_schedule, course_room_mapping,
                            course_day_components, skip_prof_check=True
                        )
                schedule_entries.append({
                    'day': day,
                    'slot_indices': list(slot_indices),
                    'comp_type': comp_type
                })
                return True
        return False

    for _ in range(lec_sessions_needed):
        if not schedule_component(LECTURE_MIN, 'LEC'):
            add_unscheduled_course(unscheduled_components, department, semester, code, str(rep_course.get('Course Name', '')).strip(), faculty, 'LEC', 0, "Could not place cross-dept LEC")
    for _ in range(tut_sessions_needed):
        if not schedule_component(TUTORIAL_MIN, 'TUT'):
            add_unscheduled_course(unscheduled_components, department, semester, code, str(rep_course.get('Course Name', '')).strip(), faculty, 'TUT', 0, "Could not place cross-dept TUT")
    for _ in range(lab_sessions_needed):
        if not schedule_component(LAB_MIN, 'LAB'):
            add_unscheduled_course(unscheduled_components, department, semester, code, str(rep_course.get('Course Name', '')).strip(), faculty, 'LAB', 0, "Could not place cross-dept LAB")

    return schedule_entries

def apply_crossdept_schedule(timetable, group_courses, schedule_entries, professor_schedule,
                             room_schedule, course_room_mapping, course_day_components):
    if not schedule_entries:
        return
    for _, row in group_courses.iterrows():
        for entry in schedule_entries:
            place_course_on_slots(
                row, timetable, entry['day'], entry['slot_indices'], entry['comp_type'],
                professor_schedule, room_schedule, course_room_mapping, course_day_components,
                skip_prof_check=True
            )

def schedule_combined_courses(timetable, combined_courses, semester, professor_schedule,
                              room_schedule, course_room_mapping, course_day_components,
                              unscheduled_components, department, basket_slots_by_semester=None):
    combined_schedule = []
    for _, course in combined_courses.iterrows():
        code = str(course.get('Course Code', '')).strip()
        if not code:
            continue
        base_code = get_base_course_code(code)
        name = str(course.get('Course Name', '')).strip()
        faculty = select_faculty(course.get('Faculty', 'TBD'))
        student_strength = int(course.get('total_students', 50))

        if faculty not in professor_schedule:
            professor_schedule[faculty] = {d: set() for d in range(len(DAYS))}
        if base_code not in course_day_components:
            course_day_components[base_code] = {}

        lec_count, tut_count, lab_count, _ = calculate_required_minutes(course)
        lec_sessions_needed = int(lec_count * 60 / LECTURE_MIN) if lec_count > 0 else 0
        tut_sessions_needed = int(tut_count * 60 / TUTORIAL_MIN) if tut_count > 0 else 0
        lab_sessions_needed = int(lab_count * 60 / LAB_MIN) if lab_count > 0 else 0

        def schedule_component(required_minutes, comp_type):
            room_type = get_required_room_type(comp_type)
            for attempt in range(5000):
                day = random.randint(0, len(DAYS)-1)
                starts = get_all_possible_start_indices()
                for start_idx in starts:
                    slot_indices, candidate_room = find_consecutive_slots_for_minutes(
                        timetable, day, start_idx, required_minutes, semester,
                        professor_schedule, faculty, room_schedule, room_type,
                        base_code, course_room_mapping, comp_type, course_day_components,
                        student_strength
                    )
                    if slot_indices is None or candidate_room is None:
                        continue
                    if not check_professor_availability(professor_schedule, faculty, day, slot_indices[0], len(slot_indices)):
                        continue
                    if basket_slots_by_semester:
                        locked = basket_slots_by_semester.get(int(semester), {}).get(day, set())
                        if any(si in locked for si in slot_indices):
                            continue

                    for si_idx, si in enumerate(slot_indices):
                        timetable[day][si]['type'] = comp_type
                        timetable[day][si]['code'] = code if si_idx == 0 else ''
                        timetable[day][si]['name'] = name if si_idx == 0 else ''
                        timetable[day][si]['faculty'] = faculty if si_idx == 0 else ''
                        timetable[day][si]['classroom'] = candidate_room if si_idx == 0 else ''
                        professor_schedule[faculty][day].add(si)
                        if candidate_room not in room_schedule:
                            room_schedule[candidate_room] = {d: set() for d in range(len(DAYS))}
                        room_schedule[candidate_room][day].add(si)

                    if day not in course_day_components[base_code]:
                        course_day_components[base_code][day] = []
                    course_day_components[base_code][day].append(comp_type)

                    combined_schedule.append({
                        'code': code,
                        'name': name,
                        'faculty': faculty,
                        'comp_type': comp_type,
                        'day': day,
                        'slot_indices': list(slot_indices),
                        'room': candidate_room,
                        'base_code': base_code
                    })
                    return True
            return False

        for _ in range(lec_sessions_needed):
            if not schedule_component(LECTURE_MIN, 'LEC'):
                add_unscheduled_course(unscheduled_components, department, semester, code, name, faculty, 'LEC', 0, "Could not place combined LEC")
        for _ in range(tut_sessions_needed):
            if not schedule_component(TUTORIAL_MIN, 'TUT'):
                add_unscheduled_course(unscheduled_components, department, semester, code, name, faculty, 'TUT', 0, "Could not place combined TUT")
        for _ in range(lab_sessions_needed):
            if not schedule_component(LAB_MIN, 'LAB'):
                add_unscheduled_course(unscheduled_components, department, semester, code, name, faculty, 'LAB', 0, "Could not place combined LAB")

    return combined_schedule

def apply_combined_schedule(timetable, combined_schedule, professor_schedule, room_schedule, course_day_components):
    for entry in combined_schedule:
        day = entry['day']
        slot_indices = entry['slot_indices']
        comp_type = entry['comp_type']
        code = entry['code']
        name = entry['name']
        faculty = entry['faculty']
        room = entry['room']
        base_code = entry['base_code']

        if faculty not in professor_schedule:
            professor_schedule[faculty] = {d: set() for d in range(len(DAYS))}
        if base_code not in course_day_components:
            course_day_components[base_code] = {}

        for si_idx, si in enumerate(slot_indices):
            timetable[day][si]['type'] = comp_type
            timetable[day][si]['code'] = code if si_idx == 0 else ''
            timetable[day][si]['name'] = name if si_idx == 0 else ''
            timetable[day][si]['faculty'] = faculty if si_idx == 0 else ''
            timetable[day][si]['classroom'] = room if si_idx == 0 else ''
            professor_schedule[faculty][day].add(si)
            if room not in room_schedule:
                room_schedule[room] = {d: set() for d in range(len(DAYS))}
            room_schedule[room][day].add(si)

        if day not in course_day_components[base_code]:
            course_day_components[base_code][day] = []
        course_day_components[base_code][day].append(comp_type)

# ---------------------------
# Basket slot enforcement
# ---------------------------
def enforce_basket_slots(timetable, semester, global_basket_schedule):
    """Ensure basket slots are present in the timetable for this semester."""
    for (sem_key, basket), basket_schedule in global_basket_schedule.items():
        if int(sem_key) != int(semester):
            continue
        for day, slot_indices, comp_type in basket_schedule:
            for si_idx, si in enumerate(slot_indices):
                # If slot is already a basket entry, keep it.
                if timetable[day][si].get('is_basket', False):
                    continue
                # Force basket placeholder in this slot
                timetable[day][si]['type'] = comp_type
                timetable[day][si]['is_basket'] = True
                timetable[day][si]['code'] = f"{basket}\n" if si_idx == 0 else ''
                timetable[day][si]['name'] = basket if si_idx == 0 else ''
                timetable[day][si]['faculty'] = '' if si_idx == 0 else ''
                timetable[day][si]['classroom'] = '' if si_idx == 0 else ''

# ---------------------------
# Global elective basket scheduling
# ---------------------------
def schedule_global_elective_baskets(df_input, professor_schedule, room_schedule, course_room_mapping):
    """Pre-schedule elective baskets globally"""
    print("\n" + "="*80)
    print("🎓 GLOBAL ELECTIVE BASKET SCHEDULING")
    print("="*80)
    
    basket_groups = {}
    
    # *** MODIFIED FIX: B1/B2 OVERLAP ***
    # This dictionary will hold separate time locks for each basket type
    # e.g., {'B1': {0: set(), 1: set()}, 'B2': {0: set(), 1: set()}}
    # This allows B1 and B2 to be scheduled at the same time.
    global_basket_slots_by_type = {}

    for _, course in df_input.iterrows():
        code = str(course.get('Course Code', '')).strip()
        basket = extract_elective_basket(code)
        
        if basket and pd.notna(basket):
            if 'Schedule' in course and str(course.get('Schedule', 'Yes')).strip().upper() != 'YES':
                continue
            
            # *** NEW: Initialize the lock for this basket type if it's new ***
            if basket not in global_basket_slots_by_type:
                global_basket_slots_by_type[basket] = {d: set() for d in range(len(DAYS))}

            semester = int(course.get('Semester', 0))
            key = (semester, basket)
            
            if key not in basket_groups:
                basket_groups[key] = []
            basket_groups[key].append(course)
    
    global_schedule = {}
    # Track lecture/lab slots per semester to prevent consecutive LEC/LAB across baskets
    global_semester_adj = {}
    
    for (semester, basket_name), basket_courses in sorted(basket_groups.items()):
        print(f"\n📚 Semester {semester}, Basket {basket_name}: {len(basket_courses)} courses")
        
        # Get the correct global slot lock for this basket type (e.g., 'B1's lock)
        current_basket_global_slots = global_basket_slots_by_type[basket_name]

        first_course = basket_courses[0]
        lec_count, tut_count, lab_count, _ = calculate_required_minutes(first_course)
        
        print(f"    Structure: L={lec_count}h, T={tut_count}h, P={lab_count}h")
        
        lec_sessions = int(lec_count * 60 / LECTURE_MIN) if lec_count > 0 else 0
        tut_sessions = int(tut_count * 60 / TUTORIAL_MIN) if tut_count > 0 else 0
        lab_sessions = int(lab_count * 60 / LAB_MIN) if lab_count > 0 else 0
        
        basket_schedule = []
        if semester not in global_semester_adj:
            global_semester_adj[semester] = {
                'LEC': {d: set() for d in range(len(DAYS))},
                'LAB': {d: set() for d in range(len(DAYS))}
            }
        
        # Schedule lectures
        for session_num in range(lec_sessions):
            scheduled = False
            for attempt in range(5000):
                day = random.randint(0, len(DAYS)-1)
                start_idx = random.randint(0, max(0, len(TIME_SLOTS)-3))
                
                conflict = False
                for prev_day, prev_slots, prev_type in basket_schedule:
                    if prev_day == day and prev_type in ['LEC', 'TUT']:
                        # Do not allow multiple LEC/TUT on the same day for a basket
                        conflict = True
                        break
                
                if conflict:
                    continue
                
                slot_indices = []
                accumulated = 0
                valid = True
                
                for i in range(start_idx, len(TIME_SLOTS)):
                    if is_minor_slot(TIME_SLOTS[i]) or is_break_time_slot(TIME_SLOTS[i], semester):
                        valid = False
                        break
                    
                    slot_indices.append(i)
                    accumulated += slot_minutes(TIME_SLOTS[i])
                    if accumulated == LECTURE_MIN:
                        break
                    if accumulated > LECTURE_MIN:
                        valid = False
                        break
                
                # Prevent consecutive lectures across baskets for the same semester
                lec_slots = global_semester_adj[semester]['LEC'][day]
                adj_conflict = False
                for si in slot_indices:
                    if si in lec_slots or (si - 1) in lec_slots or (si + 1) in lec_slots:
                        adj_conflict = True
                        break
                if adj_conflict:
                    continue

                # *** MODIFIED: Check only against the current basket type's global slots ***
                if valid and accumulated == LECTURE_MIN and len(slot_indices) > 0:
                    if any(s in current_basket_global_slots[day] for s in slot_indices):
                        continue # Slot already taken by another 'B1' course, try again
                    
                    # This slot is free. Book it.
                    basket_schedule.append((day, slot_indices, 'LEC'))
                    current_basket_global_slots[day].update(slot_indices) # Add to 'B1' lock
                    global_semester_adj[semester]['LEC'][day].update(slot_indices)
                    scheduled = True
                    slot_time = TIME_SLOTS[slot_indices[0]][0].strftime('%H:%M')
                    print(f"    ✅ Lecture {session_num+1}/{lec_sessions}: {DAYS[day]} at {slot_time}")
                    break
            
            if not scheduled:
                # Fallback: relax adjacency constraint for this basket/session
                for attempt in range(5000):
                    day = random.randint(0, len(DAYS)-1)
                    start_idx = random.randint(0, max(0, len(TIME_SLOTS)-3))
                    conflict = False
                    for prev_day, _, prev_type in basket_schedule:
                        if prev_day == day and prev_type in ['LEC', 'TUT']:
                            conflict = True
                            break
                    if conflict:
                        continue
                    slot_indices = []
                    accumulated = 0
                    valid = True
                    for i in range(start_idx, len(TIME_SLOTS)):
                        if is_minor_slot(TIME_SLOTS[i]) or is_break_time_slot(TIME_SLOTS[i], semester):
                            valid = False
                            break
                        slot_indices.append(i)
                        accumulated += slot_minutes(TIME_SLOTS[i])
                        if accumulated == LECTURE_MIN:
                            break
                        if accumulated > LECTURE_MIN:
                            valid = False
                            break
                    if valid and accumulated == LECTURE_MIN and len(slot_indices) > 0:
                        if any(s in current_basket_global_slots[day] for s in slot_indices):
                            continue
                        basket_schedule.append((day, slot_indices, 'LEC'))
                        current_basket_global_slots[day].update(slot_indices)
                        # do not update global_semester_adj in relaxed mode
                        scheduled = True
                        slot_time = TIME_SLOTS[slot_indices[0]][0].strftime('%H:%M')
                        print(f"    ✅ Lecture {session_num+1}/{lec_sessions} (relaxed): {DAYS[day]} at {slot_time}")
                        break
                if not scheduled:
                    print(f"    ⚠️ Could not schedule Lecture {session_num+1}/{lec_sessions}")
        
        # Schedule tutorials
        for session_num in range(tut_sessions):
            scheduled = False
            for attempt in range(5000):
                day = random.randint(0, len(DAYS)-1)
                start_idx = random.randint(0, max(0, len(TIME_SLOTS)-2))
                
                lec_or_tut_on_day = any(d == day and ct in ['LEC', 'TUT'] for d, _, ct in basket_schedule)
                if lec_or_tut_on_day:
                    continue
                
                conflict = False
                for prev_day, prev_slots, prev_type in basket_schedule:
                    if prev_day == day and prev_type == 'TUT':
                        if any(s in prev_slots for s in range(start_idx, start_idx + 2)):
                            conflict = True
                            break
                
                if conflict:
                    continue
                
                slot_indices = []
                accumulated = 0
                valid = True
                
                for i in range(start_idx, len(TIME_SLOTS)):
                    if is_minor_slot(TIME_SLOTS[i]) or is_break_time_slot(TIME_SLOTS[i], semester):
                        valid = False
                        break
                    
                    slot_indices.append(i)
                    accumulated += slot_minutes(TIME_SLOTS[i])
                    if accumulated == TUTORIAL_MIN:
                        break
                    if accumulated > TUTORIAL_MIN:
                        valid = False
                        break
                
                # *** MODIFIED: Check only against the current basket type's global slots ***
                if valid and accumulated == TUTORIAL_MIN and len(slot_indices) > 0:
                    if any(s in current_basket_global_slots[day] for s in slot_indices):
                        continue # Slot already taken by another 'B1' course, try again

                    # This slot is free. Book it.
                    basket_schedule.append((day, slot_indices, 'TUT'))
                    current_basket_global_slots[day].update(slot_indices) # Add to 'B1' lock
                    scheduled = True
                    slot_time = TIME_SLOTS[slot_indices[0]][0].strftime('%H:%M')
                    print(f"    ✅ Tutorial {session_num+1}/{tut_sessions}: {DAYS[day]} at {slot_time}")
                    break
            
            if not scheduled:
                # Fallback: relaxed adjacency not applicable to TUT, retry without extra constraints
                for attempt in range(5000):
                    day = random.randint(0, len(DAYS)-1)
                    start_idx = random.randint(0, max(0, len(TIME_SLOTS)-2))
                    lec_or_tut_on_day = any(d == day and ct in ['LEC', 'TUT'] for d, _, ct in basket_schedule)
                    if lec_or_tut_on_day:
                        continue
                    slot_indices = []
                    accumulated = 0
                    valid = True
                    for i in range(start_idx, len(TIME_SLOTS)):
                        if is_minor_slot(TIME_SLOTS[i]) or is_break_time_slot(TIME_SLOTS[i], semester):
                            valid = False
                            break
                        slot_indices.append(i)
                        accumulated += slot_minutes(TIME_SLOTS[i])
                        if accumulated == TUTORIAL_MIN:
                            break
                        if accumulated > TUTORIAL_MIN:
                            valid = False
                            break
                    if valid and accumulated == TUTORIAL_MIN and len(slot_indices) > 0:
                        if any(s in current_basket_global_slots[day] for s in slot_indices):
                            continue
                        basket_schedule.append((day, slot_indices, 'TUT'))
                        current_basket_global_slots[day].update(slot_indices)
                        scheduled = True
                        slot_time = TIME_SLOTS[slot_indices[0]][0].strftime('%H:%M')
                        print(f"    ✅ Tutorial {session_num+1}/{tut_sessions} (relaxed): {DAYS[day]} at {slot_time}")
                        break
                if not scheduled:
                    print(f"    ⚠️ Could not schedule Tutorial {session_num+1}/{tut_sessions}")
        
        # Schedule labs
        for session_num in range(lab_sessions):
            scheduled = False
            for attempt in range(5000):
                day = random.randint(0, len(DAYS)-1)
                start_idx = random.randint(0, max(0, len(TIME_SLOTS)-3))
                
                conflict = False
                for prev_day, prev_slots, prev_type in basket_schedule:
                    if prev_day == day:
                        if any(s in prev_slots for s in range(start_idx, start_idx + 3)):
                            conflict = True
                            break
                
                if conflict:
                    continue
                
                slot_indices = []
                accumulated = 0
                valid = True
                
                for i in range(start_idx, len(TIME_SLOTS)):
                    if is_minor_slot(TIME_SLOTS[i]) or is_break_time_slot(TIME_SLOTS[i], semester):
                        valid = False
                        break
                    
                    slot_indices.append(i)
                    accumulated += slot_minutes(TIME_SLOTS[i])
                    if accumulated == LAB_MIN:
                        break
                    if accumulated > LAB_MIN:
                        valid = False
                        break
                
                # Prevent consecutive labs across baskets for the same semester
                lab_slots = global_semester_adj[semester]['LAB'][day]
                adj_conflict = False
                for si in slot_indices:
                    if si in lab_slots or (si - 1) in lab_slots or (si + 1) in lab_slots:
                        adj_conflict = True
                        break
                if adj_conflict:
                    continue

                # *** MODIFIED: Check only against the current basket type's global slots ***
                if valid and accumulated == LAB_MIN and len(slot_indices) > 0:
                    # Prevent two LABs back-to-back for the same section/timetable
                    prev_idx = slot_indices[0] - 1
                    if prev_idx >= 0 and timetable[day][prev_idx]['type'] == 'LAB':
                        continue
                    next_idx = slot_indices[-1] + 1
                    if next_idx < len(TIME_SLOTS) and timetable[day][next_idx]['type'] == 'LAB':
                        continue

                    if any(s in current_basket_global_slots[day] for s in slot_indices):
                        continue # Slot already taken by another 'B1' course, try again
                        
                    # This slot is free. Book it.
                    basket_schedule.append((day, slot_indices, 'LAB'))
                    current_basket_global_slots[day].update(slot_indices) # Add to 'B1' lock
                    global_semester_adj[semester]['LAB'][day].update(slot_indices)
                    scheduled = True
                    slot_time = TIME_SLOTS[slot_indices[0]][0].strftime('%H:%M')
                    print(f"    ✅ Lab {session_num+1}/{lab_sessions}: {DAYS[day]} at {slot_time}")
                    break
            
            if not scheduled:
                # Fallback: relax adjacency constraint for this basket/session
                for attempt in range(5000):
                    day = random.randint(0, len(DAYS)-1)
                    start_idx = random.randint(0, max(0, len(TIME_SLOTS)-3))
                    conflict = False
                    for prev_day, prev_slots, _ in basket_schedule:
                        if prev_day == day:
                            if any(s in prev_slots for s in range(start_idx, start_idx + 3)):
                                conflict = True
                                break
                    if conflict:
                        continue
                    slot_indices = []
                    accumulated = 0
                    valid = True
                    for i in range(start_idx, len(TIME_SLOTS)):
                        if is_minor_slot(TIME_SLOTS[i]) or is_break_time_slot(TIME_SLOTS[i], semester):
                            valid = False
                            break
                        slot_indices.append(i)
                        accumulated += slot_minutes(TIME_SLOTS[i])
                        if accumulated == LAB_MIN:
                            break
                        if accumulated > LAB_MIN:
                            valid = False
                            break
                    if valid and accumulated == LAB_MIN and len(slot_indices) > 0:
                        if any(s in current_basket_global_slots[day] for s in slot_indices):
                            continue
                        basket_schedule.append((day, slot_indices, 'LAB'))
                        current_basket_global_slots[day].update(slot_indices)
                        scheduled = True
                        slot_time = TIME_SLOTS[slot_indices[0]][0].strftime('%H:%M')
                        print(f"    ✅ Lab {session_num+1}/{lab_sessions} (relaxed): {DAYS[day]} at {slot_time}")
                        break
                if not scheduled:
                    print(f"    ⚠️ Could not schedule Lab {session_num+1}/{lab_sessions}")
        
        global_schedule[(semester, basket_name)] = basket_schedule
        print(f"    📋 Total sessions scheduled: {len(basket_schedule)}")
    
    print("\n" + "="*80)
    print(f"✅ Global basket scheduling complete: {len(global_schedule)} baskets")
    print("="*80 + "\n")
    
    return global_schedule

def add_unscheduled_course(unscheduled_components, department, semester, code, name,
                           faculty, comp_type, section, reason):
    existing = next((u for u in unscheduled_components if u.code == code), None)
    if existing:
        if comp_type not in existing.component_type:
            existing.component_type += f", {comp_type}"
        if reason not in existing.reason:
            existing.reason += f"; {reason}"
    else:
        unscheduled_components.append(
            UnscheduledComponent(department, semester, code, name, faculty,
                                 comp_type, 1, section, reason)
        )

def is_7th_semester(department, semester):
    """Check if this is 7th semester from CSE/DSAI/ECE"""
    dept_upper = str(department).strip().upper()
    return int(semester) == 7 and dept_upper in ['CSE', 'DSAI', 'ECE']

# ---------------------------
# Main generation function
# ---------------------------
def generate_all_timetables():
    global TIME_SLOTS
    TIME_SLOTS = generate_time_slots()
    
    batch_info = load_batch_data()
    room_schedule = {}
    professor_schedule = {}
    course_room_mapping = {}
    
    # Pre-schedule elective baskets
    global_basket_schedule = schedule_global_elective_baskets(df, professor_schedule, room_schedule, course_room_mapping)
    
    if global_basket_schedule is None:
        global_basket_schedule = {}

    # Build per-semester basket slot locks so non-basket courses never take them
    basket_slots_by_semester = {}
    for (sem, _basket), sched in global_basket_schedule.items():
        sem = int(sem)
        basket_slots_by_semester.setdefault(sem, {d: set() for d in range(len(DAYS))})
        for day, slot_indices, _ctype in sched:
            basket_slots_by_semester[sem][day].update(slot_indices)

    # *** NEW: List to store B1 course-room mappings ***
    b1_schedule_list = []

    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    overview = wb.create_sheet("Overview")
    overview.append(["Combined Timetable for All Departments and Semesters"])
    overview.append(["Generated on:", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    overview.append([])
    overview.append(["Department", "Semester", "Sheet Name"])
    row_index = 5

    unscheduled_components = []

    SUBJECT_COLORS = [
        # Slightly darker pastel palette for improved contrast
        "FFDAB3", "C8E6C9", "BBDEFB", "FFD0E6", "FFF3BF",
        "E1BEE7", "CDEFEA", "FFD6D6", "F0F4C3", "DFF3D6",
        "FFD6D1", "E6D6FF", "E6F0A8", "DDEEFF", "E6D6EE"
    ]

    seventh_sem_processed = False
    seventh_sem_course_data = []
    crossdept_schedule = {}

    for department in df['Department'].unique():
        sems = sorted(df[df['Department'] == department]['Semester'].unique())
        for semester in sems:
            if is_7th_semester(department, semester):
                if not seventh_sem_processed:
                    for dept in ['CSE', 'DSAI', 'ECE']:
                        dept_courses = df[(df['Department'] == dept) & (df['Semester'] == 7)]
                        if not dept_courses.empty:
                            seventh_sem_course_data.append(dept_courses)
                    seventh_sem_processed = True
                continue

            dept_upper = str(department).strip().upper()
            num_sections = 2 if dept_upper == "CSE" else 1

            courses = df[(df['Department'] == department) & (df['Semester'] == semester)]
            if 'Schedule' in courses.columns:
                courses = courses[(courses['Schedule'].fillna('Yes').str.upper() == 'YES') | (courses['Schedule'].isna())]
            if courses.empty:
                continue
            courses = normalize_section_info(courses)
            courses = normalize_crossdept_info(courses)
            # Build cross-dept groups for this semester across all departments
            semester_courses_all = df[df['Semester'] == semester]
            if 'Schedule' in semester_courses_all.columns:
                semester_courses_all = semester_courses_all[(semester_courses_all['Schedule'].fillna('Yes').str.upper() == 'YES') | (semester_courses_all['Schedule'].isna())]
            semester_courses_all = normalize_crossdept_info(semester_courses_all)
            crossdept_groups_all = get_crossdept_groups(semester_courses_all)
            for g in list(crossdept_groups_all.keys()):
                crossdept_groups_all[g] = crossdept_groups_all[g][crossdept_groups_all[g].apply(is_elective, axis=1) == False]
                if crossdept_groups_all[g].empty:
                    del crossdept_groups_all[g]
            combined_courses_all = get_combined_courses_all_sections(courses)
            if not combined_courses_all.empty:
                combined_courses_all = combined_courses_all[combined_courses_all.apply(is_elective, axis=1) == False]
            combined_schedule = None

            for section in range(num_sections):
                section_title = f"{department}_{semester}" if num_sections == 1 else f"{department}_{semester}_{chr(65 + section)}"
                ws = wb.create_sheet(title=section_title)

                overview.cell(row=row_index, column=1, value=department)
                overview.cell(row=row_index, column=2, value=str(semester))
                overview.cell(row=row_index, column=3, value=section_title)
                row_index += 1

                timetable = {d: {s: {'type': None, 'code': '', 'name': '', 'faculty': '', 'classroom': '', 'is_basket': False} for s in range(len(TIME_SLOTS))} for d in range(len(DAYS))}
                course_day_components = {}

                section_label = chr(65 + section) if num_sections > 1 else None
                section_courses = courses
                if dept_upper == "CSE" and section_label:
                    section_courses = filter_courses_for_section(courses, section_label)
                if dept_upper == "CSE" and combined_courses_all is not None and not combined_courses_all.empty:
                    section_courses = section_courses[~section_courses.index.isin(combined_courses_all.index)]
                dept_crossdept_groups = {}
                if crossdept_groups_all:
                    for gname, gdf in crossdept_groups_all.items():
                        dept_rows = gdf[gdf['Department'] == department]
                        if not dept_rows.empty:
                            dept_crossdept_groups[gname] = dept_rows
                    if dept_crossdept_groups:
                        crossdept_idx = pd.concat(dept_crossdept_groups.values()).index
                        section_courses = section_courses[~section_courses.index.isin(crossdept_idx)]

                if 'P' in section_courses.columns:
                    lab_courses = section_courses[section_courses['P'] > 0].copy()
                    non_lab_courses = section_courses[section_courses['P'] == 0].copy()
                else:
                    lab_courses = section_courses.head(0)
                    non_lab_courses = section_courses.copy()

                if not lab_courses.empty:
                    lab_courses['priority'] = lab_courses.apply(get_course_priority, axis=1)
                    lab_courses = lab_courses.sort_values('priority', ascending=False)
                non_lab_courses['priority'] = non_lab_courses.apply(get_course_priority, axis=1)
                non_lab_courses = non_lab_courses.sort_values('priority', ascending=False)

                combined = pd.concat([lab_courses, non_lab_courses])
                combined['is_elective'] = combined.apply(is_elective, axis=1)
                combined['elective_basket'] = combined['Course Code'].apply(extract_elective_basket)

                courses_combined = combined.sort_values(
                    by=['is_elective', 'priority'], 
                    ascending=[False, False]
                ).drop_duplicates()

                legend_courses = courses_combined
                if combined_courses_all is not None and not combined_courses_all.empty:
                    legend_courses = pd.concat([courses_combined, combined_courses_all]).drop_duplicates()
                if dept_crossdept_groups:
                    legend_courses = pd.concat([legend_courses] + [g for g in dept_crossdept_groups.values()]).drop_duplicates()

                section_subject_color = {}
                basket_color_map = {}
                color_iter = iter(SUBJECT_COLORS)
                course_faculty_map = {}
                basket_scheduled_courses = set()

                for _, c in legend_courses.iterrows():
                    code = str(c.get('Course Code', '')).strip()
                    if code and code not in section_subject_color:
                        basket_label = extract_elective_basket(code)
                        if basket_label:
                            if basket_label not in basket_color_map:
                                try:
                                    basket_color_map[basket_label] = next(color_iter)
                                except StopIteration:
                                    basket_color_map[basket_label] = random.choice(SUBJECT_COLORS)
                            if basket_label not in section_subject_color:
                                section_subject_color[basket_label] = basket_color_map[basket_label]
                            section_subject_color[code] = basket_color_map[basket_label]
                            base_code = get_base_course_code(code)
                            if base_code and base_code not in section_subject_color:
                                section_subject_color[base_code] = basket_color_map[basket_label]
                        else:
                            try:
                                section_subject_color[code] = next(color_iter)
                            except StopIteration:
                                section_subject_color[code] = random.choice(SUBJECT_COLORS)
                        course_faculty_map[code] = select_faculty(c.get('Faculty', 'TBD'))

                # Apply global basket schedules (per semester, across all branches)
                print(f"\n📋 Applying basket schedules for {section_title}...")

                # Map basket -> courses in this section
                basket_courses_map = {}
                elective_courses_in_section = courses_combined[courses_combined['is_elective'] == True]
                for _, course in elective_courses_in_section.iterrows():
                    b = course.get('elective_basket')
                    if not b or pd.isna(b):
                        continue
                    basket_courses_map.setdefault(b, []).append(course)

                for (sem_key, basket), basket_schedule in global_basket_schedule.items():
                    if int(sem_key) != int(semester):
                        continue
                    courses_in_basket = basket_courses_map.get(basket, [])

                    # If this section has no courses for the basket, still reserve the slot
                    if not courses_in_basket:
                        for day, slot_indices, comp_type in basket_schedule:
                            for si_idx, si in enumerate(slot_indices):
                                if timetable[day][si]['type'] is None:
                                    timetable[day][si]['type'] = comp_type
                                    timetable[day][si]['is_basket'] = True
                                    timetable[day][si]['code'] = f"{basket}\n" if si_idx == 0 else ''
                                    timetable[day][si]['name'] = basket if si_idx == 0 else ''
                                    timetable[day][si]['faculty'] = '' if si_idx == 0 else ''
                                    timetable[day][si]['classroom'] = '' if si_idx == 0 else ''
                        continue

                    for _, course in enumerate(courses_in_basket):
                        code = str(course.get('Course Code', '')).strip()
                        if not code:
                            continue
                        base_code = get_base_course_code(code)
                        name = str(course.get('Course Name', '')).strip()
                        faculty = select_faculty(course.get('Faculty', 'TBD'))
                        student_strength = int(course.get('total_students', 50))

                        if faculty not in professor_schedule:
                            professor_schedule[faculty] = {d: set() for d in range(len(DAYS))}
                        if base_code not in course_day_components:
                            course_day_components[base_code] = {}

                        for day, slot_indices, comp_type in basket_schedule:
                            room_type = get_required_room_type(comp_type)
                            candidate_room = find_suitable_room_for_slot(
                                code, room_type, day, slot_indices,
                                room_schedule, course_room_mapping, comp_type,
                                student_strength
                            )
                            if candidate_room is None:
                                add_unscheduled_course(unscheduled_components, department, semester, code, name, faculty, comp_type, section, f"No suitable room found (Needs {student_strength} capacity)")
                                continue

                            for si in slot_indices:
                                professor_schedule[faculty][day].add(si)
                                if candidate_room not in room_schedule:
                                    room_schedule[candidate_room] = {d: set() for d in range(len(DAYS))}
                                room_schedule[candidate_room][day].add(si)

                            for si_idx, si in enumerate(slot_indices):
                                timetable[day][si]['type'] = comp_type
                                timetable[day][si]['is_basket'] = True
                                timetable[day][si]['code'] = f"{basket}\n{base_code}" if si_idx == 0 else ''
                                timetable[day][si]['name'] = name if si_idx == 0 else ''
                                timetable[day][si]['faculty'] = faculty if si_idx == 0 else ''
                                timetable[day][si]['classroom'] = candidate_room if si_idx == 0 else ''

                            if day not in course_day_components[base_code]:
                                course_day_components[base_code][day] = []
                            course_day_components[base_code][day].append(comp_type)

                        basket_scheduled_courses.add(code)

                    print(f"    ✅ Applied {basket} schedule to {len(courses_in_basket)} courses")

                # Schedule combined courses once (A), then apply same slots to B
                if dept_upper == "CSE" and section_label in ["A", "B"] and combined_courses_all is not None and not combined_courses_all.empty:
                    if section_label == "A":
                        combined_schedule = schedule_combined_courses(
                            timetable, combined_courses_all, semester, professor_schedule,
                            room_schedule, course_room_mapping, course_day_components,
                            unscheduled_components, department, basket_slots_by_semester
                        )
                    elif section_label == "B" and combined_schedule is not None:
                        apply_combined_schedule(
                            timetable, combined_schedule, professor_schedule,
                            room_schedule, course_day_components
                        )

                if section_label is None and dept_crossdept_groups:
                    for group_name, group_courses in dept_crossdept_groups.items():
                        key = (int(semester), group_name)
                        if key not in crossdept_schedule:
                            crossdept_schedule[key] = schedule_crossdept_group(
                                timetable, group_courses, semester, professor_schedule,
                                room_schedule, course_room_mapping, course_day_components,
                                unscheduled_components, department, basket_slots_by_semester
                            )
                        else:
                            apply_crossdept_schedule(
                                timetable, group_courses, crossdept_schedule[key],
                                professor_schedule, room_schedule, course_room_mapping,
                                course_day_components
                            )

                # Schedule non-elective courses
                print(f"\n📖 Scheduling non-elective courses for {section_title}...")
                for _, course in courses_combined.iterrows():
                    code = str(course.get('Course Code', '')).strip()
                    
                    if code in basket_scheduled_courses:
                        continue
                    
                    # This check is now correct: it skips B1/B2 electives AND B1/B2 core courses
                    if course.get('is_elective'):
                        continue
                    
                    base_code = get_base_course_code(code)
                    name = str(course.get('Course Name', '')).strip()
                    faculty = select_faculty(course.get('Faculty', 'TBD'))
                    student_strength = int(course.get('total_students', 50)) # <-- ADDED

                    if faculty not in professor_schedule:
                        professor_schedule[faculty] = {d: set() for d in range(len(DAYS))}
                    
                    if base_code not in course_day_components:
                        course_day_components[base_code] = {}

                    lec_count, tut_count, lab_count, ss_count = calculate_required_minutes(course)
                    lec_sessions_needed = int(lec_count * 60 / LECTURE_MIN) if lec_count > 0 else 0
                    tut_sessions_needed = int(tut_count * 60 / TUTORIAL_MIN) if tut_count > 0 else 0
                    lab_sessions_needed = int(lab_count * 60 / LAB_MIN) if lab_count > 0 else 0

                    def schedule_component(required_minutes, comp_type, student_strength, attempts_limit=5000):
                        room_type = get_required_room_type(comp_type)
                        
                        for attempt in range(attempts_limit):
                            day = random.randint(0, len(DAYS)-1)
                            starts = get_all_possible_start_indices()
                            
                            for start_idx in starts:
                                slot_indices, candidate_room = find_consecutive_slots_for_minutes(
                                    timetable, day, start_idx, required_minutes, semester,
                                    professor_schedule, faculty, room_schedule, room_type,
                                    base_code, course_room_mapping, comp_type, course_day_components,
                                    student_strength # <-- ADDED
                                )

                                if slot_indices is None:
                                    continue
                                if not check_professor_availability(professor_schedule, faculty, day, slot_indices[0], len(slot_indices)):
                                    continue
                                if candidate_room is None:
                                    continue
                                if basket_slots_by_semester:
                                    locked = basket_slots_by_semester.get(int(semester), {}).get(day, set())
                                    if any(si in locked for si in slot_indices):
                                        continue
                                
                                for si_idx, si in enumerate(slot_indices):
                                    timetable[day][si]['type'] = comp_type
                                    timetable[day][si]['code'] = code if si_idx == 0 else ''
                                    timetable[day][si]['name'] = name if si_idx == 0 else ''
                                    timetable[day][si]['faculty'] = faculty if si_idx == 0 else ''
                                    timetable[day][si]['classroom'] = candidate_room if si_idx == 0 else ''
                                    professor_schedule[faculty][day].add(si)
                                    if candidate_room not in room_schedule:
                                        room_schedule[candidate_room] = {d: set() for d in range(len(DAYS))}
                                    room_schedule[candidate_room][day].add(si)
                                
                                if day not in course_day_components[base_code]:
                                    course_day_components[base_code][day] = []
                                course_day_components[base_code][day].append(comp_type)
                                
                                return True
                        return False

                    # Schedule lectures
                    for _ in range(lec_sessions_needed):
                        ok = schedule_component(LECTURE_MIN, 'LEC', student_strength)
                        if not ok:
                            add_unscheduled_course(unscheduled_components, department, semester, code, name, faculty, 'LEC', section, f"Could not find suitable slot (Needs {student_strength} capacity)")

                    # Schedule tutorials
                    for _ in range(tut_sessions_needed):
                        ok = schedule_component(TUTORIAL_MIN, 'TUT', student_strength)
                        if not ok:
                            add_unscheduled_course(unscheduled_components, department, semester, code, name, faculty, 'TUT', section, f"Could not find suitable slot (Needs {student_strength} capacity)")

                    # Schedule labs
                    for _ in range(lab_sessions_needed):
                        ok = schedule_component(LAB_MIN, 'LAB', student_strength)
                        if not ok:
                            add_unscheduled_course(unscheduled_components, department, semester, code, name, faculty, 'LAB', section, f"No computer lab available (Needs {student_strength} capacity)")

                # Ensure basket slots are present (and override non-basket slots if needed)
                enforce_basket_slots(timetable, semester, global_basket_schedule)

                # Write timetable to sheet
                write_timetable_to_sheet(ws, timetable, section_subject_color, course_faculty_map, 
                                         legend_courses, course_room_mapping, semester)

    # *** NEW: Print B1 course list ***
    if b1_schedule_list:
        print("\n" + "="*80)
        print("B1 Basket Room Allocations")
        print("="*80)
        
        # Use a dictionary to store unique code-room pairs
        unique_b1_courses = {}
        for code, room in b1_schedule_list:
            if code not in unique_b1_courses:
                unique_b1_courses[code] = room
                
        for code, room in sorted(unique_b1_courses.items()):
            print(f"    - Course: {code.ljust(25)} Room: {room}")
        print("="*80 + "\n")
    # *** END NEW ***

    # Generate common 7th semester timetable
    if seventh_sem_course_data:
        generate_7th_sem_common_timetable(wb, seventh_sem_course_data, overview, row_index, 
                                          unscheduled_components, professor_schedule, 
                                          room_schedule, course_room_mapping, SUBJECT_COLORS)

    # Format overview sheet
    format_overview_sheet(overview, row_index)

    # Write meta entries into a hidden sheet so teacher generator can read
    if META_ENTRIES:
        try:
            meta_ws = wb.create_sheet('_META')
            headers = ['sheet', 'row', 'start_col', 'end_col', 'faculty', 'classroom', 'typ', 'code']
            meta_ws.append(headers)
            for me in META_ENTRIES:
                meta_ws.append([
                    me.get('sheet'), me.get('row'), me.get('start_col'), me.get('end_col'),
                    me.get('faculty'), me.get('classroom'), me.get('typ'), me.get('code')
                ])
            meta_ws.sheet_state = 'hidden'
        except Exception:
            pass

    # Save workbook
    out_filename = os.path.join(OUTPUT_DIR, "timetable_all_departments.xlsx")
    try:
        wb.save(out_filename)
        print(f"\n✅ Combined timetable saved as {out_filename}")
    except Exception as e:
        print(f"❌ Failed to save timetable: {e}")
        traceback.print_exc()

    # Generate teacher and unscheduled workbooks
    try:
        create_teacher_and_unscheduled_from_combined(out_filename, unscheduled_components)
    except Exception as e:
        print("❌ Failed to generate teacher/unscheduled workbooks:", e)
        traceback.print_exc()

    return out_filename

# ---------------------------
# 7th semester common timetable
# ---------------------------
def generate_7th_sem_common_timetable(wb, course_data_list, overview, row_index, 
                                      unscheduled_components, professor_schedule, 
                                      room_schedule, course_room_mapping, SUBJECT_COLORS):
    """Generate common timetable for 7th semester"""
    
    all_7th_courses = pd.concat(course_data_list, ignore_index=True)
    if 'Schedule' in all_7th_courses.columns:
        all_7th_courses = all_7th_courses[(all_7th_courses['Schedule'].fillna('Yes').str.upper() == 'YES') | 
                                          (all_7th_courses['Schedule'].isna())]
    
    if all_7th_courses.empty:
        return
    
    section_title = "Common_7th_Semester"
    ws = wb.create_sheet(title=section_title)
    
    overview.cell(row=row_index, column=1, value="CSE/DSAI/ECE")
    overview.cell(row=row_index, column=2, value="7")
    overview.cell(row=row_index, column=3, value=section_title)
    
    if 'P' in all_7th_courses.columns:
        lab_courses = all_7th_courses[all_7th_courses['P'] > 0].copy()
        non_lab_courses = all_7th_courses[all_7th_courses['P'] == 0].copy()
    else:
        lab_courses = all_7th_courses.head(0)
        non_lab_courses = all_7th_courses.copy()
    
    if not lab_courses.empty:
        lab_courses['priority'] = lab_courses.apply(get_course_priority, axis=1)
        lab_courses = lab_courses.sort_values('priority', ascending=False)
    non_lab_courses['priority'] = non_lab_courses.apply(get_course_priority, axis=1)
    non_lab_courses = non_lab_courses.sort_values('priority', ascending=False)
    
    combined = pd.concat([lab_courses, non_lab_courses])
    combined['is_elective'] = combined.apply(is_elective, axis=1)
    combined['elective_basket'] = combined['Course Code'].apply(extract_elective_basket)
    
    courses_combined = combined.sort_values(
        by=['is_elective', 'priority'], 
        ascending=[False, False]
    ).drop_duplicates()
    
    timetable = {d: {s: {'type': None, 'code': '', 'name': '', 'faculty': '', 'classroom': '', 'is_basket': False} 
                       for s in range(len(TIME_SLOTS))} for d in range(len(DAYS))}
    course_day_components = {}
    
    section_subject_color = {}
    basket_color_map = {}
    color_iter = iter(SUBJECT_COLORS)
    course_faculty_map = {}
    
    for _, c in courses_combined.iterrows():
        code = str(c.get('Course Code', '')).strip()
        if code and code not in section_subject_color:
            basket_label = extract_elective_basket(code)
            if basket_label:
                if basket_label not in basket_color_map:
                    try:
                        basket_color_map[basket_label] = next(color_iter)
                    except StopIteration:
                        basket_color_map[basket_label] = random.choice(SUBJECT_COLORS)
                if basket_label not in section_subject_color:
                    section_subject_color[basket_label] = basket_color_map[basket_label]
                section_subject_color[code] = basket_color_map[basket_label]
                base_code = get_base_course_code(code)
                if base_code and base_code not in section_subject_color:
                    section_subject_color[base_code] = basket_color_map[basket_label]
            else:
                try:
                    section_subject_color[code] = next(color_iter)
                except StopIteration:
                    section_subject_color[code] = random.choice(SUBJECT_COLORS)
            course_faculty_map[code] = select_faculty(c.get('Faculty', 'TBD'))
    
    # Schedule non-electives as before, and schedule elective baskets together
    non_electives = courses_combined[courses_combined['is_elective'] == False]
    electives = courses_combined[courses_combined['is_elective'] == True]

    # Helper to schedule a single course (reuse existing pattern)
    def schedule_single_course(course_row):
        code = str(course_row.get('Course Code', '')).strip()
        base_code = get_base_course_code(code)
        name = str(course_row.get('Course Name', '')).strip()
        faculty = select_faculty(course_row.get('Faculty', 'TBD'))
        student_strength = int(course_row.get('total_students', 50))

        if faculty not in professor_schedule:
            professor_schedule[faculty] = {d: set() for d in range(len(DAYS))}
        if base_code not in course_day_components:
            course_day_components[base_code] = {}

        lec_count, tut_count, lab_count, _ = calculate_required_minutes(course_row)
        lec_sessions_needed = int(lec_count * 60 / LECTURE_MIN) if lec_count > 0 else 0
        tut_sessions_needed = int(tut_count * 60 / TUTORIAL_MIN) if tut_count > 0 else 0
        lab_sessions_needed = int(lab_count * 60 / LAB_MIN) if lab_count > 0 else 0

        def schedule_component(required_minutes, comp_type, student_strength_local, faculty_local, code_local, name_local):
            room_type = get_required_room_type(comp_type)
            for attempt in range(5000):
                day = random.randint(0, len(DAYS)-1)
                starts = get_all_possible_start_indices()
                for start_idx in starts:
                    slot_indices, candidate_room = find_consecutive_slots_for_minutes(
                        timetable, day, start_idx, required_minutes, 7,
                        professor_schedule, faculty_local, room_schedule, room_type,
                        code_local, course_room_mapping, comp_type, course_day_components,
                        student_strength_local
                    )
                    if slot_indices is None or candidate_room is None:
                        continue
                    if not check_professor_availability(professor_schedule, faculty_local, day, slot_indices[0], len(slot_indices)):
                        continue
                    for si_idx, si in enumerate(slot_indices):
                        timetable[day][si]['type'] = comp_type
                        timetable[day][si]['code'] = code_local if si_idx == 0 else ''
                        timetable[day][si]['name'] = name_local if si_idx == 0 else ''
                        timetable[day][si]['faculty'] = faculty_local if si_idx == 0 else ''
                        timetable[day][si]['classroom'] = candidate_room if si_idx == 0 else ''
                        professor_schedule[faculty_local][day].add(si)
                        if candidate_room not in room_schedule:
                            room_schedule[candidate_room] = {d: set() for d in range(len(DAYS))}
                        room_schedule[candidate_room][day].add(si)
                    if day not in course_day_components[get_base_course_code(code_local)]:
                        course_day_components[get_base_course_code(code_local)][day] = []
                    course_day_components[get_base_course_code(code_local)][day].append(comp_type)
                    return True
            return False

        for _ in range(lec_sessions_needed):
            ok = schedule_component(LECTURE_MIN, 'LEC', student_strength, faculty, code, name)
            if not ok:
                add_unscheduled_course(unscheduled_components, "Common_7th", 7, code, name, faculty, 'LEC', 0, f"Could not find suitable slot (Needs {student_strength} capacity)")
        for _ in range(tut_sessions_needed):
            ok = schedule_component(TUTORIAL_MIN, 'TUT', student_strength, faculty, code, name)
            if not ok:
                add_unscheduled_course(unscheduled_components, "Common_7th", 7, code, name, faculty, 'TUT', 0, f"Could not find suitable slot (Needs {student_strength} capacity)")
        for _ in range(lab_sessions_needed):
            ok = schedule_component(LAB_MIN, 'LAB', student_strength, faculty, code, name)
            if not ok:
                add_unscheduled_course(unscheduled_components, "Common_7th", 7, code, name, faculty, 'LAB', 0, f"No computer lab available (Needs {student_strength} capacity)")

    # Group electives by basket and schedule each basket as a unit
    if not electives.empty:
        baskets = {}
        for _, row in electives.iterrows():
            basket = extract_elective_basket(str(row.get('Course Code', '')).strip())
            if not basket:
                continue
            baskets.setdefault(basket, []).append(row)

        for basket_label, rows in baskets.items():
            # Use first row as representative for LTPS structure
            first = rows[0]
            lec_count, tut_count, lab_count, _ = calculate_required_minutes(first)
            lec_sessions_needed = int(lec_count * 60 / LECTURE_MIN) if lec_count > 0 else 0
            tut_sessions_needed = int(tut_count * 60 / TUTORIAL_MIN) if tut_count > 0 else 0
            lab_sessions_needed = int(lab_count * 60 / LAB_MIN) if lab_count > 0 else 0

            # Aggregate faculties and student strengths
            facs = [select_faculty(r.get('Faculty', 'TBD')) for r in rows]
            facs = [f for f in facs if f]
            unique_facs = list(dict.fromkeys(facs))
            agg_faculty = '/'.join(unique_facs) if unique_facs else 'TBD'
            # For 7th-semester basket scheduling, use the maximum course size
            # among basket options rather than summing all student counts.
            total_strength = max((int(r.get('total_students', 50)) for r in rows), default=50)

            # Representative base code (used for meta); prefer first course's base
            rep_code = str(first.get('Course Code', '')).strip()
            rep_base = get_base_course_code(rep_code)

            # Ensure professor_schedule entries exist for all faculties
            for f in unique_facs:
                if f not in professor_schedule:
                    professor_schedule[f] = {d: set() for d in range(len(DAYS))}
            # Track which days have already been used for this basket (so
            # multiple sessions of the same component type aren't placed
            # on the same weekday).
            scheduled_days_for_basket = set()

            def schedule_basket_component(required_minutes, comp_type, student_strength_local, faculties_local):
                room_type = get_required_room_type(comp_type)
                for attempt in range(5000):
                    day = random.randint(0, len(DAYS)-1)
                    starts = get_all_possible_start_indices()
                    for start_idx in starts:
                        # Build slot_indices manually (same slot for all courses)
                        slot_indices_local = []
                        acc = 0
                        i = start_idx
                        while i < len(TIME_SLOTS) and acc < required_minutes:
                            if is_minor_slot(TIME_SLOTS[i]) or is_break_time_slot(TIME_SLOTS[i], 7):
                                break
                            if timetable[day][i]['type'] is not None:
                                break
                            slot_indices_local.append(i)
                            acc += slot_minutes(TIME_SLOTS[i])
                            if acc > required_minutes:
                                break
                            i += 1
                        if acc != required_minutes:
                            continue
                        # Prevent back-to-back lectures in 7th semester common timetable
                        if comp_type == 'LEC':
                            prev_idx = slot_indices_local[0] - 1
                            if prev_idx >= 0 and timetable[day][prev_idx]['type'] == 'LEC':
                                continue
                            next_idx = slot_indices_local[-1] + 1
                            if next_idx < len(TIME_SLOTS) and timetable[day][next_idx]['type'] == 'LEC':
                                continue
                        # Prevent two LABs back-to-back when scheduling basket labs
                        if comp_type == 'LAB':
                            prev_idx = slot_indices_local[0] - 1
                            if prev_idx >= 0 and timetable[day][prev_idx]['type'] == 'LAB':
                                continue
                            next_idx = slot_indices_local[-1] + 1
                            if next_idx < len(TIME_SLOTS) and timetable[day][next_idx]['type'] == 'LAB':
                                continue
                        # Don't schedule multiple sessions of the same basket
                        # on the same weekday; skip if this day already used.
                        if day in scheduled_days_for_basket:
                            continue

                        # Try to find a separate room for each course in this basket for the same slot_indices
                        course_room_map = {}
                        failed = False
                        rooms_assigned = []
                        faculties_assigned = []

                        for row in rows:
                            course_code = str(row.get('Course Code', '')).strip()
                            base = get_base_course_code(course_code)
                            faculty_c = select_faculty(row.get('Faculty', 'TBD'))
                            strength_c = int(row.get('total_students', 50))
                            room_type_c = get_required_room_type(comp_type)

                            # Check professor availability
                            if not check_professor_availability(professor_schedule, faculty_c, day, slot_indices_local[0], len(slot_indices_local)):
                                failed = True
                                break

                            # Find a suitable room for this course for these exact slot indices
                            candidate_room_c = find_suitable_room_for_slot(course_code, room_type_c, day, slot_indices_local, room_schedule, course_room_mapping, comp_type, strength_c)
                            if candidate_room_c is None:
                                failed = True
                                break

                            # Tentatively record assignment
                            course_room_map[course_code] = (candidate_room_c, faculty_c, base)
                            rooms_assigned.append(candidate_room_c)
                            faculties_assigned.append(faculty_c)

                        if failed:
                            # revert any tentative course_room_mapping done by find_suitable_room_for_slot? (mapping is persistent)
                            continue

                        # All courses in basket can be placed in these slot_indices; commit assignments
                        for si_idx, si in enumerate(slot_indices_local):
                            timetable[day][si]['type'] = comp_type
                            timetable[day][si]['is_basket'] = True
                            # Visible code shows basket label and a representative base (first course)
                            timetable[day][si]['code'] = f"{basket_label}\n{rep_base}" if si_idx == 0 else ''
                            timetable[day][si]['name'] = basket_label if si_idx == 0 else ''
                            # Aggregate faculties and rooms for meta/display
                            timetable[day][si]['faculty'] = '/'.join(unique_facs) if si_idx == 0 else ''
                            timetable[day][si]['classroom'] = '/'.join(rooms_assigned) if si_idx == 0 else ''

                            # Mark professors and rooms busy per course
                            for course_code, (room_c, fac_c, base_c) in course_room_map.items():
                                professor_schedule[fac_c][day].add(si)
                                if room_c not in room_schedule:
                                    room_schedule[room_c] = {d: set() for d in range(len(DAYS))}
                                room_schedule[room_c][day].add(si)

                        if rep_base not in course_day_components:
                            course_day_components[rep_base] = {}
                        if day not in course_day_components[rep_base]:
                            course_day_components[rep_base][day] = []
                        course_day_components[rep_base][day].append(comp_type)
                        # Mark this day as used for this basket so subsequent
                        # sessions (LEC/TUT/LAB) for the same basket won't
                        # be scheduled again on the same weekday.
                        scheduled_days_for_basket.add(day)
                        return True
                return False

            for _ in range(lec_sessions_needed):
                ok = schedule_basket_component(LECTURE_MIN, 'LEC', total_strength, unique_facs)
                if not ok:
                    add_unscheduled_course(unscheduled_components, "Common_7th", 7, basket_label, basket_label, agg_faculty, 'LEC', 0, f"Could not schedule basket {basket_label} (Needs {total_strength} capacity)")
            for _ in range(tut_sessions_needed):
                ok = schedule_basket_component(TUTORIAL_MIN, 'TUT', total_strength, unique_facs)
                if not ok:
                    add_unscheduled_course(unscheduled_components, "Common_7th", 7, basket_label, basket_label, agg_faculty, 'TUT', 0, f"Could not schedule basket {basket_label} (Needs {total_strength} capacity)")
            for _ in range(lab_sessions_needed):
                ok = schedule_basket_component(LAB_MIN, 'LAB', total_strength, unique_facs)
                if not ok:
                    add_unscheduled_course(unscheduled_components, "Common_7th", 7, basket_label, basket_label, agg_faculty, 'LAB', 0, f"Could not schedule basket {basket_label} labs (Needs {total_strength} capacity)")

    # Schedule non-elective courses normally
    for _, course in non_electives.iterrows():
        schedule_single_course(course)
    
    write_timetable_to_sheet(ws, timetable, section_subject_color, course_faculty_map, 
                             courses_combined, course_room_mapping, 7)

# ---------------------------
# Write timetable to Excel
# ---------------------------
def write_timetable_to_sheet(ws, timetable, section_subject_color, course_faculty_map, 
                             courses_combined, course_room_mapping, semester):
    """Write formatted timetable to worksheet"""
    
    # Header row
    header = ['Day'] + [f"{slot[0].strftime('%H:%M')}-{slot[1].strftime('%H:%M')}" for slot in TIME_SLOTS]
    ws.append(header)
    
    # Styling
    # Darker, still soft header styling for better contrast
    header_fill = PatternFill(start_color="BBDEFB", end_color="BBDEFB", fill_type="solid")
    header_font = Font(bold=True, size=11, color="000000")
    header_alignment = Alignment(horizontal='center', vertical='center')
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                    top=Side(style='thin'), bottom=Side(style='thin'))
    # Slightly stronger fills for component type defaults (better readability)
    lec_fill_default = PatternFill(start_color="FFDAB3", end_color="FFDAB3", fill_type="solid")
    lab_fill_default = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
    tut_fill_default = PatternFill(start_color="BBDEFB", end_color="BBDEFB", fill_type="solid")
    break_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
    minor_fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
    
    # Write timetable rows
    for day_idx, day_name in enumerate(DAYS):
        ws.append([day_name] + [''] * len(TIME_SLOTS))
        row_num = ws.max_row
        merges = []
        
        for slot_idx in range(len(TIME_SLOTS)):
            cell_obj = ws.cell(row=row_num, column=slot_idx + 2)
            
            if is_minor_slot(TIME_SLOTS[slot_idx]):
                cell_obj.value = "Minor Slot"
                cell_obj.fill = minor_fill
                cell_obj.font = Font(bold=True)
                cell_obj.alignment = Alignment(horizontal='center', vertical='center')
                cell_obj.border = border
                continue
            
            if is_break_time_slot(TIME_SLOTS[slot_idx], semester):
                cell_obj.value = "LUNCH BREAK"  # Updated label
                cell_obj.fill = break_fill
                cell_obj.font = Font(bold=True)
                cell_obj.alignment = Alignment(horizontal='center', vertical='center')
                cell_obj.border = border
                continue
            
            if timetable[day_idx][slot_idx]['type'] is None:
                cell_obj.border = border
                continue
            
            typ = timetable[day_idx][slot_idx]['type']
            code = timetable[day_idx][slot_idx]['code']
            cls = timetable[day_idx][slot_idx]['classroom']
            fac = timetable[day_idx][slot_idx]['faculty']
            
            if code:
                span = [slot_idx]
                j = slot_idx + 1
                while j < len(TIME_SLOTS) and timetable[day_idx][j]['type'] is not None and timetable[day_idx][j]['code'] == '':
                    span.append(j)
                    j += 1
                import re
                # Detect basket entries even when stored as 'B1\nBASECODE'
                raw_code = str(code) if code is not None else ''
                parts = [p.strip() for p in raw_code.splitlines() if p.strip()]
                basket_label = None
                base_code = ''
                if parts and re.match(r'^B\d+$', parts[0].upper()):
                    basket_label = parts[0].upper()
                    if len(parts) > 1:
                        base_code = parts[1]
                elif re.match(r'^B\d+$', raw_code.strip().upper()):
                    basket_label = raw_code.strip().upper()

                if basket_label:
                    # Record metadata so teacher generator can still pick up
                    # faculty and room info, but display only basket+type.
                    start_col = slot_idx + 2
                    end_col = slot_idx + 2 + len(span) - 1
                    # Prefer per-course faculty entries for the basket so
                    # each faculty sees their own course in teacher timetable.
                    basket_rows = courses_combined
                    if 'elective_basket' in courses_combined.columns:
                        basket_rows = courses_combined[courses_combined['elective_basket'] == basket_label]
                    else:
                        basket_rows = courses_combined[courses_combined['Course Code'].apply(extract_elective_basket) == basket_label]
                    if basket_rows is not None and len(basket_rows) > 0:
                        for _, brow in basket_rows.iterrows():
                            full_code = str(brow.get('Course Code', '')).strip()
                            if not full_code:
                                continue
                            base_c = get_base_course_code(full_code)
                            fac_b = select_faculty(brow.get('Faculty', 'TBD'))
                            room_b = course_room_mapping.get(f"{full_code}_{typ}") or \
                                     course_room_mapping.get(f"{base_c}_{typ}") or cls
                            META_ENTRIES.append({
                                'sheet': ws.title,
                                'row': row_num,
                                'start_col': start_col,
                                'end_col': end_col,
                                'faculty': fac_b,
                                'classroom': room_b,
                                'typ': typ,
                                'code': base_c or basket_label
                            })
                    else:
                        META_ENTRIES.append({
                            'sheet': ws.title,
                            'row': row_num,
                            'start_col': start_col,
                            'end_col': end_col,
                            'faculty': fac,
                            'classroom': cls,
                            'typ': typ,
                            'code': base_code or basket_label
                        })
                    display = f"{basket_label}\n{typ}"
                else:
                    display = f"{code}\n{typ}\nRoom: {cls}\n{fac}"
                    # Record metadata for merged ranges so teacher timetable
                    # can fill every slot in the span (merged cells only keep
                    # the value in the first column).
                    if len(span) > 1:
                        start_col = slot_idx + 2
                        end_col = slot_idx + 2 + len(span) - 1
                        META_ENTRIES.append({
                            'sheet': ws.title,
                            'row': row_num,
                            'start_col': start_col,
                            'end_col': end_col,
                            'faculty': fac,
                            'classroom': cls,
                            'typ': typ,
                            'code': code
                        })
                
                # Use basket label color for all basket slots
                if basket_label and basket_label in section_subject_color:
                    subj_color = section_subject_color.get(basket_label)
                    fill = PatternFill(start_color=subj_color, end_color=subj_color, fill_type="solid")
                else:
                    # Use the full code (including B1/B2) for coloring
                    full_code = code.replace('\n', '-') if '\n' in code else code
                    if full_code in section_subject_color:
                        subj_color = section_subject_color.get(full_code)
                        fill = PatternFill(start_color=subj_color, end_color=subj_color, fill_type="solid")
                    elif code.split('\n')[-1] in section_subject_color:
                        # Fallback for base code
                        actual_code = code.split('\n')[-1]
                        subj_color = section_subject_color.get(actual_code)
                        fill = PatternFill(start_color=subj_color, end_color=subj_color, fill_type="solid")
                    else:
                        fill = {'LEC': lec_fill_default, 'LAB': lab_fill_default, 
                                'TUT': tut_fill_default}.get(typ, lec_fill_default)
                
                cell_obj.value = display
                cell_obj.fill = fill
                merges.append((slot_idx + 2, slot_idx + 2 + len(span) - 1, display, fill))
            
            cell_obj.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
            cell_obj.border = border
        
        # Apply merges
        for start_col, end_col, val, fill in merges:
            if end_col > start_col:
                rng = f"{get_column_letter(start_col)}{row_num}:{get_column_letter(end_col)}{row_num}"
                try:
                    ws.merge_cells(rng)
                    mc = ws[f"{get_column_letter(start_col)}{row_num}"]
                    mc.value = val
                    mc.fill = fill
                    mc.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
                    mc.border = border
                except Exception:
                    pass
    
    # Set column widths
    for col_idx in range(1, len(TIME_SLOTS)+2):
        try:
            ws.column_dimensions[get_column_letter(col_idx)].width = 15
        except Exception:
            pass
    
    # Set row heights
    for row in ws.iter_rows(min_row=2, max_row=len(DAYS)+1):
        ws.row_dimensions[row[0].row].height = 40
    # Freeze header row so time headers stay visible
    try:
        ws.freeze_panes = 'A2'
    except Exception:
        pass
    # Give the sheet a light tab color to look nicer in Excel
    try:
        ws.sheet_properties.tabColor = "E3F2FD"
    except Exception:
        pass
    
    # Add self-study courses section
    current_row = len(DAYS) + 4
    ss_courses = []
    for _, course in courses_combined.iterrows():
        l = int(course['L']) if pd.notna(course['L']) else 0
        t = int(course['T']) if pd.notna(course['T']) else 0
        p = int(course['P']) if pd.notna(course['P']) else 0
        s = int(course['S']) if pd.notna(course['S']) else 0
        if s > 0 and l == 0 and t == 0 and p == 0:
            ss_courses.append({
                'code': str(course['Course Code']),
                'name': str(course['Course Name']),
                'faculty': str(course['Faculty'])
            })
    
    if ss_courses:
        ws.cell(row=current_row, column=1, value="Self-Study Only Courses")
        ws.cell(row=current_row, column=1).font = Font(bold=True)
        current_row += 1
        
        headers = ['Course Code', 'Course Name', 'Faculty']
        for col, header in enumerate(headers, 1):
            ws.cell(row=current_row, column=col, value=header)
            ws.cell(row=current_row, column=col).font = Font(bold=True)
        current_row += 1
        
        for course in ss_courses:
            ws.cell(row=current_row, column=1, value=course['code'])
            ws.cell(row=current_row, column=2, value=course['name'])
            ws.cell(row=current_row, column=3, value=course['faculty'])
            current_row += 1
        
        current_row += 2
    
    # Add Course Details
    legend_title = ws.cell(row=current_row, column=1, value="Course Details")
    legend_title.font = Font(bold=True, size=12)
    current_row += 2
    
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    
    legend_headers = ['Course Code', 'Color', 'Name of the Course', 'Course Co-ordinator', 'LTPSC', 'Room Number']
    for col, header in enumerate(legend_headers, 1):
        cell = ws.cell(row=current_row, column=col, value=header)
        cell.font = Font(bold=True)
        cell.border = border
        cell.fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    current_row += 1
    
    # Add legend entries
    for code, color in section_subject_color.items():
        # Use the code itself (which is unique, e.g., "B1-MA161")
        assigned_room = course_room_mapping.get(f"{code}_LEC") or \
                        course_room_mapping.get(f"{code}_LAB") or \
                        course_room_mapping.get(f"{code}_TUT") or "—"
        
        if not assigned_room or assigned_room == "—":
            # Try to find a room for the base code if full code fails
            base_code = get_base_course_code(code)
            assigned_room = course_room_mapping.get(f"{base_code}_LEC") or \
                            course_room_mapping.get(f"{base_code}_LAB") or \
                            course_room_mapping.get(f"{base_code}_TUT") or "—"
            if not assigned_room or assigned_room == "—":
                continue # Skip if no room is mapped at all
        
        ws.row_dimensions[current_row].height = 30
        
        ltps_value = ""
        course_name = ''
        fac_name = ''
        for _, course_row in courses_combined.iterrows():
            if str(course_row['Course Code']) == code:
                l = str(int(course_row['L'])) if pd.notna(course_row['L']) else "0"
                t = str(int(course_row['T'])) if pd.notna(course_row['T']) else "0"
                p = str(int(course_row['P'])) if pd.notna(course_row['P']) else "0"
                s = str(int(course_row['S'])) if pd.notna(course_row['S']) and 'S' in course_row else "0"
                c = str(int(course_row['C'])) if pd.notna(course_row['C']) else "0"
                ltps_value = f"{l}-{t}-{p}-{s}-{c}"
                course_name = str(course_row['Course Name'])
                fac_name = course_faculty_map.get(code, '')
                break
        
        cells = [
            (code, None),
            ('', PatternFill(start_color=color, end_color=color, fill_type="solid")),
            (course_name, None),
            (fac_name, None),
            (ltps_value, None),
            (assigned_room, None)
        ]
        
        for col, (value, fill) in enumerate(cells, 1):
            cell = ws.cell(row=current_row, column=col, value=value)
            cell.border = border
            if fill:
                cell.fill = fill
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True, indent=2)
        
        current_row += 1

def format_overview_sheet(overview, row_index):
    """Format overview sheet with improved visual styling."""
    # Merge title and date rows for a cleaner header
    try:
        overview.merge_cells('A1:C1')
    except Exception:
        pass
    try:
        overview.merge_cells('A2:C2')
    except Exception:
        pass

    # Title styling (row 1)
    title_cell = overview.cell(row=1, column=1)
    title_cell.font = Font(bold=True, size=14, color='0B3954')
    title_cell.alignment = Alignment(horizontal='center', vertical='center')

    # Date/info styling (row 2)
    date_cell = overview.cell(row=2, column=1)
    date_cell.font = Font(italic=True, size=10, color='333333')
    date_cell.alignment = Alignment(horizontal='center', vertical='center')

    # Column widths
    overview.column_dimensions[get_column_letter(1)].width = 30
    overview.column_dimensions[get_column_letter(2)].width = 12
    overview.column_dimensions[get_column_letter(3)].width = 30

    # Header row (row 4) styling
    header_fill = PatternFill(start_color="1976D2", end_color="1976D2", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal='center', vertical='center')
    header_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                           top=Side(style='thin'), bottom=Side(style='thin'))

    for cell in overview[4]:
        try:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = header_border
        except Exception:
            pass

    # Apply alternating row fills and borders for table body
    alt_fill = PatternFill(start_color="F7F9FC", end_color="F7F9FC", fill_type="solid")
    body_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    start_row = 5
    end_row = max(start_row, row_index - 1)
    for r in range(start_row, end_row + 1):
        row_fill = alt_fill if (r % 2 == 0) else None
        for c in range(1, 4):
            cell = overview.cell(row=r, column=c)
            if row_fill:
                try:
                    cell.fill = row_fill
                except Exception:
                    pass
            try:
                cell.border = body_border
                cell.alignment = Alignment(horizontal='left', vertical='center')
            except Exception:
                pass

    # Enable autofilter for the table and freeze panes so header stays visible
    try:
        overview.auto_filter.ref = f"A4:C{max(4, row_index-1)}"
    except Exception:
        pass
    try:
        overview.freeze_panes = 'A5'
    except Exception:
        pass

    # Give the overview sheet a subtle tab color
    try:
        overview.sheet_properties.tabColor = "1976D2"
    except Exception:
        pass

# ---------------------------
# Teacher and unscheduled workbooks
# ---------------------------
def split_faculty_names(fac_str):
    """Split faculty string into separate names"""
    if fac_str is None:
        return []
    s = str(fac_str).strip()
    if s == '' or s.lower() in ['nan', 'none']:
        return []
    parts = [s]
    for sep in ['/', ',', '&', ';']:
        if sep in s:
            parts = [p.strip() for p in s.split(sep) if p.strip()]
            break
    return parts if parts else [s]

def parse_cell_for_course(cell_value):
    """Parse cell value to extract course info"""
    if cell_value is None:
        return ('', '', '', '')
    text = str(cell_value).strip()
    if text == '':
        return ('', '', '', '')
    
    lines = [ln.strip() for ln in text.splitlines() if ln.strip()]
    faculty = ''
    room = ''
    code = ''
    typ = ''
    
    for ln in lines:
        if 'room' in ln.lower():
            parts = ln.split(':')
            if len(parts) >= 2:
                room = parts[-1].strip()
    
    if len(lines) >= 1:
        last = lines[-1]
        if 'room' not in last.lower() and ':' not in last:
            faculty = last
    
    first = lines[0] if lines else ''
    if first:
        # Handle basket format (B1\nCODE)
        if '\n' in text and first.startswith('B'):
            code = lines[1] if len(lines) > 1 else first # This gets the base code
        else:
            code = first
        
        for t in ['LEC', 'LAB', 'TUT']:
            if t in text.upper():
                typ = t
                break
    
    return (code, typ, room, faculty)

def create_teacher_and_unscheduled_from_combined(timetable_filename, unscheduled_components):
    """Build teacher_timetables.xlsx and unscheduled_courses.xlsx"""
    try:
        wb = load_workbook(timetable_filename, data_only=True)
    except Exception as e:
        print(f"❌ Failed to open {timetable_filename}: {e}")
        return
    
    teacher_slots = {}
    teacher_slot_types = {}
    slot_headers = []

    # Build elective base-code set to simplify teacher display
    elective_base_codes = set()
    try:
        for _, row in df.iterrows():
            code = str(row.get('Course Code', '')).strip()
            if extract_elective_basket(code):
                elective_base_codes.add(get_base_course_code(code))
    except Exception:
        elective_base_codes = set()

    def format_teacher_entry(code, typ, sheetname, room):
        if code and code in elective_base_codes:
            return f"{code}\nRoom: {room}"
        return f"{code} {typ}\n({sheetname})\nRoom: {room}" if code else ''

    # Load meta information if present so teacher workbook can be built
    meta_map = {}
    if '_META' in wb.sheetnames:
        meta_ws = wb['_META']
        try:
            for r in range(2, meta_ws.max_row + 1):
                m_sheet = meta_ws.cell(r, 1).value
                m_row = meta_ws.cell(r, 2).value
                m_start = meta_ws.cell(r, 3).value
                m_end = meta_ws.cell(r, 4).value
                m_fac = meta_ws.cell(r, 5).value
                m_room = meta_ws.cell(r, 6).value
                m_typ = meta_ws.cell(r, 7).value
                m_code = meta_ws.cell(r, 8).value
                if not m_sheet or not m_row or not m_start:
                    continue
                for col in range(int(m_start), int(m_end) + 1):
                    key = (m_sheet, int(m_row), int(col))
                    meta_map.setdefault(key, []).append({
                        'faculty': m_fac,
                        'room': m_room,
                        'typ': m_typ,
                        'code': m_code
                    })
        except Exception:
            meta_map = {}
    
    for sheetname in wb.sheetnames:
        if sheetname.lower() == 'overview':
            continue
        ws = wb[sheetname]
        header = [str(ws.cell(1, c).value).strip() if ws.cell(1, c).value else '' for c in range(2, ws.max_column + 1)]
        if len(header) > len(slot_headers):
            slot_headers = header

        for r in range(2, ws.max_row + 1):
            day = ws.cell(r, 1).value
            if not day or str(day) not in DAYS:
                break
            day_idx = DAYS.index(day)

            for c in range(2, ws.max_column + 1):
                # If there's meta info for this exact cell use it
                if (sheetname, r, c) in meta_map:
                    for m in meta_map[(sheetname, r, c)]:
                        faculty = m.get('faculty')
                        room = m.get('room')
                        typ = m.get('typ')
                        code = m.get('code')
                        fac_list = split_faculty_names(faculty)
                        for f in fac_list:
                            if not f or str(f).strip().upper() in ["BREAK", "MINOR SLOT", "NAN", "NONE", "", "LUNCH BREAK"]:
                                continue
                            teacher_slots.setdefault(f, {d: {i: '' for i in range(len(slot_headers))} for d in range(len(DAYS))})
                            teacher_slot_types.setdefault(f, {d: {i: '' for i in range(len(slot_headers))} for d in range(len(DAYS))})
                            teacher_slots[f][day_idx][c - 2] = format_teacher_entry(code, typ, sheetname, room)
                            teacher_slot_types[f][day_idx][c - 2] = typ or ''
                    continue

                code, typ, room, faculty = parse_cell_for_course(ws.cell(r, c).value)
                for f in split_faculty_names(faculty):
                    if not f or str(f).strip().upper() in ["BREAK", "MINOR SLOT", "NAN", "NONE", "", "LUNCH BREAK"]:
                        continue
                    teacher_slots.setdefault(f, {d: {i: '' for i in range(len(slot_headers))} for d in range(len(DAYS))})
                    teacher_slot_types.setdefault(f, {d: {i: '' for i in range(len(slot_headers))} for d in range(len(DAYS))})
                    teacher_slots[f][day_idx][c - 2] = format_teacher_entry(code, typ, sheetname, room)
                    teacher_slot_types[f][day_idx][c - 2] = typ or ''
    
    # Create teacher workbook
    twb = Workbook()
    if "Sheet" in twb.sheetnames:
        twb.remove(twb["Sheet"])
    
    # Softer header and light component fills for teacher sheets
    header_fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
    alt_fill = PatternFill(start_color="FBFCFD", end_color="FBFCFD", fill_type="solid")
    lec_fill_default = PatternFill(start_color="FFDAB3", end_color="FFDAB3", fill_type="solid")
    lab_fill_default = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
    tut_fill_default = PatternFill(start_color="BBDEFB", end_color="BBDEFB", fill_type="solid")
    break_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
    minor_fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
    border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
    header_font = Font(bold=True, size=12)
    title_font = Font(bold=True, size=14)
    cell_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    for teacher in sorted(teacher_slots.keys()):
        safe_name = teacher[:31] or "Unknown"
        ws = twb.create_sheet(title=safe_name)
        
        ws.merge_cells("A1:{}1".format(get_column_letter(len(slot_headers) + 1)))
        title_cell = ws.cell(row=1, column=1, value=f"{teacher} — Weekly Timetable")
        title_cell.font = title_font
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        
        ws.append(["Day"] + slot_headers)
        for cell in ws[2]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
        
        for d, day in enumerate(DAYS):
            row = [day] + [teacher_slots[teacher][d][i] for i in range(len(slot_headers))]
            ws.append(row)
            row_idx = ws.max_row
            for col_idx, cell in enumerate(ws[row_idx], start=1):
                val = str(cell.value).upper() if cell.value is not None else ""
                if col_idx == 1:
                    # Day column zebra fill only
                    if d % 2 == 0:
                        cell.fill = alt_fill
                else:
                    slot_type = ''
                    try:
                        slot_type = teacher_slot_types.get(teacher, {}).get(d, {}).get(col_idx - 2, '')
                    except Exception:
                        slot_type = ''
                    if "LUNCH BREAK" in val:
                        cell.fill = break_fill
                    elif "MINOR SLOT" in val:
                        cell.fill = minor_fill
                    elif slot_type == 'LAB' or " LAB" in val or "LAB" in val:
                        cell.fill = lab_fill_default
                    elif slot_type == 'TUT' or " TUT" in val or "TUT" in val:
                        cell.fill = tut_fill_default
                    elif slot_type == 'LEC' or " LEC" in val or "LEC" in val:
                        cell.fill = lec_fill_default
                    elif d % 2 == 0 and val == "":
                        cell.fill = alt_fill
                cell.alignment = cell_align
                cell.border = border
            ws.row_dimensions[row_idx].height = 35
        
        ws.column_dimensions["A"].width = 15
        for col in range(2, len(slot_headers) + 2):
            ws.column_dimensions[get_column_letter(col)].width = 20
    
    twb.save(os.path.join(OUTPUT_DIR, "teacher_timetables.xlsx"))
    print("✅ Saved teacher_timetables.xlsx")
    
    # Create unscheduled workbook
    uwb = Workbook()
    ws = uwb.active
    ws.title = "Unscheduled Courses"
    
    headers = ["Course Code", "Department", "Semester", "Component Type", "Reason"]
    ws.append(headers)
    
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    unscheduled_unique = {}
    
    for u in unscheduled_components:
        if u.code not in unscheduled_unique:
            if hasattr(u, "reason") and u.reason and len(str(u.reason).strip()) > 0:
                reason_text = str(u.reason).strip()
            else:
                reason_text = "Unspecified scheduling issue"
            
            unscheduled_unique[u.code] = {
                "Course Code": u.code,
                "Department": u.department,
                "Semester": u.semester,
                "Component Type": u.component_type,
                "Reason": reason_text
            }
    
    for entry in unscheduled_unique.values():
        ws.append([entry["Course Code"], entry["Department"], entry["Semester"], 
                   entry["Component Type"], entry["Reason"]])
    
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 50
    
    uwb.save(os.path.join(OUTPUT_DIR, "unscheduled_courses.xlsx"))
    print(f"✅ Saved unscheduled_courses.xlsx with {len(unscheduled_unique)} courses")


# ---------------------------
# Main execution
# ---------------------------
if __name__ == "__main__":
    try:
        print("\n" + "="*80)
        print("🎓 IIIT DHARWAD TIMETABLE GENERATOR")
        print("="*80)
        print("\n🔧 Configuration:")
        print(f"    - No morning break (removed 10:30-10:45)")
        print(f"    - Lunch break: 13:15-14:00 (extended)")
        print(f"    - Lecture duration: {LECTURE_MIN} minutes")
        print(f"    - Tutorial duration: {TUTORIAL_MIN} minutes")
        print(f"    - Lab duration: {LAB_MIN} minutes")
        print(f"    - Min gap between lectures: {MIN_GAP_BETWEEN_LECTURES} minutes")
        print("="*80 + "\n")
        
        generate_all_timetables()
        
        
        print("\n" + "="*80)
        print("✅ TIMETABLE GENERATION COMPLETE!")
        print("="*80)
        print(f"\n📁 Output files saved in: {OUTPUT_DIR}")
        print("    1. timetable_all_departments.xlsx - Main timetable")
        print("    2. teacher_timetables.xlsx - Faculty schedules")
        print("    3. unscheduled_courses.xlsx - Courses that couldn't be scheduled")
        print("\n💡 Tips:")
        print("    - Check unscheduled_courses.xlsx to see which courses failed")
        print("    - If many courses are unscheduled, consider:")
        print("      * Adding more computer labs or larger lecture rooms")
        print("      * Relaxing some constraints")
        print("      * Extending time slots into evening")
        print("="*80 + "\n")
        
    except Exception as e:
        print("\n" + "="*80)
        print("❌ ERROR DURING TIMETABLE GENERATION")
        print("="*80)
        print(f"Error: {e}")
        traceback.print_exc()
        print("="*80 + "\n")
