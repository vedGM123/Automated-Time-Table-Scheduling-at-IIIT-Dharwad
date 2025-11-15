# comprehensive_timetable.py -- Enhanced with all TT_gen.py features
# Run: python comprehensive_timetable.py
# Requires: pandas, openpyxl

import pandas as pd
import random
from datetime import datetime, time, timedelta
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter
from dataclasses import dataclass
import traceback
import os
import json
from pathlib import Path

# ---------------------------
# Setup Directories and Config
# ---------------------------
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
INPUT_DIR = os.path.join(BASE_DIR, "data")
OUTPUT_DIR = os.path.join(BASE_DIR, "output")
CONFIG_PATH = os.path.join(BASE_DIR, "config.json")

os.makedirs(OUTPUT_DIR, exist_ok=True)

# Load configuration
try:
    with open(CONFIG_PATH, "r", encoding="utf-8") as f:
        config = json.load(f)
except Exception:
    config = {}

# ---------------------------
# Constants from config
# ---------------------------
DAYS = config.get("days", ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday'])
LECTURE_MIN = config.get("LECTURE_MIN", 90)
LAB_MIN = config.get("LAB_MIN", 120)
TUTORIAL_MIN = config.get("TUTORIAL_MIN", 60)
SELF_STUDY_MIN = config.get("SELF_STUDY_MIN", 60)

# Break windows
MORNING_BREAK_START = time(10, 30)
MORNING_BREAK_END = time(10, 45)
LUNCH_BREAK_START = time(13, 0)
LUNCH_BREAK_END = time(13, 45)

# Lunch configuration
LUNCH_WINDOW_START = time(12, 30)
LUNCH_WINDOW_END = time(14, 0)
LUNCH_DURATION = 60

# ---------------------------
# Global Variables
# ---------------------------
TIME_SLOTS = []
lunch_breaks = {}
GLOBAL_BASKET_SCHEDULE = {}
SCHEDULED_BASKET_COURSE_CODES = set()

# ---------------------------
# Dataclasses
# ---------------------------
@dataclass
class UnscheduledComponent:
    department: str
    semester: int
    code: str
    name: str
    faculty: str
    component_type: str
    sessions: int
    section: int
    reason: str

# ---------------------------
# Time Slot Generation
# ---------------------------
def generate_time_slots():
    """Generate irregular time slots matching TT_gen.py"""
    slots = [
        (time(7, 30), time(9, 0)),
        (time(9, 0), time(10, 0)),
        (time(10, 0), time(10, 30)),
        (time(10, 30), time(10, 45)),
        (time(10, 45), time(11, 0)),
        (time(11, 0), time(11, 30)),
        (time(11, 30), time(12, 0)),
        (time(12, 0), time(12, 15)),
        (time(12, 15), time(12, 30)),
        (time(12, 30), time(13, 15)),
        (time(13, 15), time(13, 30)),
        (time(13, 30), time(14, 0)),
        (time(14, 0), time(14, 30)),
        (time(14, 30), time(15, 30)),
        (time(15, 30), time(15, 40)),
        (time(16, 0), time(16, 30)),
        (time(16, 30), time(17, 10)),
        (time(17, 10), time(17, 30)),
        (time(17, 30), time(18, 30)),
        (time(18, 30), time(23, 59)),
    ]
    return slots

def initialize_time_slots():
    global TIME_SLOTS
    TIME_SLOTS = generate_time_slots()

def slot_minutes(slot):
    """Calculate duration of a slot in minutes"""
    s, e = slot
    s_m = s.hour * 60 + s.minute
    e_m = e.hour * 60 + e.minute
    if e_m < s_m:
        e_m += 24 * 60
    return e_m - s_m

# ---------------------------
# Load Data
# ---------------------------
def load_data():
    """Load courses data with robust error handling"""
    try:
        csv_path = os.path.join(INPUT_DIR, 'combined.csv')
        encodings_to_try = ['utf-8-sig', 'utf-8', 'cp1252']
        
        for encoding in encodings_to_try:
            try:
                df = pd.read_csv(csv_path, encoding=encoding)
                df = df.replace(r'^\s*$', pd.NA, regex=True)
                df = df.replace('nan', pd.NA)
                return df
            except UnicodeDecodeError:
                continue
        
        raise Exception("Unable to read with any encoding")
    except Exception as e:
        print(f"Error loading combined.csv: {e}")
        return pd.DataFrame(columns=['Department', 'Semester', 'Course Code', 'Course Name', 
                                    'L', 'T', 'P', 'S', 'C', 'Faculty', 'Schedule', 'total_students'])

def load_rooms():
    """Load room data"""
    rooms = {}
    try:
        csv_path = os.path.join(INPUT_DIR, 'rooms.csv')
        rooms_df = pd.read_csv(csv_path)
        
        for _, row in rooms_df.iterrows():
            room_id = str(row['roomNumber'])
            rooms[room_id] = {
                'capacity': int(row.get('capacity', 60)),
                'type': str(row['type']),
                'roomNumber': room_id,
                'schedule': {day: set() for day in range(len(DAYS))}
            }
    except Exception as e:
        print(f"Warning: Error loading rooms.csv: {e}")
        # Default rooms
        rooms = {
            "L101": {'capacity': 70, 'type': 'LECTURE_ROOM', 'roomNumber': 'L101', 
                    'schedule': {day: set() for day in range(len(DAYS))}},
            "C201": {'capacity': 40, 'type': 'COMPUTER_LAB', 'roomNumber': 'C201',
                    'schedule': {day: set() for day in range(len(DAYS))}},
        }
    return rooms

def load_batch_data(df):
    """Load batch information from total_students column"""
    batch_info = {}
    
    try:
        grouped = df.groupby(['Department', 'Semester'])
        
        for (dept, sem), group in grouped:
            if 'total_students' in group.columns and not group['total_students'].isna().all():
                try:
                    total_students = int(group['total_students'].max())
                except ValueError:
                    continue
                
                max_batch_size = 85
                num_sections = (total_students + max_batch_size - 1) // max_batch_size
                section_size = (total_students + num_sections - 1) // num_sections
                
                batch_info[(dept, sem)] = {
                    'total': total_students,
                    'num_sections': num_sections,
                    'section_size': section_size
                }
        
        # Process basket courses
        basket_courses = df[df['Course Code'].astype(str).str.contains('^B[0-9]-', na=False)]
        for _, course in basket_courses.iterrows():
            code = str(course['Course Code'])
            if 'total_students' in df.columns and pd.notna(course['total_students']):
                try:
                    total_students = int(course['total_students'])
                except ValueError:
                    total_students = 35
            else:
                total_students = 35
            
            batch_info[('ELECTIVE', code)] = {
                'total': total_students,
                'num_sections': 1,
                'section_size': total_students
            }
    except Exception as e:
        print(f"Warning: Error processing batch sizes: {e}")
    
    return batch_info

# ---------------------------
# Lunch Break Calculation
# ---------------------------
def calculate_lunch_breaks(semesters):
    """Dynamically calculate staggered lunch breaks"""
    global lunch_breaks
    lunch_breaks = {}
    total_semesters = len(semesters)
    
    if total_semesters == 0:
        return lunch_breaks
    
    total_window_minutes = (
        LUNCH_WINDOW_END.hour * 60 + LUNCH_WINDOW_END.minute -
        LUNCH_WINDOW_START.hour * 60 - LUNCH_WINDOW_START.minute
    )
    stagger_interval = (total_window_minutes - LUNCH_DURATION) / (total_semesters - 1) if total_semesters > 1 else 0
    
    sorted_semesters = sorted(semesters)
    
    for i, semester in enumerate(sorted_semesters):
        start_minutes = (LUNCH_WINDOW_START.hour * 60 + LUNCH_WINDOW_START.minute + 
                        int(i * stagger_interval))
        start_hour = start_minutes // 60
        start_min = start_minutes % 60
        
        end_minutes = start_minutes + LUNCH_DURATION
        end_hour = end_minutes // 60
        end_min = end_minutes % 60
        
        lunch_breaks[semester] = (
            time(start_hour, start_min),
            time(end_hour, end_min)
        )
    
    return lunch_breaks

# ---------------------------
# Helper Functions
# ---------------------------
def overlaps(a_start, a_end, b_start, b_end):
    """Check if two time ranges overlap"""
    a_s_min = a_start.hour * 60 + a_start.minute
    a_e_min = a_end.hour * 60 + a_end.minute
    b_s_min = b_start.hour * 60 + b_start.minute
    b_e_min = b_end.hour * 60 + b_end.minute
    return (a_s_min < b_e_min) and (b_s_min < a_e_min)

def is_break_time_slot(slot, semester=None):
    """Check if slot is a break time"""
    start, end = slot
    
    if overlaps(start, end, MORNING_BREAK_START, MORNING_BREAK_END):
        return True
    
    if overlaps(start, end, LUNCH_BREAK_START, LUNCH_BREAK_END):
        return True
    
    # Staggered lunch
    if semester and semester in lunch_breaks:
        lunch_start, lunch_end = lunch_breaks[semester]
        if overlaps(start, end, lunch_start, lunch_end):
            return True
    
    return False

def is_minor_slot(slot):
    """Check if slot is a minor slot"""
    start, end = slot
    if start == time(7, 30) and end == time(9, 0):
        return True
    if start == time(18, 30):
        return True
    return False

def select_faculty(faculty_field):
    """Select faculty from potentially multiple options"""
    if pd.isna(faculty_field) or str(faculty_field).strip().lower() in ['nan', 'none', '']:
        return "TBD"
    s = str(faculty_field).strip()
    for sep in ['/', ',', '&', ';']:
        if sep in s:
            return s.split(sep)[0].strip()
    return s

def get_course_priority(row):
    """Calculate course priority based on L, T, P values"""
    try:
        l = int(row.get('L', 0)) if pd.notna(row.get('L', 0)) else 0
        t = int(row.get('T', 0)) if pd.notna(row.get('T', 0)) else 0
        p = int(row.get('P', 0)) if pd.notna(row.get('P', 0)) else 0
        return -(l + t + p)
    except Exception:
        return 0

def calculate_required_minutes(course_row):
    """Calculate required minutes for L, T, P, S"""
    l = int(course_row['L']) if ('L' in course_row and pd.notna(course_row['L'])) else 0
    t = int(course_row['T']) if ('T' in course_row and pd.notna(course_row['T'])) else 0
    p = int(course_row['P']) if ('P' in course_row and pd.notna(course_row['P'])) else 0
    s = int(course_row['S']) if ('S' in course_row and pd.notna(course_row['S'])) else 0
    return (l, t, p, s)

def get_required_room_type(course_row):
    """Determine room type needed"""
    try:
        p = int(course_row['P']) if ('P' in course_row and pd.notna(course_row['P'])) else 0
        return 'COMPUTER_LAB' if p > 0 else 'LECTURE_ROOM'
    except Exception:
        return 'LECTURE_ROOM'

def is_basket_course(code):
    """Check if course is a basket course"""
    return str(code).startswith('B') and '-' in str(code)

def get_basket_group(code):
    """Get basket group identifier"""
    if is_basket_course(code):
        return code.split('-')[0]
    return None

# ---------------------------
# Room Allocation
# ---------------------------
def find_suitable_room_for_slot(course_code, room_type, day, slot_indices, room_schedule, 
                                course_room_mapping, rooms, batch_info):
    """Find suitable room ensuring same course gets same room"""
    if course_code in course_room_mapping:
        fixed_room = course_room_mapping[course_code]
        for si in slot_indices:
            if si in room_schedule.get(fixed_room, {}).get(day, set()):
                return None
        return fixed_room
    
    # Get available rooms
    if room_type == 'COMPUTER_LAB':
        pool = [rid for rid, r in rooms.items() if 'COMPUTER' in r['type'].upper() or 'LAB' in r['type'].upper()]
    else:
        pool = [rid for rid, r in rooms.items() if 'LECTURE' in r['type'].upper() or 'SEATER' in r['type'].upper()]
    
    if not pool:
        return None
    
    random.shuffle(pool)
    for room_id in pool:
        if room_id not in room_schedule:
            room_schedule[room_id] = {d: set() for d in range(len(DAYS))}
        
        if all(si not in room_schedule[room_id][day] for si in slot_indices):
            course_room_mapping[course_code] = room_id
            return room_id
    
    return None

def find_consecutive_slots_for_minutes(timetable, day, start_idx, required_minutes,
                                      semester, professor_schedule, faculty,
                                      room_schedule, room_type, course_code, 
                                      course_room_mapping, rooms, batch_info):
    """Find consecutive slots with enough minutes"""
    n = len(TIME_SLOTS)
    slot_indices = []
    i = start_idx
    accumulated = 0
    
    while i < n and accumulated < required_minutes:
        if is_minor_slot(TIME_SLOTS[i]) or is_break_time_slot(TIME_SLOTS[i], semester):
            return None, None
        
        if timetable[day][i]['type'] is not None:
            return None, None
        
        if faculty in professor_schedule and i in professor_schedule[faculty][day]:
            return None, None
        
        slot_indices.append(i)
        accumulated += slot_minutes(TIME_SLOTS[i])
        i += 1
    
    if accumulated >= required_minutes:
        room = find_suitable_room_for_slot(course_code, room_type, day, slot_indices, 
                                          room_schedule, course_room_mapping, rooms, batch_info)
        if room is not None:
            return slot_indices, room
    
    return None, None

def get_all_possible_start_indices():
    """Get shuffled list of possible start indices"""
    idxs = list(range(len(TIME_SLOTS)))
    random.shuffle(idxs)
    return idxs

def check_professor_availability(professor_schedule, faculty, day, start_idx, duration_slots):
    """Check if professor has sufficient gap between classes"""
    if faculty not in professor_schedule:
        return True
    if not professor_schedule[faculty][day]:
        return True
    
    new_start = TIME_SLOTS[start_idx][0]
    new_start_m = new_start.hour * 60 + new_start.minute
    MIN_GAP = 180
    
    for s in professor_schedule[faculty][day]:
        exist_start = TIME_SLOTS[s][0]
        exist_m = exist_start.hour * 60 + exist_start.minute
        if abs(exist_m - new_start_m) < MIN_GAP:
            return False
    
    return True

# ---------------------------
# Basket Course Scheduling
# ---------------------------
def schedule_basket_groups_globally(all_courses, professor_schedule, rooms, batch_info):
    """Pre-schedule all basket groups to common time slots"""
    global GLOBAL_BASKET_SCHEDULE, SCHEDULED_BASKET_COURSE_CODES
    
    basket_groups_to_schedule = {}
    room_schedule = {rid: {d: set() for d in range(len(DAYS))} for rid in rooms.keys()}
    course_room_mapping = {}
    
    basket_courses_df = all_courses[
        all_courses['Course Code'].astype(str).str.contains('^B[0-9]-', na=False) &
        ((all_courses['Schedule'].fillna('Yes').str.upper() == 'YES') | (all_courses['Schedule'].isna()))
    ].copy()
    
    for _, course in basket_courses_df.iterrows():
        code = str(course['Course Code'])
        group = get_basket_group(code)
        
        if group:
            if group not in basket_groups_to_schedule:
                basket_groups_to_schedule[group] = {
                    'courses': {},
                    'sessions_needed': 0
                }
            
            l, _, _, _ = calculate_required_minutes(course)
            sessions = l
            
            course_details = course.to_dict()
            course_details['sessions_needed'] = sessions
            basket_groups_to_schedule[group]['courses'][code] = course_details
            basket_groups_to_schedule[group]['sessions_needed'] = max(
                basket_groups_to_schedule[group]['sessions_needed'], sessions
            )
    
    for group, group_data in basket_groups_to_schedule.items():
        courses = list(group_data['courses'].values())
        sessions_needed = group_data['sessions_needed']
        
        if sessions_needed == 0:
            continue
        
        for session_num in range(int(sessions_needed)):
            scheduled = False
            attempts = 0
            
            while not scheduled and attempts < 2000:
                attempts += 1
                
                day = random.randint(0, len(DAYS) - 1)
                start_idx = random.randint(0, len(TIME_SLOTS) - 3)
                
                # Calculate duration in slots
                required_minutes = LECTURE_MIN
                duration = 0
                accumulated = 0
                temp_idx = start_idx
                
                while temp_idx < len(TIME_SLOTS) and accumulated < required_minutes:
                    if is_minor_slot(TIME_SLOTS[temp_idx]) or is_break_time_slot(TIME_SLOTS[temp_idx]):
                        break
                    accumulated += slot_minutes(TIME_SLOTS[temp_idx])
                    duration += 1
                    temp_idx += 1
                
                if accumulated < required_minutes:
                    continue
                
                # Check faculty availability
                faculty_map = {str(c['Course Code']): select_faculty(str(c['Faculty'])) for c in courses}
                faculty_free = True
                
                for code, f in faculty_map.items():
                    if f not in professor_schedule:
                        professor_schedule[f] = {d: set() for d in range(len(DAYS))}
                    
                    for i in range(duration):
                        if (start_idx + i) in professor_schedule.get(f, {}).get(day, set()):
                            faculty_free = False
                            break
                    if not faculty_free:
                        break
                
                if not faculty_free:
                    continue
                
                # Check room availability
                potential_rooms = []
                for room_id, room in rooms.items():
                    if 'LECTURE' in room['type'].upper() or 'SEATER' in room['type'].upper():
                        is_free = True
                        for i in range(duration):
                            if (start_idx + i) in room_schedule[room_id][day]:
                                is_free = False
                                break
                        if is_free:
                            potential_rooms.append(room_id)
                
                if len(potential_rooms) < len(courses):
                    continue
                
                # Assign rooms and reserve slots
                assigned_rooms_map = {}
                sorted_courses = sorted(courses, key=lambda c: str(c['Course Code']))
                
                for i, course_detail in enumerate(sorted_courses):
                    course_code = str(course_detail['Course Code'])
                    room_id = potential_rooms[i]
                    faculty = faculty_map[course_code]
                    
                    for j in range(duration):
                        room_schedule[room_id][day].add(start_idx + j)
                        professor_schedule[faculty][day].add(start_idx + j)
                    
                    assigned_rooms_map[course_code] = room_id
                
                GLOBAL_BASKET_SCHEDULE[(group, session_num)] = {
                    'day': day,
                    'start_slot': start_idx,
                    'duration': duration,
                    'rooms': assigned_rooms_map
                }
                
                for course_detail in courses:
                    SCHEDULED_BASKET_COURSE_CODES.add(str(course_detail['Course Code']))
                
                scheduled = True
            
            if not scheduled:
                print(f"Warning: Failed to schedule Basket Group {group} Session {session_num + 1}")

def fill_basket_schedule(timetable, department, semester, global_basket_schedule, basket_courses_df):
    """Fill timetable with pre-scheduled basket courses"""
    section_courses = basket_courses_df[
        (basket_courses_df['Department'] == department) & 
        (basket_courses_df['Semester'] == semester)
    ]
    
    relevant_basket_map = {}
    for _, course in section_courses.iterrows():
        code = str(course['Course Code'])
        group = get_basket_group(code)
        if group:
            if group not in relevant_basket_map:
                relevant_basket_map[group] = []
            relevant_basket_map[group].append(course.to_dict())
    
    for (group, session_num), schedule_data in global_basket_schedule.items():
        if group in relevant_basket_map:
            day = schedule_data['day']
            start_slot = schedule_data['start_slot']
            duration = schedule_data['duration']
            assigned_rooms = schedule_data['rooms']
            
            all_course_details = []
            for course_details in relevant_basket_map[group]:
                course_code = str(course_details['Course Code'])
                if course_code in assigned_rooms:
                    all_course_details.append({
                        'code': course_code,
                        'name': str(course_details['Course Name']),
                        'faculty': select_faculty(str(course_details['Faculty'])),
                        'classroom': assigned_rooms[course_code]
                    })
            
            if not all_course_details:
                continue
            
            all_course_details.sort(key=lambda x: x['code'])
            main_marker_course = all_course_details[0]
            
            for i in range(duration):
                current_slot = start_slot + i
                
                if current_slot < len(TIME_SLOTS):
                    timetable[day][current_slot]['type'] = 'LEC'
                    
                    if i == 0:
                        timetable[day][current_slot]['code'] = main_marker_course['code']
                        timetable[day][current_slot]['name'] = main_marker_course['name']
                        timetable[day][current_slot]['faculty'] = main_marker_course['faculty']
                        timetable[day][current_slot]['classroom'] = main_marker_course['classroom']
                        timetable[day][current_slot]['basket_group_members'] = all_course_details
                    else:
                        timetable[day][current_slot]['code'] = ''
                        timetable[day][current_slot]['name'] = ''
                        timetable[day][current_slot]['faculty'] = ''
                        timetable[day][current_slot]['classroom'] = ''
                        timetable[day][current_slot]['basket_group_members'] = []

# ---------------------------
# Main Timetable Generation
# ---------------------------
def generate_all_timetables():
    global lunch_breaks, GLOBAL_BASKET_SCHEDULE, SCHEDULED_BASKET_COURSE_CODES
    
    initialize_time_slots()
    
    df = load_data()
    if df.empty:
        print("No data to process")
        return []
    
    rooms = load_rooms()
    batch_info = load_batch_data(df)
    
    professor_schedule = {}
    room_schedule = {rid: {d: set() for d in range(len(DAYS))} for rid in rooms.keys()}
    course_room_mapping = {}
    unscheduled_components = []
    
    # Calculate lunch breaks
    all_semesters = sorted(set(
        int(str(sem)[0]) for sem in df['Semester'].unique() 
        if str(sem) and str(sem)[0].isdigit()
    ))
    calculate_lunch_breaks(all_semesters)
    
    # Pre-schedule basket groups
    schedule_basket_groups_globally(df, professor_schedule, rooms, batch_info)
    
    # Create workbook
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    
    overview = wb.create_sheet("Overview")
    overview.append(["Combined Timetable for All Departments and Semesters"])
    overview.append(["Generated on:", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    overview.append([])
    overview.append(["Department", "Semester", "Sheet Name"])
    row_index = 5
    
    # Color palettes
    SUBJECT_COLORS = [
        "FF6B6B", "4ECDC4", "FF9F1C", "5D5FEF", "45B7D1",
        "F72585", "7209B7", "3A0CA3", "4361EE", "4CC9F0",
        "06D6A0", "FFD166", "EF476F", "118AB2", "073B4C"
    ]
    
    basket_group_colors = {
        'B1': "FF9999", 'B2': "99FF99", 'B3': "9999FF", 
        'B4': "FFFF99", 'B5': "FF99FF", 'B6': "99FFFF"
    }
    
    for department in df['Department'].unique():
        sems = sorted(df[df['Department'] == department]['Semester'].unique())
        
        for semester in sems:
            # Determine number of sections (2 for CSE sem 2,4,6)
            dept_upper = str(department).strip().upper()
            num_sections = 2 if (dept_upper == "CSE" and int(semester) in [2, 4, 6]) else 1
            
            courses = df[
                (df['Department'] == department) & 
                (df['Semester'] == semester) &
                (~df['Course Code'].isin(SCHEDULED_BASKET_COURSE_CODES))
            ]
            
            if 'Schedule' in courses.columns:
                courses = courses[
                    (courses['Schedule'].fillna('Yes').str.upper() == 'YES') | 
                    (courses['Schedule'].isna())
                ]
            
            if courses.empty:
                continue
            
            # Priority sorting
            if 'P' in courses.columns:
                lab_courses = courses[courses['P'] > 0].copy()
                non_lab_courses = courses[courses['P'] == 0].copy()
            else:
                lab_courses = courses.head(0)
                non_lab_courses = courses.copy()
            
            if not lab_courses.empty:
                lab_courses['priority'] = lab_courses.apply(get_course_priority, axis=1)
                lab_courses = lab_courses.sort_values('priority', ascending=False)
            
            non_lab_courses['priority'] = non_lab_courses.apply(get_course_priority, axis=1)
            non_lab_courses = non_lab_courses.sort_values('priority', ascending=False)
            
            courses_combined = pd.concat([lab_courses, non_lab_courses])
            
            for section in range(num_sections):
                section_title = f"{department}_{semester}" if num_sections == 1 else f"{department}_{semester}_{chr(65 + section)}"
                ws = wb.create_sheet(title=section_title)
                
                overview.cell(row=row_index, column=1, value=department)
                overview.cell(row=row_index, column=2, value=str(semester))
                overview.cell(row=row_index, column=3, value=section_title)
                row_index += 1
                
                timetable = {
                    d: {
                        s: {'type': None, 'code': '', 'name': '', 'faculty': '', 
                            'classroom': '', 'basket_group_members': []}
                        for s in range(len(TIME_SLOTS))
                    } for d in range(len(DAYS))
                }
                
                # Fill basket courses first
                basket_courses_df = df[df['Course Code'].astype(str).str.contains('^B[0-9]-', na=False)].copy()
                fill_basket_schedule(timetable, department, semester, GLOBAL_BASKET_SCHEDULE, basket_courses_df)
                
                # Create color mapping
                section_subject_color = {}
                color_iter = iter(SUBJECT_COLORS)
                course_faculty_map = {}
                
                for _, c in courses_combined.iterrows():
                    code = str(c.get('Course Code', '')).strip()
                    if code and code not in section_subject_color:
                        if is_basket_course(code):
                            basket_group = get_basket_group(code)
                            section_subject_color[code] = basket_group_colors.get(
                                basket_group, random.choice(SUBJECT_COLORS)
                            )
                        else:
                            try:
                                section_subject_color[code] = next(color_iter)
                            except StopIteration:
                                section_subject_color[code] = random.choice(SUBJECT_COLORS)
                        course_faculty_map[code] = select_faculty(c.get('Faculty', 'TBD'))
                
                # Schedule remaining courses
                for _, course in courses_combined.iterrows():
                    code = str(course.get('Course Code', '')).strip()
                    name = str(course.get('Course Name', '')).strip()
                    faculty = select_faculty(course.get('Faculty', 'TBD'))
                    
                    if faculty not in professor_schedule:
                        professor_schedule[faculty] = {d: set() for d in range(len(DAYS))}
                    
                    lec_count, tut_count, lab_count, ss_count = calculate_required_minutes(course)
                    room_type = get_required_room_type(course)
                    
                    def schedule_component(required_minutes, comp_type, attempts_limit=800):
                        for attempt in range(attempts_limit):
                            day = random.randint(0, len(DAYS) - 1)
                            starts = get_all_possible_start_indices()
                            
                            for start_idx in starts:
                                slot_indices, candidate_room = find_consecutive_slots_for_minutes(
                                    timetable, day, start_idx, required_minutes, semester,
                                    professor_schedule, faculty, room_schedule, room_type,
                                    code, course_room_mapping, rooms, batch_info
                                )
                                
                                if slot_indices is None:
                                    continue
                                
                                if not check_professor_availability(
                                    professor_schedule, faculty, day, slot_indices[0], len(slot_indices)
                                ):
                                    continue
                                
                                if candidate_room is None:
                                    continue
                                
                                # Schedule the component
                                for si_idx, si in enumerate(slot_indices):
                                    timetable[day][si]['type'] = comp_type
                                    timetable[day][si]['code'] = code if si_idx == 0 else ''
                                    timetable[day][si]['name'] = name if si_idx == 0 else ''
                                    timetable[day][si]['faculty'] = faculty if si_idx == 0 else ''
                                    timetable[day][si]['classroom'] = candidate_room if si_idx == 0 else ''
                                    professor_schedule[faculty][day].add(si)
                                    if candidate_room not in room_schedule:
                                        room_schedule[candidate_room] = {d: set() for d in range(len(DAYS))}
                                    room_schedule[candidate_room][day].add(si)
                                
                                return True
                        return False
                    
                    # Schedule lectures
                    for _ in range(lec_count):
                        ok = schedule_component(LECTURE_MIN, 'LEC', attempts_limit=800)
                        if not ok:
                            unscheduled_components.append(
                                UnscheduledComponent(department, semester, code, name, faculty,
                                                    'LEC', 1, section, "No slot available")
                            )
                    
                    # Schedule tutorials
                    for _ in range(tut_count):
                        ok = schedule_component(TUTORIAL_MIN, 'TUT', attempts_limit=600)
                        if not ok:
                            unscheduled_components.append(
                                UnscheduledComponent(department, semester, code, name, faculty,
                                                    'TUT', 1, section, "No slot available")
                            )
                    
                    # Schedule labs
                    for _ in range(lab_count):
                        ok = schedule_component(LAB_MIN, 'LAB', attempts_limit=800)
                        if not ok:
                            unscheduled_components.append(
                                UnscheduledComponent(department, semester, code, name, faculty,
                                                    'LAB', 1, section, "Lab not scheduled")
                            )
                    
                    # Schedule self-study
                    for _ in range(ss_count):
                        ok = schedule_component(SELF_STUDY_MIN, 'SS', attempts_limit=400)
                        if not ok:
                            unscheduled_components.append(
                                UnscheduledComponent(department, semester, code, name, faculty,
                                                    'SS', 1, section, "Self-study not scheduled")
                            )
                
                # Write to worksheet
                header = ['Day'] + [f"{slot[0].strftime('%H:%M')}-{slot[1].strftime('%H:%M')}" 
                                   for slot in TIME_SLOTS]
                ws.append(header)
                
                # Header formatting
                header_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
                header_font = Font(bold=True)
                header_alignment = Alignment(horizontal='center', vertical='center')
                
                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = header_alignment
                
                border = Border(
                    left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin')
                )
                
                break_fill = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")
                minor_fill = PatternFill(start_color="9ACD32", end_color="9ACD32", fill_type="solid")
                
                for day_idx, day_name in enumerate(DAYS):
                    ws.append([day_name] + [''] * len(TIME_SLOTS))
                    row_num = ws.max_row
                    merges = []
                    
                    for slot_idx in range(len(TIME_SLOTS)):
                        cell_obj = ws.cell(row=row_num, column=slot_idx + 2)
                        
                        if is_minor_slot(TIME_SLOTS[slot_idx]):
                            cell_obj.value = "Minor Slot"
                            cell_obj.fill = minor_fill
                            cell_obj.font = Font(bold=True)
                            cell_obj.alignment = Alignment(horizontal='center', vertical='center')
                            cell_obj.border = border
                            continue
                        
                        if is_break_time_slot(TIME_SLOTS[slot_idx], semester):
                            cell_obj.value = "BREAK"
                            cell_obj.fill = break_fill
                            cell_obj.font = Font(bold=True)
                            cell_obj.alignment = Alignment(horizontal='center', vertical='center')
                            cell_obj.border = border
                            continue
                        
                        if timetable[day_idx][slot_idx]['type'] is None:
                            cell_obj.border = border
                            continue
                        
                        typ = timetable[day_idx][slot_idx]['type']
                        code = timetable[day_idx][slot_idx]['code']
                        cls = timetable[day_idx][slot_idx]['classroom']
                        fac = timetable[day_idx][slot_idx]['faculty']
                        
                        if code:
                            span = [slot_idx]
                            j = slot_idx + 1
                            while j < len(TIME_SLOTS) and timetable[day_idx][j]['type'] is not None and timetable[day_idx][j]['code'] == '':
                                span.append(j)
                                j += 1
                            
                            # Check for basket group
                            if 'basket_group_members' in timetable[day_idx][slot_idx] and timetable[day_idx][slot_idx]['basket_group_members']:
                                basket_group_members = timetable[day_idx][slot_idx]['basket_group_members']
                                basket_group = get_basket_group(code)
                                basket_header = f"{basket_group} Courses\n"
                                course_details = [
                                    f"{c['code']}: {c['faculty']} ({c['classroom']})"
                                    for c in basket_group_members
                                ]
                                display = basket_header + "\n".join(course_details)
                            else:
                                display = f"{typ}\nroom no. :{cls}\n{fac}"
                            
                            if code in section_subject_color:
                                subj_color = section_subject_color[code]
                                fill = PatternFill(start_color=subj_color, end_color=subj_color, fill_type="solid")
                            else:
                                fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")
                            
                            cell_obj.value = display
                            cell_obj.fill = fill
                            merges.append((slot_idx + 2, slot_idx + 2 + len(span) - 1, display, fill))
                        
                        cell_obj.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
                        cell_obj.border = border
                    
                    for start_col, end_col, val, fill in merges:
                        if end_col > start_col:
                            rng = f"{get_column_letter(start_col)}{row_num}:{get_column_letter(end_col)}{row_num}"
                            try:
                                ws.merge_cells(rng)
                                mc = ws[f"{get_column_letter(start_col)}{row_num}"]
                                mc.value = val
                                mc.fill = fill
                                mc.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
                                mc.border = border
                            except Exception:
                                pass
                
                for col_idx in range(1, len(TIME_SLOTS) + 2):
                    try:
                        ws.column_dimensions[get_column_letter(col_idx)].width = 15
                    except Exception:
                        pass
                
                for row in ws.iter_rows(min_row=2, max_row=len(DAYS) + 1):
                    ws.row_dimensions[row[0].row].height = 40
                
                # Add self-study only courses section
                current_row = len(DAYS) + 4
                
                ss_courses_for_this_section = []
                for _, course in courses_combined.iterrows():
                    l = int(course['L']) if pd.notna(course['L']) else 0
                    t = int(course['T']) if pd.notna(course['T']) else 0
                    p = int(course['P']) if pd.notna(course['P']) else 0
                    s = int(course['S']) if pd.notna(course['S']) else 0
                    if s > 0 and l == 0 and t == 0 and p == 0:
                        ss_courses_for_this_section.append({
                            'code': str(course['Course Code']),
                            'name': str(course['Course Name']),
                            'faculty': str(course['Faculty'])
                        })
                
                if ss_courses_for_this_section:
                    ws.cell(row=current_row, column=1, value="Self-Study Only Courses")
                    ws.cell(row=current_row, column=1).font = Font(bold=True)
                    current_row += 1
                    
                    headers = ['Course Code', 'Course Name', 'Faculty']
                    for col, header in enumerate(headers, 1):
                        ws.cell(row=current_row, column=col, value=header)
                        ws.cell(row=current_row, column=col).font = Font(bold=True)
                    current_row += 1
                    
                    for course in ss_courses_for_this_section:
                        ws.cell(row=current_row, column=1, value=course['code'])
                        ws.cell(row=current_row, column=2, value=course['name'])
                        ws.cell(row=current_row, column=3, value=course['faculty'])
                        current_row += 1
                    
                    current_row += 2
                
                # Add legend
                legend_title = ws.cell(row=current_row, column=1, value="Legend")
                legend_title.font = Font(bold=True, size=12)
                current_row += 2
                
                ws.column_dimensions['A'].width = 20
                ws.column_dimensions['B'].width = 10
                ws.column_dimensions['C'].width = 40
                ws.column_dimensions['D'].width = 30
                ws.column_dimensions['E'].width = 15
                ws.column_dimensions['F'].width = 15
                
                legend_headers = ['Subject Code', 'Color', 'Subject Name', 'Faculty', 'LTPS', 'Room']
                for col, header in enumerate(legend_headers, 1):
                    cell = ws.cell(row=current_row, column=col, value=header)
                    cell.font = Font(bold=True)
                    cell.border = border
                    cell.fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                current_row += 1
                
                for code, color in section_subject_color.items():
                    assigned_room = course_room_mapping.get(code, "—")
                    if not assigned_room or assigned_room == "—":
                        continue
                    
                    ws.row_dimensions[current_row].height = 30
                    
                    ltps_value = ""
                    for _, course_row in courses_combined.iterrows():
                        if str(course_row['Course Code']) == code:
                            l = str(int(course_row['L'])) if pd.notna(course_row['L']) else "0"
                            t = str(int(course_row['T'])) if pd.notna(course_row['T']) else "0"
                            p = str(int(course_row['P'])) if pd.notna(course_row['P']) else "0"
                            s = str(int(course_row['S'])) if pd.notna(course_row['S']) and 'S' in course_row else "0"
                            ltps_value = f"{l}-{t}-{p}-{s}"
                            break
                    
                    course_name = ''
                    fac_name = ''
                    if code in course_faculty_map:
                        fac_name = course_faculty_map[code]
                        for _, cr in courses_combined.iterrows():
                            if str(cr['Course Code']) == code:
                                course_name = str(cr['Course Name'])
                                break
                    
                    cells = [
                        (code, None),
                        ('', PatternFill(start_color=color, end_color=color, fill_type="solid")),
                        (course_name, None),
                        (fac_name, None),
                        (ltps_value, None),
                        (assigned_room, None)
                    ]
                    
                    for col, (value, fill) in enumerate(cells, 1):
                        cell = ws.cell(row=current_row, column=col, value=value)
                        cell.border = border
                        if fill:
                            cell.fill = fill
                        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True, indent=2)
                    
                    current_row += 1
    
    # Format overview sheet
    for col in range(1, 4):
        overview.column_dimensions[get_column_letter(col)].width = 20
    
    for row_ in overview.iter_rows(min_row=1, max_row=4):
        for cell in row_:
            cell.font = Font(bold=True)
    
    for cell in overview[4]:
        cell.fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
        cell.font = Font(bold=True)
        cell.border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
    
    for row_ in overview.iter_rows(min_row=5, max_row=row_index - 1):
        for cell in row_:
            cell.border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
    
    # Save workbook
    out_filename = os.path.join(OUTPUT_DIR, "timetable_all_departments.xlsx")
    try:
        wb.save(out_filename)
        print(f"Combined timetable saved as {out_filename}")
    except Exception as e:
        print(f"Failed to save combined timetable: {e}")
        traceback.print_exc()
    
    # Create teacher and unscheduled workbooks
    try:
        create_teacher_and_unscheduled_from_combined(out_filename, unscheduled_components)
    except Exception as e:
        print("Failed to generate teacher/unscheduled workbooks:", e)
        traceback.print_exc()
    
    return out_filename

# ---------------------------
# Teacher and Unscheduled Helper
# ---------------------------
def split_faculty_names(fac_str):
    """Split faculty string into separate names"""
    if fac_str is None:
        return []
    s = str(fac_str).strip()
    if s == '' or s.lower() in ['nan', 'none']:
        return []
    parts = [s]
    for sep in ['/', ',', '&', ';']:
        if sep in s:
            parts = [p.strip() for p in s.split(sep) if p.strip()]
            break
    return parts if parts else [s]

def parse_cell_for_course(cell_value):
    """Extract course info from timetable cell"""
    if cell_value is None:
        return ('', '', '', '')
    text = str(cell_value).strip()
    if text == '':
        return ('', '', '', '')
    
    lines = [ln.strip() for ln in text.splitlines() if ln.strip()]
    faculty = ''
    room = ''
    code = ''
    typ = ''
    
    for ln in lines:
        if 'room no' in ln.lower():
            parts = ln.split(':')
            if len(parts) >= 2:
                room = parts[-1].strip()
            else:
                room = ln.strip()
    
    if len(lines) >= 1:
        last = lines[-1]
        if 'room no' not in last.lower() and 'courses' not in last.lower() and ':' not in last:
            faculty = last
    
    first = lines[0] if lines else ''
    if first:
        tokens = first.split()
        if len(tokens) >= 2 and tokens[1].upper() in ['LEC', 'LAB', 'TUT', 'SS']:
            code = tokens[0].strip()
            typ = tokens[1].strip().upper()
        else:
            code = tokens[0].strip() if tokens else ''
            for t in ['LEC', 'LAB', 'TUT', 'SS']:
                if t in text.upper():
                    typ = t
                    break
    
    if not faculty and len(lines) >= 2:
        for cand in lines[1:]:
            if any(ch.isalpha() for ch in cand) and 'room no' not in cand.lower() and 'courses' not in cand.lower() and ':' not in cand:
                faculty = cand
                break
    
    return (code, typ, room, faculty)

def create_teacher_and_unscheduled_from_combined(timetable_filename, unscheduled_components):
    """Create teacher timetables and unscheduled courses workbooks"""
    try:
        wb = load_workbook(timetable_filename, data_only=True)
    except Exception as e:
        print(f"Failed to open {timetable_filename}: {e}")
        return
    
    teacher_slots = {}
    slot_headers = []
    
    for sheetname in wb.sheetnames:
        if sheetname.lower() == 'overview':
            continue
        ws = wb[sheetname]
        header = [str(ws.cell(1, c).value).strip() if ws.cell(1, c).value else '' 
                 for c in range(2, ws.max_column + 1)]
        if len(header) > len(slot_headers):
            slot_headers = header
        
        for r in range(2, ws.max_row + 1):
            day = ws.cell(r, 1).value
            if not day or str(day) not in DAYS:
                break
            day_idx = DAYS.index(day)
            
            for c in range(2, ws.max_column + 1):
                code, typ, room, faculty = parse_cell_for_course(ws.cell(r, c).value)
                for f in split_faculty_names(faculty):
                    if not f:
                        continue
                    if str(f).strip().upper() in ["BREAK", "MINOR SLOT", "NAN", "NONE", ""]:
                        continue
                    
                    teacher_slots.setdefault(f, {
                        d: {i: '' for i in range(len(slot_headers))} 
                        for d in range(len(DAYS))
                    })
                    teacher_slots[f][day_idx][c - 2] = f"{code} {typ}\n({sheetname})\nRoom: {room}" if code else ''
    
    # Create teacher workbook
    twb = Workbook()
    if "Sheet" in twb.sheetnames:
        twb.remove(twb["Sheet"])
    
    header_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
    alt_fill = PatternFill(start_color="FFF8DC", end_color="FFF8DC", fill_type="solid")
    border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    header_font = Font(bold=True, size=12)
    title_font = Font(bold=True, size=14)
    cell_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    for teacher in sorted(teacher_slots.keys()):
        safe_name = teacher[:31] or "Unknown"
        ws = twb.create_sheet(title=safe_name)
        
        ws.merge_cells("A1:{}1".format(get_column_letter(len(slot_headers) + 1)))
        title_cell = ws.cell(row=1, column=1, value=f"{teacher} — Weekly Timetable")
        title_cell.font = title_font
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        
        ws.append(["Day"] + slot_headers)
        for cell in ws[2]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
        
        for d, day in enumerate(DAYS):
            row = [day] + [teacher_slots[teacher][d][i] for i in range(len(slot_headers))]
            ws.append(row)
            row_idx = ws.max_row
            
            if d % 2 == 0:
                for cell in ws[row_idx]:
                    cell.fill = alt_fill
            
            for cell in ws[row_idx]:
                cell.alignment = cell_align
                cell.border = border
            ws.row_dimensions[row_idx].height = 35
        
        ws.column_dimensions["A"].width = 15
        for col in range(2, len(slot_headers) + 2):
            ws.column_dimensions[get_column_letter(col)].width = 20
    
    twb.save(os.path.join(OUTPUT_DIR, "teacher_timetables.xlsx"))
    print("✅ Saved formatted teacher_timetables.xlsx")
    
    # Unscheduled courses workbook
    uwb = Workbook()
    ws = uwb.active
    ws.title = "Unscheduled Courses"
    
    headers = ["Course Code", "Department", "Semester", "Reason"]
    ws.append(headers)
    
    unscheduled_unique = {}
    
    for u in unscheduled_components:
        if u.code not in unscheduled_unique:
            if hasattr(u, "reason") and u.reason and len(str(u.reason).strip()) > 0:
                reason_text = str(u.reason).strip()
            else:
                if "collision" in str(u.component_type).lower():
                    reason_text = "Number of collisions exceeded limit"
                elif "slot" in str(u.component_type).lower() or "no slot" in str(u.reason).lower():
                    reason_text = "No slot available"
                elif "faculty" in str(u.reason).lower():
                    reason_text = "Faculty unavailable"
                else:
                    reason_text = "Unspecified scheduling issue"
            
            unscheduled_unique[u.code] = {
                "Course Code": u.code,
                "Department": u.department,
                "Semester": u.semester,
                "Reason": reason_text
            }
    
    for entry in unscheduled_unique.values():
        ws.append([entry[h] for h in headers])
    
    uwb.save(os.path.join(OUTPUT_DIR, "unscheduled_courses.xlsx"))
    print(f"✅ Saved unscheduled_courses.xlsx with {len(unscheduled_unique)} unique courses")

# ---------------------------
# Exam Timetable Generator
# ---------------------------
def exam_generator():
    """Generate exam timetable with room allocation and invigilation schedule"""
    try:
        df_courses = pd.read_csv(os.path.join(INPUT_DIR, "combined.csv"))
        df_rooms = pd.read_csv(os.path.join(INPUT_DIR, "rooms.csv"))
    except FileNotFoundError as e:
        print(f"❌ Missing file: {e}")
        return None
    
    df_courses = df_courses.dropna(subset=["Course Code", "Course Name", "Faculty", "Department", "Semester"])
    if "total_students" not in df_courses.columns:
        df_courses["total_students"] = 50
    df_courses["total_students"] = df_courses["total_students"].fillna(50).astype(int)
    
    df_rooms.columns = [c.strip().lower() for c in df_rooms.columns]
    
    def find_col(keywords):
        for c in df_rooms.columns:
            if any(k in c for k in keywords):
                return c
        return None
    
    room_col = find_col(["room", "num", "id"])
    cap_col = find_col(["cap", "seat"])
    type_col = find_col(["type"])
    
    if not room_col or not cap_col:
        print("❌ rooms.csv must have columns for room number and capacity.")
        return None
    
    df_rooms = df_rooms.rename(columns={room_col: "room", cap_col: "capacity"})
    if type_col:
        df_rooms = df_rooms.rename(columns={type_col: "type"})
    else:
        df_rooms["type"] = "LECTURE_ROOM"
    
    df_rooms["room"] = df_rooms["room"].astype(str).str.strip()
    df_rooms["capacity"] = pd.to_numeric(df_rooms["capacity"], errors="coerce").fillna(0).astype(int)
    df_rooms = df_rooms[df_rooms["capacity"] > 0].sort_values(by="capacity").reset_index(drop=True)
    
    print(f"✅ Loaded {len(df_rooms)} rooms for exam scheduling")
    
    faculty_list = list(set(sum([str(f).replace(" and ", "/").replace(",", "/").split("/") 
                                 for f in df_courses["Faculty"]], [])))
    faculty_list = [f.strip() for f in faculty_list if f.strip()]
    
    start_date = datetime(2025, 11, 20)
    num_days = min(10, len(df_courses))
    dates = [start_date + timedelta(days=i) for i in range(num_days)]
    days = [d.strftime("%A") for d in dates]
    
    shuffled = df_courses.sample(frac=1, random_state=42).reset_index(drop=True)
    course_date_map = {row["Course Code"]: dates[i % len(dates)] for i, row in shuffled.iterrows()}
    
    date_room_usage = {d.strftime("%d-%b-%Y"): set() for d in dates}
    invigilation_entries = []
    
    for date in dates:
        date_str = date.strftime("%d-%b-%Y")
        today_courses = shuffled[shuffled["Course Code"].isin(
            [c for c, dt in course_date_map.items() if dt == date]
        )]
        
        for _, course in today_courses.iterrows():
            code = course["Course Code"]
            name = course["Course Name"]
            dept = course["Department"]
            sem = course["Semester"]
            teacher = str(course["Faculty"]).strip()
            students = int(course["total_students"])
            time_slot = "03:00 PM—04:30 PM"
            
            assigned_rooms = []
            remaining = students
            
            available = df_rooms[~df_rooms["room"].isin(date_room_usage[date_str])].copy()
            available = available.sort_values(by="capacity", ascending=True)
            
            suitable = available[available["capacity"] >= remaining]
            if not suitable.empty:
                best = suitable.iloc[0]
                assigned_rooms = [best["room"]]
                date_room_usage[date_str].add(best["room"])
            else:
                total_cap = 0
                for _, room_row in available.iterrows():
                    assigned_rooms.append(room_row["room"])
                    total_cap += room_row["capacity"]
                    date_room_usage[date_str].add(room_row["room"])
                    if total_cap >= remaining:
                        break
                if total_cap < remaining:
                    print(f"⚠️ Not enough capacity for {code} ({students} students)")
            
            available_teachers = [f for f in faculty_list if f.lower() not in teacher.lower()]
            for room in assigned_rooms:
                invigilator = random.choice(available_teachers) if available_teachers else "TBD"
                invigilation_entries.append({
                    "Faculty": invigilator,
                    "Date": date_str,
                    "Time": time_slot,
                    "Course Code": code,
                    "Course Name": name,
                    "Department": dept,
                    "Semester": sem,
                    "Room": room,
                    "Strength": students
                })
    
    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Exam Timetable"
    
    bold_center = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    header_fill = PatternFill(start_color="BFBFBF", end_color="BFBFBF", fill_type="solid")
    border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    
    session_title = "Jan-April 03:00 PM to 04:30 PM"
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(dates) + 1)
    title = ws.cell(row=1, column=1, value=session_title)
    title.font = Font(bold=True, size=14)
    title.alignment = center
    title.fill = header_fill
    title.border = border
    
    ws.cell(row=2, column=1, value="Date").font = bold_center
    for i, d in enumerate(dates):
        c = ws.cell(row=2, column=i + 2, value=d.strftime("%d-%b-%Y"))
        c.font = bold_center
        c.alignment = center
        c.fill = header_fill
        c.border = border
    
    ws.cell(row=3, column=1, value="Days").font = bold_center
    for i, day in enumerate(days):
        c = ws.cell(row=3, column=i + 2, value=day)
        c.font = bold_center
        c.alignment = center
        c.fill = header_fill
        c.border = border
    
    grouped_by_date = {}
    for e in invigilation_entries:
        grouped_by_date.setdefault(e["Date"], set()).add(e["Course Code"])
    max_rows = max(len(v) for v in grouped_by_date.values())
    
    for r in range(max_rows):
        for i, d in enumerate(dates):
            code_list = list(grouped_by_date.get(d.strftime("%d-%b-%Y"), []))
            val = code_list[r] if r < len(code_list) else ""
            cell = ws.cell(row=r + 4, column=i + 2, value=val)
            cell.alignment = center
            cell.border = border
    
    ws.column_dimensions["A"].width = 15
    for col in range(2, len(dates) + 2):
        ws.column_dimensions[get_column_letter(col)].width = 16
    
    # Invigilation schedule
    ws2 = wb.create_sheet("Exam Invigilation Schedule")
    headers = ["Faculty", "Date", "Time", "Course Code", "Course Name", 
               "Department", "Semester", "Room", "Strength"]
    ws2.append(headers)
    
    for i, h in enumerate(headers, 1):
        cell = ws2.cell(row=1, column=i, value=h)
        cell.font = bold_center
        cell.alignment = center
        cell.fill = header_fill
        cell.border = border
    
    for entry in invigilation_entries:
        ws2.append([entry[h] for h in headers])
    
    for col in range(1, len(headers) + 1):
        ws2.column_dimensions[get_column_letter(col)].width = 22
    for r in range(2, ws2.max_row + 1):
        for c in ws2[r]:
            c.alignment = center
            c.border = border
    
    exam_file = os.path.join(OUTPUT_DIR, "Exam_timetable.xlsx")
    wb.save(exam_file)
    print(f"✅ Exam timetable saved → {exam_file}")
    return exam_file

# ---------------------------
# Main Execution
# ---------------------------
if __name__ == "__main__":
    try:
        print("=" * 60)
        print("Starting Enhanced Timetable Generation")
        print("=" * 60)
        
        # Generate class timetables
        print("\n[1/2] Generating class and teacher timetables...")
        generate_all_timetables()
        
        # Generate exam timetable
        print("\n[2/2] Generating exam timetable...")
        exam_generator()
        
        print("\n" + "=" * 60)
        print("✅ All timetables generated successfully!")
        print("=" * 60)
        print(f"\nOutput files saved in: {OUTPUT_DIR}")
        print("  - timetable_all_departments.xlsx")
        print("  - teacher_timetables.xlsx")
        print("  - unscheduled_courses.xlsx")
        print("  - Exam_timetable.xlsx")
        
    except Exception as e:
        print("\n" + "=" * 60)
        print("❌ Error running timetable generation:")
        print("=" * 60)
        print(str(e))
        traceback.print_exc()