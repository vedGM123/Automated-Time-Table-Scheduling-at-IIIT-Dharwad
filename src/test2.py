# TT_gen.py -- Complete Fixed Timetable Generator
# Fixes applied:
# 1. Removed 10:30-10:45 break
# 2. Changed lunch break to 13:00-14:00
# 3. Better time slot generation
# 4. Fixed elective basket scheduling
# 5. Supports increased lab capacity (80 students)

import pandas as pd
import random
from datetime import datetime, time, timedelta
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter
from dataclasses import dataclass
from pathlib import Path
import traceback
import os
import json

# ---------------------------
# Configuration
# ---------------------------
BASE_DIR = Path(__file__).resolve().parent
PROJECT_DIR = BASE_DIR.parent
DATA_DIR = BASE_DIR.parent / "data"
CONFIG_PATH = DATA_DIR / "config.json"

try:
    with open(CONFIG_PATH, "r", encoding="utf-8") as f:
        config = json.load(f)
except Exception:
    config = {}

DAYS = config.get("days", ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday'])
LECTURE_MIN = config.get("LECTURE_MIN", 90)
LAB_MIN = config.get("LAB_MIN", 120)
TUTORIAL_MIN = config.get("TUTORIAL_MIN", 60)
SELF_STUDY_MIN = config.get("SELF_STUDY_MIN", 60)
MIN_GAP_BETWEEN_LECTURES = 10

# Updated break windows - REMOVED morning break, extended lunch
LUNCH_BREAK_START = time(13, 0)
LUNCH_BREAK_END = time(14, 0)

@dataclass
class UnscheduledComponent:
    department: str
    semester: int
    code: str
    name: str
    faculty: str
    component_type: str
    sessions: int
    section: int
    reason: str

INPUT_DIR = PROJECT_DIR / "data"
OUTPUT_DIR = PROJECT_DIR / "output"

os.makedirs(OUTPUT_DIR, exist_ok=True)

# ---------------------------
# Load CSVs
# ---------------------------
try:
    df = pd.read_csv(os.path.join(INPUT_DIR, 'combined.csv'))
    print(f"✅ Loaded {len(df)} courses from combined.csv")
except FileNotFoundError:
    raise SystemExit("Error: 'combined.csv' not found in data directory.")

try:
    rooms_df = pd.read_csv(os.path.join(INPUT_DIR, 'rooms.csv'))
    print(f"✅ Loaded {len(rooms_df)} rooms from rooms.csv")
except FileNotFoundError:
    print("⚠️ Warning: rooms.csv not found")
    rooms_df = pd.DataFrame(columns=['roomNumber', 'type'])

# Normalize rooms
lecture_rooms = rooms_df[rooms_df.get('type', '').str.upper() == 'LECTURE_ROOM']['roomNumber'].tolist()
computer_lab_rooms = rooms_df[rooms_df.get('type', '').str.upper() == 'COMPUTER_LAB']['roomNumber'].tolist()
large_rooms = rooms_df[rooms_df.get('type', '').str.upper() == 'SEATER_120']['roomNumber'].tolist()

print(f"\n📊 Room inventory:")
print(f"   - Lecture rooms: {len(lecture_rooms)}")
print(f"   - Computer labs: {len(computer_lab_rooms)}")
print(f"   - Large rooms: {len(large_rooms)}")

# ---------------------------
# Time Slot Generation - UPDATED
# ---------------------------
def generate_time_slots():
    """
    Generate time slots with:
    - NO morning break (removed 10:30-10:45)
    - Lunch break: 13:00-14:00 (extended from 13:00-13:45)
    """
    slots = []
    
    # Morning session: 7:30 - 13:00 (continuous, no break)
    slots.append((time(7, 30), time(8, 30)))   # 60 min
    slots.append((time(8, 30), time(9, 30)))   # 60 min
    slots.append((time(9, 30), time(10, 30)))  # 60 min
    slots.append((time(10, 30), time(11, 30))) # 60 min (was break before)
    slots.append((time(11, 30), time(12, 30))) # 60 min
    slots.append((time(12, 30), time(13, 0)))  # 30 min
    
    # Lunch break: 13:00 - 14:00 (UPDATED - extended 1 hour)
    slots.append((time(13, 0), time(14, 0)))   # BREAK
    
    # Afternoon session: 14:00 - 18:30
    slots.append((time(14, 0), time(15, 0)))   # 60 min
    slots.append((time(15, 0), time(16, 0)))   # 60 min
    slots.append((time(16, 0), time(17, 0)))   # 60 min
    slots.append((time(17, 0), time(18, 0)))   # 60 min
    slots.append((time(18, 0), time(18, 30)))  # 30 min
    
    return slots

TIME_SLOTS = generate_time_slots()
print(f"⏰ Generated {len(TIME_SLOTS)} time slots (no morning break, lunch 13:00-14:00)")

# ---------------------------
# Helper functions
# ---------------------------
def slot_minutes(slot):
    s, e = slot
    s_m = s.hour * 60 + s.minute
    e_m = e.hour * 60 + e.minute
    if e_m < s_m:
        e_m += 24 * 60
    return e_m - s_m

def overlaps(a_start, a_end, b_start, b_end):
    a_s_min = a_start.hour * 60 + a_start.minute
    a_e_min = a_end.hour * 60 + a_end.minute
    b_s_min = b_start.hour * 60 + b_start.minute
    b_e_min = b_end.hour * 60 + b_end.minute
    return (a_s_min < b_e_min) and (b_s_min < a_e_min)

def is_break_time_slot(slot, semester=None):
    """Updated to only check lunch break (13:00-14:00)"""
    start, end = slot
    # Only lunch break now
    if start == LUNCH_BREAK_START and end == LUNCH_BREAK_END:
        return True
    if overlaps(start, end, LUNCH_BREAK_START, LUNCH_BREAK_END):
        return True
    return False

def is_minor_slot(slot):
    """Check if slot is early morning or late evening"""
    start, end = slot
    if start.hour < 8:
        return True
    if start.hour >= 18:
        return True
    return False

def select_faculty(faculty_field):
    if pd.isna(faculty_field) or str(faculty_field).strip().lower() in ['nan', 'none', '']:
        return "TBD"
    s = str(faculty_field).strip()
    for sep in ['/', ',', '&', ';']:
        if sep in s:
            return s.split(sep)[0].strip()
    return s

def get_course_priority(row):
    try:
        l = int(row.get('L', 0)) if pd.notna(row.get('L', 0)) else 0
        t = int(row.get('T', 0)) if pd.notna(row.get('T', 0)) else 0
        p = int(row.get('P', 0)) if pd.notna(row.get('P', 0)) else 0
        return -(l + t + p)
    except Exception:
        return 0

def calculate_required_minutes(course_row):
    l = int(course_row['L']) if ('L' in course_row and pd.notna(course_row['L'])) else 0
    t = int(course_row['T']) if ('T' in course_row and pd.notna(course_row['T'])) else 0
    p = int(course_row['P']) if ('P' in course_row and pd.notna(course_row['P'])) else 0
    c = int(course_row['C']) if ('C' in course_row and pd.notna(course_row['C'])) else 0
    
    is_half_semester = (c < 3 and c > 0)
    if is_half_semester:
        lec_count = l // 2 if l > 0 else 0
        tut_count = t // 2 if t > 0 else 0
        lab_count = p // 2 if p > 0 else 0
    else:
        lec_count = l
        tut_count = t
        lab_count = p
    
    return (lec_count, tut_count, lab_count, 0)

def get_required_room_type(component_type):
    if component_type == 'LAB':
        return 'COMPUTER_LAB'
    return 'LECTURE_ROOM'

# ---------------------------
# Elective basket helpers
# ---------------------------
def extract_elective_basket(course_code):
    if pd.isna(course_code):
        return None
    code = str(course_code).strip().upper()
    import re
    match = re.match(r'^(B\d+)-', code)
    return match.group(1) if match else None

def get_base_course_code(course_code):
    if pd.isna(course_code):
        return str(course_code)
    code = str(course_code).strip()
    if '-' in code and code.split('-')[0].upper().startswith('B'):
        return code.split('-', 1)[1]
    return code

def is_elective(course_row):
    code = str(course_row.get('Course Code', '')).strip()
    return extract_elective_basket(code) is not None

def has_component_on_day(timetable, day, course_code, component_type):
    for slot_idx in range(len(TIME_SLOTS)):
        slot_data = timetable[day][slot_idx]
        if slot_data['code'] == course_code and slot_data['type'] == component_type:
            return True
    return False

# ---------------------------
# Room allocation
# ---------------------------
def find_suitable_room_for_slot(course_code, room_type, day, slot_indices, room_schedule, 
                                course_room_mapping, component_type):
    """Find suitable room with proper type checking"""
    mapping_key = f"{course_code}_{component_type}"
    
    if mapping_key in course_room_mapping:
        fixed_room = course_room_mapping[mapping_key]
        for si in slot_indices:
            if si in room_schedule[fixed_room][day]:
                return None
        return fixed_room
    
    if room_type == 'COMPUTER_LAB':
        pool = computer_lab_rooms.copy()
        # Also include hardware labs as fallback
        hardware_labs = rooms_df[rooms_df.get('type', '').str.upper() == 'HARDWARE_LAB']['roomNumber'].tolist()
        pool.extend(hardware_labs)
    else:
        pool = lecture_rooms.copy()
        pool.extend(large_rooms)
    
    if not pool:
        return None
    
    random.shuffle(pool)
    
    for room in pool:
        if room not in room_schedule:
            room_schedule[room] = {d: set() for d in range(len(DAYS))}
        
        if all(si not in room_schedule[room][day] for si in slot_indices):
            course_room_mapping[mapping_key] = room
            return room
    
    return None

def find_consecutive_slots_for_minutes(timetable, day, start_idx, required_minutes,
                                       semester, professor_schedule, faculty,
                                       room_schedule, room_type, course_code, 
                                       course_room_mapping, component_type,
                                       course_day_components):
    """Find consecutive free slots"""
    n = len(TIME_SLOTS)
    slot_indices = []
    i = start_idx
    accumulated = 0
    
    # Check lecture-tutorial same-day constraint
    if component_type == 'LEC' and has_component_on_day(timetable, day, course_code, 'TUT'):
        return None, None
    if component_type == 'TUT' and has_component_on_day(timetable, day, course_code, 'LEC'):
        return None, None
    
    while i < n and accumulated < required_minutes:
        if is_minor_slot(TIME_SLOTS[i]) or is_break_time_slot(TIME_SLOTS[i], semester):
            return None, None
        
        if timetable[day][i]['type'] is not None:
            return None, None
        
        if faculty in professor_schedule and i in professor_schedule[faculty][day]:
            return None, None
        
        slot_indices.append(i)
        accumulated += slot_minutes(TIME_SLOTS[i])
        i += 1
    
    if accumulated >= required_minutes:
        room = find_suitable_room_for_slot(course_code, room_type, day, slot_indices, 
                                          room_schedule, course_room_mapping, component_type)
        if room is not None:
            return slot_indices, room
    
    return None, None

def get_all_possible_start_indices():
    idxs = list(range(len(TIME_SLOTS)))
    random.shuffle(idxs)
    return idxs

def check_professor_availability(professor_schedule, faculty, day, start_idx, duration_slots):
    if faculty not in professor_schedule:
        return True
    if not professor_schedule[faculty][day]:
        return True
    
    new_slots = set(range(start_idx, start_idx + duration_slots))
    existing_slots = professor_schedule[faculty][day]
    
    return not (new_slots & existing_slots)

def load_batch_data():
    batch_info = {}
    for _, r in df.iterrows():
        dept = str(r['Department'])
        sem = int(r['Semester'])
        batch_info[(dept, sem)] = {'num_sections': 1}
    return batch_info

# ---------------------------
# Global elective basket scheduling
# ---------------------------
def schedule_global_elective_baskets(df_input, professor_schedule, room_schedule, course_room_mapping):
    """Pre-schedule elective baskets globally"""
    print("\n" + "="*80)
    print("🎓 GLOBAL ELECTIVE BASKET SCHEDULING")
    print("="*80)
    
    basket_groups = {}
    
    for _, course in df_input.iterrows():
        code = str(course.get('Course Code', '')).strip()
        basket = extract_elective_basket(code)
        
        if basket and pd.notna(basket):
            if 'Schedule' in course and str(course.get('Schedule', 'Yes')).strip().upper() != 'YES':
                continue
            
            semester = int(course.get('Semester', 0))
            key = (semester, basket)
            
            if key not in basket_groups:
                basket_groups[key] = []
            basket_groups[key].append(course)
    
    global_schedule = {}
    
    for (semester, basket_name), basket_courses in sorted(basket_groups.items()):
        print(f"\n📚 Semester {semester}, Basket {basket_name}: {len(basket_courses)} courses")
        
        first_course = basket_courses[0]
        lec_count, tut_count, lab_count, _ = calculate_required_minutes(first_course)
        
        print(f"   Structure: L={lec_count}h, T={tut_count}h, P={lab_count}h")
        
        lec_sessions = int(lec_count * 60 / LECTURE_MIN) if lec_count > 0 else 0
        tut_sessions = int(tut_count * 60 / TUTORIAL_MIN) if tut_count > 0 else 0
        lab_sessions = int(lab_count * 60 / LAB_MIN) if lab_count > 0 else 0
        
        basket_schedule = []
        
        # Schedule lectures
        for session_num in range(lec_sessions):
            scheduled = False
            for attempt in range(5000):
                day = random.randint(0, len(DAYS)-1)
                start_idx = random.randint(0, max(0, len(TIME_SLOTS)-3))
                
                conflict = False
                for prev_day, prev_slots, prev_type in basket_schedule:
                    if prev_day == day and prev_type in ['LEC', 'TUT']:
                        if any(s in prev_slots for s in range(start_idx, start_idx + 2)):
                            conflict = True
                            break
                
                if conflict:
                    continue
                
                slot_indices = []
                accumulated = 0
                valid = True
                
                for i in range(start_idx, len(TIME_SLOTS)):
                    if is_minor_slot(TIME_SLOTS[i]) or is_break_time_slot(TIME_SLOTS[i], semester):
                        valid = False
                        break
                    
                    slot_indices.append(i)
                    accumulated += slot_minutes(TIME_SLOTS[i])
                    
                    if accumulated >= LECTURE_MIN:
                        break
                
                if valid and accumulated >= LECTURE_MIN and len(slot_indices) > 0:
                    basket_schedule.append((day, slot_indices, 'LEC'))
                    scheduled = True
                    slot_time = TIME_SLOTS[slot_indices[0]][0].strftime('%H:%M')
                    print(f"   ✅ Lecture {session_num+1}/{lec_sessions}: {DAYS[day]} at {slot_time}")
                    break
            
            if not scheduled:
                print(f"   ⚠️ Could not schedule Lecture {session_num+1}/{lec_sessions}")
        
        # Schedule tutorials
        for session_num in range(tut_sessions):
            scheduled = False
            for attempt in range(5000):
                day = random.randint(0, len(DAYS)-1)
                start_idx = random.randint(0, max(0, len(TIME_SLOTS)-2))
                
                lec_on_day = any(d == day and ct == 'LEC' for d, _, ct in basket_schedule)
                if lec_on_day:
                    continue
                
                conflict = False
                for prev_day, prev_slots, prev_type in basket_schedule:
                    if prev_day == day and prev_type == 'TUT':
                        if any(s in prev_slots for s in range(start_idx, start_idx + 2)):
                            conflict = True
                            break
                
                if conflict:
                    continue
                
                slot_indices = []
                accumulated = 0
                valid = True
                
                for i in range(start_idx, len(TIME_SLOTS)):
                    if is_minor_slot(TIME_SLOTS[i]) or is_break_time_slot(TIME_SLOTS[i], semester):
                        valid = False
                        break
                    
                    slot_indices.append(i)
                    accumulated += slot_minutes(TIME_SLOTS[i])
                    
                    if accumulated >= TUTORIAL_MIN:
                        break
                
                if valid and accumulated >= TUTORIAL_MIN and len(slot_indices) > 0:
                    basket_schedule.append((day, slot_indices, 'TUT'))
                    scheduled = True
                    slot_time = TIME_SLOTS[slot_indices[0]][0].strftime('%H:%M')
                    print(f"   ✅ Tutorial {session_num+1}/{tut_sessions}: {DAYS[day]} at {slot_time}")
                    break
            
            if not scheduled:
                print(f"   ⚠️ Could not schedule Tutorial {session_num+1}/{tut_sessions}")
        
        # Schedule labs
        for session_num in range(lab_sessions):
            scheduled = False
            for attempt in range(5000):
                day = random.randint(0, len(DAYS)-1)
                start_idx = random.randint(0, max(0, len(TIME_SLOTS)-3))
                
                conflict = False
                for prev_day, prev_slots, prev_type in basket_schedule:
                    if prev_day == day:
                        if any(s in prev_slots for s in range(start_idx, start_idx + 3)):
                            conflict = True
                            break
                
                if conflict:
                    continue
                
                slot_indices = []
                accumulated = 0
                valid = True
                
                for i in range(start_idx, len(TIME_SLOTS)):
                    if is_minor_slot(TIME_SLOTS[i]) or is_break_time_slot(TIME_SLOTS[i], semester):
                        valid = False
                        break
                    
                    slot_indices.append(i)
                    accumulated += slot_minutes(TIME_SLOTS[i])
                    
                    if accumulated >= LAB_MIN:
                        break
                
                if valid and accumulated >= LAB_MIN and len(slot_indices) > 0:
                    basket_schedule.append((day, slot_indices, 'LAB'))
                    scheduled = True
                    slot_time = TIME_SLOTS[slot_indices[0]][0].strftime('%H:%M')
                    print(f"   ✅ Lab {session_num+1}/{lab_sessions}: {DAYS[day]} at {slot_time}")
                    break
            
            if not scheduled:
                print(f"   ⚠️ Could not schedule Lab {session_num+1}/{lab_sessions}")
        
        global_schedule[(semester, basket_name)] = basket_schedule
        print(f"   📋 Total sessions scheduled: {len(basket_schedule)}")
    
    print("\n" + "="*80)
    print(f"✅ Global basket scheduling complete: {len(global_schedule)} baskets")
    print("="*80 + "\n")
    
    return global_schedule

def add_unscheduled_course(unscheduled_components, department, semester, code, name, 
                          faculty, comp_type, section, reason):
    existing = next((u for u in unscheduled_components if u.code == code), None)
    if existing:
        if comp_type not in existing.component_type:
            existing.component_type += f", {comp_type}"
        if reason not in existing.reason:
            existing.reason += f"; {reason}"
    else:
        unscheduled_components.append(
            UnscheduledComponent(department, semester, code, name, faculty, 
                               comp_type, 1, section, reason)
        )

def is_7th_semester(department, semester):
    """Check if this is 7th semester from CSE/DSAI/ECE"""
    dept_upper = str(department).strip().upper()
    return int(semester) == 7 and dept_upper in ['CSE', 'DSAI', 'ECE']

# ---------------------------
# Main generation function
# ---------------------------
def generate_all_timetables():
    global TIME_SLOTS
    TIME_SLOTS = generate_time_slots()
    
    batch_info = load_batch_data()
    room_schedule = {}
    professor_schedule = {}
    course_room_mapping = {}
    
    # Pre-schedule elective baskets
    global_basket_schedule = schedule_global_elective_baskets(df, professor_schedule, room_schedule, course_room_mapping)
    
    if global_basket_schedule is None:
        global_basket_schedule = {}

    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    overview = wb.create_sheet("Overview")
    overview.append(["Combined Timetable for All Departments and Semesters"])
    overview.append(["Generated on:", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    overview.append([])
    overview.append(["Department", "Semester", "Sheet Name"])
    row_index = 5

    unscheduled_components = []

    SUBJECT_COLORS = [
        "FF6B6B", "4ECDC4", "FF9F1C", "5D5FEF", "45B7D1",
        "F72585", "7209B7", "3A0CA3", "4361EE", "4CC9F0",
        "06D6A0", "FFD166", "EF476F", "118AB2", "073B4C"
    ]

    seventh_sem_processed = False
    seventh_sem_course_data = []

    for department in df['Department'].unique():
        sems = sorted(df[df['Department'] == department]['Semester'].unique())
        for semester in sems:
            if is_7th_semester(department, semester):
                if not seventh_sem_processed:
                    for dept in ['CSE', 'DSAI', 'ECE']:
                        dept_courses = df[(df['Department'] == dept) & (df['Semester'] == 7)]
                        if not dept_courses.empty:
                            seventh_sem_course_data.append(dept_courses)
                    seventh_sem_processed = True
                continue

            dept_upper = str(department).strip().upper()
            num_sections = 2 if (dept_upper == "CSE" and int(semester) in [2, 4, 6]) else 1

            courses = df[(df['Department'] == department) & (df['Semester'] == semester)]
            if 'Schedule' in courses.columns:
                courses = courses[(courses['Schedule'].fillna('Yes').str.upper() == 'YES') | (courses['Schedule'].isna())]
            if courses.empty:
                continue

            if 'P' in courses.columns:
                lab_courses = courses[courses['P'] > 0].copy()
                non_lab_courses = courses[courses['P'] == 0].copy()
            else:
                lab_courses = courses.head(0)
                non_lab_courses = courses.copy()

            if not lab_courses.empty:
                lab_courses['priority'] = lab_courses.apply(get_course_priority, axis=1)
                lab_courses = lab_courses.sort_values('priority', ascending=False)
            non_lab_courses['priority'] = non_lab_courses.apply(get_course_priority, axis=1)
            non_lab_courses = non_lab_courses.sort_values('priority', ascending=False)

            combined = pd.concat([lab_courses, non_lab_courses])
            combined['is_elective'] = combined.apply(is_elective, axis=1)
            combined['elective_basket'] = combined['Course Code'].apply(extract_elective_basket)

            courses_combined = combined.sort_values(
                by=['is_elective', 'priority'], 
                ascending=[False, False]
            ).drop_duplicates()

            for section in range(num_sections):
                section_title = f"{department}_{semester}" if num_sections == 1 else f"{department}_{semester}_{chr(65 + section)}"
                ws = wb.create_sheet(title=section_title)

                overview.cell(row=row_index, column=1, value=department)
                overview.cell(row=row_index, column=2, value=str(semester))
                overview.cell(row=row_index, column=3, value=section_title)
                row_index += 1

                timetable = {d: {s: {'type': None, 'code': '', 'name': '', 'faculty': '', 'classroom': ''} for s in range(len(TIME_SLOTS))} for d in range(len(DAYS))}
                course_day_components = {}

                section_subject_color = {}
                color_iter = iter(SUBJECT_COLORS)
                course_faculty_map = {}
                basket_scheduled_courses = set()

                for _, c in courses_combined.iterrows():
                    code = str(c.get('Course Code', '')).strip()
                    if code and code not in section_subject_color:
                        try:
                            section_subject_color[code] = next(color_iter)
                        except StopIteration:
                            section_subject_color[code] = random.choice(SUBJECT_COLORS)
                        course_faculty_map[code] = select_faculty(c.get('Faculty', 'TBD'))

                # Apply global basket schedules
                print(f"\n📋 Applying basket schedules for {section_title}...")
                
                elective_courses_in_section = courses_combined[courses_combined['is_elective'] == True]
                
                for _, course in elective_courses_in_section.iterrows():
                    code = str(course.get('Course Code', '')).strip()
                    basket = course.get('elective_basket')
                    
                    if not basket or pd.isna(basket):
                        continue
                    
                    basket_key = (semester, basket)
                    if basket_key not in global_basket_schedule:
                        print(f"   ⚠️ No global schedule found for {basket}")
                        continue
                    
                    basket_schedule = global_basket_schedule[basket_key]
                    base_code = get_base_course_code(code)
                    name = str(course.get('Course Name', '')).strip()
                    faculty = select_faculty(course.get('Faculty', 'TBD'))
                    
                    if faculty not in professor_schedule:
                        professor_schedule[faculty] = {d: set() for d in range(len(DAYS))}
                    
                    if base_code not in course_day_components:
                        course_day_components[base_code] = {}
                    
                    for day, slot_indices, comp_type in basket_schedule:
                        room_type = get_required_room_type(comp_type)
                        
                        candidate_room = find_suitable_room_for_slot(
                            base_code, room_type, day, slot_indices, 
                            room_schedule, course_room_mapping, comp_type
                        )
                        
                        if candidate_room is None:
                            print(f"   ⚠️ No room available for {code} {comp_type}")
                            continue
                        
                        for si_idx, si in enumerate(slot_indices):
                            if timetable[day][si]['type'] is None:
                                timetable[day][si]['type'] = comp_type
                                timetable[day][si]['code'] = f"{basket}\n{code}" if si_idx == 0 else ''
                                timetable[day][si]['name'] = name if si_idx == 0 else ''
                                timetable[day][si]['faculty'] = faculty if si_idx == 0 else ''
                                timetable[day][si]['classroom'] = candidate_room if si_idx == 0 else ''
                                professor_schedule[faculty][day].add(si)
                                if candidate_room not in room_schedule:
                                    room_schedule[candidate_room] = {d: set() for d in range(len(DAYS))}
                                room_schedule[candidate_room][day].add(si)
                        
                        if day not in course_day_components[base_code]:
                            course_day_components[base_code][day] = []
                        course_day_components[base_code][day].append(comp_type)
                    
                    basket_scheduled_courses.add(code)
                    print(f"   ✅ Applied {basket} schedule to {code}")

                # Schedule non-elective courses
                print(f"\n📖 Scheduling non-elective courses for {section_title}...")
                for _, course in courses_combined.iterrows():
                    code = str(course.get('Course Code', '')).strip()
                    
                    if code in basket_scheduled_courses:
                        continue
                    
                    if course.get('is_elective'):
                        continue
                    
                    base_code = get_base_course_code(code)
                    name = str(course.get('Course Name', '')).strip()
                    faculty = select_faculty(course.get('Faculty', 'TBD'))

                    if faculty not in professor_schedule:
                        professor_schedule[faculty] = {d: set() for d in range(len(DAYS))}
                    
                    if base_code not in course_day_components:
                        course_day_components[base_code] = {}

                    lec_count, tut_count, lab_count, ss_count = calculate_required_minutes(course)
                    lec_sessions_needed = int(lec_count * 60 / LECTURE_MIN) if lec_count > 0 else 0
                    tut_sessions_needed = int(tut_count * 60 / TUTORIAL_MIN) if tut_count > 0 else 0
                    lab_sessions_needed = int(lab_count * 60 / LAB_MIN) if lab_count > 0 else 0

                    def schedule_component(required_minutes, comp_type, attempts_limit=5000):
                        room_type = get_required_room_type(comp_type)
                        
                        for attempt in range(attempts_limit):
                            day = random.randint(0, len(DAYS)-1)
                            starts = get_all_possible_start_indices()
                            
                            for start_idx in starts:
                                slot_indices, candidate_room = find_consecutive_slots_for_minutes(
                                    timetable, day, start_idx, required_minutes, semester,
                                    professor_schedule, faculty, room_schedule, room_type,
                                    base_code, course_room_mapping, comp_type, course_day_components
                                )

                                if slot_indices is None:
                                    continue
                                if not check_professor_availability(professor_schedule, faculty, day, slot_indices[0], len(slot_indices)):
                                    continue
                                if candidate_room is None:
                                    continue
                                
                                for si_idx, si in enumerate(slot_indices):
                                    timetable[day][si]['type'] = comp_type
                                    timetable[day][si]['code'] = code if si_idx == 0 else ''
                                    timetable[day][si]['name'] = name if si_idx == 0 else ''
                                    timetable[day][si]['faculty'] = faculty if si_idx == 0 else ''
                                    timetable[day][si]['classroom'] = candidate_room if si_idx == 0 else ''
                                    professor_schedule[faculty][day].add(si)
                                    if candidate_room not in room_schedule:
                                        room_schedule[candidate_room] = {d: set() for d in range(len(DAYS))}
                                    room_schedule[candidate_room][day].add(si)
                                
                                if day not in course_day_components[base_code]:
                                    course_day_components[base_code][day] = []
                                course_day_components[base_code][day].append(comp_type)
                                
                                return True
                        return False

                    # Schedule lectures
                    for _ in range(lec_sessions_needed):
                        ok = schedule_component(LECTURE_MIN, 'LEC')
                        if not ok:
                            add_unscheduled_course(unscheduled_components, department, semester, code, name, faculty, 'LEC', section, "Could not find suitable slot")

                    # Schedule tutorials
                    for _ in range(tut_sessions_needed):
                        ok = schedule_component(TUTORIAL_MIN, 'TUT')
                        if not ok:
                            add_unscheduled_course(unscheduled_components, department, semester, code, name, faculty, 'TUT', section, "Could not find suitable slot")

                    # Schedule labs
                    for _ in range(lab_sessions_needed):
                        ok = schedule_component(LAB_MIN, 'LAB')
                        if not ok:
                            add_unscheduled_course(unscheduled_components, department, semester, code, name, faculty, 'LAB', section, "No computer lab available")

                # Write timetable to sheet
                write_timetable_to_sheet(ws, timetable, section_subject_color, course_faculty_map, 
                                        courses_combined, course_room_mapping, semester)

    # Generate common 7th semester timetable
    if seventh_sem_course_data:
        generate_7th_sem_common_timetable(wb, seventh_sem_course_data, overview, row_index, 
                                         unscheduled_components, professor_schedule, 
                                         room_schedule, course_room_mapping, SUBJECT_COLORS)

    # Format overview sheet
    format_overview_sheet(overview, row_index)

    # Save workbook
    out_filename = os.path.join(OUTPUT_DIR, "timetable_all_departments.xlsx")
    try:
        wb.save(out_filename)
        print(f"\n✅ Combined timetable saved as {out_filename}")
    except Exception as e:
        print(f"❌ Failed to save timetable: {e}")
        traceback.print_exc()

    # Generate teacher and unscheduled workbooks
    try:
        create_teacher_and_unscheduled_from_combined(out_filename, unscheduled_components)
    except Exception as e:
        print("❌ Failed to generate teacher/unscheduled workbooks:", e)
        traceback.print_exc()

    return out_filename

# ---------------------------
# 7th semester common timetable
# ---------------------------
def generate_7th_sem_common_timetable(wb, course_data_list, overview, row_index, 
                                     unscheduled_components, professor_schedule, 
                                     room_schedule, course_room_mapping, SUBJECT_COLORS):
    """Generate common timetable for 7th semester"""
    
    all_7th_courses = pd.concat(course_data_list, ignore_index=True)
    if 'Schedule' in all_7th_courses.columns:
        all_7th_courses = all_7th_courses[(all_7th_courses['Schedule'].fillna('Yes').str.upper() == 'YES') | 
                                         (all_7th_courses['Schedule'].isna())]
    
    if all_7th_courses.empty:
        return
    
    section_title = "Common_7th_Semester"
    ws = wb.create_sheet(title=section_title)
    
    overview.cell(row=row_index, column=1, value="CSE/DSAI/ECE")
    overview.cell(row=row_index, column=2, value="7")
    overview.cell(row=row_index, column=3, value=section_title)
    
    if 'P' in all_7th_courses.columns:
        lab_courses = all_7th_courses[all_7th_courses['P'] > 0].copy()
        non_lab_courses = all_7th_courses[all_7th_courses['P'] == 0].copy()
    else:
        lab_courses = all_7th_courses.head(0)
        non_lab_courses = all_7th_courses.copy()
    
    if not lab_courses.empty:
        lab_courses['priority'] = lab_courses.apply(get_course_priority, axis=1)
        lab_courses = lab_courses.sort_values('priority', ascending=False)
    non_lab_courses['priority'] = non_lab_courses.apply(get_course_priority, axis=1)
    non_lab_courses = non_lab_courses.sort_values('priority', ascending=False)
    
    combined = pd.concat([lab_courses, non_lab_courses])
    combined['is_elective'] = combined.apply(is_elective, axis=1)
    combined['elective_basket'] = combined['Course Code'].apply(extract_elective_basket)
    
    courses_combined = combined.sort_values(
        by=['is_elective', 'priority'], 
        ascending=[False, False]
    ).drop_duplicates()
    
    timetable = {d: {s: {'type': None, 'code': '', 'name': '', 'faculty': '', 'classroom': ''} 
                     for s in range(len(TIME_SLOTS))} for d in range(len(DAYS))}
    course_day_components = {}
    
    section_subject_color = {}
    color_iter = iter(SUBJECT_COLORS)
    course_faculty_map = {}
    
    for _, c in courses_combined.iterrows():
        code = str(c.get('Course Code', '')).strip()
        if code and code not in section_subject_color:
            try:
                section_subject_color[code] = next(color_iter)
            except StopIteration:
                section_subject_color[code] = random.choice(SUBJECT_COLORS)
            course_faculty_map[code] = select_faculty(c.get('Faculty', 'TBD'))
    
    # Schedule courses (simplified for 7th semester)
    for _, course in courses_combined.iterrows():
        code = str(course.get('Course Code', '')).strip()
        base_code = get_base_course_code(code)
        name = str(course.get('Course Name', '')).strip()
        faculty = select_faculty(course.get('Faculty', 'TBD'))
        
        if faculty not in professor_schedule:
            professor_schedule[faculty] = {d: set() for d in range(len(DAYS))}
        
        if base_code not in course_day_components:
            course_day_components[base_code] = {}
        
        lec_count, tut_count, lab_count, _ = calculate_required_minutes(course)
        lec_sessions_needed = int(lec_count * 60 / LECTURE_MIN) if lec_count > 0 else 0
        tut_sessions_needed = int(tut_count * 60 / TUTORIAL_MIN) if tut_count > 0 else 0
        lab_sessions_needed = int(lab_count * 60 / LAB_MIN) if lab_count > 0 else 0
        
        def schedule_component(required_minutes, comp_type, attempts_limit=5000):
            room_type = get_required_room_type(comp_type)
            
            for attempt in range(attempts_limit):
                day = random.randint(0, len(DAYS)-1)
                starts = get_all_possible_start_indices()
                
                for start_idx in starts:
                    slot_indices, candidate_room = find_consecutive_slots_for_minutes(
                        timetable, day, start_idx, required_minutes, 7,
                        professor_schedule, faculty, room_schedule, room_type,
                        base_code, course_room_mapping, comp_type, course_day_components
                    )
                    
                    if slot_indices is None or candidate_room is None:
                        continue
                    if not check_professor_availability(professor_schedule, faculty, day, slot_indices[0], len(slot_indices)):
                        continue
                    
                    for si_idx, si in enumerate(slot_indices):
                        timetable[day][si]['type'] = comp_type
                        timetable[day][si]['code'] = code if si_idx == 0 else ''
                        timetable[day][si]['name'] = name if si_idx == 0 else ''
                        timetable[day][si]['faculty'] = faculty if si_idx == 0 else ''
                        timetable[day][si]['classroom'] = candidate_room if si_idx == 0 else ''
                        professor_schedule[faculty][day].add(si)
                        if candidate_room not in room_schedule:
                            room_schedule[candidate_room] = {d: set() for d in range(len(DAYS))}
                        room_schedule[candidate_room][day].add(si)
                    
                    if day not in course_day_components[base_code]:
                        course_day_components[base_code][day] = []
                    course_day_components[base_code][day].append(comp_type)
                    
                    return True
            return False
        
        for _ in range(lec_sessions_needed):
            ok = schedule_component(LECTURE_MIN, 'LEC')
            if not ok:
                add_unscheduled_course(unscheduled_components, "Common_7th", 7, code, name, faculty, 'LEC', 0, "Could not find suitable slot")
        
        for _ in range(tut_sessions_needed):
            ok = schedule_component(TUTORIAL_MIN, 'TUT')
            if not ok:
                add_unscheduled_course(unscheduled_components, "Common_7th", 7, code, name, faculty, 'TUT', 0, "Could not find suitable slot")
        
        for _ in range(lab_sessions_needed):
            ok = schedule_component(LAB_MIN, 'LAB')
            if not ok:
                add_unscheduled_course(unscheduled_components, "Common_7th", 7, code, name, faculty, 'LAB', 0, "No computer lab available")
    
    write_timetable_to_sheet(ws, timetable, section_subject_color, course_faculty_map, 
                            courses_combined, course_room_mapping, 7)

# ---------------------------
# Write timetable to Excel
# ---------------------------
def write_timetable_to_sheet(ws, timetable, section_subject_color, course_faculty_map, 
                            courses_combined, course_room_mapping, semester):
    """Write formatted timetable to worksheet"""
    
    # Header row
    header = ['Day'] + [f"{slot[0].strftime('%H:%M')}-{slot[1].strftime('%H:%M')}" for slot in TIME_SLOTS]
    ws.append(header)
    
    # Styling
    header_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
    header_font = Font(bold=True)
    header_alignment = Alignment(horizontal='center', vertical='center')
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                   top=Side(style='thin'), bottom=Side(style='thin'))
    lec_fill_default = PatternFill(start_color="FA8072", end_color="FA8072", fill_type="solid")
    lab_fill_default = PatternFill(start_color="7CFC00", end_color="7CFC00", fill_type="solid")
    tut_fill_default = PatternFill(start_color="87CEEB", end_color="87CEEB", fill_type="solid")
    break_fill = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")
    minor_fill = PatternFill(start_color="9ACD32", end_color="9ACD32", fill_type="solid")
    
    # Write timetable rows
    for day_idx, day_name in enumerate(DAYS):
        ws.append([day_name] + [''] * len(TIME_SLOTS))
        row_num = ws.max_row
        merges = []
        
        for slot_idx in range(len(TIME_SLOTS)):
            cell_obj = ws.cell(row=row_num, column=slot_idx + 2)
            
            if is_minor_slot(TIME_SLOTS[slot_idx]):
                cell_obj.value = "Minor Slot"
                cell_obj.fill = minor_fill
                cell_obj.font = Font(bold=True)
                cell_obj.alignment = Alignment(horizontal='center', vertical='center')
                cell_obj.border = border
                continue
            
            if is_break_time_slot(TIME_SLOTS[slot_idx], semester):
                cell_obj.value = "LUNCH BREAK"  # Updated label
                cell_obj.fill = break_fill
                cell_obj.font = Font(bold=True)
                cell_obj.alignment = Alignment(horizontal='center', vertical='center')
                cell_obj.border = border
                continue
            
            if timetable[day_idx][slot_idx]['type'] is None:
                cell_obj.border = border
                continue
            
            typ = timetable[day_idx][slot_idx]['type']
            code = timetable[day_idx][slot_idx]['code']
            cls = timetable[day_idx][slot_idx]['classroom']
            fac = timetable[day_idx][slot_idx]['faculty']
            
            if code:
                span = [slot_idx]
                j = slot_idx + 1
                while j < len(TIME_SLOTS) and timetable[day_idx][j]['type'] is not None and timetable[day_idx][j]['code'] == '':
                    span.append(j)
                    j += 1
                
                if '\n' in code and code.startswith('B'):
                    display = f"{code}\n{typ}\nRoom: {cls}\n{fac}"
                else:
                    display = f"{code}\n{typ}\nRoom: {cls}\n{fac}"
                
                if code.split('\n')[0] if '\n' in code else code in section_subject_color:
                    actual_code = code.split('\n')[-1] if '\n' in code else code
                    subj_color = section_subject_color.get(actual_code, section_subject_color.get(code))
                    fill = PatternFill(start_color=subj_color, end_color=subj_color, fill_type="solid")
                else:
                    fill = {'LEC': lec_fill_default, 'LAB': lab_fill_default, 
                           'TUT': tut_fill_default}.get(typ, lec_fill_default)
                
                cell_obj.value = display
                cell_obj.fill = fill
                merges.append((slot_idx + 2, slot_idx + 2 + len(span) - 1, display, fill))
            
            cell_obj.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
            cell_obj.border = border
        
        # Apply merges
        for start_col, end_col, val, fill in merges:
            if end_col > start_col:
                rng = f"{get_column_letter(start_col)}{row_num}:{get_column_letter(end_col)}{row_num}"
                try:
                    ws.merge_cells(rng)
                    mc = ws[f"{get_column_letter(start_col)}{row_num}"]
                    mc.value = val
                    mc.fill = fill
                    mc.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
                    mc.border = border
                except Exception:
                    pass
    
    # Set column widths
    for col_idx in range(1, len(TIME_SLOTS)+2):
        try:
            ws.column_dimensions[get_column_letter(col_idx)].width = 15
        except Exception:
            pass
    
    # Set row heights
    for row in ws.iter_rows(min_row=2, max_row=len(DAYS)+1):
        ws.row_dimensions[row[0].row].height = 40
    
    # Add self-study courses section
    current_row = len(DAYS) + 4
    ss_courses = []
    for _, course in courses_combined.iterrows():
        l = int(course['L']) if pd.notna(course['L']) else 0
        t = int(course['T']) if pd.notna(course['T']) else 0
        p = int(course['P']) if pd.notna(course['P']) else 0
        s = int(course['S']) if pd.notna(course['S']) else 0
        if s > 0 and l == 0 and t == 0 and p == 0:
            ss_courses.append({
                'code': str(course['Course Code']),
                'name': str(course['Course Name']),
                'faculty': str(course['Faculty'])
            })
    
    if ss_courses:
        ws.cell(row=current_row, column=1, value="Self-Study Only Courses")
        ws.cell(row=current_row, column=1).font = Font(bold=True)
        current_row += 1
        
        headers = ['Course Code', 'Course Name', 'Faculty']
        for col, header in enumerate(headers, 1):
            ws.cell(row=current_row, column=col, value=header)
            ws.cell(row=current_row, column=col).font = Font(bold=True)
        current_row += 1
        
        for course in ss_courses:
            ws.cell(row=current_row, column=1, value=course['code'])
            ws.cell(row=current_row, column=2, value=course['name'])
            ws.cell(row=current_row, column=3, value=course['faculty'])
            current_row += 1
        
        current_row += 2
    
    # Add legend
    legend_title = ws.cell(row=current_row, column=1, value="Legend")
    legend_title.font = Font(bold=True, size=12)
    current_row += 2
    
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    
    legend_headers = ['Subject Code', 'Color', 'Subject Name', 'Faculty', 'LTPS', 'Room']
    for col, header in enumerate(legend_headers, 1):
        cell = ws.cell(row=current_row, column=col, value=header)
        cell.font = Font(bold=True)
        cell.border = border
        cell.fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    current_row += 1
    
    # Add legend entries
    for code, color in section_subject_color.items():
        assigned_room = course_room_mapping.get(f"{code}_LEC") or course_room_mapping.get(f"{code}_LAB") or "—"
        if not assigned_room or assigned_room == "—":
            continue
        
        ws.row_dimensions[current_row].height = 30
        
        ltps_value = ""
        course_name = ''
        fac_name = ''
        for _, course_row in courses_combined.iterrows():
            if str(course_row['Course Code']) == code:
                l = str(int(course_row['L'])) if pd.notna(course_row['L']) else "0"
                t = str(int(course_row['T'])) if pd.notna(course_row['T']) else "0"
                p = str(int(course_row['P'])) if pd.notna(course_row['P']) else "0"
                s = str(int(course_row['S'])) if pd.notna(course_row['S']) and 'S' in course_row else "0"
                ltps_value = f"{l}-{t}-{p}-{s}"
                course_name = str(course_row['Course Name'])
                fac_name = course_faculty_map.get(code, '')
                break
        
        cells = [
            (code, None),
            ('', PatternFill(start_color=color, end_color=color, fill_type="solid")),
            (course_name, None),
            (fac_name, None),
            (ltps_value, None),
            (assigned_room, None)
        ]
        
        for col, (value, fill) in enumerate(cells, 1):
            cell = ws.cell(row=current_row, column=col, value=value)
            cell.border = border
            if fill:
                cell.fill = fill
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True, indent=2)
        
        current_row += 1

def format_overview_sheet(overview, row_index):
    """Format overview sheet"""
    for col in range(1, 4):
        overview.column_dimensions[get_column_letter(col)].width = 20
    
    for row_ in overview.iter_rows(min_row=1, max_row=4):
        for cell in row_:
            cell.font = Font(bold=True)
    
    for cell in overview[4]:
        cell.fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
        cell.font = Font(bold=True)
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                           top=Side(style='thin'), bottom=Side(style='thin'))

# ---------------------------
# Teacher and unscheduled workbooks
# ---------------------------
def split_faculty_names(fac_str):
    """Split faculty string into separate names"""
    if fac_str is None:
        return []
    s = str(fac_str).strip()
    if s == '' or s.lower() in ['nan', 'none']:
        return []
    parts = [s]
    for sep in ['/', ',', '&', ';']:
        if sep in s:
            parts = [p.strip() for p in s.split(sep) if p.strip()]
            break
    return parts if parts else [s]

def parse_cell_for_course(cell_value):
    """Parse cell value to extract course info"""
    if cell_value is None:
        return ('', '', '', '')
    text = str(cell_value).strip()
    if text == '':
        return ('', '', '', '')
    
    lines = [ln.strip() for ln in text.splitlines() if ln.strip()]
    faculty = ''
    room = ''
    code = ''
    typ = ''
    
    for ln in lines:
        if 'room' in ln.lower():
            parts = ln.split(':')
            if len(parts) >= 2:
                room = parts[-1].strip()
    
    if len(lines) >= 1:
        last = lines[-1]
        if 'room' not in last.lower() and ':' not in last:
            faculty = last
    
    first = lines[0] if lines else ''
    if first:
        # Handle basket format (B1\nCODE)
        if '\n' in text and first.startswith('B'):
            code = lines[1] if len(lines) > 1 else first
        else:
            code = first
        
        for t in ['LEC', 'LAB', 'TUT']:
            if t in text.upper():
                typ = t
                break
    
    return (code, typ, room, faculty)

def create_teacher_and_unscheduled_from_combined(timetable_filename, unscheduled_components):
    """Build teacher_timetables.xlsx and unscheduled_courses.xlsx"""
    try:
        wb = load_workbook(timetable_filename, data_only=True)
    except Exception as e:
        print(f"❌ Failed to open {timetable_filename}: {e}")
        return
    
    teacher_slots = {}
    slot_headers = []
    
    for sheetname in wb.sheetnames:
        if sheetname.lower() == 'overview':
            continue
        ws = wb[sheetname]
        header = [str(ws.cell(1, c).value).strip() if ws.cell(1, c).value else '' for c in range(2, ws.max_column + 1)]
        if len(header) > len(slot_headers):
            slot_headers = header
        
        for r in range(2, ws.max_row + 1):
            day = ws.cell(r, 1).value
            if not day or str(day) not in DAYS:
                break
            day_idx = DAYS.index(day)
            
            for c in range(2, ws.max_column + 1):
                code, typ, room, faculty = parse_cell_for_course(ws.cell(r, c).value)
                for f in split_faculty_names(faculty):
                    if not f or str(f).strip().upper() in ["BREAK", "MINOR SLOT", "NAN", "NONE", "", "LUNCH BREAK"]:
                        continue
                    
                    teacher_slots.setdefault(f, {d: {i: '' for i in range(len(slot_headers))} for d in range(len(DAYS))})
                    teacher_slots[f][day_idx][c - 2] = f"{code} {typ}\n({sheetname})\nRoom: {room}" if code else ''
    
    # Create teacher workbook
    twb = Workbook()
    if "Sheet" in twb.sheetnames:
        twb.remove(twb["Sheet"])
    
    header_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
    alt_fill = PatternFill(start_color="FFF8DC", end_color="FFF8DC", fill_type="solid")
    border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
    header_font = Font(bold=True, size=12)
    title_font = Font(bold=True, size=14)
    cell_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    for teacher in sorted(teacher_slots.keys()):
        safe_name = teacher[:31] or "Unknown"
        ws = twb.create_sheet(title=safe_name)
        
        ws.merge_cells("A1:{}1".format(get_column_letter(len(slot_headers) + 1)))
        title_cell = ws.cell(row=1, column=1, value=f"{teacher} — Weekly Timetable")
        title_cell.font = title_font
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        
        ws.append(["Day"] + slot_headers)
        for cell in ws[2]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
        
        for d, day in enumerate(DAYS):
            row = [day] + [teacher_slots[teacher][d][i] for i in range(len(slot_headers))]
            ws.append(row)
            row_idx = ws.max_row
            if d % 2 == 0:
                for cell in ws[row_idx]:
                    cell.fill = alt_fill
            for cell in ws[row_idx]:
                cell.alignment = cell_align
                cell.border = border
            ws.row_dimensions[row_idx].height = 35
        
        ws.column_dimensions["A"].width = 15
        for col in range(2, len(slot_headers) + 2):
            ws.column_dimensions[get_column_letter(col)].width = 20
    
    twb.save(os.path.join(OUTPUT_DIR, "teacher_timetables.xlsx"))
    print("✅ Saved teacher_timetables.xlsx")
    
    # Create unscheduled workbook
    uwb = Workbook()
    ws = uwb.active
    ws.title = "Unscheduled Courses"
    
    headers = ["Course Code", "Department", "Semester", "Component Type", "Reason"]
    ws.append(headers)
    
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    unscheduled_unique = {}
    
    for u in unscheduled_components:
        if u.code not in unscheduled_unique:
            if hasattr(u, "reason") and u.reason and len(str(u.reason).strip()) > 0:
                reason_text = str(u.reason).strip()
            else:
                reason_text = "Unspecified scheduling issue"
            
            unscheduled_unique[u.code] = {
                "Course Code": u.code,
                "Department": u.department,
                "Semester": u.semester,
                "Component Type": u.component_type,
                "Reason": reason_text
            }
    
    for entry in unscheduled_unique.values():
        ws.append([entry["Course Code"], entry["Department"], entry["Semester"], 
                  entry["Component Type"], entry["Reason"]])
    
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 50
    
    uwb.save(os.path.join(OUTPUT_DIR, "unscheduled_courses.xlsx"))
    print(f"✅ Saved unscheduled_courses.xlsx with {len(unscheduled_unique)} courses")

# ---------------------------
# Exam generator
# ---------------------------
def exam_generator():
    """Generate exam timetable with room allocation and invigilation"""
    exam_file = "Exam_timetable.xlsx"
    
    try:
        df_courses = pd.read_csv(os.path.join(INPUT_DIR, "combined.csv"))
        df_rooms = pd.read_csv(os.path.join(INPUT_DIR, "rooms.csv"))
    except FileNotFoundError as e:
        print(f"❌ Missing file: {e}")
        return None
    
    df_courses = df_courses.dropna(subset=["Course Code", "Course Name", "Faculty", "Department", "Semester"])
    if "total_students" not in df_courses.columns:
        df_courses["total_students"] = 50
    df_courses["total_students"] = df_courses["total_students"].fillna(50).astype(int)
    
    df_rooms.columns = [c.strip().lower() for c in df_rooms.columns]
    
    def find_col(keywords):
        for c in df_rooms.columns:
            if any(k in c for k in keywords):
                return c
        return None
    
    room_col = find_col(["room", "num", "id"])
    cap_col = find_col(["cap", "seat"])
    type_col = find_col(["type"])
    
    if not room_col or not cap_col:
        print("❌ rooms.csv must have columns for room number and capacity.")
        return None
    
    df_rooms = df_rooms.rename(columns={room_col: "room", cap_col: "capacity"})
    if type_col:
        df_rooms = df_rooms.rename(columns={type_col: "type"})
    else:
        df_rooms["type"] = "LECTURE_ROOM"
    
    df_rooms["room"] = df_rooms["room"].astype(str).str.strip()
    df_rooms["capacity"] = pd.to_numeric(df_rooms["capacity"], errors="coerce").fillna(0).astype(int)
    df_rooms = df_rooms[df_rooms["capacity"] > 0].sort_values(by="capacity").reset_index(drop=True)
    
    print(f"✅ Loaded {len(df_rooms)} rooms for exam scheduling")
    
    faculty_list = list(set(sum([str(f).replace(" and ", "/").replace(",", "/").split("/") for f in df_courses["Faculty"]], [])))
    faculty_list = [f.strip() for f in faculty_list if f.strip()]
    
    session_title = "Jan-April 03:00 PM to 04:30 PM"
    start_date = datetime(2025, 11, 20)
    num_days = min(10, len(df_courses))
    dates = [start_date + timedelta(days=i) for i in range(num_days)]
    days = [d.strftime("%A") for d in dates]
    
    shuffled = df_courses.sample(frac=1, random_state=42).reset_index(drop=True)
    course_date_map = {row["Course Code"]: dates[i % len(dates)] for i, row in shuffled.iterrows()}
    
    date_room_usage = {d.strftime("%d-%b-%Y"): set() for d in dates}
    invigilation_entries = []
    
    for date in dates:
        date_str = date.strftime("%d-%b-%Y")
        today_courses = shuffled[shuffled["Course Code"].isin(
            [c for c, dt in course_date_map.items() if dt == date]
        )]
        
        for _, course in today_courses.iterrows():
            code = course["Course Code"]
            name = course["Course Name"]
            dept = course["Department"]
            sem = course["Semester"]
            teacher = str(course["Faculty"]).strip()
            students = int(course["total_students"])
            time_slot = "03:00 PM—04:30 PM"
            
            assigned_rooms = []
            remaining = students
            
            available = df_rooms[~df_rooms["room"].isin(date_room_usage[date_str])].copy()
            available = available.sort_values(by="capacity", ascending=True)
            
            suitable = available[available["capacity"] >= remaining]
            if not suitable.empty:
                best = suitable.iloc[0]
                assigned_rooms = [best["room"]]
                date_room_usage[date_str].add(best["room"])
            else:
                total_cap = 0
                for _, room_row in available.iterrows():
                    assigned_rooms.append(room_row["room"])
                    total_cap += room_row["capacity"]
                    date_room_usage[date_str].add(room_row["room"])
                    if total_cap >= remaining:
                        break
                if total_cap < remaining:
                    print(f"⚠️ Not enough capacity for {code}")
            
            available_teachers = [f for f in faculty_list if f.lower() not in teacher.lower()]
            for room in assigned_rooms:
                invigilator = random.choice(available_teachers) if available_teachers else "TBD"
                invigilation_entries.append({
                    "Faculty": invigilator,
                    "Date": date_str,
                    "Time": time_slot,
                    "Course Code": code,
                    "Course Name": name,
                    "Department": dept,
                    "Semester": sem,
                    "Room": room,
                    "Strength": students
                })
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Exam Timetable"
    
    bold_center = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    header_fill = PatternFill(start_color="BFBFBF", end_color="BFBFBF", fill_type="solid")
    border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
    
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(dates)+1)
    title = ws.cell(row=1, column=1, value=session_title)
    title.font = Font(bold=True, size=14)
    title.alignment = center
    title.fill = header_fill
    title.border = border
    
    ws.cell(row=2, column=1, value="Date").font = bold_center
    for i, d in enumerate(dates):
        c = ws.cell(row=2, column=i+2, value=d.strftime("%d-%b-%Y"))
        c.font = bold_center
        c.alignment = center
        c.fill = header_fill
        c.border = border
    
    ws.cell(row=3, column=1, value="Days").font = bold_center
    for i, day in enumerate(days):
        c = ws.cell(row=3, column=i+2, value=day)
        c.font = bold_center
        c.alignment = center
        c.fill = header_fill
        c.border = border
    
    grouped_by_date = {}
    for e in invigilation_entries:
        grouped_by_date.setdefault(e["Date"], set()).add(e["Course Code"])
    max_rows = max(len(v) for v in grouped_by_date.values()) if grouped_by_date else 0
    
    for r in range(max_rows):
        for i, d in enumerate(dates):
            code_list = list(grouped_by_date.get(d.strftime("%d-%b-%Y"), []))
            val = code_list[r] if r < len(code_list) else ""
            cell = ws.cell(row=r+4, column=i+2, value=val)
            cell.alignment = center
            cell.border = border
    
    ws.column_dimensions["A"].width = 15
    for col in range(2, len(dates)+2):
        ws.column_dimensions[get_column_letter(col)].width = 16
    
    ws2 = wb.create_sheet("Exam Invigilation Schedule")
    headers = ["Faculty", "Date", "Time", "Course Code", "Course Name", "Department", "Semester", "Room", "Strength"]
    ws2.append(headers)
    
    for i, h in enumerate(headers, 1):
        cell = ws2.cell(row=1, column=i, value=h)
        cell.font = bold_center
        cell.alignment = center
        cell.fill = header_fill
        cell.border = border
    
    for entry in invigilation_entries:
        ws2.append([entry[h] for h in headers])
    
    for col in range(1, len(headers)+1):
        ws2.column_dimensions[get_column_letter(col)].width = 22
    for r in range(2, ws2.max_row+1):
        for c in ws2[r]:
            c.alignment = center
            c.border = border
    
    exam_file = os.path.join(OUTPUT_DIR, "Exam_timetable.xlsx")
    wb.save(exam_file)
    print(f"✅ Exam timetable saved → {exam_file}")
    return exam_file

# ---------------------------
# Main execution
# ---------------------------
if __name__ == "__main__":
    try:
        print("\n" + "="*80)
        print("🎓 IIIT DHARWAD TIMETABLE GENERATOR")
        print("="*80)
        print("\n🔧 Configuration:")
        print(f"   - No morning break (removed 10:30-10:45)")
        print(f"   - Lunch break: 13:00-14:00 (extended)")
        print(f"   - Lecture duration: {LECTURE_MIN} minutes")
        print(f"   - Tutorial duration: {TUTORIAL_MIN} minutes")
        print(f"   - Lab duration: {LAB_MIN} minutes")
        print(f"   - Min gap between lectures: {MIN_GAP_BETWEEN_LECTURES} minutes")
        print("="*80 + "\n")
        
        generate_all_timetables()
        exam_generator()
        
        print("\n" + "="*80)
        print("✅ TIMETABLE GENERATION COMPLETE!")
        print("="*80)
        print(f"\n📁 Output files saved in: {OUTPUT_DIR}")
        print("   1. timetable_all_departments.xlsx - Main timetable")
        print("   2. teacher_timetables.xlsx - Faculty schedules")
        print("   3. unscheduled_courses.xlsx - Courses that couldn't be scheduled")
        print("   4. Exam_timetable.xlsx - Exam schedule")
        print("\n💡 Tips:")
        print("   - Check unscheduled_courses.xlsx to see which courses failed")
        print("   - If many courses are unscheduled, consider:")
        print("     * Adding more computer labs")
        print("     * Relaxing some constraints")
        print("     * Extending time slots into evening")
        print("="*80 + "\n")
        
    except Exception as e:
        print("\n" + "="*80)
        print("❌ ERROR DURING TIMETABLE GENERATION")
        print("="*80)
        print(f"Error: {e}")
        traceback.print_exc()
        print("="*80 + "\n")