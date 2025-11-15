import pandas as pd
import random
from datetime import datetime, time, timedelta
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter
import csv
import glob
import os
import json
from pathlib import Path
from collections import namedtuple

# Load duration constants from config
def load_config():
    try:
        with open('config.json', 'r') as f:
            config = json.load(f)
            return config['duration_constants']
    except:
        # Return defaults if config file not found
        return {
            'hour_slots': 2,
            'lecture_duration': 3,
            'lab_duration': 4,
            'tutorial_duration': 2,
            'self_study_duration': 2, 
            'break_duration': 1
        }

# Initialize duration constants
durations = load_config()
HOUR_SLOTS = durations['hour_slots']
LECTURE_DURATION = durations['lecture_duration']
LAB_DURATION = durations['lab_duration']
TUTORIAL_DURATION = durations['tutorial_duration']
SELF_STUDY_DURATION = durations['self_study_duration']
BREAK_DURATION = durations['break_duration']

# Constants
DAYS = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday']
START_TIME = time(9, 0)
END_TIME = time(18, 30)

# Lunch break parameters
LUNCH_WINDOW_START = time(12, 30)  # Lunch breaks can start from 12:30
LUNCH_WINDOW_END = time(14, 0)     # Last lunch break must end by 14:00 
LUNCH_DURATION = 60              # Each semester gets 60 min lunch (4 slots * 15 min = 60 min, changed from 45 min)

# Initialize global variables
TIME_SLOTS = []
lunch_breaks = {}  # Global lunch breaks dictionary
GLOBAL_BASKET_SCHEDULE = {} # Key: (basket_group, session_num), Value: {'day': day, 'start_slot': slot, 'duration': dur, 'rooms': {course_code: room_id, ...}}
SCHEDULED_BASKET_COURSE_CODES = set() # Course codes successfully scheduled

def calculate_lunch_breaks(semesters):
    """Dynamically calculate staggered lunch breaks for semesters"""
    global lunch_breaks
    lunch_breaks = {}  # Reset global lunch_breaks
    total_semesters = len(semesters)
    
    if total_semesters == 0:
        return lunch_breaks
        
    # Calculate time between breaks to distribute them evenly
    total_window_minutes = (
        LUNCH_WINDOW_END.hour * 60 + LUNCH_WINDOW_END.minute -
        LUNCH_WINDOW_START.hour * 60 - LUNCH_WINDOW_START.minute
    )
    stagger_interval = (total_window_minutes - LUNCH_DURATION) / (total_semesters - 1) if total_semesters > 1 else 0
    
    # Sort semesters to ensure consistent assignment
    sorted_semesters = sorted(semesters)
    
    for i, semester in enumerate(sorted_semesters):
        start_minutes = (LUNCH_WINDOW_START.hour * 60 + LUNCH_WINDOW_START.minute + 
                         int(i * stagger_interval))
        start_hour = start_minutes // 60
        start_min = start_minutes % 60
        
        end_minutes = start_minutes + LUNCH_DURATION
        end_hour = end_minutes // 60
        end_min = end_minutes % 60
        
        lunch_breaks[semester] = (
            time(start_hour, start_min),
            time(end_hour, end_min)
        )
    
    return lunch_breaks

def initialize_time_slots():
    global TIME_SLOTS
    TIME_SLOTS = generate_time_slots()

def generate_time_slots():
    slots = []
    current_time = datetime.combine(datetime.today(), START_TIME)
    end_time = datetime.combine(datetime.today(), END_TIME)
    
    while current_time < end_time:
        current = current_time.time()
        next_time = current_time + timedelta(minutes=30)
        
        # Keep all time slots but we'll mark break times later
        slots.append((current, next_time.time()))
        current_time = next_time
    
    return slots

def load_rooms():
    rooms = {}
    try:
        # Assuming rooms.csv is now accessible in the current directory or specified path if provided
        csv_path = Path(r"C:\Users\LENOVO\Automated-Time-Table-Scheduling-At-IIIT-Dharwad\data\rooms.csv") 
        # For a robust solution, you should change the absolute path in the original code 
        # (C:\Users\LENOVO\Automated-Time-Table-Scheduling-At-IIIT-Dharwad\data\rooms.csv) to a relative path or expect the file to be uploaded/passed.
        # Since I can't access local drive paths, I'll rely on a placeholder/upload logic here.
        
        # For simplicity, I'll use a placeholder structure if not found.
        if not csv_path.exists():
            print("Warning: rooms.csv not found, creating default room allocation.")
            return {
                "L101": {'capacity': 70, 'type': 'LECTURE_ROOM', 'roomNumber': 'L101', 'schedule': {day: set() for day in range(len(DAYS))}},
                "L102": {'capacity': 70, 'type': 'LECTURE_ROOM', 'roomNumber': 'L102', 'schedule': {day: set() for day in range(len(DAYS))}},
                "C201": {'capacity': 40, 'type': 'COMPUTER_LAB', 'roomNumber': 'C201', 'schedule': {day: set() for day in range(len(DAYS))}},
                "S301": {'capacity': 120, 'type': 'SEATER_120', 'roomNumber': 'S301', 'schedule': {day: set() for day in range(len(DAYS))}},
            }

        with csv_path.open('r') as f:
            reader = csv.DictReader(f)
            for row in reader:
                rooms[row['id']] = {
                    'capacity': int(row['capacity']),
                    'type': row['type'],
                    'roomNumber': row['roomNumber'],
                    'schedule': {day: set() for day in range(len(DAYS))}
                }
    except Exception as e:
        print(f"Error loading rooms.csv: {e}")
        return None
    return rooms


def load_batch_data():
    """Load batch information directly from total_students column in combined.csv"""
    batch_info = {}
    
    # Use the uploaded file name directly
    csv_path ="C:/Users/LENOVO/Automated-Time-Table-Scheduling-At-IIIT-Dharwad/data/combined.csv"
    
    try:
        # Ensure the combined.csv is loaded from the current context
        df_batch = pd.read_csv(csv_path)
        
        # Group by Department and Semester to get total students
        grouped = df_batch.groupby(['Department', 'Semester'])
        
        for (dept, sem), group in grouped:
            # Check if total_students column exists and has values
            if 'total_students' in group.columns and not group['total_students'].isna().all():
                # Get the max number of students for this department/semester
                try:
                    total_students = int(group['total_students'].max())
                except ValueError:
                    continue # Skip if max is not convertible to int
                    
                # Default max batch size of 85
                max_batch_size = 85
                
                # Calculate number of sections needed
                num_sections = (total_students + max_batch_size - 1) // max_batch_size
                section_size = (total_students + num_sections - 1) // num_sections

                batch_info[(dept, sem)] = {
                    'total': total_students,
                    'num_sections': num_sections,
                    'section_size': section_size
                }
                
        # Process basket/elective courses individually
        basket_courses = df_batch[df_batch['Course Code'].astype(str).str.contains('^B[0-9]-')]
        
        # Process each basket course
        for _, course in basket_courses.iterrows():
            code = str(course['Course Code'])
            # Use total_students column if available
            if 'total_students' in df_batch.columns and pd.notna(course['total_students']):
                try:
                    total_students = int(course['total_students'])
                except ValueError:
                    total_students = 35 # Default
            else:
                # Default to 35 students for basket courses if not specified
                total_students = 35
                
            batch_info[('ELECTIVE', code)] = {
                'total': total_students,
                'num_sections': 1,  # Electives are typically single section for scheduling
                'section_size': total_students
            }
    except FileNotFoundError:
        print("Warning: combined.csv not found, using default batch sizes")
    except Exception as e:
        print(f"Warning: Error processing batch sizes from combined.csv: {e}")
        
    return batch_info

def find_adjacent_lab_room(room_id, rooms):
    """Find an adjacent lab room based on room numbering"""
    if not room_id or room_id not in rooms:
        return None
    
    # Get room number and extract base info
    current_num_str = ''.join(filter(str.isdigit, rooms[room_id]['roomNumber']))
    if not current_num_str: return None
    current_num = int(current_num_str)
    current_floor = current_num // 100
    
    # Look for adjacent room with same type
    for rid, room in rooms.items():
        if rid != room_id and room['type'] == rooms[room_id]['type']:
            room_num_str = ''.join(filter(str.isdigit, room['roomNumber']))
            if not room_num_str: continue
            room_num = int(room_num_str)
            
            # Check if on same floor and adjacent number
            if room_num // 100 == current_floor and abs(room_num - current_num) == 1:
                return rid
    return None

def find_suitable_room(course_type, department, semester, day, start_slot, duration, rooms, batch_info, timetable, course_code="", used_rooms=None):
    """Find suitable room(s) considering student numbers and avoiding room conflicts"""
    if not rooms:
        return "DEFAULT_ROOM"
    
    required_capacity = 60  # Default fallback
    is_basket = is_basket_course(course_code)
    
    # --- Determine Required Capacity ---
    # Load required capacity from combined.csv (using the loaded df)
    try:
        csv_path= "C:/Users/LENOVO/Automated-Time-Table-Scheduling-At-IIIT-Dharwad/data/combined.csv"

        df_room = pd.read_csv(csv_path) 
        if course_code and 'total_students' in df_room.columns:
            course_row = df_room[df_room['Course Code'].astype(str) == course_code]
            if not course_row.empty and pd.notna(course_row['total_students'].iloc[0]):
                required_capacity = int(course_row['total_students'].iloc[0])
    except Exception:
        pass # Fallback to 60 if reading fails
        
    if required_capacity == 60:
         # Fallback to batch info for group size
        if is_basket:
            elective_info = batch_info.get(('ELECTIVE', course_code))
            if elective_info:
                required_capacity = elective_info['section_size']
        else:
            dept_info = batch_info.get((department, semester))
            if dept_info:
                required_capacity = dept_info['section_size']
                

    used_room_ids = set() if used_rooms is None else used_rooms

    # --- Special handling for large classes ---
    if course_type in ['LEC', 'TUT', 'SS'] and required_capacity > 70:
        # Priority 1: 240-seater for > 120 students
        if required_capacity > 120:
            seater_240_rooms = {rid: room for rid, room in rooms.items() if 'SEATER_240' in room['type'].upper()}
            room_id = try_room_allocation(seater_240_rooms, 'LEC', required_capacity, day, start_slot, duration, used_room_ids)
            if room_id: return room_id
            
        # Priority 2: 120-seater for > 70 students
        seater_120_rooms = {rid: room for rid, room in rooms.items() if 'SEATER_120' in room['type'].upper()}
        room_id = try_room_allocation(seater_120_rooms, 'LEC', required_capacity, day, start_slot, duration, used_room_ids)
        if room_id: return room_id

    # --- Special handling for labs (split needed) ---
    if course_type in ['COMPUTER_LAB', 'HARDWARE_LAB']:
        # Check if the class size requires two adjacent labs (e.g., > 35 students per section)
        if required_capacity > 35: 
            for room_id, room in rooms.items():
                if room_id in used_room_ids or room['type'].upper() != course_type:
                    continue
                    
                slots_free = True
                for i in range(duration):
                    if start_slot + i in room['schedule'][day]:
                        slots_free = False
                        break
                
                if slots_free:
                    adjacent_room = find_adjacent_lab_room(room_id, rooms)
                    if adjacent_room and adjacent_room not in used_room_ids:
                        adjacent_free = True
                        for i in range(duration):
                            if start_slot + i in rooms[adjacent_room]['schedule'][day]:
                                adjacent_free = False
                                break
                                
                        if adjacent_free:
                            # Reserve both rooms globally
                            for i in range(duration):
                                room['schedule'][day].add(start_slot + i)
                                rooms[adjacent_room]['schedule'][day].add(start_slot + i)
                            return f"{room_id},{adjacent_room}"  # Return both room IDs
                            
        # If no split needed or split failed, try normal allocation
        return try_room_allocation(rooms, course_type, required_capacity, day, start_slot, duration, used_room_ids)

    # --- Normal Lecture/Tutorial Allocation (including basket courses) ---
    # For lectures and basket courses, try different room types in priority order
    if course_type in ['LEC', 'TUT', 'SS'] or is_basket:
        # Collect all suitable room candidates (Lecture rooms and Seater rooms)
        all_lec_rooms = {
            rid: room for rid, room in rooms.items() 
            if 'LECTURE_ROOM' in room['type'].upper() or 'SEATER' in room['type'].upper()
        }
        
        # 1. Filter rooms by capacity
        filtered_rooms = {
            rid: room for rid, room in all_lec_rooms.items() 
            if room['capacity'] >= required_capacity
        }

        # 2. Sort the filtered rooms by capacity (smallest first)
        # We use a standard list of tuples here, as try_room_allocation can iterate over it.
        sorted_room_list = sorted(filtered_rooms.items(), key=lambda item: item[1]['capacity'])
        
        # Convert back to a dictionary (optional, but maintains the structure for try_room_allocation)
        sorted_rooms_dict = dict(sorted_room_list)
        
        return try_room_allocation(sorted_rooms_dict, 'LEC', required_capacity, day, start_slot, duration, used_room_ids)

    # Fallback for unexpected course types
    return try_room_allocation(rooms, course_type, required_capacity, day, start_slot, duration, used_room_ids)

def try_room_allocation(rooms, course_type, required_capacity, day, start_slot, duration, used_room_ids):
    """Helper function to try allocating rooms of a certain type"""
    for room_id, room in rooms.items():
        if room_id in used_room_ids or room['type'].upper() == 'LIBRARY':
            continue
            
        # For lectures and tutorials, only use lecture rooms and seater rooms
        if course_type in ['LEC', 'TUT', 'SS']:
            if not ('LECTURE_ROOM' in room['type'].upper() or 'SEATER' in room['type'].upper()):
                continue
        # For labs, match lab type exactly
        elif course_type == 'COMPUTER_LAB' and room['type'].upper() != 'COMPUTER_LAB':
            continue
        elif course_type == 'HARDWARE_LAB' and room['type'].upper() != 'HARDWARE_LAB':
            continue
            
        # Check capacity except for labs which can be split into batches
        if course_type not in ['COMPUTER_LAB', 'HARDWARE_LAB'] and room['capacity'] < required_capacity:
            continue

        # Check availability (global room schedule)
        slots_free = True
        for i in range(duration):
            if start_slot + i in room['schedule'][day]:
                slots_free = False
                break
                
        if slots_free:
            # Reserve the slots in the room's global schedule
            for i in range(duration):
                room['schedule'][day].add(start_slot + i)
            return room_id
            
    return None

def get_required_room_type(course):
    """Determine required room type based on course attributes"""
    if pd.notna(course['P']) and course['P'] > 0:
        course_code = str(course['Course Code']).upper()
        # For CS courses, use computer labs
        if 'CS' in course_code or 'DS' in course_code:
            return 'COMPUTER_LAB'
        # For EC courses, use hardware labs
        elif 'EC' in course_code:
            return 'HARDWARE_LAB'
        return 'COMPUTER_LAB'  # Default to computer lab if unspecified
    else:
        # For lectures, tutorials, and self-study
        return 'LECTURE_ROOM'

# Add this function to help identify basket courses
def is_basket_course(code):
    """Check if course is part of a basket based on code prefix"""
    return code.startswith('B') and '-' in code

def get_basket_group(code):
    """Get the basket group (B1, B2 etc) from course code"""
    if is_basket_course(code):
        return code.split('-')[0]
    return None

def get_basket_group_slots(timetable, day, basket_group):
    """Find existing slots with courses from same basket group"""
    basket_slots = []
    for slot_idx, slot in timetable[day].items():
        code = slot.get('code', '')
        if code and get_basket_group(code) == basket_group:
            basket_slots.append(slot_idx)
    return basket_slots

# Load data from CSV with robust error handling
try:
    # Try different encodings and handle BOM
    encodings_to_try = ['utf-8-sig', 'utf-8', 'cp1252']
    df = None
    last_error = None
    
    for encoding in encodings_to_try:
        try:
            # Use the uploaded file name
            csv_path = "C:/Users/LENOVO/Automated-Time-Table-Scheduling-At-IIIT-Dharwad/data/combined.csv" 
            df = pd.read_csv(csv_path, encoding=encoding)
            # Convert empty strings and 'nan' strings to actual NaN
            df = df.replace(r'^\s*$', pd.NA, regex=True)
            df = df.replace('nan', pd.NA)
            break
        except UnicodeDecodeError:
            continue
        except Exception as e:
            last_error = e
            continue
            
    if df is None:
        print(f"Error: Unable to read combined.csv. Please check the file format.\nDetails: {str(last_error)}")
        # If unable to read, create an empty DataFrame to prevent downstream crashes
        df = pd.DataFrame(columns=['Department', 'Semester', 'Course Code', 'Course Name', 'L', 'T', 'P', 'S', 'C', 'Faculty', 'Schedule', 'total_students'])

except Exception as e:
    print(f"Error: Failed to load combined.csv.\nDetails: {str(e)}")
    df = pd.DataFrame(columns=['Department', 'Semester', 'Course Code', 'Course Name', 'L', 'T', 'P', 'S', 'C', 'Faculty', 'Schedule', 'total_students'])

if df.empty:
    print("Warning: No data found in combined.csv or failed to load.")


def is_break_time(slot, semester=None):
    """Check if a time slot falls within break times"""
    global lunch_breaks
    start, end = slot
    
    # Morning break: 10:30-11:00 (slots 30 min long, so checks 10:30, 11:00 start)
    morning_break = (time(10, 30) <= start < time(11, 0))
    
    # Staggered lunch breaks based on semester
    lunch_break = False
    if semester:
        try:
            base_sem = int(str(semester)[0])  # Get base semester number (e.g., 4 from 4A)
            if base_sem in lunch_breaks:
                lunch_start, lunch_end = lunch_breaks[base_sem]
                lunch_break = (lunch_start <= start < lunch_end)
        except:
             pass # Handle cases where semester is not numeric 
    else:
        # For general checks without semester info, block all potential lunch periods
        lunch_break = any(lunch_start <= start < lunch_end 
                          for lunch_start, lunch_end in lunch_breaks.values())
    
    return morning_break or lunch_break

def is_lecture_scheduled(timetable, day, start_slot, end_slot):
    """Check if there's a lecture scheduled in the given time range"""
    for slot in range(start_slot, end_slot):
        if (slot < len(TIME_SLOTS) and 
            timetable[day][slot]['type'] and 
            timetable[day][slot]['type'] in ['LEC', 'LAB', 'TUT']):
            return True
    return False

def calculate_required_slots(course):
    """Calculate how many slots needed based on L, T, P, S values and credits"""
    l = float(course['L']) if pd.notna(course['L']) else 0  # Lecture credits
    t = int(course['T']) if pd.notna(course['T']) else 0    # Tutorial hours
    p = int(course['P']) if pd.notna(course['P']) else 0    # Lab hours
    s = int(course['S']) if pd.notna(course['S']) else 0    # Self study hours
    c = int(course['C']) if pd.notna(course['C']) else 0    # Total credits
    
    # Check if course is self-study only
    if s > 0 and l == 0 and t == 0 and p == 0:
        return 0, 0, 0, 0
        
    # Calculate number of lecture sessions based on credits
    # Each lecture slot is 1.5 hours (3 * 30 min slots)
    lecture_sessions = 0
    if l > 0:
         # Assuming 3 credits = 2 sessions (3 hours total), 2 credits = 1 session (1.5 hours)
         # Using L as the number of *sessions* to schedule
         lecture_sessions = int(l) # Using L directly for number of sessions
        
    # Other calculations remain the same
    tutorial_sessions = t  
    lab_sessions = p // 2 # 2 hours per lab session (4 * 30 min slots)
    self_study_sessions = s // 4 if (l > 0 or t > 0 or p > 0) else 0 # 2 hours per self study session (4 * 30 min slots)
    
    return lecture_sessions, tutorial_sessions, lab_sessions, self_study_sessions

def select_faculty(faculty_str):
    """Select a faculty from potentially multiple options."""
    if pd.isna(faculty_str):
        return ""
    if '/' in faculty_str:
        # Split by slash and strip whitespace
        faculty_options = [f.strip() for f in faculty_str.split('/')]
        return faculty_options[0]  # Take first faculty as default
    return faculty_str

def check_faculty_daily_components(professor_schedule, faculty, day, department, semester, section, timetable, course_code=None, activity_type=None):
    """Check faculty/course scheduling constraints for the day"""
    component_count = 0
    
    # Check all slots for this day
    for slot_data in timetable[day].values():
        if slot_data.get('faculty', '') == faculty and slot_data.get('type') in ['LEC', 'LAB', 'TUT']:
            slot_code = slot_data.get('code', '')
            if slot_code:
                # Count only the starting slot of a session
                # For non-basket courses, count the session once
                if not is_basket_course(slot_code):
                    component_count += 1
                # For basket courses, each course in the group counts as a component
                elif 'basket_group_members' in slot_data:
                    # Only count if the faculty is teaching one of the courses in the group
                    for member in slot_data['basket_group_members']:
                        if member['faculty'] == faculty:
                            component_count += 1
                            break

    # Special handling for basket courses - allow parallel scheduling of basket electives
    if course_code and is_basket_course(course_code):
        # Allow more flexibility for basket courses (e.g., max 3 components)
        return component_count < 3 
    
    # Check the component count for the faculty in general
    return component_count < 2  # Keep max 2 components per day limit for regular courses

def check_faculty_course_gap(professor_schedule, timetable, faculty, course_code, day, start_slot):
    """Check if there is sufficient gap (3 hours) between sessions of same course"""
    min_gap_hours = 3
    slots_per_hour = 2  # Assuming 30-min slots
    required_gap = min_gap_hours * slots_per_hour
    
    # Check a window before the start slot
    for i in range(max(0, start_slot - required_gap), start_slot):
        if i < len(TIME_SLOTS):
            slot_data = timetable[day].get(i)
            if slot_data and slot_data['code'] == course_code and slot_data['type'] in ['LEC', 'TUT']:
                return False
                
    # Check a window after the scheduled end slot (start_slot + duration)
    duration = LECTURE_DURATION if timetable[day].get(start_slot, {}).get('type') == 'LEC' else TUTORIAL_DURATION
    end_slot = start_slot + duration
    for i in range(end_slot, min(len(TIME_SLOTS), end_slot + required_gap)):
        slot_data = timetable[day].get(i)
        if slot_data and slot_data['code'] == course_code and slot_data['type'] in ['LEC', 'TUT']:
            return False
    
    return True

def load_reserved_slots():
    """Load reserved time slots from CSV file"""
    try:
        # Assuming the user uploads reserved_slots.csv as well
        reserved_slots_path = 'reserved_slots.csv' 
        if not os.path.exists(reserved_slots_path):
            print("Warning: reserved_slots.csv not found, no slots will be reserved")
            return {day: {} for day in DAYS}
            
        df_reserved = pd.read_csv(reserved_slots_path)
        reserved = {day: {} for day in DAYS}
        
        for _, row in df_reserved.iterrows():
            day = row['Day']
            # Safely handle time parsing
            try:
                start = datetime.strptime(row['Start Time'], '%H:%M').time()
                end = datetime.strptime(row['End Time'], '%H:%M').time()
            except ValueError:
                print(f"Warning: Invalid time format in reserved_slots.csv for row: {row}")
                continue
                
            department = str(row['Department'])
            # Handle semester sections (e.g., "4" matches "4A" and "4B")
            semesters = []
            sem_str = str(row['Semester']).replace(' ', '')
            for s in sem_str.split(';'):
                if s.isdigit():  # If just a number like "4"
                    base_sem = int(s)  
                    semesters.extend([f"{base_sem}A", f"{base_sem}B", str(base_sem)])
                elif s:  # If already has section like "4A" or is ALL
                    semesters.append(s)
            
            key = (department, tuple(semesters))
            if day not in reserved:
                reserved[day] = {}
            if key not in reserved[day]:
                reserved[day][key] = []
                
            reserved[day][key].append((start, end))
        return reserved
    except Exception as e:
        print(f"Warning: Error loading reserved slots: {str(e)}")
        return {day: {} for day in DAYS}

def is_slot_reserved(slot, day, semester, department, reserved_slots):
    """Check if a time slot is reserved for this semester and department"""
    if day not in reserved_slots:
        return False
        
    slot_start, slot_end = slot
    
    # Check each reservation
    for (dept, semesters), slots in reserved_slots[day].items():
        # Match if department is ALL or matches exactly
        if dept.upper() == 'ALL' or dept == department:
            # Match if semester is in the expanded semester list
            if 'ALL' in semesters or str(semester) in semesters or any(str(semester).startswith(s) for s in semesters if s and s.isdigit()):
                for reserved_start, reserved_end in slots:
                    # Check for overlap
                    if (slot_start >= reserved_start and slot_start < reserved_end) or \
                       (slot_end > reserved_start and slot_end <= reserved_end) or \
                       (reserved_start >= slot_start and reserved_end <= slot_end): # Full overlap case
                        return True
    return False

def load_faculty_preferences():
    """Load faculty scheduling preferences from CSV"""
    preferences = {}
    try:
        # Assuming FACULTY.csv is uploaded or is in the expected 'tt data' subfolder
        # I'll modify the path to check the current directory first for robustness
        faculty_path = 'FACULTY.csv'
        if not os.path.exists(faculty_path):
            faculty_path = os.path.join('tt data', 'FACULTY.csv') # Fallback to original path
            if not os.path.exists(faculty_path):
                print("Warning: FACULTY.csv not found, proceeding without faculty preferences")
                return {}
        
        with open(faculty_path, 'r') as f:
            reader = csv.DictReader(f)
            for row in reader:
                preferred_days = [d.strip() for d in row['Preferred Days'].split(';')] if row['Preferred Days'] else []
                preferred_times = []
                if row['Preferred Times']:
                    time_ranges = row['Preferred Times'].split(';')
                    for time_range in time_ranges:
                        try:
                            start, end = time_range.split('-')
                            start_time = datetime.strptime(start.strip(), '%H:%M').time()
                            end_time = datetime.strptime(end.strip(), '%H:%M').time()
                            preferred_times.append((start_time, end_time))
                        except Exception as e:
                            print(f"Warning: Invalid time format '{time_range}' in FACULTY.csv for {row['Name']}")
                            continue
                            
                preferences[row['Name']] = {
                    'preferred_days': preferred_days,
                    'preferred_times': preferred_times
                }
    except Exception as e:
        print(f"Warning: Error loading FACULTY.csv: {e}. Proceeding without faculty preferences.")
        return {}
    return preferences

def is_preferred_slot(faculty, day, time_slot, faculty_preferences):
    """Check if a time slot is within faculty's preferences"""
    if faculty not in faculty_preferences:
        return True  # No preferences specified, any slot is fine
        
    prefs = faculty_preferences[faculty]
    
    # Check if day is preferred
    if prefs['preferred_days'] and DAYS[day] not in prefs['preferred_days']:
        return False
        
    # Check if time is within preferred ranges
    if prefs['preferred_times']:
        slot_start, slot_end = time_slot
        for pref_start, pref_end in prefs['preferred_times']:
            # Check for full slot containment
            if (slot_start >= pref_start and slot_end <= pref_end):
                return True
        return False
        
    return True  # No time preferences specified

def get_course_priority(course):
    """Calculate course scheduling priority based on constraints"""
    priority = 0
    code = str(course['Course Code'])
    
    # Give regular course labs highest priority
    if pd.notna(course['P']) and course['P'] > 0 and not is_basket_course(code):
        priority += 10 
        if 'CS' in code or 'EC' in code: 
            priority += 2 # Extra priority for specialized labs
    elif is_basket_course(code):
        priority += 5 # High priority for Basket courses to ensure they get synchronized slots first
    elif pd.notna(course['L']) and float(course['L']) > 0:
        priority += 3  
    elif pd.notna(course['T']) and course['T'] > 0:
        priority += 2  
    return priority

def get_best_slots(timetable, professor_schedule, faculty, day, duration, reserved_slots, semester, department, faculty_preferences):
    """Find best available consecutive slots in a day considering faculty preferences"""
    best_slots = []
    preferred_slots = []
    
    for start_slot in range(len(TIME_SLOTS) - duration + 1):
        slots_free = True
        
        # Check faculty daily component limit before checking all slots
        # For a full check, we assume that if we take this block, it adds a component.
        # This is a heuristic and might cause issues if a component starts but fails room allocation later.
        # But for efficiency, we check here:
        # if not check_faculty_daily_components(professor_schedule, faculty, day, department, semester, 0, timetable, course_code="dummy", activity_type="dummy"):
        #    continue # Skip the entire block check for this day if components maxed out
        
        for i in range(duration):
            current_slot = start_slot + i
            
            # Check for conflict in the section timetable (already occupied by LEC/LAB/TUT/SS or a Basket course)
            if timetable[day][current_slot]['type'] is not None:
                slots_free = False
                break
                
            # Check faculty availability (globally tracked)
            if current_slot in professor_schedule.get(faculty, {}).get(day, set()):
                slots_free = False
                break
                
            # Check for breaks
            if is_break_time(TIME_SLOTS[current_slot], semester) or \
               is_slot_reserved(TIME_SLOTS[current_slot], DAYS[day], semester, department, reserved_slots):
                slots_free = False
                break
                
            # Soft Constraint: Breaks between lectures/tutorials are important
            # Ensure one break slot (30 min = 1 slot) before and after if not a lab/tutorial
            if duration == LECTURE_DURATION:
                # Check slot before (if available and not already a break/end of activity)
                if current_slot > 0 and i == 0:
                    prev_slot_data = timetable[day][current_slot - 1]
                    if (prev_slot_data['type'] in ['LEC', 'TUT'] or 
                        (current_slot - 1) in professor_schedule.get(faculty, {}).get(day, set())):
                        slots_free = False
                        break
                
                # Check slot after (handled implicitly by requiring LECTURE_DURATION slots, 
                # but an explicit check for the slot *after* the entire block helps)
                if current_slot + 1 < len(TIME_SLOTS) and i == duration - 1:
                    next_slot_data = timetable[day][current_slot + 1]
                    if (next_slot_data['type'] in ['LEC', 'TUT'] or 
                        (current_slot + 1) in professor_schedule.get(faculty, {}).get(day, set())):
                        slots_free = False
                        break


        if slots_free:
            # Prioritize morning slots for labs or preferred slots for lectures
            if duration == LAB_DURATION:
                slot_time = TIME_SLOTS[start_slot][0]
                if slot_time < time(12, 30):  # Before lunch
                    preferred_slots.append(start_slot)
                else:
                    best_slots.append(start_slot)
            elif is_preferred_slot(faculty, day, TIME_SLOTS[start_slot], faculty_preferences):
                preferred_slots.append(start_slot)
            else:
                best_slots.append(start_slot)
    
    return preferred_slots + best_slots

class UnscheduledComponent:
    def __init__(self, department, semester, code, name, faculty, component_type, sessions, section='', reason=''):
        self.department = department
        self.semester = semester
        self.code = code
        self.name = name
        self.faculty = faculty  
        self.component_type = component_type
        self.sessions = sessions
        self.section = section
        self.reason = reason
        
    def __eq__(self, other):
        if not isinstance(other, UnscheduledComponent):
            return False
        return (self.department == other.department and
                self.semester == other.semester and
                self.code == other.code and
                self.component_type == other.component_type and
                self.section == other.section)
        
    def __hash__(self):
        return hash((self.department, self.semester, self.code, self.component_type, self.section))

def unscheduled_reason(course, department, semester, professor_schedule, rooms, component_type, check_attempts):
    """Generate detailed reason why a course component couldn't be scheduled"""
    faculty = select_faculty(course['Faculty'])
    code = str(course['Course Code'])
    
    # Check faculty availability
    faculty_slots_used = 0
    for day in range(len(DAYS)):
        faculty_slots_used += len(professor_schedule.get(faculty, {}).get(day, set()))
    
    # If faculty is heavily scheduled
    if faculty_slots_used > 20:  # Threshold: 10 hours of teaching per week
        return f"Faculty '{faculty}' already has {faculty_slots_used/2:.1f} hours of teaching scheduled"
    
    # Check room availability issues
    if component_type == 'LAB':
        lab_rooms_free_slots = 0
        for rid, room in rooms.items():
            if 'LAB' in room['type'].upper() or 'COMPUTER' in room['type'].upper():
                total_slots = len(DAYS) * (len(TIME_SLOTS))
                used_slots = sum(len(room['schedule'].get(day, [])) for day in range(len(DAYS)))
                lab_rooms_free_slots += (total_slots - used_slots)
        
        if lab_rooms_free_slots < 5:  # Very few lab slots left
            return f"Lab rooms almost fully booked ({lab_rooms_free_slots} slots left)"
    
    # Check for large classes with insufficient large rooms
    if 'total_students' in course and pd.notna(course['total_students']):
        try:
            total_students = int(course['total_students'])
            if total_students > 100:
                large_rooms_available = any(room['type'].upper() == 'SEATER_120' or room['type'].upper() == 'SEATER_240' 
                                            for room in rooms.values())
                
                if not large_rooms_available:
                    return f"No large rooms available with capacity for {total_students} students"
        except (ValueError, TypeError):
            pass
    
    # Check timeslot conflicts with other courses in same department/semester
    if check_attempts > 800:  # If we made many attempts but still couldn't find a slot
        return f"No suitable timeslot found after {check_attempts} attempts - heavy scheduling conflicts"
        
    # Default reason
    duration_map = {
        'LEC': f"{LECTURE_DURATION/2:.1f} hour",
        'LAB': f"{LAB_DURATION/2:.1f} hour",
        'TUT': f"{TUTORIAL_DURATION/2:.1f} hour"
    }
    duration_str = duration_map.get(component_type, "")
    
    return f"Could not find compatible {duration_str} timeslot for {code} {component_type} with faculty {faculty}"


# --- NEW BASKET SCHEDULING LOGIC ---

def schedule_basket_groups_globally(all_courses, professor_schedule, rooms, batch_info):
    """
    Finds a single common time slot for all courses belonging to the same basket group
    and assigns a unique room/faculty to each course in the group globally.
    Updates professor_schedule and rooms directly.
    """
    global GLOBAL_BASKET_SCHEDULE, SCHEDULED_BASKET_COURSE_CODES
    
    basket_groups_to_schedule = {}
    
    # Filter for basket courses (e.g., B1-CourseCode) that should be scheduled
    basket_courses_df = all_courses[
        all_courses['Course Code'].astype(str).str.contains('^B[0-9]-(?!nan)') &
        ((all_courses['Schedule'].fillna('Yes').str.upper() == 'YES') | (all_courses['Schedule'].isna()))
    ].copy()
    
    # 1. Group courses by Basket Group and consolidate their requirements
    for _, course in basket_courses_df.iterrows():
        code = str(course['Course Code'])
        group = get_basket_group(code)
        
        if group:
            if group not in basket_groups_to_schedule:
                basket_groups_to_schedule[group] = {
                    'courses': {}, # Stores unique course details: {code: course_dict}
                    'sessions_needed': 0
                }
            
            l, _, _, _ = calculate_required_slots(course)
            sessions = l 
            
            course_details = course.to_dict()
            course_details['sessions_needed'] = sessions
            basket_groups_to_schedule[group]['courses'][code] = course_details
            
            basket_groups_to_schedule[group]['sessions_needed'] = max(basket_groups_to_schedule[group]['sessions_needed'], sessions)
            
    # 2. Schedule each group globally
    for group, group_data in basket_groups_to_schedule.items():
        courses = list(group_data['courses'].values())
        sessions_needed = group_data['sessions_needed']
        
        if sessions_needed == 0: continue
            
        duration = LECTURE_DURATION
        required_rooms = len(courses)
        
        # Calculate max capacity needed
        max_required_capacity = 0
        for course_detail in courses:
             if course_detail.get('total_students') and pd.notna(course_detail['total_students']):
                  try:
                      max_required_capacity = max(max_required_capacity, int(course_detail['total_students']))
                  except: pass
        if max_required_capacity == 0: max_required_capacity = 35

        # For each required session
        for session_num in range(int(sessions_needed)):
            scheduled = False
            attempts = 0
            
            while not scheduled and attempts < 2000:
                attempts += 1
                
                day = random.randint(0, len(DAYS)-1)
                start_slot = random.randint(0, len(TIME_SLOTS) - duration)
                
                # --- A. Check Faculty Availability & Daily Component Limit ---
                faculty_free = True
                faculty_map = {str(c['Course Code']): select_faculty(str(c['Faculty'])) for c in courses}
                
                for code, f in faculty_map.items():
                    if f not in professor_schedule:
                        professor_schedule[f] = {d: set() for d in range(len(DAYS))}
                    
                    # Check for conflicts with existing schedule
                    for i in range(duration):
                         if (start_slot + i) in professor_schedule.get(f, {}).get(day, set()):
                             faculty_free = False
                             break
                    if not faculty_free: break
                    
                if not faculty_free: continue
                
                # --- B. Check for General Time Slot Conflicts (Breaks, Reserved) ---
                is_blocked = False
                for i in range(duration):
                    if is_break_time(TIME_SLOTS[start_slot + i]):
                        is_blocked = True
                        break
                    # We skip the semester-specific reserved slot check here as it's too restrictive for a global slot.
                if is_blocked: continue
                
                # --- C. Check for Room Availability ---
                potential_room_ids = []
                for room_id, room in rooms.items():
                    if 'LECTURE_ROOM' in room['type'].upper() or 'SEATER' in room['type'].upper():
                        if room.get('capacity', 0) >= max_required_capacity:
                            is_room_free = True
                            for i in range(duration):
                                if (start_slot + i) in room['schedule'][day]:
                                    is_room_free = False
                                    break
                            
                            if is_room_free:
                                potential_room_ids.append(room_id)
                
                if len(potential_room_ids) < required_rooms:
                    continue 
                
                # --- Schedule the Basket Group Session ---
                assigned_rooms_map = {} 
                potential_room_ids.sort(key=lambda rid: rooms[rid]['capacity']) 
                sorted_courses = sorted(courses, key=lambda c: str(c['Course Code']))
                
                for i, course_detail in enumerate(sorted_courses):
                    course_code = str(course_detail['Course Code'])
                    room_id = potential_room_ids[i]
                    faculty = faculty_map[course_code]
                    
                    # Reserve the room slots (globally)
                    for j in range(duration):
                         rooms[room_id]['schedule'][day].add(start_slot + j)
                         
                    # Reserve faculty slots (globally)
                    for j in range(duration):
                         professor_schedule[faculty][day].add(start_slot + j)
                         
                    assigned_rooms_map[course_code] = room_id
                
                # Store the global schedule for this session
                GLOBAL_BASKET_SCHEDULE[(group, session_num)] = {
                    'day': day, 
                    'start_slot': start_slot, 
                    'duration': duration,
                    'rooms': assigned_rooms_map 
                }
                
                for course_detail in courses:
                    SCHEDULED_BASKET_COURSE_CODES.add(str(course_detail['Course Code']))
                    
                scheduled = True
            
            if not scheduled:
                 print(f"Warning: Failed to schedule a common slot for Basket Group {group} Session {session_num+1} after {attempts} attempts.")
# --- END NEW BASKET SCHEDULING LOGIC ---

def fill_basket_schedule(timetable, department, semester, global_basket_schedule, basket_courses_df):
    """Fills the section's timetable with pre-scheduled basket courses."""
    
    # 1. Identify which unique basket course codes belong to this Dept/Sem section
    section_courses = basket_courses_df[(basket_courses_df['Department'] == department) & 
                                        (basket_courses_df['Semester'] == semester)]
    
    relevant_basket_map = {}
    for _, course in section_courses.iterrows():
        code = str(course['Course Code'])
        group = get_basket_group(code)
        if group:
            if group not in relevant_basket_map:
                relevant_basket_map[group] = []
            relevant_basket_map[group].append(course.to_dict())

    # 2. Iterate through all globally scheduled basket sessions
    for (group, session_num), schedule_data in global_basket_schedule.items():
        if group in relevant_basket_map:
            day = schedule_data['day']
            start_slot = schedule_data['start_slot']
            duration = schedule_data['duration']
            assigned_rooms = schedule_data['rooms']
            
            all_course_details = []
            for course_details in relevant_basket_map[group]:
                 course_code = str(course_details['Course Code'])
                 if course_code in assigned_rooms:
                    all_course_details.append({
                        'code': course_code, 
                        'name': str(course_details['Course Name']),
                        'faculty': select_faculty(str(course_details['Faculty'])),
                        'classroom': assigned_rooms[course_code]
                    })
            
            if not all_course_details:
                continue
                
            all_course_details.sort(key=lambda x: x['code'])
            main_marker_course = all_course_details[0]

            for i in range(duration):
                current_slot = start_slot + i
                
                # Check for conflicts before writing (should only happen if a regular course slipped through or failed late checks)
                if timetable[day][current_slot]['type'] is None or is_basket_course(timetable[day][current_slot]['code']):
                     
                     timetable[day][current_slot]['type'] = 'LEC' 
                     
                     if i == 0:
                        # Mark first slot with the first course's details and the group's members list
                        timetable[day][current_slot]['code'] = main_marker_course['code']
                        timetable[day][current_slot]['name'] = main_marker_course['name']
                        timetable[day][current_slot]['faculty'] = main_marker_course['faculty']
                        timetable[day][current_slot]['classroom'] = main_marker_course['classroom']
                        timetable[day][current_slot]['basket_group_members'] = all_course_details 
                     else:
                        # Mark continuation slots as empty placeholders, but keep the LEC type
                        timetable[day][current_slot]['type'] = 'LEC'
                        timetable[day][current_slot]['code'] = ''
                        timetable[day][current_slot]['name'] = ''
                        timetable[day][current_slot]['faculty'] = ''
                        timetable[day][current_slot]['classroom'] = ''
                        timetable[day][current_slot]['basket_group_members'] = []

def generate_all_timetables():
    global lunch_breaks, GLOBAL_BASKET_SCHEDULE, SCHEDULED_BASKET_COURSE_CODES
    initialize_time_slots()  # Initialize time slots before using
    
    reserved_slots = load_reserved_slots()
    faculty_preferences = load_faculty_preferences()
    workbooks = {}  
    
    # Initialize tracking objects
    professor_schedule = {}    
    rooms = load_rooms()
    batch_info = load_batch_data()
    unscheduled_components = set()

    # Color palette and basket colors (copied from original code)
    subject_colors = [
        "FFB6C1", "98FB98", "87CEFA", "DDA0DD", "F0E68C",  
        "E6E6FA", "FFDAB9", "B0E0E6", "FFA07A", "D8BFD8",
        "AFEEEE", "F08080", "90EE90", "ADD8E6", "FFB6C1"
    ]
    basket_group_colors = {
        'B1': "FF9999", 'B2': "99FF99", 'B3': "9999FF", 'B4': "FFFF99", 'B5': "FF99FF", 
        'B6': "99FFFF", 'B7': "FFB366", 'B8': "B366FF", 'B9': "66FFB3"
    }
    self_study_courses = []

    # Calculate lunch breaks dynamically
    all_semesters = sorted(set(int(str(sem)[0]) for sem in df['Semester'].unique() if str(sem) and str(sem)[0].isdigit()))
    lunch_breaks = calculate_lunch_breaks(all_semesters)

    # --- NEW STEP: Pre-schedule all Basket Groups globally ---
    schedule_basket_groups_globally(df, professor_schedule, rooms, batch_info)
    # --- END NEW STEP ---

    for department in df['Department'].unique():
        # Create new workbook for each department
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet
        workbooks[department] = wb
        
        # Track assigned faculty for courses
        course_faculty_assignments = {}
        
        # Process all semesters for this department
        for semester in df[df['Department'] == department]['Semester'].unique():
            # Filter out courses marked as not to be scheduled OR already scheduled as part of a basket group
            courses = df[(df['Department'] == department) &  
                         (df['Semester'] == semester) &
                         ((df['Schedule'].fillna('Yes').str.upper() == 'YES') | (df['Schedule'].isna())) &
                         (~df['Course Code'].isin(SCHEDULED_BASKET_COURSE_CODES)) # <-- New filter
                        ].copy()
            
            if courses.empty:
                continue

            # First separate and handle lab scheduling
            lab_courses = courses[courses['P'] > 0].copy()
            lab_courses['priority'] = lab_courses.apply(get_course_priority, axis=1)
            lab_courses = lab_courses.sort_values('priority', ascending=False)

            # Handle remaining courses (non-lab, non-basket)
            non_lab_courses = courses[courses['P'] == 0].copy()
            non_lab_courses['priority'] = non_lab_courses.apply(get_course_priority, axis=1)
            non_lab_courses = non_lab_courses.sort_values('priority', ascending=False)

            # Combine sorted courses with labs first
            courses = pd.concat([lab_courses, non_lab_courses])

            # Get section info
            dept_info = batch_info.get((department, semester))
            num_sections = dept_info['num_sections'] if dept_info else 1

            # First identify self-study only courses
            for _, course in courses.iterrows():
                l = float(course['L']) if pd.notna(course['L']) else 0
                t = int(course['T']) if pd.notna(course['T']) else 0
                p = int(course['P']) if pd.notna(course['P']) else 0
                s = int(course['S']) if pd.notna(course['S']) else 0
                
                if s > 0 and l == 0 and t == 0 and p == 0:
                    self_study_courses.append({
                        'code': str(course['Course Code']),
                        'name': str(course['Course Name']),
                        'faculty': select_faculty(str(course['Faculty'])),
                        'department': department,
                        'semester': semester
                    })
                    courses = courses[courses['Course Code'] != course['Course Code']] # Remove from main scheduling

            for section in range(num_sections):
                section_title = f"{department}_{semester}" if num_sections == 1 else f"{department}_{semester}_{chr(65+section)}"
                ws = wb.create_sheet(title=section_title)
                
                # Initialize timetable structure
                timetable = {day: {slot: {'type': None, 'code': '', 'name': '', 'faculty': '', 'classroom': '', 'basket_group_members': []}  # Added new field
                             for slot in range(len(TIME_SLOTS))} for day in range(len(DAYS))}
                
                # --- NEW STEP: Fill timetable with pre-scheduled basket sessions ---
                basket_courses_df = df[df['Course Code'].astype(str).str.contains('^B[0-9]-(?!nan)')].copy()
                fill_basket_schedule(timetable, department, semester, GLOBAL_BASKET_SCHEDULE, basket_courses_df)
                # --- END NEW STEP ---
                
                # Create a mapping for subject colors and faculty
                subject_color_map = {}
                course_faculty_map = {} 
                color_idx = 0
                
                # Initialize map with all courses (including those scheduled/filtered out) for legend/color consistency
                for _, course in df.iterrows():
                    code = str(course['Course Code'])
                    if code not in subject_color_map and code and code != 'nan':
                        if is_basket_course(code):
                            basket_group = get_basket_group(code)
                            subject_color_map[code] = basket_group_colors.get(basket_group, subject_colors[color_idx % len(subject_colors)])
                        else:
                            subject_color_map[code] = subject_colors[color_idx % len(subject_colors)]
                        course_faculty_map[code] = {
                            'name': str(course['Course Name']),
                            'faculty': str(course['Faculty'])
                        }
                        color_idx += 1
                
                # Process remaining courses (non-basket, non-self-study, non-scheduled)
                for _, course in courses.iterrows():
                    code = str(course['Course Code'])
                    name = str(course['Course Name'])
                    faculty_options = str(course['Faculty'])
                    
                    # Determine faculty for this section
                    if code in course_faculty_assignments:
                        if '/' in faculty_options:
                             # Logic to pick unassigned faculty for the section
                            available_faculty = [f.strip() for f in faculty_options.split('/') if f.strip() not in course_faculty_assignments[code]]
                            faculty = available_faculty[0] if available_faculty else select_faculty(faculty_options)
                            if faculty not in course_faculty_assignments[code]:
                                course_faculty_assignments[code].append(faculty)
                        else:
                            faculty = select_faculty(faculty_options)
                    else:
                        faculty = select_faculty(faculty_options)
                        course_faculty_assignments[code] = [faculty]
                        
                    if faculty not in professor_schedule:
                        professor_schedule[faculty] = {day: set() for day in range(len(DAYS))}

                    lecture_sessions, tutorial_sessions, lab_sessions, self_study_sessions_req = calculate_required_slots(course)

                    # --- Schedule Labs --- (High Priority)
                    if lab_sessions > 0:
                        room_type = get_required_room_type(course)
                        for _ in range(lab_sessions):
                            scheduled = False
                            attempts = 0
                            days = list(range(len(DAYS)))
                            random.shuffle(days)
                            
                            for day in days:
                                # Get all possible slots (this uses best_slots which checks for faculty/time conflicts)
                                possible_slots = get_best_slots(timetable, professor_schedule, faculty, day, LAB_DURATION, reserved_slots, semester, department, faculty_preferences)
                                random.shuffle(possible_slots)
                                
                                for start_slot in possible_slots:
                                    # Find room (reserves room globally if successful)
                                    room_id = find_suitable_room(room_type, department, semester,
                                                                 day, start_slot, LAB_DURATION,
                                                                 rooms, batch_info, timetable, code)
                                    
                                    if room_id:
                                        classroom = room_id if ',' not in str(room_id) else f"{room_id.split(',')[0]}+{room_id.split(',')[1]}"
                                        
                                        for i in range(LAB_DURATION):
                                            professor_schedule[faculty][day].add(start_slot+i)
                                            timetable[day][start_slot+i]['type'] = 'LAB'
                                            timetable[day][start_slot+i]['code'] = code if i == 0 else ''
                                            timetable[day][start_slot+i]['name'] = name if i == 0 else ''
                                            timetable[day][start_slot+i]['faculty'] = faculty if i == 0 else ''
                                            timetable[day][start_slot+i]['classroom'] = classroom if i == 0 else ''
                                        scheduled = True
                                        break
                                if scheduled: break

                            if not scheduled:
                                unscheduled_components.add(
                                    UnscheduledComponent(department, semester, code, name,
                                                         faculty, 'LAB', 1, section,
                                                         "Could not find suitable room and time slot combination")
                                )

                    # --- Schedule Lectures (LEC) and Tutorials (TUT) ---
                    components_to_schedule = [('LEC', LECTURE_DURATION, lecture_sessions), 
                                              ('TUT', TUTORIAL_DURATION, tutorial_sessions)]
                    
                    for activity_type, duration, sessions_to_schedule in components_to_schedule:
                        for _ in range(sessions_to_schedule):
                            scheduled = False
                            attempts = 0
                            
                            while not scheduled and attempts < 1000:
                                attempts += 1
                                day = random.randint(0, len(DAYS)-1)
                                
                                if not check_faculty_daily_components(professor_schedule, faculty, day, department, semester, section, timetable, code, activity_type):
                                    continue
                                
                                # Use get_best_slots to find a good starting slot
                                possible_slots = get_best_slots(timetable, professor_schedule, faculty, day, duration, reserved_slots, semester, department, faculty_preferences)
                                random.shuffle(possible_slots)
                                
                                for start_slot in possible_slots:
                                    # Final check for faculty-course gap after selecting slot
                                    if not check_faculty_course_gap(professor_schedule, timetable, faculty, code, day, start_slot):
                                         continue
                                         
                                    # Find room (reserves room globally if successful)
                                    room_id = find_suitable_room('LECTURE_ROOM', department, semester, 
                                                                 day, start_slot, duration, 
                                                                 rooms, batch_info, timetable, code)
                                    
                                    if room_id:
                                        classroom = room_id
                                        
                                        for i in range(duration):
                                            professor_schedule[faculty][day].add(start_slot+i)
                                            timetable[day][start_slot+i]['type'] = activity_type
                                            timetable[day][start_slot+i]['code'] = code if i == 0 else ''
                                            timetable[day][start_slot+i]['name'] = name if i == 0 else ''
                                            timetable[day][start_slot+i]['faculty'] = faculty if i == 0 else ''
                                            timetable[day][start_slot+i]['classroom'] = classroom if i == 0 else ''
                                        scheduled = True
                                        break
                                if scheduled: break
                                
                            if not scheduled:
                                unscheduled_components.add(
                                    UnscheduledComponent(department, semester, code, name, 
                                                         faculty, activity_type, 1, section)
                                )

                    # --- Schedule Self-Study (SS) ---
                    for _ in range(self_study_sessions_req):
                        scheduled = False
                        attempts = 0
                        while not scheduled and attempts < 1000:
                            attempts += 1
                            day = random.randint(0, len(DAYS)-1)
                            
                            possible_slots = get_best_slots(timetable, professor_schedule, faculty, day, SELF_STUDY_DURATION, reserved_slots, semester, department, faculty_preferences)
                            random.shuffle(possible_slots)
                            
                            for start_slot in possible_slots:
                                room_id = find_suitable_room('LECTURE_ROOM', department, semester, 
                                                             day, start_slot, SELF_STUDY_DURATION, 
                                                             rooms, batch_info, timetable, code)
                                
                                if room_id:
                                    classroom = room_id
                                    
                                    for i in range(SELF_STUDY_DURATION):
                                        professor_schedule[faculty][day].add(start_slot+i)
                                        timetable[day][start_slot+i]['type'] = 'SS' 
                                        timetable[day][start_slot+i]['code'] = code if i == 0 else ''
                                        timetable[day][start_slot+i]['name'] = name if i == 0 else ''
                                        timetable[day][start_slot+i]['faculty'] = faculty if i == 0 else ''
                                        timetable[day][start_slot+i]['classroom'] = classroom if i == 0 else ''
                                    scheduled = True
                                    break
                            if scheduled: break

                # --- Write timetable to worksheet ---
                header = ['Day'] + [f"{slot[0].strftime('%H:%M')}-{slot[1].strftime('%H:%M')}" for slot in TIME_SLOTS]
                ws.append(header)
                
                header_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
                header_font = Font(bold=True)
                header_alignment = Alignment(horizontal='center', vertical='center')
                
                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = header_alignment
                
                lec_fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")
                lab_fill = PatternFill(start_color="98FB98", end_color="98FB98", fill_type="solid")
                tut_fill = PatternFill(start_color="FFE4E1", end_color="FFE4E1", fill_type="solid")
                ss_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
                break_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
                border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                top=Side(style='thin'), bottom=Side(style='thin'))
                
                for day_idx, day in enumerate(DAYS):
                    row_num = day_idx + 2
                    ws.append([day])
                    
                    merge_ranges = []  # Track merge ranges for this row
                    
                    for slot_idx in range(len(TIME_SLOTS)):
                        cell_value = ''
                        cell_fill = None
                        
                        if is_break_time(TIME_SLOTS[slot_idx], semester):
                            cell_value = "BREAK"
                            cell_fill = break_fill
                        elif timetable[day_idx][slot_idx]['type']:
                            activity_type = timetable[day_idx][slot_idx]['type']
                            code = timetable[day_idx][slot_idx]['code']
                            classroom = timetable[day_idx][slot_idx]['classroom']
                            faculty = timetable[day_idx][slot_idx]['faculty']
                            
                            # Only create content for start of activity (i.e., when code is present)
                            if code:
                                # Get duration based on activity type
                                duration = {
                                    'LEC': LECTURE_DURATION,
                                    'LAB': LAB_DURATION,
                                    'TUT': TUTORIAL_DURATION,
                                    'SS': SELF_STUDY_DURATION
                                }.get(activity_type, 1)
                                
                                
                                # --- NEW BASKET RENDERING LOGIC ---
                                if 'basket_group_members' in timetable[day_idx][slot_idx] and timetable[day_idx][slot_idx]['basket_group_members']:
                                    basket_group_members = timetable[day_idx][slot_idx]['basket_group_members']
                                    basket_group = get_basket_group(code)
                                    
                                    # Use group color
                                    color_code = subject_color_map.get(code, "FFB6C1")
                                    cell_fill = PatternFill(start_color=color_code, end_color=color_code, fill_type="solid")
                                    
                                    # Build cell value for the merged block
                                    basket_header = f"{basket_group} Courses ({activity_type})\n"
                                    course_details = [
                                        f"{c['code']}: {c['faculty']} ({c['classroom']})"
                                        for c in basket_group_members
                                    ]
                                    cell_value = f"{basket_header}" + "\n".join(course_details)

                                # --- END NEW BASKET RENDERING LOGIC ---
                                
                                else:
                                    # Original logic for non-basket courses
                                    
                                    # Use subject-specific color if available
                                    if code in subject_color_map:
                                        color_code = subject_color_map[code]
                                        cell_fill = PatternFill(start_color=color_code, end_color=color_code, fill_type="solid")
                                    else:
                                        cell_fill = {
                                            'LAB': lab_fill,
                                            'TUT': tut_fill,
                                            'SS': ss_fill,
                                            'LEC': lec_fill
                                        }.get(activity_type, lec_fill)
                                    
                                    cell_value = f"{code} {activity_type}\n{classroom}\n{faculty}"
                                    
                                
                                # Create merge range
                                if duration > 1:
                                    start_col = get_column_letter(slot_idx + 2)
                                    end_col = get_column_letter(slot_idx + duration + 1)
                                    merge_range = f"{start_col}{row_num}:{end_col}{row_num}"
                                    merge_ranges.append((merge_range, cell_fill))
                            
                        cell = ws.cell(row=row_num, column=slot_idx+2, value=cell_value)
                        if cell_fill:
                            cell.fill = cell_fill
                        cell.border = border
                        cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
                        
                    # Apply merges after creating all cells in the row
                    for merge_range, fill in merge_ranges:
                        ws.merge_cells(merge_range)
                        merged_cell = ws[merge_range.split(':')[0]]
                        merged_cell.fill = fill
                        merged_cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')

                for col_idx in range(1, len(TIME_SLOTS)+2):
                    col_letter = get_column_letter(col_idx)
                    ws.column_dimensions[col_letter].width = 15
                
                for row in ws.iter_rows(min_row=2, max_row=len(DAYS)+1):
                    ws.row_dimensions[row[0].row].height = 40

                # ... (Rest of self-study, unscheduled component, and legend logic remains unchanged)
                
                current_row = len(DAYS) + 4 

                if self_study_courses:
                    ws.cell(row=current_row, column=1, value="Self-Study Only Courses")
                    ws.cell(row=current_row, column=1).font = Font(bold=True)
                    current_row += 1
                    
                    headers = ['Course Code', 'Course Name', 'Faculty']
                    for col, header in enumerate(headers, 1):
                        ws.cell(row=current_row, column=col, value=header)
                        ws.cell(row=current_row, column=col).font = Font(bold=True)
                    current_row += 1
                    
                    for course in self_study_courses:
                        if course['department'] == department and course['semester'] == semester:
                            ws.cell(row=current_row, column=1, value=course['code'])
                            ws.cell(row=current_row, column=2, value=course['name'])
                            ws.cell(row=current_row, column=3, value=course['faculty'])
                            current_row += 1
                            
                    current_row += 2 

                dept_unscheduled = [c for c in unscheduled_components  
                                    if c.department == department and  
                                    c.semester == semester and
                                    (c.section == section if num_sections > 1 else True)]

                if dept_unscheduled:
                    current_row += 2 
                    unsch_title = ws.cell(row=current_row, column=1, value="Unscheduled Components")
                    unsch_title.font = Font(bold=True, size=12, color="FF0000")
                    current_row += 2

                    headers = ['Course Code', 'Course Name', 'Faculty', 'Component', 'Sessions', 'Reason']
                    for col, header in enumerate(headers, 1):
                        cell = ws.cell(row=current_row, column=col, value=header)
                        cell.font = Font(bold=True)
                        cell.border = border
                        cell.fill = PatternFill(start_color="FFE0E0", end_color="FFE0E0", fill_type="solid")
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        ws.column_dimensions[get_column_letter(col)].width = 20
                    current_row += 1

                    for comp in dept_unscheduled:
                        cells = [
                            (comp.code, None),
                            (comp.name, None),
                            (comp.faculty, None),
                            (comp.component_type, None),
                            (comp.sessions, None),
                            (comp.reason or "Could not find suitable slot", None)
                        ]
                        
                        for col, (value, fill) in enumerate(cells, 1):
                            cell = ws.cell(row=current_row, column=col, value=value)
                            cell.border = border
                            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                        current_row += 1
                        
                    current_row += 2 

                legend_title = ws.cell(row=current_row, column=1, value="Legend")
                legend_title.font = Font(bold=True, size=12)
                current_row += 2

                ws.column_dimensions['A'].width = 20 
                ws.column_dimensions['B'].width = 40 
                ws.column_dimensions['C'].width = 30 
                ws.column_dimensions['D'].width = 15 

                legend_headers = ['Subject Code', 'Subject Name', 'Faculty', 'Color']
                for col, header in enumerate(legend_headers, 1):
                    cell = ws.cell(row=current_row, column=col, value=header)
                    cell.font = Font(bold=True)
                    cell.border = border
                    cell.fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                current_row += 1

                all_course_codes = sorted(list(subject_color_map.keys()))
                for code in all_course_codes:
                    color = subject_color_map[code]
                    if code in course_faculty_map:
                        ws.row_dimensions[current_row].height = 25
                        
                        cells = [
                            (code, None),
                            (course_faculty_map[code]['name'], None),
                            (select_faculty(course_faculty_map[code]['faculty']), None),
                            ('', PatternFill(start_color=color, end_color=color, fill_type="solid"))
                        ]
                        
                        for col, (value, fill) in enumerate(cells, 1):
                            cell = ws.cell(row=current_row, column=col, value=value)
                            cell.border = border
                            if fill:
                                cell.fill = fill
                            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                        
                        current_row += 1

    # Save separate workbook for each department
    for department, wb in workbooks.items():
        filename = f"timetable_{department}.xlsx"
        try:
            wb.save(filename)
            print(f"Timetable for {department} saved as {filename}")
        except Exception as e:
            print(f"Error saving file {filename}: {e}")

    return [f"timetable_{dept}.xlsx" for dept in workbooks.keys()]

if __name__ == "__main__":
    generate_all_timetables()