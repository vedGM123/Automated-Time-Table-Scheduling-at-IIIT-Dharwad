import pandas as pd
import os
import random
import math
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment, Font

# ==============================
# CONFIGURATION
# ==============================
COURSE_FILE = "data/course_data.csv"
FACULTY_FILE = "data/faculty.csv"
ROOM_FILE = "data/rooms.csv"
OUTPUT_FLAT = "output/generated_timetable.csv"
OUTPUT_EXCEL = "output/structured_timetable.xlsx"

days = ["Mon", "Tue", "Wed", "Thu", "Fri"]

LECTURE_SLOTS = [
    ("09:00", "10:30"),
    ("10:40", "12:10"),
    ("14:00", "15:30"),
    ("15:40", "17:10"),
    ("17:20", "18:50")
]

TUTORIAL_SLOTS = [
    ("09:00", "10:00"),
    ("10:10", "11:10"),
    ("11:20", "12:20"),
    ("14:00", "15:00"),
    ("15:10", "16:10"),
    ("16:20", "17:20"),
    ("17:30", "18:30")
]

LAB_SLOTS = [
    ("09:00", "11:00"),
    ("11:10", "13:10"),
    ("14:00", "16:00"),
    ("16:10", "18:10")
]

OPEN_ELECTIVE_DAY = "Wed"
OPEN_ELECTIVE_SLOT = ("14:30", "15:30")

# Lunch slot (always blocked)
LUNCH_SLOT = ("13:00", "14:00")

# ==============================
# LOAD DATA
# ==============================
courses = pd.read_csv(COURSE_FILE)
faculty = pd.read_csv(FACULTY_FILE)
rooms = pd.read_csv(ROOM_FILE)
room_list = rooms["Room"].tolist() if "Room" in rooms.columns else ["R1", "R2", "R3", "R4", "R5"]

# ==============================
# HELPER FUNCTIONS
# ==============================
def time_to_minutes(t):
    h, m = map(int, t.split(":"))
    return h * 60 + m

def has_conflict(existing_schedule, new_class):
    """Prevent overlapping courses for same room, instructor, semester, and lunch break."""
    new_start = time_to_minutes(new_class['Start-Time'])
    new_end = time_to_minutes(new_class['End-Time'])

    # Block lunch slot
    lunch_start = time_to_minutes(LUNCH_SLOT[0])
    lunch_end = time_to_minutes(LUNCH_SLOT[1])
    if not (new_end <= lunch_start or new_start >= lunch_end):
        return True

    for cls in existing_schedule:
        if cls['Day'] == new_class['Day']:
            cls_start = time_to_minutes(cls['Start-Time'])
            cls_end = time_to_minutes(cls['End-Time'])
            overlap = not (new_end <= cls_start or new_start >= cls_end)
            if overlap:
                if cls['Room'] == new_class['Room']:
                    return True
                if cls['Instructor'] == new_class['Instructor']:
                    return True
                if 'Semester' in cls and 'Semester' in new_class:
                    if cls['Semester'] == new_class['Semester']:
                        return True
    return False

def get_session_counts(course):
    """Calculate number of lecture/tutorial/lab sessions."""
    L, T, P = map(int, course['L-T-P-S-C'].split('-')[:3])
    lectures = math.ceil(L / 1.5)  # each lecture 1.5 hr
    tutorials = T  # 1 hr each
    labs = math.ceil(P / 2)  # each lab 2 hrs
    return lectures, tutorials, labs

def schedule_course(course, slots, suffix, timetable, type_name="Lecture"):
    """Schedule a course session in available slots without repeating the same day."""
    used_days = course.get('ScheduledDays', set())
    scheduled_day = None

    for day in days:
        if day in used_days:
            continue
        for start_time, end_time in slots:
            for room in room_list:
                new_class = {
                    'Course Code': course['Course Code'],
                    'Course-Name': course['Course-Name'] + suffix,
                    'Instructor': course['Instructor'],
                    'Room': room,
                    'Day': day,
                    'Start-Time': start_time,
                    'End-Time': end_time,
                    'Semester': course['Semester'],
                    'Branch': course['Branch'],
                    'Type': type_name
                }
                if not has_conflict(timetable, new_class):
                    timetable.append(new_class)
                    scheduled_day = day
                    used_days.add(day)
                    course['ScheduledDays'] = used_days
                    break
            if scheduled_day:
                break
        if scheduled_day:
            break

    if not scheduled_day:
        print(f"⚠️ Could not schedule {course['Course Code']} {suffix}")
    return scheduled_day

def generate_color(code):
    """Deterministic hex color for course code."""
    random.seed(code)
    r = random.randint(120, 200)
    g = random.randint(120, 200)
    b = random.randint(120, 200)
    return f"{r:02X}{g:02X}{b:02X}"

# ==============================
# SCHEDULING
# ==============================
timetable = []

core_courses = courses[courses["Type"].str.lower() == "core"].sort_values(by="Semester")
elective_courses = courses[courses["Type"].str.lower() == "elective"]

for _, course in core_courses.iterrows():
    lectures, tutorials, labs = get_session_counts(course)
    for _ in range(lectures):
        schedule_course(course, LECTURE_SLOTS, " (Lecture)", timetable, type_name="Lecture")
    for _ in range(tutorials):
        schedule_course(course, TUTORIAL_SLOTS, " (Tutorial)", timetable, type_name="Tutorial")
    for _ in range(labs):
        schedule_course(course, LAB_SLOTS, " (Lab)", timetable, type_name="Lab")

# Open electives
for _, course in elective_courses.iterrows():
    timetable.append({
        'Course Code': course['Course Code'],
        'Course-Name': "Open Elective - " + course['Course-Name'],
        'Instructor': course['Instructor'],
        'Room': random.choice(room_list),
        'Day': OPEN_ELECTIVE_DAY,
        'Start-Time': OPEN_ELECTIVE_SLOT[0],
        'End-Time': OPEN_ELECTIVE_SLOT[1],
        'Semester': course['Semester'],
        'Branch': course['Branch'],
        'Type': "Elective"
    })

# ==============================
# EXPORT — CSV
# ==============================
df = pd.DataFrame(timetable)
df = df.sort_values(by=["Day", "Start-Time"])
os.makedirs(os.path.dirname(OUTPUT_FLAT), exist_ok=True)
df.to_csv(OUTPUT_FLAT, index=False)
print(f"✅ Flat timetable saved to: {OUTPUT_FLAT}")

# ==============================
# EXPORT — STRUCTURED EXCEL
# ==============================
df['Slot'] = df['Start-Time'] + " - " + df['End-Time']
df['Display'] = df['Course Code'] + "\n" + df['Course-Name']

structured_df = df.pivot_table(
    index='Day',
    columns='Slot',
    values='Display',
    aggfunc=lambda x: "\n---\n".join(x)
)

structured_df = structured_df.reindex(days)

# Add lunch break column if not present
lunch_slot_str = f"{LUNCH_SLOT[0]} - {LUNCH_SLOT[1]}"
structured_df[lunch_slot_str] = "Lunch Break"

# Reorder columns
structured_df = structured_df[sorted(structured_df.columns)]

os.makedirs(os.path.dirname(OUTPUT_EXCEL), exist_ok=True)
structured_df.to_excel(OUTPUT_EXCEL)
wb = load_workbook(OUTPUT_EXCEL)
ws = wb.active

# Apply colors, fonts, alignment
colors = {}
for i in range(2, ws.max_row + 1):
    for j in range(2, ws.max_column + 1):
        cell = ws.cell(row=i, column=j)
        if cell.value:
            cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
            if cell.value == "Lunch Break":
                cell.fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
                cell.font = Font(size=11, bold=True, color="000000")
            else:
                code = cell.value.split("\n")[0].strip()
                if code not in colors:
                    colors[code] = generate_color(code)
                fill_color = colors[code]
                cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                cell.font = Font(size=11, bold=True)

# Larger cells for better visibility
for i in range(2, ws.max_row + 1):
    ws.row_dimensions[i].height = 80
for j in range(2, ws.max_column + 1):
    ws.column_dimensions[ws.cell(row=1, column=j).column_letter].width = 30
for j in range(1, ws.max_column + 1):
    ws.cell(row=1, column=j).font = Font(bold=True, size=12)

wb.save(OUTPUT_EXCEL)
print(f"✅ Structured Excel saved to: {OUTPUT_EXCEL} ")
